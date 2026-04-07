"""
Bio Bitumen Master Consulting System — Executive Dashboard (UPGRADED)
======================================================================
India state map, pipeline value, trend metrics, market ticker, revenue forecast.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import datetime
from state_manager import get_config, init_state
from database import init_db, get_all_customers, get_all_packages, get_all_communications, get_customer_count_by_status
from document_index import build_index
from interpolation_engine import get_all_known_plants
from config import (COMPANY, CUSTOMER_STATUSES, STATE_SCORES, LOCATION_WEIGHTS,
                    STATES, NHAI_TENDERS, INDUSTRY_NETWORK, ENVIRONMENTAL_FACTORS)

st.set_page_config(page_title="Executive Dashboard", page_icon="🏭", layout="wide")
init_db()
init_state()
cfg = get_config()
# Fix metric truncation
try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass


st.sidebar.markdown("---")
if st.sidebar.button("Print This Page", key="print_page"):
    import streamlit.components.v1 as _stc; _stc.html('<script>window.print();</script>', height=0)

# ── Greeting ─────────────────────────────────────────────────────────
now = datetime.datetime.now()
hour = now.hour
if hour < 5: greeting = "Good Night"
elif hour < 12: greeting = "Good Morning"
elif hour < 17: greeting = "Good Afternoon"
elif hour < 21: greeting = "Good Evening"
else: greeting = "Good Night"
user_name = st.session_state.get("user_name", "Admin")

st.markdown(f"""
## {greeting}, {user_name}!
### {COMPANY['trade_name']} — Executive Dashboard
**ONE-POINT SOLUTION: Land Selection > Plant Setup > Financial Closure > Sales Network**
""")
st.caption(f"{now.strftime('%A, %d %B %Y')} | Last refreshed: {now.strftime('%I:%M %p IST')}")
try:
    from engines.visual_content_engine import show_reference_image
    show_reference_image(st, "road_bitumen", "Bio-Bitumen for India's Roads — NHAI Green Infrastructure")
except Exception:
    pass
st.markdown("---")

cfg = get_config()
customers = get_all_customers()
packages = get_all_packages()
comms = get_all_communications()
doc_df = build_index()
plants = get_all_known_plants()
status_counts = get_customer_count_by_status()

# ══════════════════════════════════════════════════════════════════════
# TOP METRICS ROW
# ══════════════════════════════════════════════════════════════════════
c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
c1.metric("Capacities", "7 Standard")
c2.metric("Documents", f"{len(doc_df):,}")
c3.metric("States", "18")
with c4:
    st.metric("Customers", len(customers))
    if len(customers) == 0:
        st.caption("[Add →](pages/11_👥_Customers.py)")
with c5:
    st.metric("Packages", len(packages))
    if len(packages) == 0:
        st.caption("[Create →](pages/61_📁_Document_Hub.py)")
c6.metric("Network", f"{INDUSTRY_NETWORK['total']:,}")

# Pipeline value
pipeline_value = sum(float(c.get("budget_cr", 0) or 0) for c in customers
                     if c.get("status") not in ("Closed Won", "Closed Lost"))
c7.metric("Pipeline", f"Rs {pipeline_value:.1f} Cr")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# CURRENT CONFIG + MARKET TICKER
# ══════════════════════════════════════════════════════════════════════
config_col, market_col = st.columns([1, 1])

with config_col:
    st.subheader("Selected Plant Configuration")
    cc1, cc2, cc3, cc4 = st.columns(4)
    cc1.metric("Capacity", f"{cfg['capacity_tpd']:.0f} TPD")
    cc2.metric("Investment", f"Rs {cfg['investment_cr']:.2f} Cr")
    cc3.metric("ROI", f"{cfg['roi_pct']:.1f}%")
    cc4.metric("Break-Even", f"{cfg['break_even_months']} months")

    cc5, cc6, cc7, cc8 = st.columns(4)
    cc5.metric("Monthly Profit", f"Rs {cfg['monthly_profit_lac']:.1f} Lac")
    cc6.metric("IRR", f"{cfg['irr_pct']:.1f}%")
    cc7.metric("DSCR Yr3", f"{cfg['dscr_yr3']:.2f}x")
    cc8.metric("Rev Yr5", f"Rs {cfg['revenue_yr5_lac']:.0f} Lac")

    # Plant Engineering Specs
    try:
        from engines.plant_engineering import compute_all
        pc = compute_all(cfg)
        with st.expander("Plant Specifications (for drawings)"):
            sp1, sp2, sp3, sp4 = st.columns(4)
            sp1.metric("Reactor", f"Ø{pc['reactor_dia_m']}m × {pc['reactor_ht_m']}m")
            sp2.metric("Dryer", f"Ø{pc['dryer_dia_m']}m × {pc['dryer_len_m']}m")
            sp3.metric("Oil Tank", f"Ø{pc['bio_oil_tank_dia_m']}m × {pc['bio_oil_tank_ht_m']}m")
            sp4.metric("Plot", f"{pc['plot_l_m']}m × {pc['plot_w_m']}m")
            sp5, sp6, sp7, sp8 = st.columns(4)
            sp5.metric("Feed/hr", f"{pc['feed_per_hour_kg']:.0f} kg")
            sp6.metric("Bio-Oil", f"{pc['bio_oil_tpd']:.1f} T/day")
            sp7.metric("Bio-Char", f"{pc['bio_char_tpd']:.1f} T/day")
            sp8.metric("Blend Output", f"{pc['blend_output_tpd']:.1f} T/day")
            from state_manager import calculate_boq
            boq = calculate_boq(cfg['capacity_tpd'])
            st.caption(f"BOQ: {len(boq)} items | Total motor load: {sum(m.get('motor_kw',0)*m.get('qty',1) for m in ([] if not hasattr(pc,'items') else []))}")
    except Exception:
        pass

with market_col:
    st.subheader("Live Market Data")
    try:
        from engines.market_data_api import get_market_summary
        market = get_market_summary()
        vg30 = market.get("vg30_estimate", {})
        fx = market.get("usd_inr", {})
        crude_data = market.get("crude_oil")
        if isinstance(crude_data, dict):
            crude_price = crude_data.get("latest_price", 0)
        elif isinstance(crude_data, list) and crude_data:
            crude_price = crude_data[-1].get("price_usd", 0)
        else:
            crude_price = 0
        # Fallback: re-fetch if zero
        if crude_price == 0:
            try:
                import yfinance as yf
                crude_price = round(yf.Ticker("BZ=F").history(period="5d")["Close"].iloc[-1], 2)
            except Exception:
                crude_price = 75.0  # Safe fallback

        mk1, mk2, mk3, mk4 = st.columns(4)
        mk1.metric("Crude Oil", f"${crude_price:.1f}/bbl")
        mk2.metric("USD/INR", f"Rs {fx.get('rate', 84):.2f}")
        mk3.metric("VG30 Est.", f"Rs {vg30.get('vg30_estimated', 38000):,.0f}")
        mk4.metric("Gold", f"${market.get('gold', {}).get('price_usd', 0):,.0f}")
    except Exception:
        st.info("Market data will load on refresh")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# INDIA STATE MAP + SALES FUNNEL
# ══════════════════════════════════════════════════════════════════════
map_col, funnel_col = st.columns(2)

with map_col:
    st.subheader("State Feasibility Scores (18 States)")
    state_data = []
    for state in STATES:
        scores = STATE_SCORES.get(state, {})
        total = sum(scores.get(k, 50) * LOCATION_WEIGHTS[k] for k in LOCATION_WEIGHTS)
        sc = {"State": state, "Score": round(total, 1)}
        for k, v in scores.items():
            sc[k.title()] = v
        state_data.append(sc)
    state_df = pd.DataFrame(state_data).sort_values("Score", ascending=True)

    fig_map = px.bar(state_df, y="State", x="Score", orientation="h",
                      color="Score", color_continuous_scale="Blues",
                      title="State Feasibility Ranking (Higher = Better)")
    fig_map.update_layout(template="plotly_white", height=500, yaxis={"categoryorder": "total ascending"})
    st.plotly_chart(fig_map, width="stretch")

with funnel_col:
    st.subheader("Customer Pipeline")
    if status_counts:
        funnel_data = [{"Status": s, "Count": status_counts.get(s, 0)} for s in CUSTOMER_STATUSES]
        fig_funnel = go.Figure(go.Funnel(
            y=[d["Status"] for d in funnel_data],
            x=[d["Count"] for d in funnel_data],
            textinfo="value+percent initial",
            marker=dict(color=["#003366", "#004c8c", "#006699", "#0088cc", "#00aadd", "#00cc44", "#cc4444"]),
        ))
        fig_funnel.update_layout(height=350, template="plotly_white", title="Sales Funnel")
        st.plotly_chart(fig_funnel, width="stretch")
    else:
        st.info("Add customers to see pipeline funnel")

    # NHAI Opportunity Summary
    st.subheader("NHAI Market Opportunity")
    open_tenders = [t for t in NHAI_TENDERS if t["status"] == "Open"]
    nt1, nt2, nt3 = st.columns(3)
    nt1.metric("Open Tenders", len(open_tenders))
    nt2.metric("Total Value", f"Rs {sum(t['budget_cr'] for t in open_tenders):,} Cr")
    nt3.metric("Bitumen Demand", f"{sum(t['bitumen_mt'] for t in open_tenders):,} MT")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# PLANT CAPACITIES AT A GLANCE
# ══════════════════════════════════════════════════════════════════════
st.subheader("Plant Capacities at a Glance")
plant_count = len(plants)
cols = st.columns(min(plant_count, 8))
for i, (key, p) in enumerate(plants.items()):
    with cols[i % len(cols)]:
        st.markdown(f"""
**{p.get('label', key)}**
- Invest: **Rs {p['inv_cr']} Cr**
- Rev Yr5: **Rs {p['rev_yr5_cr']} Cr**
- IRR: **{p['irr_pct']}%**
- DSCR: **{p['dscr_yr3']}x**
        """)

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# RECENT ACTIVITY + QUICK ACTIONS
# ══════════════════════════════════════════════════════════════════════
act_col, action_col = st.columns([2, 1])

with act_col:
    st.subheader("Recent Activity")
    tab_cust, tab_pkg, tab_comm = st.tabs(["Customers", "Packages", "Communications"])

    with tab_cust:
        if customers:
            for c in customers[:8]:
                status_color = {"New": "blue", "Contacted": "orange", "Proposal Sent": "violet",
                                "Negotiation": "red", "Closed Won": "green"}.get(c.get("status", ""), "gray")
                st.markdown(f"- :{status_color}[**{c.get('status', 'New')}**] {c['name']} ({c.get('company', '')})"
                            f" — {c.get('state', '')} | Rs {c.get('budget_cr', 0)} Cr")
        else:
            st.info("No customers yet. Use Customer Manager to add.")

    with tab_pkg:
        if packages:
            for p in packages[:8]:
                cust = next((c for c in customers if c["id"] == p["customer_id"]), None)
                name = cust["name"] if cust else "Unknown"
                st.markdown(f"- **{name}** — {p['capacity']} | {p['recipient_type']}")
        else:
            st.info("No packages yet")

    with tab_comm:
        if comms:
            for comm in comms[:8]:
                st.markdown(f"- **{comm.get('channel', 'Email')}** — {comm.get('subject', '')} ({comm.get('sent_at', '')[:10]})")
        else:
            st.info("No communications yet")

with action_col:
    st.subheader("Quick Actions")
    st.page_link("pages/36_ROI_Quick_Calc.py", label="ROI Calculator", icon="🎯")
    st.page_link("pages/33_Loan_EMI.py", label="Loan EMI Calculator", icon="🏦")
    st.page_link("pages/37_Capacity_Compare.py", label="Compare Capacities", icon="⚖️")
    st.page_link("pages/30_💰_Financial.py", label="Financial Model", icon="💰")
    st.page_link("pages/60_DPR_Generator.py", label="Generate DPR", icon="📄")
    st.page_link("pages/43_🛣️_NHAI_Tenders.py", label="NHAI Tenders", icon="🛣️")
    st.page_link("pages/81_🤖_AI_Advisor.py", label="AI Advisor", icon="🤖")

# ── Footer ───────────────────────────────────────────────────────────
st.markdown("---")
st.caption(f"{COMPANY['name']} | {COMPANY['owner']} | {COMPANY['phone']} | {COMPANY['hq']} | {COMPANY['experience']}")


# ── AI Skill: Executive Summary ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: Executive Summary"):
            if st.button("Generate", type="primary", key="ai_02📊Das"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: Executive Summary. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_02_📊_D", type="primary")
with _ex2:
    if st.button("Print", key="exp_prt_02📊D"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)
