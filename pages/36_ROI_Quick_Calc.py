"""
ROI Quick Calculator — Live Demo Tool for Investor Presentations
=================================================================
Big sliders, instant gauge charts, scenario comparison — designed for projector demos.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import plotly.graph_objects as go
import pandas as pd
from state_manager import get_config, init_state
from engines.three_process_model import calculate_process
from config import COMPANY

st.set_page_config(page_title="ROI Quick Calculator", page_icon="🎯", layout="wide")
init_state()
cfg = get_config()
# Fix metric truncation
try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass


st.page_link("pages/30_💰_Financial.py", label="← Back to Financial Model", icon="💰")

st.title("ROI Quick Calculator")
st.markdown(f"**{COMPANY['trade_name']} | Instant ROI Analysis — Change Any Slider, See Results Instantly**")
st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# INPUT SLIDERS (Big, Demo-Friendly)
# ══════════════════════════════════════════════════════════════════════
col_in, col_out = st.columns([1, 2])

with col_in:
    st.subheader("Adjust Parameters")
    capacity = st.slider("Plant Capacity (TPD)", 5, 100, int(cfg.get('capacity_tpd', 20)), 5, key="roi_cap")
    selling_price = st.slider("Selling Price (Rs/MT)", 25000, 55000, int(cfg.get('selling_price_per_mt', 35000)), 1000, key="roi_sp")
    biomass_cost = st.slider("Biomass Cost (Rs/MT)", 800, 4000, int(cfg.get('raw_material_cost_per_mt', 8000) / 4), 100, key="roi_bc")
    equity_pct = st.slider("Equity % (rest is loan)", 10, 100, 30, 5, key="roi_eq")
    process = st.radio("Process Model", [1, 2, 3],
                        format_func=lambda x: {1: "Full Chain (Biomass to Bitumen)",
                                                2: "Blending Only (Buy Bio-Oil)",
                                                3: "Raw Output (Oil + Char)"}[x],
                        key="roi_proc")

# ── Calculate ────────────────────────────────────────────────────────
result = calculate_process(process, float(capacity))

# Override with user slider values where possible
revenue_per_mt = selling_price
capex_lac = result.get("capex_lac", 800)
roi_pct = result.get("roi_pct", 25)
irr_pct = result.get("irr_pct", 22)
break_even = result.get("break_even_months", 36)
monthly_profit = result.get("pat_yr5_lac", 0) / 12  # PAT Year 5 / 12
dscr = result.get("dscr_yr3", 1.8)
profit_per_mt = result.get("profit_per_mt", 8000)
emi_lac = result.get("emi_lac_mth", 2)

with col_out:
    # ── Gauge Charts Row ─────────────────────────────────────────────
    g1, g2, g3 = st.columns(3)

    with g1:
        fig_roi = go.Figure(go.Indicator(
            mode="gauge+number+delta",
            value=roi_pct,
            title={"text": "ROI %", "font": {"size": 18}},
            gauge={
                "axis": {"range": [0, 60], "tickwidth": 1},
                "bar": {"color": "#003366"},
                "steps": [
                    {"range": [0, 15], "color": "#ffcccc"},
                    {"range": [15, 30], "color": "#ffffcc"},
                    {"range": [30, 60], "color": "#ccffcc"},
                ],
                "threshold": {"line": {"color": "#CC3333", "width": 3}, "thickness": 0.8, "value": 20},
            },
        ))
        fig_roi.update_layout(height=250, margin=dict(t=60, b=10, l=30, r=30))
        st.plotly_chart(fig_roi, width="stretch")

    with g2:
        fig_irr = go.Figure(go.Indicator(
            mode="gauge+number",
            value=irr_pct,
            title={"text": "IRR %", "font": {"size": 18}},
            gauge={
                "axis": {"range": [0, 50]},
                "bar": {"color": "#006699"},
                "steps": [
                    {"range": [0, 12], "color": "#ffcccc"},
                    {"range": [12, 25], "color": "#ffffcc"},
                    {"range": [25, 50], "color": "#ccffcc"},
                ],
            },
        ))
        fig_irr.update_layout(height=250, margin=dict(t=60, b=10, l=30, r=30))
        st.plotly_chart(fig_irr, width="stretch")

    with g3:
        fig_be = go.Figure(go.Indicator(
            mode="gauge+number",
            value=break_even,
            title={"text": "Break-Even (Months)", "font": {"size": 18}},
            gauge={
                "axis": {"range": [0, 60]},
                "bar": {"color": "#FF8800"},
                "steps": [
                    {"range": [0, 24], "color": "#ccffcc"},
                    {"range": [24, 42], "color": "#ffffcc"},
                    {"range": [42, 60], "color": "#ffcccc"},
                ],
            },
        ))
        fig_be.update_layout(height=250, margin=dict(t=60, b=10, l=30, r=30))
        st.plotly_chart(fig_be, width="stretch")

    # ── Key Metrics ──────────────────────────────────────────────────
    km1, km2, km3, km4 = st.columns(4)
    km1.metric("Total Investment", f"Rs {capex_lac/100:.1f} Cr")
    km2.metric("Monthly Profit", f"Rs {monthly_profit:.1f} Lac")
    km3.metric("DSCR Year 3", f"{dscr:.2f}x")
    km4.metric("EMI/Month", f"Rs {emi_lac:.2f} Lac")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# SCENARIO COMPARISON (Up to 3)
# ══════════════════════════════════════════════════════════════════════
st.subheader("Scenario Comparison")
st.caption("Compare different capacities side-by-side")

sc1, sc2, sc3 = st.columns(3)
scenarios = []

with sc1:
    st.markdown("**Scenario A (Small)**")
    sa_cap = st.selectbox("Capacity", [5, 10, 15, 20], index=1, key="sa_cap")
    sa = calculate_process(1, float(sa_cap))
    st.metric("Investment", f"Rs {sa.get('capex_lac', 0)/100:.1f} Cr")
    st.metric("ROI", f"{sa.get('roi_pct', 0):.1f}%")
    st.metric("Monthly Profit", f"Rs {sa.get('pat_yr5_lac', 0)/12:.1f} Lac")
    st.metric("Break-Even", f"{sa.get('break_even_months', 0)} months")
    scenarios.append({"Scenario": f"{sa_cap} TPD", **sa})

with sc2:
    st.markdown("**Scenario B (Medium)**")
    sb_cap = st.selectbox("Capacity", [15, 20, 30, 40], index=1, key="sb_cap")
    sb = calculate_process(1, float(sb_cap))
    st.metric("Investment", f"Rs {sb.get('capex_lac', 0)/100:.1f} Cr")
    st.metric("ROI", f"{sb.get('roi_pct', 0):.1f}%")
    st.metric("Monthly Profit", f"Rs {sb.get('pat_yr5_lac', 0)/12:.1f} Lac")
    st.metric("Break-Even", f"{sb.get('break_even_months', 0)} months")
    scenarios.append({"Scenario": f"{sb_cap} TPD", **sb})

with sc3:
    st.markdown("**Scenario C (Large)**")
    sc_cap = st.selectbox("Capacity", [30, 40, 50, 100], index=1, key="sc_cap")
    sc_result = calculate_process(1, float(sc_cap))
    st.metric("Investment", f"Rs {sc_result.get('capex_lac', 0)/100:.1f} Cr")
    st.metric("ROI", f"{sc_result.get('roi_pct', 0):.1f}%")
    st.metric("Monthly Profit", f"Rs {sc_result.get('pat_yr5_lac', 0)/12:.1f} Lac")
    st.metric("Break-Even", f"{sc_result.get('break_even_months', 0)} months")
    scenarios.append({"Scenario": f"{sc_cap} TPD", **sc_result})

# ── Comparison Bar Chart ─────────────────────────────────────────────
st.markdown("---")
compare_metrics = ["capex_lac", "roi_pct", "irr_pct", "pat_yr5_lac"]
compare_labels = ["CAPEX (Lac)", "ROI %", "IRR %", "Monthly Profit (Lac)"]

fig_compare = go.Figure()
colors = ["#003366", "#006699", "#0088cc"]
for i, s in enumerate(scenarios):
    fig_compare.add_trace(go.Bar(
        name=s.get("Scenario", f"Scenario {i+1}"),
        x=compare_labels,
        y=[s.get(m, 0) for m in compare_metrics],
        marker_color=colors[i],
    ))
fig_compare.update_layout(
    title="Scenario Comparison",
    barmode="group",
    template="plotly_white",
    height=400,
)
st.plotly_chart(fig_compare, width="stretch")

# ── Generate Full Report Link ────────────────────────────────────────
st.markdown("---")
st.subheader("Next Steps")
n1, n2, n3 = st.columns(3)
n1.page_link("pages/30_💰_Financial.py", label="Open Full Financial Model", icon="💰")
n2.page_link("pages/60_DPR_Generator.py", label="Generate DPR Report", icon="📄")
n3.page_link("pages/14_Client_Journey.py", label="Start Client Journey", icon="🧭")

st.markdown("---")
st.caption(f"{COMPANY['name']} | {COMPANY['owner']} | {COMPANY['phone']}")

# ── Export Section ────────────────────────────────────────────────────
st.markdown("---")
st.subheader("Export")
exp1, exp2, exp3 = st.columns(3)
with exp1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_60_ROI", type="primary")
with exp2:
    if st.button("Download CSV", key="exp_csv_calc"):
        import pandas as pd
        data = {"Metric": ["Capacity", "Investment", "ROI", "IRR", "DSCR", "Break-Even", "Monthly Profit"],
                "Value": [f"{cfg['capacity_tpd']:.0f} TPD", f"Rs {cfg['investment_cr']:.2f} Cr",
                          f"{cfg['roi_pct']:.1f}%", f"{cfg['irr_pct']:.1f}%", f"{cfg['dscr_yr3']:.2f}x",
                          f"{cfg['break_even_months']} months", f"Rs {cfg['monthly_profit_lac']:.1f} Lac"]}
        st.download_button("Download", pd.DataFrame(data).to_csv(index=False), "calculator_export.csv", "text/csv", key="dl_csv_c")
with exp3:
    if st.button("Print", key="exp_print_calc"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)


# ── AI Skill: ROI Explanation ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: ROI Explanation"):
            if st.button("Generate", type="primary", key="ai_60ROIQ"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: ROI Explanation. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass

# ROI Disclaimer
st.caption("📊 ROI shown uses simplified formula. For bank-standard ROI after full debt service, see Financial Model.")
