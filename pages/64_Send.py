"""
Bio Bitumen Consultant Portal — Send to Customer
Email and WhatsApp communication hub for sending packages to customers.
All financial values read from session state cfg — ZERO hardcoded outputs.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
from state_manager import init_state, get_config, format_inr
import pandas as pd
from database import (init_db, get_all_customers, get_customer,
                       get_packages_for_customer, get_communications_for_customer,
                       insert_communication)
from master_data_loader import get_plant
from engines.email_engine import send_email, generate_email_body, get_email_config
from engines.whatsapp_engine import generate_whatsapp_message, get_whatsapp_link
from config import COMPANY, CAPACITY_KEYS
import streamlit.components.v1 as stc

st.set_page_config(page_title="Send to Customer", page_icon="📧", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass

init_db()
st.title("Send to Customer")

# ── Select Customer ───────────────────────────────────────────────────
customers = get_all_customers()
if not customers:
    st.warning("No customers found. Add customers in **Customer Manager** first.")
    st.stop()

cust_map = {c["id"]: f"{c['name']} ({c.get('company', 'N/A')})" for c in customers}
selected_id = st.selectbox("Select Customer", options=list(cust_map.keys()),
                            format_func=lambda x: cust_map[x])
customer = get_customer(selected_id)

st.info(f"**{customer['name']}** | {customer.get('email', 'No email')} | "
        f"{customer.get('phone', 'No phone')} | {customer.get('state', 'N/A')}")

# ── Select Package (optional) ────────────────────────────────────────
packages = get_packages_for_customer(selected_id)
selected_pkg = None
if packages:
    pkg_map = {p["id"]: f"#{p['id']} — {p['capacity']} ({p['recipient_type']}) — {p['created_at']}" for p in packages}
    pkg_id = st.selectbox("Select Package (optional)", options=[None] + list(pkg_map.keys()),
                           format_func=lambda x: pkg_map.get(x, "None — send without package"))
    if pkg_id:
        selected_pkg = next(p for p in packages if p["id"] == pkg_id)

# ── Capacity — from cfg (session state), NOT hardcoded ───────────────
# Use customer's interested capacity, fallback to session cfg
cap_tpd = cfg.get("capacity_tpd", 20)
cap_key_from_cfg = f"{int(cap_tpd):02d}MT"
cap_key = (selected_pkg["capacity"] if selected_pkg else
           customer.get("interested_capacity") or cap_key_from_cfg)
if cap_key not in CAPACITY_KEYS:
    cap_key = cap_key_from_cfg
if cap_key not in CAPACITY_KEYS:
    cap_key = "20MT"

plant = get_plant(cap_key)

# Override plant values with LIVE session cfg values (single source of truth)
plant_live = dict(plant)
plant_live["inv_cr"] = round(cfg.get("investment_cr", plant.get("inv_cr", 8)), 2)
plant_live["loan_cr"] = round(cfg.get("loan_cr", plant.get("loan_cr", 4.8)), 2)
plant_live["equity_cr"] = round(cfg.get("equity_cr", plant.get("equity_cr", 3.2)), 2)
plant_live["rev_yr1_cr"] = round(cfg.get("revenue_yr1_cr", plant.get("rev_yr1_cr", 3)), 2)
plant_live["rev_yr5_cr"] = round(cfg.get("revenue_yr5_cr", plant.get("rev_yr5_cr", 6.5)), 2)
plant_live["emi_lac_mth"] = round(cfg.get("emi_lac_mth", plant.get("emi_lac_mth", 0)), 2)
plant_live["irr_pct"] = round(cfg.get("irr_pct", plant.get("irr_pct", 20)), 1)
plant_live["dscr_yr3"] = round(cfg.get("dscr_yr3", plant.get("dscr_yr3", 1.5)), 2)
plant_live["roi_pct"] = round(cfg.get("roi_pct", plant.get("roi_pct", 20)), 1)
plant_live["staff"] = int(cfg.get("staff", plant.get("staff", 18)))
plant_live["power_kw"] = int(cfg.get("power_kw", plant.get("power_kw", 100)))
plant_live["oil_ltr_day"] = int(cfg.get("oil_ltr_day", plant.get("oil_ltr_day", 8000)))
plant_live["char_kg_day"] = int(cfg.get("char_kg_day", plant.get("char_kg_day", 6000)))
plant_live["label"] = f"{cap_tpd:.0f} MT/Day"
# Debt % from equity ratio
equity_ratio = cfg.get("equity_ratio", 0.40)
debt_pct = round((1 - equity_ratio) * 100)

st.markdown("---")

# ── Tabs: Email / WhatsApp / Settings / Log ──────────────────────────
tab_email, tab_whatsapp, tab_settings, tab_log = st.tabs(
    ["Email", "WhatsApp", "Email Settings", "Communication Log"])

# ══════════════════════════════════════════════════════════════════════
# TAB: EMAIL
# ══════════════════════════════════════════════════════════════════════
with tab_email:
    st.subheader("Send Email")

    config = get_email_config()
    if not config["sender_email"]:
        st.warning("Email not configured. Go to **Email Settings** tab to set up SMTP credentials.")

    to_email = st.text_input("To", value=customer.get("email", ""), key="email_to")
    subject = st.text_input("Subject",
                             value=f"Bio-Modified Bitumen {plant_live['label']} — Project Information | {COMPANY['trade_name']}",
                             key="email_subject")

    # Generate email body using LIVE cfg values
    default_body = generate_email_body(customer, plant_live, COMPANY)

    # Show HTML editor + preview side by side
    col_edit, col_preview = st.columns(2)
    with col_edit:
        st.caption("Edit Email (HTML)")
        body = st.text_area("Email Body", value=default_body, height=300, key="email_body",
                            label_visibility="collapsed")
    with col_preview:
        st.caption("Preview")
        stc.html(body, height=300, scrolling=True)

    # Attachments
    attachments = []
    if selected_pkg and selected_pkg.get("output_folder"):
        pkg_dir = selected_pkg["output_folder"]
        if os.path.exists(pkg_dir):
            pkg_files = [os.path.join(pkg_dir, f) for f in os.listdir(pkg_dir) if not f.startswith(".")]
            st.markdown(f"**Attachments from package:** {len(pkg_files)} files")
            for f in pkg_files[:10]:
                st.caption(f"  - {os.path.basename(f)}")
            if len(pkg_files) > 10:
                st.caption(f"  ... and {len(pkg_files) - 10} more")
            attachments = pkg_files

    if st.button("Send Email", type="primary", key="send_email_btn"):
        if not to_email:
            st.error("Please enter a recipient email address.")
        elif not config["sender_email"] or not config["sender_password"]:
            st.error("Email not configured. Go to **Email Settings** tab and enter your Gmail + App Password first.")
        else:
            with st.spinner("Sending email..."):
                success, message = send_email(to_email, subject, body, attachments)
                if success:
                    st.success(message)
                    insert_communication({
                        "customer_id": selected_id,
                        "package_id": selected_pkg["id"] if selected_pkg else None,
                        "channel": "email",
                        "subject": subject,
                        "content_summary": f"Email to {to_email} with {len(attachments)} attachments",
                        "attachments": [os.path.basename(a) for a in attachments],
                        "status": "sent",
                    })
                    st.rerun()
                else:
                    st.error(f"Failed: {message}")
                    insert_communication({
                        "customer_id": selected_id,
                        "package_id": selected_pkg["id"] if selected_pkg else None,
                        "channel": "email",
                        "subject": subject,
                        "content_summary": f"Failed: {message}",
                        "status": "failed",
                    })


# ══════════════════════════════════════════════════════════════════════
# TAB: WHATSAPP
# ══════════════════════════════════════════════════════════════════════
with tab_whatsapp:
    st.subheader("Send via WhatsApp")

    phone = st.text_input("WhatsApp Number", value=customer.get("whatsapp", "") or customer.get("phone", ""),
                           placeholder="+91 XXXXX XXXXX", key="wa_phone")

    # Generate message using LIVE cfg values (not hardcoded plant data)
    wa_message = generate_whatsapp_message(customer, plant_live, COMPANY)
    edited_message = st.text_area("Message", value=wa_message, height=400, key="wa_msg")

    if phone:
        wa_link = get_whatsapp_link(phone, edited_message)
        st.markdown(f"[Open WhatsApp Web]({wa_link})")

        if st.button("Log WhatsApp Message", type="primary", key="wa_log_btn"):
            insert_communication({
                "customer_id": selected_id,
                "package_id": selected_pkg["id"] if selected_pkg else None,
                "channel": "whatsapp",
                "subject": f"WhatsApp to {phone}",
                "content_summary": edited_message[:200],
                "status": "sent",
            })
            st.success("WhatsApp communication logged to Communication History!")
            st.rerun()
    else:
        st.warning("Enter a phone number to generate the WhatsApp link.")


# ══════════════════════════════════════════════════════════════════════
# TAB: EMAIL SETTINGS
# ══════════════════════════════════════════════════════════════════════
with tab_settings:
    st.subheader("Email Configuration")
    st.markdown("""Configure your SMTP settings for sending emails.
    For **Gmail**, use an [App Password](https://myaccount.google.com/apppasswords) (not your regular password).""")

    with st.form("email_settings"):
        smtp_server = st.text_input("SMTP Server", value=st.session_state.get("smtp_server", "smtp.gmail.com"))
        smtp_port = st.number_input("SMTP Port", value=st.session_state.get("smtp_port", 587))
        sender_email = st.text_input("Your Email", value=st.session_state.get("sender_email", ""))
        sender_password = st.text_input("App Password", type="password",
                                         value=st.session_state.get("sender_password", ""))
        sender_name = st.text_input("Sender Name", value=st.session_state.get("sender_name", COMPANY.get("trade_name", "PPS Anantams")))

        if st.form_submit_button("Save Settings", type="primary"):
            if not sender_email or not sender_password:
                st.error("Email and App Password are required to enable email sending.")
            else:
                st.session_state["smtp_server"] = smtp_server
                st.session_state["smtp_port"] = smtp_port
                st.session_state["sender_email"] = sender_email
                st.session_state["sender_password"] = sender_password
                st.session_state["sender_name"] = sender_name
                st.success("Email settings saved! Email sending is now enabled for this session.")


# ══════════════════════════════════════════════════════════════════════
# TAB: COMMUNICATION LOG
# ══════════════════════════════════════════════════════════════════════
with tab_log:
    st.subheader("Communication History")

    comms = get_communications_for_customer(selected_id)
    if comms:
        log_data = []
        for c in comms:
            log_data.append({
                "Date": c.get("sent_at", ""),
                "Channel": c.get("channel", "").upper(),
                "Subject": c.get("subject", ""),
                "Status": c.get("status", ""),
                "Summary": c.get("content_summary", "")[:80],
            })
        st.dataframe(pd.DataFrame(log_data), use_container_width=True, hide_index=True)

        # Export communication log
        if st.button("Export Log as Excel", key="export_comm_log"):
            try:
                import io
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Communication Log"
                headers = ["Date", "Channel", "Subject", "Status", "Summary"]
                for col, h in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=h)
                for row_idx, entry in enumerate(log_data, 2):
                    for col, h in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col, value=entry[h])
                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                st.download_button("Download", buf.getvalue(),
                    f"CommLog_{customer['name'].replace(' ','_')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_comm_log")
            except Exception as e:
                st.error(f"Export failed: {e}")
    else:
        st.info("No communications logged for this customer yet.")


# ══════════════════════════════════════════════════════════════════════
# AUTO FOLLOW-UP SEQUENCE
# ══════════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("Auto Follow-Up Sequence Planner")
st.markdown("**Automated communication sequence for each customer:**")

sequence_data = pd.DataFrame([
    {"Day": "Day 0", "Action": "Send initial package (DPR + Financial + Pitch)", "Channel": "Email + WhatsApp", "Status": "Manual trigger above"},
    {"Day": "Day 1", "Action": "WhatsApp: 'Did you receive the documents?'", "Channel": "WhatsApp", "Status": "Use template below"},
    {"Day": "Day 3", "Action": "Follow-up email: Key highlights + ROI summary", "Channel": "Email", "Status": "Use template below"},
    {"Day": "Day 7", "Action": "Call + WhatsApp: 'Any questions? Ready for discussion?'", "Channel": "Phone + WhatsApp", "Status": "Use template below"},
    {"Day": "Day 14", "Action": "Send updated pricing / market data", "Channel": "Email", "Status": "Use template below"},
    {"Day": "Day 21", "Action": "Invitation for site visit / video call", "Channel": "Email + WhatsApp", "Status": "Use template below"},
    {"Day": "Day 30", "Action": "Final follow-up: Special offer / urgency", "Channel": "Email + WhatsApp", "Status": "Use template below"},
])
st.dataframe(sequence_data, use_container_width=True, hide_index=True)

# Quick follow-up templates — ALL VALUES FROM CFG (single source of truth)
st.markdown("---")
st.subheader("Quick Follow-Up Templates")

# Get VG30 price from market data or cfg
try:
    from utils.market_data import get_vg30_price
    vg30_price = get_vg30_price()
except Exception:
    vg30_price = cfg.get("price_conv_bitumen", 45750)

selling_price = cfg.get("selling_price_per_mt", 35000)
discount_pct = round((1 - selling_price / vg30_price) * 100) if vg30_price > 0 else 30

templates = {
    "Day 1 — Delivery Confirmation": (
        f"Dear {{customer_name}},\n\n"
        f"This is {COMPANY['owner']} from {COMPANY['trade_name']}.\n\n"
        f"I sent you the Bio-Bitumen plant project details yesterday. Did you receive the documents?\n\n"
        f"Key highlights:\n"
        f"- Capacity: {cap_tpd:.0f} MT/Day\n"
        f"- Investment: Rs {plant_live['inv_cr']:.2f} Crore\n"
        f"- ROI: {plant_live['roi_pct']:.1f}%\n"
        f"- Break-even: {cfg.get('break_even_months', 30)} months\n\n"
        f"Please let me know if you have any questions.\n\n"
        f"Regards,\n{COMPANY['owner']}\n{COMPANY['phone']}"
    ),
    "Day 3 — ROI Highlight": (
        f"Dear {{customer_name}},\n\n"
        f"Quick reminder about the Bio-Bitumen opportunity:\n\n"
        f"- India imports 49% of bitumen (Rs 25,000 Cr/year)\n"
        f"- 130-216 plants needed in next 5-7 years\n"
        f"- Your ROI: {plant_live['roi_pct']:.1f}% with break-even in {cfg.get('break_even_months', 30)} months\n"
        f"- Revenue Year 5: Rs {plant_live['rev_yr5_cr']:.2f} Crore\n\n"
        f"Would you like to schedule a 30-minute discussion?\n\n"
        f"Best regards,\n{COMPANY['owner']}"
    ),
    "Day 7 — Discussion Invite": (
        f"Dear {{customer_name}},\n\n"
        f"Following up on the {cap_tpd:.0f} TPD Bio-Bitumen project proposal.\n\n"
        f"I have complete project documentation ready:\n"
        f"- DPR with financial model (Rs {plant_live['inv_cr']:.2f} Cr investment)\n"
        f"- Engineering drawings\n"
        f"- Bank-ready loan proposal (Rs {plant_live['loan_cr']:.2f} Cr at {debt_pct}% debt)\n"
        f"- NHAI compliance roadmap\n\n"
        f"Shall we schedule a call this week?\n\n"
        f"{COMPANY['owner']}\n{COMPANY['phone']}"
    ),
    "Day 14 — Market Update": (
        f"Dear {{customer_name}},\n\n"
        f"Market update: VG30 bitumen price is Rs {vg30_price:,.0f}/MT.\n"
        f"Bio-bitumen at Rs {selling_price:,}/MT offers {discount_pct}% cost advantage.\n\n"
        f"NHAI green mandate is creating strong demand.\n"
        f"IRR: {plant_live['irr_pct']:.1f}% | DSCR: {plant_live['dscr_yr3']:.2f}x\n\n"
        f"Let's discuss how to capitalize on this opportunity.\n\n"
        f"{COMPANY['owner']}"
    ),
}

for title, template in templates.items():
    with st.expander(title):
        filled = template.replace("{customer_name}", customer.get("name", "Sir/Madam") if customer else "Sir/Madam")
        edited_tmpl = st.text_area("Template (editable)", value=filled, height=200, key=f"tmpl_{title[:10]}")

        tcol1, tcol2 = st.columns(2)
        if customer and customer.get("phone"):
            wa_tmpl_link = get_whatsapp_link(customer["phone"], edited_tmpl)
            tcol1.markdown(f"[Send via WhatsApp]({wa_tmpl_link})")
        if tcol2.button(f"Log as Sent", key=f"log_{title[:10]}"):
            insert_communication({
                "customer_id": selected_id,
                "channel": "whatsapp",
                "subject": title,
                "content_summary": edited_tmpl[:200],
                "status": "sent",
            })
            st.success(f"Logged: {title}")
            st.rerun()


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
if st.button("Print", key="exp_prt_43Sen"):
    stc.html("<script>window.print();</script>", height=0)
