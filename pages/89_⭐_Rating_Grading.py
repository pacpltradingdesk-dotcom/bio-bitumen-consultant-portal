"""
89 · Rating & Grading — 4-System Dashboard
============================================
Tab 1 — Project Viability Rating  (A+ … D)
Tab 2 — DPR Completeness Grade    (% complete)
Tab 3 — IS 73:2013 Product Grade  (VG-10/30/40/80)
Tab 4 — Vendor / Supplier Rating  (A … D per vendor)
Dark gold theme: #15130F / #E8B547
"""

import sys
import json
from pathlib import Path
from datetime import datetime

import pandas as pd
import streamlit as st

sys.path.insert(0, str(Path(__file__).parent.parent))
from state_manager import init_state, get_config
from engines.rating_engine import (
    grade_project, grade_dpr, grade_product,
    grade_vendor, save_vendor_ratings, load_vendor_ratings,
    load_ratings, VENDOR_CRITERIA, IS73_GRADES,
)
from config import COMPANY

st.set_page_config(page_title="Rating & Grading · Bio Bitumen", page_icon="⭐", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass

# ── CSS ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
[data-testid="stAppViewContainer"]{background:#15130F;}
[data-testid="stSidebar"]{background:#1A1710;}
h1,h2,h3,h4,label,.stMarkdown p,.stMarkdown li{color:#F0E6D3;}
[data-testid="stMetricValue"]{color:#E8B547 !important;}
[data-testid="stMetricLabel"]{color:#B09060 !important;}
.stButton>button{background:#E8B547;color:#15130F;font-weight:700;
  border:none;border-radius:6px;padding:6px 18px;}
.stButton>button:hover{background:#F5C842;}
.stTabs [data-baseweb="tab-list"]{background:#1A1710;border-radius:8px;}
.stTabs [data-baseweb="tab"]{color:#9A8A6A !important;}
.stTabs [aria-selected="true"]{color:#E8B547 !important;}
[data-testid="stDataFrame"]{background:#1E1B14;}

.grade-hero{text-align:center;padding:24px 16px;border-radius:14px;
  background:#1E1B14;border:2px solid #3A3520;}
.grade-letter{font-size:5rem;font-weight:900;line-height:1;}
.grade-label{font-size:1rem;color:#9A8A6A;margin-top:4px;}
.grade-score{font-size:1.5rem;font-weight:700;color:#E8B547;}

.dim-row{display:flex;align-items:center;gap:10px;padding:6px 10px;
  border-radius:6px;background:#1E1B14;margin:3px 0;}
.dim-name{flex:1;font-size:13px;color:#C8B88A;}
.dim-bar-bg{flex:2;background:#2E2B1F;border-radius:4px;height:8px;}
.dim-bar-fill{height:8px;border-radius:4px;}
.dim-pts{min-width:52px;text-align:right;font-size:13px;font-weight:600;color:#E8B547;}
.dim-note{flex:2;font-size:12px;color:#9A8A6A;}

.sec-row{display:flex;align-items:center;gap:10px;padding:5px 10px;
  border-radius:6px;margin:2px 0;}
.sec-complete{background:#1B2A1E;}
.sec-partial{background:#2A2410;}
.sec-missing{background:#2A1414;}
.sec-name{flex:2;font-size:13px;color:#C8B88A;}
.sec-status{min-width:90px;font-size:12px;font-weight:600;}

.vg-card{background:#1E1B14;border:1px solid #3A3520;border-radius:10px;
  padding:16px;margin:4px;}
.vg-title{font-size:1.3rem;font-weight:800;margin-bottom:4px;}
.vg-use{font-size:12px;color:#9A8A6A;margin-bottom:10px;}
.param-row{display:flex;justify-content:space-between;padding:3px 0;
  border-bottom:1px solid #2E2B1F;font-size:12px;}
.param-label{color:#9A8A6A;}
.param-value{color:#C8B88A;font-weight:600;}

.vendor-card{background:#1E1B14;border-left:4px solid #E8B547;
  border-radius:0 10px 10px 0;padding:14px 16px;margin:6px 0;}
</style>
""", unsafe_allow_html=True)

# ── header ────────────────────────────────────────────────────────────────
st.markdown("""
<h1 style="color:#E8B547;margin-bottom:2px;">⭐ Rating & Grading</h1>
<p style="color:#9A8A6A;margin-top:0;">
  Project Viability · DPR Completeness · IS 73:2013 Product Grade · Vendor Rating
</p>
""", unsafe_allow_html=True)
st.markdown("---")

# ── helper: render grade hero ─────────────────────────────────────────────
def render_grade_hero(result: dict):
    g = result["grade"]
    c = result["color"]
    s = result["score"]
    lbl = result.get("label", "")
    st.markdown(
        f'<div class="grade-hero">'
        f'<div class="grade-letter" style="color:{c};">{g}</div>'
        f'<div class="grade-score">{s} / {result.get("max_score",100)}</div>'
        f'<div class="grade-label">{lbl}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )


def pct_bar(pct: float, color: str = "#E8B547", height: int = 8) -> str:
    return (
        f'<div style="background:#2E2B1F;border-radius:4px;height:{height}px;width:100%;">'
        f'<div style="width:{min(pct,100):.1f}%;height:{height}px;border-radius:4px;'
        f'background:{color};transition:width .4s;"></div></div>'
    )


# ══════════════════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════════════════
tab1, tab2, tab3, tab4 = st.tabs([
    "🏆 Project Viability",
    "📋 DPR Completeness",
    "🧪 Product Grade (IS 73)",
    "🏭 Vendor Rating",
])

# ─────────────────────────────────────────────────────────────────────────
# TAB 1 — PROJECT VIABILITY RATING
# ─────────────────────────────────────────────────────────────────────────
with tab1:
    st.subheader("Project Viability Rating")
    st.caption("Scores your bio-bitumen project on 7 financial & operational dimensions.")

    cached = load_ratings().get("project")

    col_run, col_info = st.columns([2, 3])
    with col_run:
        if st.button("Calculate Project Rating", type="primary", key="run_proj"):
            cached = grade_project(cfg)
            st.success("Rating calculated!")

    if cached:
        hero_col, detail_col = st.columns([1, 2])

        with hero_col:
            render_grade_hero(cached)
            st.markdown(
                f'<p style="font-size:13px;color:#9A8A6A;margin-top:12px;'
                f'text-align:center;">{cached["summary"]}</p>',
                unsafe_allow_html=True,
            )

        with detail_col:
            st.markdown("**Dimension Breakdown**")
            for dim in cached["breakdown"]:
                mx  = dim["max"]
                pts = dim["score"]
                pct = pts / mx * 100
                col = "#51CF66" if pct >= 75 else ("#FFD43B" if pct >= 50 else "#FF6B6B")
                st.markdown(
                    f'<div class="dim-row">'
                    f'<span class="dim-name">{dim["dim"]}</span>'
                    f'<div class="dim-bar-bg"><div class="dim-bar-fill" '
                    f'style="width:{pct:.0f}%;background:{col};"></div></div>'
                    f'<span class="dim-pts">{pts}/{mx}</span>'
                    f'<span class="dim-note">{dim["note"]}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        # grade scale reference
        st.markdown("---")
        st.caption("Grade Scale: A+ ≥ 90 | A ≥ 80 | B+ ≥ 70 | B ≥ 60 | C ≥ 50 | D < 50")

    # ── Live market context for project rating ────────────────────────
    st.markdown("---")
    st.markdown("**📡 Live Market Context**")
    try:
        from engines.free_apis import get_exchange_rates, get_india_gdp, get_india_holidays
        _fx89 = get_exchange_rates()
        _gdp89 = get_india_gdp(2)
        _hols89 = get_india_holidays(datetime.now().year)

        _mc1, _mc2, _mc3 = st.columns(3)
        with _mc1:
            if "error" not in _fx89:
                st.metric("USD / INR (live)", f"₹ {_fx89.get('usd_inr',84):.2f}")
                st.metric("EUR / INR (live)", f"₹ {_fx89.get('eur_inr',90):.2f}")
        with _mc2:
            if _gdp89 and len(_gdp89) >= 2:
                _g1 = _gdp89[-2]["gdp_usd_billion"]
                _g2 = _gdp89[-1]["gdp_usd_billion"]
                _gdp_growth = (_g2 - _g1) / _g1 * 100
                st.metric("India GDP Growth", f"{_gdp_growth:.1f}%",
                          help=f"Year {_gdp89[-1]['year']}: USD {_g2:,.0f}B")
                st.metric("GDP (latest)", f"USD {_g2:,.0f}B")
        with _mc3:
            if _hols89:
                _next_hol = next(
                    (h for h in _hols89
                     if h["date"] >= datetime.now().strftime("%Y-%m-%d")), None
                )
                if _next_hol:
                    st.info(f"🗓️ Next holiday: **{_next_hol['name_en']}** "
                            f"({_next_hol['date']})")
                _yr_hols = len([h for h in _hols89
                                if h["date"].startswith(str(datetime.now().year))])
                st.metric("Holidays This Year", _yr_hols,
                          help="Affects working days in financial projections")
    except Exception:
        pass

# ─────────────────────────────────────────────────────────────────────────
# TAB 2 — DPR COMPLETENESS
# ─────────────────────────────────────────────────────────────────────────
with tab2:
    st.subheader("DPR Completeness Grade")
    st.caption("Checks all 15 DPR sections — page presence + data entry status.")

    cached_dpr = load_ratings().get("dpr")

    if st.button("Check DPR Completeness", type="primary", key="run_dpr"):
        cached_dpr = grade_dpr(cfg, Path(__file__).parent)
        st.success("DPR completeness checked!")

    if cached_dpr:
        h_col, d_col = st.columns([1, 2])

        with h_col:
            render_grade_hero(cached_dpr)
            st.markdown(
                f'<p style="font-size:13px;color:#9A8A6A;margin-top:12px;'
                f'text-align:center;">{cached_dpr["summary"]}</p>',
                unsafe_allow_html=True,
            )
            miss = cached_dpr.get("missing", [])
            part = cached_dpr.get("partial", [])
            if miss:
                st.markdown(f'<p style="color:#FF6B6B;font-size:12px;">Missing: {", ".join(miss)}</p>',
                            unsafe_allow_html=True)
            if part:
                st.markdown(f'<p style="color:#FFD43B;font-size:12px;">Partial: {", ".join(part)}</p>',
                            unsafe_allow_html=True)

        with d_col:
            st.markdown("**Section Status**")
            for sec in cached_dpr["sections"]:
                status = sec["status"]
                if status == "complete":
                    bg, icon, sc = "sec-complete", "✅", "#51CF66"
                elif status == "page_only":
                    bg, icon, sc = "sec-partial", "⚠️", "#FFD43B"
                else:
                    bg, icon, sc = "sec-missing", "❌", "#FF6B6B"

                pct = sec["pct"]
                st.markdown(
                    f'<div class="sec-row {bg}">'
                    f'<span style="font-size:15px;">{icon}</span>'
                    f'<span class="sec-name">{sec["section"]}</span>'
                    f'<span class="sec-status" style="color:{sc};">'
                    f'{sec["earned"]}/{sec["weight"]} pts</span>'
                    f'<div style="flex:1;">{pct_bar(pct, sc, 6)}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
    else:
        st.info("Click **Check DPR Completeness** to analyse.")

# ─────────────────────────────────────────────────────────────────────────
# TAB 3 — IS 73:2013 PRODUCT GRADE
# ─────────────────────────────────────────────────────────────────────────
with tab3:
    st.subheader("IS 73:2013 Product Grade — Bio-Bitumen VG Classification")
    st.caption("Enter lab test results to determine which viscosity grade your product meets.")

    t3_left, t3_right = st.columns([1, 1])

    with t3_left:
        st.markdown("**Enter Lab Test Values**")
        abs_vis = st.number_input("Absolute Viscosity at 60°C (Poise)",
                                   min_value=0.0, value=0.0, step=100.0, key="av_in")
        kin_vis = st.number_input("Kinematic Viscosity at 135°C (cSt)",
                                   min_value=0.0, value=0.0, step=10.0, key="kv_in")
        soft_pt = st.number_input("Softening Point (°C)",
                                   min_value=0.0, value=0.0, step=1.0, key="sp_in")
        flash_pt = st.number_input("Flash Point (°C)",
                                    min_value=0.0, value=0.0, step=5.0, key="fp_in")
        ductility = st.number_input("Ductility at 25°C (cm)",
                                     min_value=0.0, value=0.0, step=5.0, key="duc_in")
        penetration = st.number_input("Penetration at 25°C (0.1mm)",
                                       min_value=0.0, value=0.0, step=5.0, key="pen_in")

        test_btn = st.button("Determine Grade", type="primary", key="run_product")

    with t3_right:
        # reference table always visible
        st.markdown("**IS 73:2013 Grade Reference**")
        ref_rows = []
        for g, spec in IS73_GRADES.items():
            lo, hi = spec["abs_viscosity_range"]
            plo, phi = spec["penetration_range"]
            ref_rows.append({
                "Grade": g,
                "Abs Viscosity (P)": f"{lo}–{hi}",
                "Softening Pt (°C)": f"≥ {spec['softening_pt_min']}",
                "Penetration": f"{plo}–{phi}",
                "Ductility (cm)": f"≥ {spec['ductility_min']}",
                "Use Case": spec["use"],
            })
        st.dataframe(pd.DataFrame(ref_rows), use_container_width=True, hide_index=True)

    if test_btn:
        # build params — skip zeros (means not entered)
        params = {}
        if abs_vis > 0:    params["abs_viscosity"] = abs_vis
        if kin_vis > 0:    params["kin_viscosity"] = kin_vis
        if soft_pt > 0:    params["softening_pt"] = soft_pt
        if flash_pt > 0:   params["flash_pt"] = flash_pt
        if ductility > 0:  params["ductility"] = ductility
        if penetration > 0: params["penetration"] = penetration

        if not params:
            st.warning("Enter at least one test value.")
        else:
            res = grade_product(params)
            st.markdown("---")

            if res.get("eligible"):
                st.success(f"✅ **Eligible grades: {', '.join(res['eligible'])}**")
            else:
                st.warning("⚠️ No grade fully satisfied with entered values.")

            st.markdown(f"**{res['summary']}**")

            # per-grade scorecard
            gcols = st.columns(len(IS73_GRADES))
            for col, (grade, gres) in zip(gcols, res["results"].items()):
                qual = gres["qualifies"]
                color = "#51CF66" if qual else ("#FFD43B" if gres["score"] >= 60 else "#FF6B6B")
                with col:
                    st.markdown(
                        f'<div class="vg-card" style="border-color:{color};">'
                        f'<div class="vg-title" style="color:{color};">{grade}</div>'
                        f'<div class="vg-use">{gres["use"]}</div>'
                        f'{pct_bar(gres["score"], color)}'
                        f'<p style="font-size:12px;color:#9A8A6A;margin:4px 0 0;">'
                        f'{gres["passed"]}/{gres["total_checks"]} criteria met</p>',
                        unsafe_allow_html=True,
                    )
                    for param, chk in gres["checks"].items():
                        ic = "✅" if chk["pass"] else "❌"
                        st.markdown(
                            f'<div class="param-row">'
                            f'<span class="param-label">{param[:20]}</span>'
                            f'<span class="param-value">{ic} {chk["value"]}</span>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )
                    st.markdown('</div>', unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────
# TAB 4 — VENDOR RATING
# ─────────────────────────────────────────────────────────────────────────
with tab4:
    st.subheader("Vendor / Supplier Rating")
    st.caption("Score each supplier on Price, Delivery, Quality, Support, Compliance.")

    saved_vendors = load_vendor_ratings()

    # ── Add new vendor ────────────────────────────────────────────────
    with st.expander("➕ Add / Rate a New Vendor", expanded=(not saved_vendors)):
        v_name = st.text_input("Vendor Name", placeholder="e.g. ABC Biomass Suppliers", key="v_name")
        v_type = st.selectbox("Type", ["Raw Material", "Machinery", "Civil", "Electrical",
                                        "Transport", "Other"], key="v_type")

        vc1, vc2, vc3 = st.columns(3)
        with vc1:
            v_price = st.slider("Price Competitiveness", 0, 100, 70,
                                 help="100 = best price in market", key="v_price")
            v_delivery = st.slider("Delivery Reliability", 0, 100, 70,
                                    help="100 = always on time", key="v_del")
        with vc2:
            v_quality = st.slider("Product / Service Quality", 0, 100, 70,
                                   help="100 = no defects ever", key="v_qual")
            v_support = st.slider("After-sales Support", 0, 100, 70,
                                   help="100 = immediate response", key="v_supp")
        with vc3:
            v_comply = st.slider("Compliance & Docs", 0, 100, 70,
                                  help="100 = all docs present", key="v_comp")
            v_notes = st.text_area("Notes", placeholder="Optional remarks", key="v_notes",
                                    height=80)

        if st.button("Rate This Vendor", type="primary", key="add_vendor"):
            if not v_name.strip():
                st.warning("Enter vendor name.")
            else:
                new_v = {
                    "name": v_name.strip(), "type": v_type,
                    "price": v_price, "delivery": v_delivery,
                    "quality": v_quality, "support": v_support,
                    "compliance": v_comply, "notes": v_notes,
                    "rated_at": datetime.now().strftime("%d %b %Y"),
                }
                all_v = [v.get("name","") == new_v["name"] for v in saved_vendors]
                if any(all_v):
                    # update existing
                    idx = all_v.index(True)
                    raw_vendors = [
                        {k: saved_vendors[i].get(k, v)
                         for k in ["name","type","price","delivery","quality","support","compliance","notes","rated_at"]}
                        for i, v in enumerate(saved_vendors)
                    ]
                    raw_vendors[idx] = new_v
                else:
                    raw_vendors = [
                        {k: v.get(k, 0) for k in ["name","type","price","delivery","quality","support","compliance","notes","rated_at"]}
                        for v in saved_vendors
                    ]
                    raw_vendors.append(new_v)
                saved_vendors = save_vendor_ratings(raw_vendors)
                st.success(f"Vendor **{new_v['name']}** rated!")
                st.rerun()

    # ── Vendor leaderboard ────────────────────────────────────────────
    if saved_vendors:
        st.markdown("### Vendor Leaderboard")

        # sort by score
        sorted_v = sorted(saved_vendors, key=lambda v: v["score"], reverse=True)

        # summary table
        tbl = []
        for v in sorted_v:
            tbl.append({
                "Rank": sorted_v.index(v) + 1,
                "Vendor": v["name"],
                "Grade": v["grade"],
                "Score": v["score"],
                "Label": v["label"],
            })
        df_v = pd.DataFrame(tbl)
        st.dataframe(df_v, use_container_width=True, hide_index=True)

        # detail cards
        st.markdown("### Detailed Scorecards")
        card_cols = st.columns(min(len(sorted_v), 3))
        for i, v in enumerate(sorted_v):
            with card_cols[i % 3]:
                col_c = v["color"]
                st.markdown(
                    f'<div class="vendor-card" style="border-left-color:{col_c};">'
                    f'<div style="display:flex;justify-content:space-between;align-items:center;">'
                    f'<b style="color:#E8B547;">{v["name"]}</b>'
                    f'<span style="font-size:1.8rem;font-weight:900;color:{col_c};">{v["grade"]}</span>'
                    f'</div>'
                    f'<div style="color:#9A8A6A;font-size:12px;">{v.get("type","")}</div>'
                    f'{pct_bar(v["score"], col_c)}'
                    f'<p style="font-size:13px;color:#C8B88A;margin:6px 0 8px;">{v["summary"]}</p>',
                    unsafe_allow_html=True,
                )
                for dim in v["breakdown"]:
                    raw = dim["raw"]
                    col_d = "#51CF66" if raw >= 75 else ("#FFD43B" if raw >= 50 else "#FF6B6B")
                    st.markdown(
                        f'<div style="display:flex;justify-content:space-between;'
                        f'font-size:12px;padding:2px 0;">'
                        f'<span style="color:#9A8A6A;">{dim["criterion"]}</span>'
                        f'<span style="color:{col_d};font-weight:600;">{raw:.0f}/100</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                st.markdown('</div>', unsafe_allow_html=True)

        # download
        dl_df = pd.DataFrame([{
            "Vendor": v["name"], "Grade": v["grade"], "Score": v["score"],
            **{d["criterion"]: d["raw"] for d in v["breakdown"]},
            "Summary": v["summary"],
        } for v in sorted_v])
        st.download_button(
            "⬇ Download Vendor Ratings",
            dl_df.to_csv(index=False),
            file_name=f"vendor_ratings_{datetime.now():%Y%m%d}.csv",
            mime="text/csv",
            key="dl_vendors",
        )
    else:
        st.info("No vendors rated yet. Use the form above to add your first vendor.")

# ── sidebar quick view ────────────────────────────────────────────────────
with st.sidebar:
    cached_all = load_ratings()
    for key, label, icon in [
        ("project", "Project", "🏆"),
        ("dpr",     "DPR",     "📋"),
    ]:
        r = cached_all.get(key)
        if r:
            st.markdown(
                f'<div style="background:#1E1B14;border-radius:8px;padding:10px 12px;'
                f'margin-bottom:8px;border:1px solid #3A3520;">'
                f'<span style="color:#9A8A6A;font-size:12px;">{icon} {label}</span><br>'
                f'<span style="font-size:2rem;font-weight:900;color:{r["color"]};">{r["grade"]}</span>'
                f'<span style="color:#9A8A6A;font-size:12px;margin-left:8px;">{r["score"]}/100</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

    vrs = load_vendor_ratings()
    if vrs:
        best = max(vrs, key=lambda v: v["score"])
        st.markdown(
            f'<div style="background:#1E1B14;border-radius:8px;padding:10px 12px;'
            f'border:1px solid #3A3520;">'
            f'<span style="color:#9A8A6A;font-size:12px;">🏭 Top Vendor</span><br>'
            f'<b style="color:#E8B547;">{best["name"]}</b> '
            f'<span style="color:{best["color"]};font-weight:700;">{best["grade"]}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

    st.markdown("---")
    if st.button("Print", key="print_rating"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)

# ── export ────────────────────────────────────────────────────────────────
st.markdown("---")
ex1, ex2 = st.columns(2)
with ex1:
    try:
        import io
        from openpyxl import Workbook
        wb = Workbook()
        cr = load_ratings()

        # Sheet 1: Project
        ws1 = wb.active; ws1.title = "Project Rating"
        pr = cr.get("project", {})
        ws1.append(["Dimension", "Score", "Max", "Note"])
        for d in pr.get("breakdown", []):
            ws1.append([d["dim"], d["score"], d["max"], d["note"]])
        ws1.append([])
        ws1.append(["Overall", pr.get("score",""), 100, pr.get("grade","")])

        # Sheet 2: DPR
        ws2 = wb.create_sheet("DPR Completeness")
        dr = cr.get("dpr", {})
        ws2.append(["Section", "Status", "Earned", "Max Weight"])
        for s in dr.get("sections", []):
            ws2.append([s["section"], s["status"], s["earned"], s["weight"]])

        # Sheet 3: Vendors
        ws3 = wb.create_sheet("Vendor Ratings")
        vrs = load_vendor_ratings()
        if vrs:
            ws3.append(["Vendor", "Grade", "Score"] +
                       [d["criterion"] for d in vrs[0]["breakdown"]])
            for v in vrs:
                ws3.append([v["name"], v["grade"], v["score"]] +
                           [d["raw"] for d in v["breakdown"]])

        buf = io.BytesIO(); wb.save(buf)
        st.download_button("Download Excel", buf.getvalue(),
                           f"ratings_{datetime.now():%Y%m%d}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_xl_89", type="primary")
    except Exception:
        pass

with ex2:
    if st.button("Print", key="prt_89"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)

# ── AI assist ─────────────────────────────────────────────────────────────
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("AI Assist — Rating Interpretation"):
            if st.button("Generate AI Recommendations", type="primary", key="ai_89"):
                cr = load_ratings()
                pr = cr.get("project", {})
                dr = cr.get("dpr", {})
                prompt = (
                    f"Bio-bitumen project ratings: "
                    f"Project grade {pr.get('grade','N/A')} ({pr.get('score','?')}/100), "
                    f"DPR completeness {dr.get('grade','N/A')} ({dr.get('score','?')}/100). "
                    f"Summary: {pr.get('summary','')} "
                    f"Give 3 specific actionable improvements to raise both grades. "
                    "Professional consultant format, concise."
                )
                with st.spinner("AI thinking…"):
                    resp, prov = ask_ai(prompt, "Senior bio-energy project consultant.")
                if resp:
                    st.markdown(f"**via {prov.upper()}:**")
                    st.markdown(resp)
except Exception:
    pass

st.markdown("---")
st.caption(f"Last updated: {datetime.now().strftime('%d %B %Y %H:%M')} | {COMPANY['name']}")

# ── next steps ────────────────────────────────────────────────────────────
try:
    from engines.page_navigation import add_next_steps
    add_next_steps(st, "89")
except Exception:
    pass
