"""
Weather & Site Analysis — Live Weather, 7-Day Forecast, Construction Planning
===============================================================================
Powered by Open-Meteo API (100% Free, No API Key)
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from state_manager import init_state, get_config
from engines.free_apis import (get_weather_current, get_weather_forecast,
                                get_weather_history, CITY_COORDS,
                                get_india_holidays, detect_visitor_location)
from config import COMPANY

st.set_page_config(page_title="Weather & Site Analysis", page_icon="🌤️", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass


st.title("Weather & Site Analysis")
st.markdown("**Live Weather | 7-Day Forecast | Construction Planning | Holiday Calendar**")
st.markdown("*Powered by Open-Meteo (100% Free, No API Key)*")
st.markdown("---")

# ── Auto-detect location ─────────────────────────────────────────────
try:
    loc = detect_visitor_location()
    detected_city = loc.get("city", "Vadodara")
except Exception:
    detected_city = "Vadodara"

# ══════════════════════════════════════════════════════════════════════
# CITY SELECTOR
# ══════════════════════════════════════════════════════════════════════
cities = sorted(CITY_COORDS.keys())
sel1, sel2 = st.columns([1, 2])
with sel1:
    default_idx = cities.index(detected_city) if detected_city in cities else cities.index("Vadodara")
    selected_city = st.selectbox("Select Plant Location", cities, index=default_idx, key="weather_city")

# ══════════════════════════════════════════════════════════════════════
# CURRENT WEATHER
# ══════════════════════════════════════════════════════════════════════
with st.spinner(f"Fetching weather for {selected_city}..."):
    weather = get_weather_current(selected_city)

if "error" not in weather:
    with sel2:
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #0077be, #00a8e8); padding: 20px; border-radius: 12px; color: white;">
            <h3 style="color: white; margin: 0;">{selected_city} — {weather.get('condition', 'N/A')}</h3>
            <div style="display: flex; gap: 30px; margin-top: 10px;">
                <div><span style="font-size: 2em;">{weather.get('temperature_c', 0)}°C</span><br>Temperature</div>
                <div><span style="font-size: 2em;">{weather.get('humidity_pct', 0)}%</span><br>Humidity</div>
                <div><span style="font-size: 2em;">{weather.get('wind_kmh', 0)}</span><br>Wind km/h</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # Construction feasibility
    temp = weather.get("temperature_c", 30)
    humidity = weather.get("humidity_pct", 50)
    wind = weather.get("wind_kmh", 10)
    condition = weather.get("condition", "")

    feasible = True
    warnings = []
    if temp > 45:
        warnings.append("Extreme heat — avoid outdoor work")
        feasible = False
    elif temp > 40:
        warnings.append("High heat — limit afternoon work")
    if "rain" in condition.lower() or "shower" in condition.lower():
        warnings.append("Rain — bitumen work not recommended")
        feasible = False
    if "thunder" in condition.lower():
        warnings.append("Thunderstorm — stop all outdoor work")
        feasible = False
    if wind > 50:
        warnings.append("High wind — crane operations unsafe")
        feasible = False

    if feasible and not warnings:
        st.success(f"**Construction Conditions: GOOD** — {condition}, {temp}°C")
    elif feasible:
        st.warning(f"**Construction Conditions: CAUTION** — {'; '.join(warnings)}")
    else:
        st.error(f"**Construction Conditions: NOT RECOMMENDED** — {'; '.join(warnings)}")
else:
    st.error(f"Could not fetch weather: {weather.get('error', 'Unknown error')}")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# 7-DAY FORECAST
# ══════════════════════════════════════════════════════════════════════
st.subheader("7-Day Forecast")

forecast = get_weather_forecast(selected_city, 7)
if forecast:
    fc_df = pd.DataFrame(forecast)

    fig_fc = go.Figure()
    fig_fc.add_trace(go.Scatter(x=fc_df["date"], y=fc_df["temp_max"],
                                  mode="lines+markers", name="Max Temp",
                                  line=dict(color="#CC3333", width=2)))
    fig_fc.add_trace(go.Scatter(x=fc_df["date"], y=fc_df["temp_min"],
                                  mode="lines+markers", name="Min Temp",
                                  line=dict(color="#003366", width=2)))
    fig_fc.add_trace(go.Bar(x=fc_df["date"], y=fc_df["rain_mm"],
                              name="Rainfall (mm)", marker_color="#0088cc", opacity=0.5,
                              yaxis="y2"))
    fig_fc.update_layout(
        title=f"7-Day Weather Forecast — {selected_city}",
        template="plotly_white", height=400,
        yaxis=dict(title="Temperature (°C)"),
        yaxis2=dict(title="Rainfall (mm)", overlaying="y", side="right"),
        xaxis_title="Date",
    )
    st.plotly_chart(fig_fc, width="stretch")

    # Forecast table
    for day in forecast:
        rain = day.get("rain_mm", 0)
        rain_icon = "🌧️" if rain > 5 else ("💧" if rain > 0 else "☀️")
        st.markdown(f"{rain_icon} **{day['date']}** — {day['condition']} | "
                    f"{day['temp_min']}°C - {day['temp_max']}°C | "
                    f"Rain: {rain:.1f}mm | Wind: {day['wind_max']:.0f} km/h")

    # Construction days assessment
    good_days = sum(1 for d in forecast if d.get("rain_mm", 0) < 2 and d.get("temp_max", 30) < 42)
    st.metric("Good Construction Days (next 7)", f"{good_days}/7")
else:
    st.info("Forecast loading...")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# MULTI-CITY COMPARISON
# ══════════════════════════════════════════════════════════════════════
st.subheader("Compare Weather Across Cities")
compare_cities = st.multiselect("Select cities to compare", cities,
                                  default=["Vadodara", "Pune", "Indore"], max_selections=5,
                                  key="weather_compare")

if len(compare_cities) >= 2:
    comp_data = []
    for city in compare_cities:
        w = get_weather_current(city)
        if "error" not in w:
            comp_data.append({
                "City": city,
                "Temp (°C)": w.get("temperature_c", 0),
                "Humidity (%)": w.get("humidity_pct", 0),
                "Wind (km/h)": w.get("wind_kmh", 0),
                "Condition": w.get("condition", ""),
            })

    if comp_data:
        comp_df = pd.DataFrame(comp_data)
        st.dataframe(comp_df, width="stretch", hide_index=True)

        fig_comp = px.bar(comp_df, x="City", y="Temp (°C)", color="Humidity (%)",
                           title="Temperature & Humidity Comparison",
                           color_continuous_scale="RdYlBu_r")
        fig_comp.update_layout(template="plotly_white", height=350)
        st.plotly_chart(fig_comp, width="stretch")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# INDIAN HOLIDAYS (Construction Planning)
# ══════════════════════════════════════════════════════════════════════
st.subheader("Indian Public Holidays — Construction Planning")

holidays = get_india_holidays(2026)
if holidays:
    hol_df = pd.DataFrame(holidays)
    st.dataframe(hol_df[["date", "name", "name_en", "type"]], width="stretch", hide_index=True)

    working_days_lost = len(holidays)
    st.info(f"**{working_days_lost} public holidays in 2026** — factor into construction timeline. "
            f"Effective working days: ~{300 - working_days_lost} (from 300 planned)")
else:
    st.info("Holiday data loading...")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# CONSTRUCTION SEASON GUIDE
# ══════════════════════════════════════════════════════════════════════
st.subheader("Construction Season Guide — India")

season_data = pd.DataFrame([
    {"Month": "January", "Season": "Winter", "Construction": "Excellent", "Bitumen Work": "Good", "Notes": "Cool weather, ideal for construction"},
    {"Month": "February", "Season": "Winter", "Construction": "Excellent", "Bitumen Work": "Good", "Notes": "Best month for road construction"},
    {"Month": "March", "Season": "Spring", "Construction": "Good", "Bitumen Work": "Excellent", "Notes": "Warming up, bitumen flows well"},
    {"Month": "April", "Season": "Summer", "Construction": "Good", "Bitumen Work": "Excellent", "Notes": "Peak bitumen season, hot and dry"},
    {"Month": "May", "Season": "Summer", "Construction": "Fair", "Bitumen Work": "Good", "Notes": "Very hot, limit afternoon work"},
    {"Month": "June", "Season": "Monsoon", "Construction": "Poor", "Bitumen Work": "Poor", "Notes": "Monsoon begins, rain disrupts work"},
    {"Month": "July", "Season": "Monsoon", "Construction": "Poor", "Bitumen Work": "Not Recommended", "Notes": "Peak monsoon, minimal road work"},
    {"Month": "August", "Season": "Monsoon", "Construction": "Poor", "Bitumen Work": "Not Recommended", "Notes": "Heavy rains continue"},
    {"Month": "September", "Season": "Monsoon", "Construction": "Fair", "Bitumen Work": "Fair", "Notes": "Monsoon retreating, windows of work"},
    {"Month": "October", "Season": "Post-Monsoon", "Construction": "Good", "Bitumen Work": "Good", "Notes": "Season reopens, demand surge"},
    {"Month": "November", "Season": "Post-Monsoon", "Construction": "Excellent", "Bitumen Work": "Excellent", "Notes": "Peak season for road tenders"},
    {"Month": "December", "Season": "Winter", "Construction": "Excellent", "Bitumen Work": "Good", "Notes": "Good weather, year-end rush"},
])

# Color code
def color_rating(val):
    colors = {"Excellent": "background-color: #ccffcc", "Good": "background-color: #ffffcc",
              "Fair": "background-color: #ffddcc", "Poor": "background-color: #ffcccc",
              "Not Recommended": "background-color: #ff9999"}
    return colors.get(val, "")

st.dataframe(season_data, width="stretch", hide_index=True)

st.markdown("---")
st.caption(f"{COMPANY['name']} | Weather data: Open-Meteo (free) | Holidays: Nager.Date (free)")

# ── Connect to Financial Model ────────────────────────────────────
st.markdown("---")
st.subheader("Connect to Financial Model")
current_days = cfg.get("working_days", 300)
try:
    from engines.free_apis import get_india_holidays
    holidays = get_india_holidays()
    holiday_count = len(holidays) if holidays else 15
except Exception:
    holiday_count = 15

# Monsoon adjustment
monsoon_loss = {"Mumbai":45,"Pune":35,"Chennai":30,"Kolkata":40,"Guwahati":50,
                "Bhubaneswar":35,"Hyderabad":25,"Ahmedabad":20,"Vadodara":20,
                "Jaipur":15,"Lucknow":25,"Indore":20,"Bhopal":25,"Nagpur":20,
                "Patna":30,"Ranchi":30,"Raipur":25,"Chandigarh":20,"Varanasi":25}
city = cfg.get("location", "Vadodara")
rain_days = monsoon_loss.get(city, 25)
effective_days = 365 - 52 - holiday_count - rain_days

st.markdown(f"""
| Factor | Days |
|--------|------|
| Calendar | 365 |
| Sundays | -52 |
| Holidays | -{holiday_count} |
| Monsoon ({city}) | -{rain_days} |
| **Effective** | **{effective_days}** |
""")

if effective_days != current_days:
    st.info(f"Financial Model uses {current_days} days. Weather suggests {effective_days} days.")
    if st.button(f"Update Financial Model to {effective_days} days", type="primary", key="weather_sync"):
        from state_manager import update_fields
        update_fields({"working_days": effective_days})
        st.success(f"Working days updated to {effective_days}!")
        st.rerun()
else:
    st.success(f"Financial Model already uses {current_days} days — matches weather data")


# ── AI Skill: Construction Calendar ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: Construction Calendar"):
            if st.button("Generate", type="primary", key="ai_71Weath"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: Construction Calendar. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_71Wea", type="primary")
with _ex2:
    if st.button("Print", key="exp_prt_71Wea"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)
