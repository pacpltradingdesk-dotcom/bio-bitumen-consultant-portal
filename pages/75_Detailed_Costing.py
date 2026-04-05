"""
Detailed Costing — Complete Cost-Per-Tonne Bifurcation + Finished Goods + Annual P&L
=====================================================================================
All values computed from state_manager cfg. ZERO hardcoded outputs.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from state_manager import get_config, init_state, format_inr
from config import COMPANY
from engines.detailed_costing import calculate_complete_cost_sheet, LOCATION_MULTIPLIERS
from engines.dpr_financial_engine import calculate_finished_goods

st.set_page_config(page_title="Detailed Costing", page_icon="📊", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass

st.title("Detailed Costing — Cost Per Tonne Bifurcation")
st.markdown(f"**Complete cost breakdown for {cfg['capacity_tpd']:.0f} TPD plant in {cfg.get('state', 'Maharashtra')}**")
st.markdown("---")

# Calculate everything from single source
cs = calculate_complete_cost_sheet(cfg)
fg = calculate_finished_goods(cfg, cs)

# ══════════════════════════════════════════════════════════════════════
# TOP KPIs
# ══════════════════════════════════════════════════════════════════════
k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("Net Cost/Tonne", f"₹{cs['net_cpt']:,}",
          help="All-in cost per tonne after by-product credits")
k2.metric("Sale Price/Tonne", f"₹{cs['sale_price_pt']:,}",
          help="Weighted: 70% VG30 + 30% VG40")
k3.metric("Margin/Tonne", f"₹{cs['margin_pt']:,}",
          delta=f"{cs['margin_pct']:.1f}%",
          delta_color="normal" if cs['margin_pt'] > 0 else "inverse")
k4.metric("Daily Revenue", format_inr(cs['total_rev_daily']))
k5.metric("Annual Net Profit", format_inr(cs['annual_pnl']['net_profit']),
          delta=f"DPR ROI {cs['annual_pnl']['roi_pct']:.1f}%")

st.info("**Note:** This is the DPR-grade conservative costing (landed cost of conventional bitumen + all overheads). "
        f"Financial Model ROI ({cfg.get('roi_pct', 20):.1f}%) uses EBITDA/Investment method which is the standard "
        "bank presentation metric. Both are correct for different purposes — use DPR costing for per-tonne analysis, "
        "Financial Model for investment return analysis.")
st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# 8 TABS
# ══════════════════════════════════════════════════════════════════════
tab_sheet, tab_rm, tab_bitumen, tab_landing, tab_production, tab_fg, tab_pnl, tab_location = st.tabs([
    "📊 Cost Sheet", "🌾 Raw Materials", "🛢️ Bitumen Cost",
    "🚛 Landing Cost", "⚙️ Production Cost",
    "📦 Finished Goods", "📈 Annual P&L", "📍 Location Multipliers"
])

# ── TAB: COST SHEET ──────────────────────────────────────────────────
with tab_sheet:
    st.subheader("Complete Cost Sheet — ₹ per Tonne of Bio-Bitumen Blend")
    st.caption(f"Location: {cs['state']} | Blend Output: {cs['blend_total_tpd']} T/day | "
               f"Feed: {cfg['capacity_tpd']:.0f} TPD | Yields: Oil {cfg.get('bio_oil_yield_pct',32)}% / "
               f"Char {cfg.get('bio_char_yield_pct',28)}% / Syngas {cfg.get('syngas_yield_pct',22)}%")

    d = cs["blend_total_tpd"] if cs["blend_total_tpd"] > 0 else 1
    rows = []
    for i, (head, daily_cost) in enumerate(cs["cost_heads"], 1):
        cpt = daily_cost / d
        pct = daily_cost / cs["gross_daily"] * 100 if cs["gross_daily"] > 0 else 0
        rows.append({"#": i, "Cost Head": head, "Daily ₹": f"₹{daily_cost:,.0f}",
                      "₹/Tonne": f"₹{cpt:,.0f}", "% of Gross": f"{pct:.1f}%"})

    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    # Summary box
    st.markdown(f"""
    | | Amount |
    |---|---|
    | **GROSS COST / Tonne** | **₹{cs['gross_daily']/d:,.0f}** |
    | Less: By-product Credits | -₹{cs['by_product_credit']/d:,.0f} |
    | Less: Scrap/Carbon Credits | -₹{cs['scrap_total']/d:,.0f} |
    | **NET COST / Tonne** | **₹{cs['net_cpt']:,}** |
    | Sale Price (70% VG30 + 30% VG40) | ₹{cs['sale_price_pt']:,} |
    | **MARGIN / Tonne** | **₹{cs['margin_pt']:,} ({cs['margin_pct']:.1f}%)** |
    """)

    # Cost breakdown pie chart
    c1, c2 = st.columns(2)
    with c1:
        fig = go.Figure(data=[go.Pie(
            labels=[h for h, _ in cs["cost_heads"]],
            values=[c for _, c in cs["cost_heads"]],
            hole=0.4,
            marker=dict(colors=["#f59e0b", "#ef4444", "#38bdf8", "#a78bfa", "#34d399",
                                 "#fb923c", "#64748b", "#f472b6", "#94a3b8", "#4ade80"]),
        )])
        fig.update_layout(title="Cost Breakdown (10 Heads)", template="plotly_white", height=350)
        st.plotly_chart(fig, use_container_width=True)

    with c2:
        rev_items = cs["revenue"]["items"]
        fig2 = go.Figure(data=[go.Bar(
            x=[i["product"].split("(")[0].strip() for i in rev_items],
            y=[i["daily_rev"] for i in rev_items],
            marker_color=["#f59e0b", "#fb923c", "#34d399", "#38bdf8", "#a78bfa", "#64748b"],
        )])
        fig2.update_layout(title="Revenue by Product (Daily ₹)", template="plotly_white", height=350,
                           xaxis_tickangle=-30)
        st.plotly_chart(fig2, use_container_width=True)

# ── TAB: RAW MATERIALS ──────────────────────────────────────────────
with tab_rm:
    st.subheader("Raw Material — 6 Feedstock Blended Average")
    st.caption(f"Location RM multiplier: {cs['rm']['rm_multiplier']:.2f}x ({cs['state']})")

    rm_rows = []
    for item in cs["rm"]["items"]:
        rm_rows.append({
            "Feedstock": item["feedstock"],
            "Farm Gate ₹/T": f"₹{item['farm_gate_price']:,}",
            "Mix %": f"{item['mix_pct']:.0f}%",
            "Weighted ₹": f"₹{item['weighted_price']:,}",
        })
    st.dataframe(pd.DataFrame(rm_rows), use_container_width=True, hide_index=True)

    r1, r2, r3 = st.columns(3)
    r1.metric("Blended RM Price/T", f"₹{cs['rm']['blended_price']:,}")
    r2.metric("Daily RM Cost", format_inr(cs['rm']['daily_cost']))
    r3.metric("RM % of Gross Cost", f"{cs['rm']['daily_cost']/cs['gross_daily']*100:.1f}%" if cs['gross_daily'] > 0 else "0%")

# ── TAB: BITUMEN COST ───────────────────────────────────────────────
with tab_bitumen:
    st.subheader("Conventional Bitumen — Landed Cost Breakdown")
    st.caption(f"Need {cs['bitumen']['conv_bitumen_needed']:.1f} T/day of VG30 for {cfg.get('bio_blend_pct', 20)}% blend")

    bit_rows = []
    for name, amount in cs["bitumen"]["breakdown"]:
        bit_rows.append({"Component": name, "₹/Tonne": f"₹{amount:,.0f}"})
    st.dataframe(pd.DataFrame(bit_rows), use_container_width=True, hide_index=True)

    b1, b2, b3 = st.columns(3)
    b1.metric("Landed Price/T", f"₹{cs['bitumen']['landed_per_tonne']:,}")
    b2.metric("Daily Bitumen Cost", format_inr(cs['bitumen']['daily_cost']))
    b3.metric("Bitumen % of Gross", f"{cs['bitumen']['daily_cost']/cs['gross_daily']*100:.1f}%" if cs['gross_daily'] > 0 else "0%")

# ── TAB: LANDING COST ───────────────────────────────────────────────
with tab_landing:
    st.subheader("Agro Waste Landing Cost — Farm Gate to Plant Gate")
    st.caption(f"Transport multiplier: {cs['landing']['transport_multiplier']:.2f}x ({cs['state']})")

    lc_rows = []
    for name, cost, is_transport in cs["landing"]["items"]:
        lc_rows.append({
            "Component": name,
            "₹/Tonne Feed": f"₹{cost:,.0f}",
            "Transport Adjusted": "✅" if is_transport else "",
        })
    st.dataframe(pd.DataFrame(lc_rows), use_container_width=True, hide_index=True)

    l1, l2, l3 = st.columns(3)
    l1.metric("Agro Landing/T", f"₹{cs['landing']['lc_per_tonne_agro']:,}")
    l2.metric("Agro Daily", format_inr(cs['landing']['lc_agro_daily']))
    l3.metric("Bitumen Freight Daily", format_inr(cs['landing']['lc_bitumen_daily']))

# ── TAB: PRODUCTION COST ────────────────────────────────────────────
with tab_production:
    st.subheader("Production Cost — Energy, Labour, Overheads, Chemicals")

    prod_rows = []
    for name, amount, formula in cs["production"]["items"]:
        prod_rows.append({"Component": name, "Daily ₹": f"₹{amount:,}", "Formula": formula})
    st.dataframe(pd.DataFrame(prod_rows), use_container_width=True, hide_index=True)

    p1, p2 = st.columns(2)
    p1.metric("Total Production Cost/Day", format_inr(cs['production']['total']))
    p2.metric("Syngas Credit/Day", format_inr(cs['production']['syngas_credit']),
              help="Syngas used as internal fuel reduces net energy cost")

    # Waste & Packing summary
    st.markdown("---")
    w1, w2, w3 = st.columns(3)
    w1.metric("Waste & Rejection/Day", format_inr(cs['waste']['total']),
              help=f"Includes {cs['waste']['waste_factor_pct']}% loss factor on production cost")
    w2.metric("Packing (Net)/Day", format_inr(cs['packing']['total']),
              help="Gross packing minus scrap return value")
    w3.metric("Outbound Delivery/Day", format_inr(cs['outbound']['total']),
              help=f"₹{cs['outbound']['cost_per_tonne']:,}/T × {cs['blend_total_tpd']} T/day")

# ── TAB: FINISHED GOODS ─────────────────────────────────────────────
with tab_fg:
    st.subheader("Finished Goods — Revenue by Product Stream")
    st.caption(f"7 revenue streams | {cfg.get('working_days', 300)} operating days/year")

    fg_rows = []
    for item in fg["items"]:
        fg_rows.append({
            "Product": item["product"],
            "Qty/Day": item["qty_per_day"],
            "Sale Price ₹/T": item["sale_price"],
            "Daily Revenue": f"₹{item['daily_revenue']:,}" if isinstance(item["daily_revenue"], (int, float)) else item["daily_revenue"],
            "Annual Revenue": format_inr(item["annual_revenue"]) if isinstance(item["annual_revenue"], (int, float)) else item["annual_revenue"],
            "% of Total": f"{item['pct_of_total']:.1f}%",
            "Buyer Segment": item["buyer"],
        })
    st.dataframe(pd.DataFrame(fg_rows), use_container_width=True, hide_index=True)

    f1, f2, f3 = st.columns(3)
    f1.metric("Total Daily Revenue", format_inr(fg["total_daily"]))
    f2.metric("Total Annual Revenue", format_inr(fg["total_annual"]))
    f3.metric("Annual Revenue (Cr)", f"₹ {fg['total_annual_cr']:.2f} Cr")

# ── TAB: ANNUAL P&L ─────────────────────────────────────────────────
with tab_pnl:
    st.subheader("Annual Profit & Loss Statement")
    pnl = cs["annual_pnl"]

    pnl_rows = [
        {"Line Item": "Revenue", "₹ Annual": format_inr(pnl["revenue"]), "% of Revenue": "100.0%"},
        {"Line Item": "Less: Cost of Goods Sold (COGS)", "₹ Annual": format_inr(-pnl["cogs"]), "% of Revenue": f"{pnl['cogs']/pnl['revenue']*100:.1f}%" if pnl["revenue"] > 0 else ""},
        {"Line Item": "Add: Scrap & By-product Credits", "₹ Annual": format_inr(pnl["scrap_credit"]), "% of Revenue": f"{pnl['scrap_credit']/pnl['revenue']*100:.1f}%" if pnl["revenue"] > 0 else ""},
        {"Line Item": "GROSS PROFIT", "₹ Annual": format_inr(pnl["gross_profit"]), "% of Revenue": f"{pnl['gross_margin_pct']:.1f}%"},
        {"Line Item": "---", "₹ Annual": "---", "% of Revenue": "---"},
        {"Line Item": "Less: Depreciation", "₹ Annual": format_inr(-pnl["depreciation"]), "% of Revenue": ""},
        {"Line Item": "Less: Interest on Term Loan", "₹ Annual": format_inr(-pnl["interest"]), "% of Revenue": ""},
        {"Line Item": "Less: Selling & Admin (2%)", "₹ Annual": format_inr(-pnl["sga"]), "% of Revenue": ""},
        {"Line Item": "EBT (Earnings Before Tax)", "₹ Annual": format_inr(pnl["ebt"]), "% of Revenue": ""},
        {"Line Item": f"Less: Tax @ {cfg.get('tax_rate', 0.25)*100:.0f}%", "₹ Annual": format_inr(-pnl["tax"]), "% of Revenue": ""},
        {"Line Item": "NET PROFIT (PAT)", "₹ Annual": format_inr(pnl["net_profit"]), "% of Revenue": f"{pnl['net_margin_pct']:.1f}%"},
    ]
    st.dataframe(pd.DataFrame(pnl_rows), use_container_width=True, hide_index=True)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Gross Margin", f"{pnl['gross_margin_pct']:.1f}%")
    m2.metric("Net Margin", f"{pnl['net_margin_pct']:.1f}%")
    m3.metric("ROI", f"{pnl['roi_pct']:.1f}%")
    m4.metric("Payback", f"{pnl['payback_years']:.1f} yrs" if pnl['payback_years'] < 100 else "N/A")

# ── TAB: LOCATION MULTIPLIERS ───────────────────────────────────────
with tab_location:
    st.subheader("Location Cost Multipliers — 9 States")

    loc_rows = []
    for state_name, mult in LOCATION_MULTIPLIERS.items():
        loc_rows.append({
            "State": state_name,
            "Raw Material": f"{mult['rm']:.2f}x",
            "Labour": f"{mult['lb']:.2f}x",
            "Transport In": f"{mult['tr_in']:.2f}x",
            "Transport Out": f"{mult['tr_out']:.2f}x",
            "Energy": f"{mult['energy']:.2f}x",
            "Elec Rate": f"₹{mult['elec_rate']}/kWh",
        })
    st.dataframe(pd.DataFrame(loc_rows), use_container_width=True, hide_index=True)

    current_mult = cs["multiplier"]
    st.info(f"**Current: {cs['state']}** — RM: {current_mult['rm']:.2f}x | Labour: {current_mult['lb']:.2f}x | "
            f"Transport: {current_mult['tr_in']:.2f}x | Energy: {current_mult['energy']:.2f}x")

# ── Export ───────────────────────────────────────────────────────────
st.markdown("---")
ex1, ex2 = st.columns(2)
with ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Cost Sheet"
        _ws.cell(row=1, column=1, value="Bio-Bitumen Complete Cost Sheet")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg['capacity_tpd']:.0f} TPD | State: {cs['state']}")
        _ws.cell(row=3, column=1, value="Cost Head")
        _ws.cell(row=3, column=2, value="Daily Rs")
        _ws.cell(row=3, column=3, value="Rs/Tonne")
        for _i, (_head, _cost) in enumerate(cs["cost_heads"], 4):
            _ws.cell(row=_i, column=1, value=_head)
            _ws.cell(row=_i, column=2, value=round(_cost))
            _ws.cell(row=_i, column=3, value=round(_cost / d))
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Cost Sheet Excel", _xl_data, "Detailed_Cost_Sheet.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_cs_xl", type="primary")
with ex2:
    if st.button("Print", key="prt_cost"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)

st.markdown("---")
st.caption(f"{COMPANY['name']} | DPR Detailed Costing Engine | All values auto-update from project config")
