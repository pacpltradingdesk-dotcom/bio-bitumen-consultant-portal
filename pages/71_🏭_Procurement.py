"""
Bio Bitumen Master Consulting System — Procurement Module
==========================================================
Vendor database, machinery cost comparison, import vs local options.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
from database import init_db, get_vendor_quotes, insert_vendor_quote
from state_manager import get_config, init_state
from config import CAPACITY_KEYS, CAPACITY_LABELS

st.set_page_config(page_title="Procurement", page_icon="🛒", layout="wide")
init_db()
init_state()
st.sidebar.markdown("---")
if st.sidebar.button("Print This Page", key="print_page"):
    import streamlit.components.v1 as _stc; _stc.html('<script>window.print();</script>', height=0)

st.title("Procurement & Vendor Management")
st.markdown("**Machinery vendors, cost comparison, and equipment sourcing**")
st.markdown("---")

cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass


# ── Default Equipment Database ────────────────────────────────────────
# VERIFIED supplier data from Verified_Supplier_Price_Research_March2026.md
DEFAULT_EQUIPMENT = [
    # Machinery — from verified IndiaMART/OEM research (March 2026)
    {"equipment": "Pyrolysis Reactor 10 TPD (Batch)", "vendor_name": "IndiaMART Supplier, Ahmedabad", "price_lac": 45, "delivery_weeks": 12, "warranty_months": 12, "source": "IndiaMART Verified Mar2026"},
    {"equipment": "Pyrolysis Reactor 10 TPD (Continuous)", "vendor_name": "Proctech Machineries, Ahmedabad", "price_lac": 75, "delivery_weeks": 16, "warranty_months": 12, "source": "IndiaMART Verified Mar2026"},
    {"equipment": "Pyrolysis Reactor 20 TPD (Continuous)", "vendor_name": "Custom/KGN Industries", "price_lac": 350, "delivery_weeks": 20, "warranty_months": 12, "source": "Verified Mar2026 Avg"},
    {"equipment": "Pyrolysis Reactor 30 TPD (Continuous)", "vendor_name": "Balaji Minerals, Hyderabad", "price_lac": 100, "delivery_weeks": 24, "warranty_months": 12, "source": "KGN Industries Website"},
    {"equipment": "Biomass Shredder 1-5 TPH", "vendor_name": "Sandhu Industries, Khanna Punjab", "price_lac": 7, "delivery_weeks": 6, "warranty_months": 12, "source": "IndiaMART Verified"},
    {"equipment": "Biomass Shredder 5-10 TPH", "vendor_name": "Ecostan India, Ludhiana", "price_lac": 19.75, "delivery_weeks": 8, "warranty_months": 12, "source": "IndiaMART Verified"},
    {"equipment": "Biomass Hammer Mill 4-5 TPH", "vendor_name": "Amandeep Engg, Ludhiana", "price_lac": 28.45, "delivery_weeks": 8, "warranty_months": 12, "source": "IndiaMART Verified"},
    {"equipment": "Rotary Dryer 1-50 TPH", "vendor_name": "Shiva Techno Fab, Ahmedabad", "price_lac": 12, "delivery_weeks": 8, "warranty_months": 12, "source": "IndiaMART Verified"},
    {"equipment": "Rotary Dryer 10 TPH (Industrial)", "vendor_name": "MOZER, Kolkata", "price_lac": 25, "delivery_weeks": 10, "warranty_months": 12, "source": "IndiaMART Verified"},
    {"equipment": "Shell & Tube Heat Exchanger (Set)", "vendor_name": "Platex India, Pune", "price_lac": 12, "delivery_weeks": 6, "warranty_months": 12, "source": "IndiaMART Verified"},
    {"equipment": "Vacuum Distillation Unit (Industrial)", "vendor_name": "Balaji Consultants, Pune", "price_lac": 125, "delivery_weeks": 16, "warranty_months": 12, "source": "Verified Mar2026"},
    {"equipment": "Bio-Oil Storage Tanks 3x100KL", "vendor_name": "Zaid Steel, Mumbai / Advance Eng, Ahmedabad", "price_lac": 24, "delivery_weeks": 6, "warranty_months": 6, "source": "IndiaMART Verified"},
    {"equipment": "Material Handling Conveyors (Set 6-8)", "vendor_name": "Various IndiaMART", "price_lac": 12, "delivery_weeks": 6, "warranty_months": 12, "source": "IndiaMART Avg"},
    {"equipment": "Pollution Control Scrubber+Cyclone", "vendor_name": "Entrade Engineers, Ahmedabad", "price_lac": 15, "delivery_weeks": 8, "warranty_months": 12, "source": "IndiaMART Verified"},
    {"equipment": "DG Set 500 KVA (Kirloskar)", "vendor_name": "Kirloskar, Bhilai", "price_lac": 35.75, "delivery_weeks": 4, "warranty_months": 24, "source": "IndiaMART Verified"},
    {"equipment": "DG Set 600 KVA (Cummins)", "vendor_name": "Cummins, Bengaluru", "price_lac": 38, "delivery_weeks": 4, "warranty_months": 24, "source": "IndiaMART Verified"},
    {"equipment": "Bitumen Testing Lab (Complete)", "vendor_name": "AMV Scientific / Sun LabTek", "price_lac": 8, "delivery_weeks": 4, "warranty_months": 12, "source": "IndiaMART Verified"},
    {"equipment": "HT/LT Substation & Transformer", "vendor_name": "Siemens/ABB", "price_lac": 12, "delivery_weeks": 8, "warranty_months": 24, "source": "OEM Quote"},
    {"equipment": "Weighbridge 60MT Electronic", "vendor_name": "Essae, Bangalore", "price_lac": 6, "delivery_weeks": 4, "warranty_months": 12, "source": "OEM Price"},
    {"equipment": "Fire Fighting System", "vendor_name": "Minimax/Ceasefire", "price_lac": 4, "delivery_weeks": 4, "warranty_months": 24, "source": "Fire NOC Spec"},
]

tab_db, tab_add, tab_compare = st.tabs(["Equipment Database", "Add Vendor Quote", "Cost Comparison"])

with tab_db:
    st.subheader("Equipment & Vendor Database")

    # Load from DB + defaults
    db_quotes = get_vendor_quotes()
    if not db_quotes:
        # Auto-seed with defaults
        for eq in DEFAULT_EQUIPMENT:
            insert_vendor_quote(eq)
        db_quotes = get_vendor_quotes()

    search = st.text_input("Search equipment", placeholder="e.g. reactor, condenser, dryer...")
    if search:
        db_quotes = [q for q in db_quotes if search.lower() in q.get("equipment", "").lower()
                     or search.lower() in q.get("vendor_name", "").lower()]

    if db_quotes:
        eq_df = pd.DataFrame(db_quotes)
        display = eq_df[["equipment", "vendor_name", "price_lac", "delivery_weeks", "warranty_months", "source"]].copy()
        display.columns = ["Equipment", "Vendor", "Price (Lac)", "Delivery (Weeks)", "Warranty (Months)", "Source"]
        st.dataframe(display, width="stretch", hide_index=True)

        total = sum(q.get("price_lac", 0) for q in db_quotes)
        st.metric("Total Equipment Cost", f"Rs {total:.1f} Lac ({total/100:.2f} Cr)")
    else:
        st.info("No equipment found matching search.")

with tab_add:
    st.subheader("Add New Vendor Quote")
    with st.form("add_vendor"):
        c1, c2 = st.columns(2)
        with c1:
            eq_name = st.text_input("Equipment Name")
            vendor = st.text_input("Vendor Name")
            price = st.number_input("Price (Lakhs)", min_value=0.0, step=0.5)
        with c2:
            delivery = st.number_input("Delivery (Weeks)", min_value=1, step=1, value=8)
            warranty = st.number_input("Warranty (Months)", min_value=0, step=6, value=12)
            source = st.text_input("Source/Reference")

        if st.form_submit_button("Add Quote"):
            if eq_name and vendor:
                insert_vendor_quote({
                    "equipment": eq_name, "vendor_name": vendor,
                    "price_lac": price, "delivery_weeks": delivery,
                    "warranty_months": warranty, "source": source,
                })
                st.success(f"Added: {eq_name} from {vendor}")
                st.rerun()

with tab_compare:
    st.subheader("Equipment Cost vs Plant Capacity")
    st.markdown(f"**Current config: {cfg['capacity_tpd']:.0f} MT/Day | Investment: Rs {cfg['investment_cr']:.2f} Cr**")

    plant = cfg["plant_data"]
    if plant:
        comparison = {
            "Civil Works": plant.get("civil_lac", 0),
            "Machinery": plant.get("mach_lac", 0),
            "GST on Machinery": plant.get("gst_mach_lac", 0),
            "Working Capital": plant.get("wc_lac", 0),
            "Other (IDC, Preop, Cont, Sec)": (plant.get("idc_lac", 0) + plant.get("preop_lac", 0) +
                                                plant.get("cont_lac", 0) + plant.get("sec_lac", 0)),
        }
        import plotly.express as px
        fig = px.pie(names=list(comparison.keys()), values=list(comparison.values()),
                      title=f"Investment Breakdown — {cfg['capacity_tpd']:.0f} MT/Day")
        st.plotly_chart(fig, width="stretch")

        # Machinery as % of total
        mach_pct = plant.get("mach_lac", 0) / (cfg["investment_lac"] or 1) * 100
        st.metric("Machinery as % of Total Investment", f"{mach_pct:.1f}%")

# ══════════════════════════════════════════════════════════════════════
# AI SKILLS — Right where you need them
# ══════════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("AI Procurement Tools")

try:
    from engines.ai_engine import is_ai_available
    if is_ai_available():
        from engines.ai_skills import read_vendor_quotation, generate_bom_from_description

        skill_tab1, skill_tab2 = st.tabs(["📋 Analyze Vendor Quote", "📦 Generate BOM"])

        with skill_tab1:
            st.caption("Paste a vendor quotation → AI extracts key data and rates it")
            quote_text = st.text_area("Paste vendor quotation text here:", height=150, key="vendor_quote_text")
            if st.button("Analyze Quote", type="primary", key="analyze_quote") and quote_text:
                with st.spinner("AI analyzing quotation..."):
                    result, prov = read_vendor_quotation(quote_text, cfg)
                if result:
                    st.markdown(result)
                    st.caption(f"Powered by {prov}")

        with skill_tab2:
            st.caption("Describe equipment → AI generates Bill of Materials")
            bom_desc = st.text_area("Describe the equipment/system:", height=100, key="bom_desc",
                placeholder="e.g., Piping from reactor to condenser, 6 inch MS pipe, 20m length, with isolation valves")
            if st.button("Generate BOM", type="primary", key="gen_bom") and bom_desc:
                with st.spinner("AI generating Bill of Materials..."):
                    result, prov = generate_bom_from_description(bom_desc, cfg)
                if result:
                    st.markdown(result)
    else:
        st.info("Add API keys in AI Settings to enable AI procurement tools")
except Exception:
    pass


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_10_🏭_P", type="primary")
with _ex2:
    if st.button("Print", key="exp_prt_10🏭P"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)
