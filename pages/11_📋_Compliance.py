"""
Bio Bitumen Master Consulting System — Compliance & License Tracker
===================================================================
25 license types, state-wise tracking, per-customer compliance management.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
from state_manager import init_state, get_config
import pandas as pd
from database import (init_db, get_all_customers, get_customer,
                       get_compliance_items, update_compliance_item,
                       init_compliance_for_customer)
from config import LICENSE_TYPES, STATES

st.set_page_config(page_title="Compliance Tracker", page_icon="📋", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass

init_db()
st.sidebar.markdown("---")
if st.sidebar.button("Print This Page", key="print_page"):
    import streamlit.components.v1 as _stc; _stc.html('<script>window.print();</script>', height=0)

st.title("Compliance & License Tracker")
st.markdown("**Track all licenses and approvals needed for plant setup — state-wise**")
st.markdown("---")

tab_master, tab_customer = st.tabs(["Master License List", "Customer Compliance Tracking"])

# ═══════════════════════════════════════════════════════════════════
# TAB 1: MASTER LICENSE LIST
# ═══════════════════════════════════════════════════════════════════
with tab_master:
    st.subheader(f"All {len(LICENSE_TYPES)} License Types Required for Bio-Bitumen Plant")

    # Summary metrics
    mandatory = [lt for lt in LICENSE_TYPES if lt.get("mandatory")]
    optional = [lt for lt in LICENSE_TYPES if not lt.get("mandatory")]
    c1, c2, c3 = st.columns(3)
    c1.metric("Total Licenses", len(LICENSE_TYPES))
    c2.metric("Mandatory", len(mandatory))
    c3.metric("Optional", len(optional))

    # Category filter
    categories = sorted(set(lt["category"] for lt in LICENSE_TYPES))
    selected_cats = st.multiselect("Filter by Category", categories, default=categories)

    # Display table
    filtered = [lt for lt in LICENSE_TYPES if lt["category"] in selected_cats]
    license_df = pd.DataFrame(filtered)
    license_df["Mandatory"] = license_df["mandatory"].apply(lambda x: "Yes" if x else "No")
    license_df["Timeline"] = license_df["typical_days"].apply(lambda x: f"{x} days")
    display_df = license_df[["name", "authority", "category", "Mandatory", "Timeline"]].copy()
    display_df.columns = ["License", "Issuing Authority", "Category", "Mandatory", "Typical Timeline"]
    st.dataframe(display_df, width="stretch", hide_index=True)

    # Timeline Gantt-like view
    st.markdown("---")
    st.subheader("Timeline Estimate (Sequential)")
    total_days = sum(lt["typical_days"] for lt in mandatory)
    st.info(f"Total estimated time for all mandatory licenses: **{total_days} days** (~{total_days//30} months)")
    st.caption("Many licenses can be applied for in parallel, reducing actual timeline to 3-4 months.")


# ═══════════════════════════════════════════════════════════════════
# TAB 2: CUSTOMER COMPLIANCE TRACKING
# ═══════════════════════════════════════════════════════════════════
with tab_customer:
    st.subheader("Per-Customer License Tracking")

    customers = get_all_customers()
    if not customers:
        st.warning("Add customers in Customer Manager first.")
        st.stop()

    cust_map = {c["id"]: f"{c['name']} ({c.get('company', '')})" for c in customers}
    selected_id = st.selectbox("Select Customer", list(cust_map.keys()),
                                format_func=lambda x: cust_map[x])

    # Initialize compliance items for this customer
    init_compliance_for_customer(selected_id, LICENSE_TYPES)
    items = get_compliance_items(selected_id)

    # Status summary
    status_counts = {}
    for item in items:
        s = item["status"]
        status_counts[s] = status_counts.get(s, 0) + 1

    sc1, sc2, sc3, sc4 = st.columns(4)
    sc1.metric("Not Started", status_counts.get("Not Started", 0))
    sc2.metric("Applied", status_counts.get("Applied", 0))
    sc3.metric("Received", status_counts.get("Received", 0))
    sc4.metric("Expired", status_counts.get("Expired", 0))

    # Editable compliance table
    st.markdown("---")
    for item in items:
        with st.expander(f"{'✅' if item['status'] == 'Received' else '⏳' if item['status'] == 'Applied' else '⬜'} {item['license_name']} — {item['status']}"):
            col1, col2, col3 = st.columns(3)
            with col1:
                new_status = st.selectbox("Status", ["Not Started", "Applied", "Received", "Expired"],
                                           index=["Not Started", "Applied", "Received", "Expired"].index(item["status"]),
                                           key=f"status_{item['id']}")
            with col2:
                applied = st.date_input("Applied Date", value=None, key=f"applied_{item['id']}")
                received = st.date_input("Received Date", value=None, key=f"received_{item['id']}")
            with col3:
                expiry = st.date_input("Expiry Date", value=None, key=f"expiry_{item['id']}")
                notes = st.text_input("Notes", value=item.get("notes", ""), key=f"notes_{item['id']}")

            if st.button("Update", key=f"update_{item['id']}"):
                update_compliance_item(item["id"], {
                    "status": new_status,
                    "applied_date": str(applied) if applied else None,
                    "received_date": str(received) if received else None,
                    "expiry_date": str(expiry) if expiry else None,
                    "notes": notes,
                })
                st.success(f"Updated {item['license_name']}")
                st.rerun()

        # ── Compliance Score ─────────────────────────────────────────
        st.markdown("---")
        st.subheader("Compliance Score")
        total_items = len(items)
        received_items = sum(1 for i in items if i["status"] == "Received")
        applied_items = sum(1 for i in items if i["status"] == "Applied")
        compliance_score = (received_items / total_items * 100) if total_items > 0 else 0

        cs1, cs2, cs3, cs4 = st.columns(4)
        cs1.metric("Total Licenses", total_items)
        cs2.metric("Received", received_items)
        cs3.metric("Applied", applied_items)
        cs4.metric("Compliance Score", f"{compliance_score:.0f}%")

        import plotly.express as px
        st.progress(compliance_score / 100)

        # ── Expiry Alerts ────────────────────────────────────────────
        import datetime
        today = datetime.date.today()
        expiring_soon = []
        for item in items:
            if item.get("expiry_date") and item["status"] == "Received":
                try:
                    exp_date = datetime.date.fromisoformat(item["expiry_date"])
                    days_left = (exp_date - today).days
                    if 0 < days_left <= 90:
                        expiring_soon.append((item["license_name"], days_left, item["expiry_date"]))
                except (ValueError, TypeError):
                    pass

        if expiring_soon:
            st.warning(f"**{len(expiring_soon)} licenses expiring within 90 days!**")
            for name, days, exp in sorted(expiring_soon, key=lambda x: x[1]):
                icon = "🔴" if days <= 30 else "🟡"
                st.markdown(f"{icon} **{name}** — expires {exp} ({days} days left)")

        # ── License Timeline (Gantt) ─────────────────────────────────
        st.markdown("---")
        st.subheader("License Application Timeline")

        import plotly.express as px
        gantt_data = []
        for item in items:
            if item.get("applied_date"):
                start = item["applied_date"]
                end = item.get("received_date") or item.get("expiry_date") or str(today)
                gantt_data.append({
                    "License": item["license_name"],
                    "Start": start,
                    "End": end,
                    "Status": item["status"],
                })

        if gantt_data:
            gantt_df = pd.DataFrame(gantt_data)
            gantt_df["Start"] = pd.to_datetime(gantt_df["Start"])
            gantt_df["End"] = pd.to_datetime(gantt_df["End"])
            color_map = {"Applied": "#FF8800", "Received": "#00AA44", "Expired": "#CC3333", "Not Started": "#cccccc"}
            fig_gantt = px.timeline(gantt_df, x_start="Start", x_end="End", y="License",
                                    color="Status", color_discrete_map=color_map,
                                    title="License Application Progress")
            fig_gantt.update_yaxes(autorange="reversed")
            fig_gantt.update_layout(template="plotly_white", height=400)
            st.plotly_chart(fig_gantt, width="stretch")
        else:
            st.info("Apply for licenses and add dates to see the timeline chart.")

# ── Export Section ────────────────────────────────────────────────────
st.markdown("---")
st.subheader("Export")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_11_📋_C", type="primary")
with _ex2:
    if st.button("Print Page", key="exp_print_analysis"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)

# ══════════════════════════════════════════════════════════════════════
# AI COMPLIANCE TOOLS — Draft applications automatically
# ══════════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("AI Compliance Tools")

try:
    from engines.ai_engine import is_ai_available
    if is_ai_available():
        from engines.ai_skills import draft_gpcb_application, draft_factory_license, get_compliance_checklist

        comp_tab1, comp_tab2, comp_tab3 = st.tabs(["📄 Draft GPCB Application", "🏭 Draft Factory License", "✅ State Checklist"])

        with comp_tab1:
            if st.button("Draft CTE Application", type="primary", key="draft_cte"):
                with st.spinner("AI drafting GPCB application..."):
                    result, prov = draft_gpcb_application(cfg)
                if result:
                    st.markdown(result)
                    st.download_button("Download Application", result, "GPCB_CTE_Application.txt", "text/plain", key="dl_cte")

        with comp_tab2:
            if st.button("Draft Factory License", type="primary", key="draft_fl"):
                with st.spinner("AI drafting Factory License application..."):
                    result, prov = draft_factory_license(cfg)
                if result:
                    st.markdown(result)
                    st.download_button("Download Application", result, "Factory_License_Application.txt", "text/plain", key="dl_fl")

        with comp_tab3:
            if st.button("Generate State Checklist", type="primary", key="gen_checklist"):
                with st.spinner(f"AI generating {cfg.get('state','Gujarat')} compliance checklist..."):
                    result, prov = get_compliance_checklist(cfg)
                if result:
                    st.markdown(result)
                    st.download_button("Download Checklist", result, "Compliance_Checklist.txt", "text/plain", key="dl_checklist")
    else:
        st.info("Add API keys in AI Settings to enable AI compliance tools")
except Exception:
    pass
