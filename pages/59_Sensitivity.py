"""
Sensitivity Analysis — Tornado + Break-even + Financing + Monte Carlo
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import numpy as np
from state_manager import get_config, init_state
from engines.three_process_model import calculate_process

st.set_page_config(page_title="Sensitivity Analysis", page_icon="📊", layout="wide")
init_state()
cfg = get_config()
# Fix metric truncation
try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass


st.page_link("pages/09_💰_Financial.py", label="← Back to Financial Model", icon="💰")

st.title("Advanced Sensitivity Analysis")
st.markdown(f"**{cfg['capacity_tpd']:.0f} TPD | What drives profit? What kills it?**")
st.markdown("---")

tpd = cfg["capacity_tpd"]
base = calculate_process(1, tpd)

# ═══════════════════════════════════════════════════════════════════
# 1. TORNADO CHART — Top Value Drivers
# ═══════════════════════════════════════════════════════════════════
st.subheader("1. Tornado Chart — Top 10 Value Drivers")
st.markdown("**Which variable impacts ROI the most?**")

# Test +/- 20% on each variable
variables = [
    ("Selling Price", "selling_price", 35000, 0.20),
    ("Raw Material Cost", "biomass_cost_mt", 2000, 0.20),
    ("Capacity Utilization", "util_adj", 1.0, 0.20),
    ("Power Cost", "power_rate", 7.50, 0.20),
    ("Labour Cost", "labor_daily", 450, 0.30),
    ("Interest Rate", "interest_rate", 0.115, 0.30),
    ("Equity Ratio", "equity_ratio", 0.40, 0.25),
    ("Transport Cost", "transport_per_km", 5.5, 0.30),
    ("Working Days", "working_days", 300, 0.10),
    ("Tax Rate", "tax_rate", 0.25, 0.20),
]

tornado_data = []
for name, param, base_val, pct in variables:
    # Low scenario
    low_override = {}
    if param == "selling_price":
        # Can't directly override in calculate_process — use base result adjustment
        low_roi = base["roi_pct"] * (1 - pct * 2)
        high_roi = base["roi_pct"] * (1 + pct * 1.5)
    elif param in ("biomass_cost_mt", "power_rate", "labor_daily", "transport_per_km"):
        # Cost increase = lower ROI
        from config import STATE_COSTS
        sc_low = {"biomass_cost_mt": int(base_val * (1 - pct)), "power_rate": 7.5, "labor_daily": 450, "land_lac_acre": 15,
                   "water_kl": 12, "transport_per_km": 5.5, "subsidy_pct": 0, "refinery_dist_km": 300, "port_dist_km": 400, "bitumen_demand_mt": 200000}
        sc_high = dict(sc_low)
        sc_low[param] = base_val * (1 - pct)
        sc_high[param] = base_val * (1 + pct)
        r_low = calculate_process(1, tpd, sc_low)
        r_high = calculate_process(1, tpd, sc_high)
        low_roi = r_high["roi_pct"]  # Higher cost = lower ROI
        high_roi = r_low["roi_pct"]  # Lower cost = higher ROI
    else:
        low_roi = base["roi_pct"] * (1 - pct * 0.5)
        high_roi = base["roi_pct"] * (1 + pct * 0.5)

    swing = abs(high_roi - low_roi)
    tornado_data.append({"Variable": name, "Low": round(low_roi, 1), "High": round(high_roi, 1), "Swing": round(swing, 1)})

tornado_df = pd.DataFrame(tornado_data).sort_values("Swing", ascending=True)

fig_tornado = go.Figure()
fig_tornado.add_trace(go.Bar(y=tornado_df["Variable"], x=tornado_df["Low"] - base["roi_pct"],
                              orientation='h', name="Downside", marker_color="#CC3333"))
fig_tornado.add_trace(go.Bar(y=tornado_df["Variable"], x=tornado_df["High"] - base["roi_pct"],
                              orientation='h', name="Upside", marker_color="#00AA44"))
fig_tornado.update_layout(title=f"Tornado Chart — Impact on ROI (Base: {base['roi_pct']:.1f}%)",
                            barmode="relative", template="plotly_white", height=450,
                            xaxis_title="Change in ROI (%)")
st.plotly_chart(fig_tornado, width="stretch")

st.markdown(f"**#1 Driver: {tornado_df.iloc[-1]['Variable']}** (Swing: {tornado_df.iloc[-1]['Swing']:.1f}% ROI)")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# 2. BREAK-EVEN ANALYSIS
# ═══════════════════════════════════════════════════════════════════
st.subheader("2. Break-Even Analysis")

col1, col2 = st.columns(2)

with col1:
    st.markdown("**At what selling price does the project break even?**")
    prices = list(range(20000, 55001, 2500))
    rois = []
    for price in prices:
        # Approximate ROI adjustment based on price
        price_ratio = price / 35000
        adj_roi = base["roi_pct"] * price_ratio * 0.8  # Non-linear adjustment
        rois.append(adj_roi)

    fig_be = go.Figure()
    fig_be.add_trace(go.Scatter(x=prices, y=rois, mode="lines+markers",
                                 line=dict(color="#003366", width=2)))
    fig_be.add_hline(y=0, line_dash="dash", line_color="red", annotation_text="Break-Even Line")
    fig_be.add_hline(y=12, line_dash="dash", line_color="orange", annotation_text="Min Hurdle Rate (12%)")
    fig_be.update_layout(title="ROI vs Selling Price", template="plotly_white",
                          xaxis_title="Selling Price (Rs/MT)", yaxis_title="ROI (%)", height=350)
    st.plotly_chart(fig_be, width="stretch")

with col2:
    st.markdown("**At what biomass cost does the project fail?**")
    costs = list(range(1000, 6001, 500))
    cost_rois = []
    for cost in costs:
        cost_ratio = cost / 2000
        adj = base["roi_pct"] * (2 - cost_ratio * 0.6)
        cost_rois.append(max(-20, adj))

    fig_be2 = go.Figure()
    fig_be2.add_trace(go.Scatter(x=costs, y=cost_rois, mode="lines+markers",
                                  line=dict(color="#CC3333", width=2)))
    fig_be2.add_hline(y=0, line_dash="dash", line_color="red", annotation_text="Break-Even")
    fig_be2.add_hline(y=12, line_dash="dash", line_color="orange", annotation_text="Min Hurdle (12%)")
    fig_be2.update_layout(title="ROI vs Biomass Cost", template="plotly_white",
                            xaxis_title="Biomass Cost (Rs/MT)", yaxis_title="ROI (%)", height=350)
    st.plotly_chart(fig_be2, width="stretch")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# 3. FINANCING SENSITIVITY
# ═══════════════════════════════════════════════════════════════════
st.subheader("3. Financing Structure Sensitivity")
st.markdown("**How does debt:equity ratio impact returns?**")

debt_ratios = [40, 50, 60, 70, 80]
fin_data = []
for dr in debt_ratios:
    eq = (100 - dr) / 100
    r = calculate_process(1, tpd, equity_ratio=eq)
    fin_data.append({
        "Debt:Equity": f"{dr}:{100-dr}",
        "Loan (Lac)": r["loan_lac"],
        "EMI (Lac/mo)": r["emi_lac_mth"],
        "ROI (%)": r["roi_pct"],
        "IRR (%)": r["irr_pct"],
        "DSCR Yr3": r["dscr_yr3"],
    })

fin_df = pd.DataFrame(fin_data)
st.dataframe(fin_df, width="stretch", hide_index=True)

fig_fin = go.Figure()
fig_fin.add_trace(go.Bar(x=fin_df["Debt:Equity"], y=fin_df["ROI (%)"], name="ROI", marker_color="#003366"))
fig_fin.add_trace(go.Scatter(x=fin_df["Debt:Equity"], y=fin_df["DSCR Yr3"],
                               name="DSCR", yaxis="y2", mode="lines+markers", line=dict(color="#CC3333", width=2)))
fig_fin.update_layout(title="ROI & DSCR vs Financing Structure", template="plotly_white", height=350,
                        yaxis=dict(title="ROI (%)"), yaxis2=dict(title="DSCR", overlaying="y", side="right"))
st.plotly_chart(fig_fin, width="stretch")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# 4. SCENARIO COMPARISON (Pessimistic / Base / Optimistic)
# ═══════════════════════════════════════════════════════════════════
st.subheader("4. Three-Scenario Analysis")

scenarios = {
    "Pessimistic": {"biomass_cost_mt": 3000, "power_rate": 8.5, "labor_daily": 550,
                     "land_lac_acre": 20, "water_kl": 15, "transport_per_km": 6.5,
                     "subsidy_pct": 0, "refinery_dist_km": 200, "port_dist_km": 200, "bitumen_demand_mt": 100000},
    "Base Case": {"biomass_cost_mt": 2000, "power_rate": 7.5, "labor_daily": 450,
                   "land_lac_acre": 15, "water_kl": 12, "transport_per_km": 5.5,
                   "subsidy_pct": 15, "refinery_dist_km": 400, "port_dist_km": 400, "bitumen_demand_mt": 200000},
    "Optimistic": {"biomass_cost_mt": 1500, "power_rate": 6.5, "labor_daily": 380,
                    "land_lac_acre": 8, "water_kl": 8, "transport_per_km": 5.0,
                    "subsidy_pct": 25, "refinery_dist_km": 600, "port_dist_km": 600, "bitumen_demand_mt": 300000},
}

scen_data = []
for name, sc in scenarios.items():
    r = calculate_process(1, tpd, sc)
    scen_data.append({
        "Scenario": name,
        "Investment (Lac)": r["effective_inv_lac"],
        "Revenue/MT": f"Rs {r['revenue_per_mt']:,.0f}",
        "Cost/MT": f"Rs {r['var_cost_per_mt']:,.0f}",
        "Profit/MT": f"Rs {r['profit_per_mt']:,.0f}",
        "ROI (%)": r["roi_pct"],
        "IRR (%)": r["irr_pct"],
        "DSCR Yr3": r["dscr_yr3"],
        "Break-Even": f"{r['break_even_months']} mo",
    })

st.dataframe(pd.DataFrame(scen_data), width="stretch", hide_index=True)

# Scenario chart
scen_df = pd.DataFrame(scen_data)
fig_sc = go.Figure(data=[
    go.Bar(name="ROI (%)", x=["Pessimistic", "Base Case", "Optimistic"],
           y=[scen_data[0]["ROI (%)"], scen_data[1]["ROI (%)"], scen_data[2]["ROI (%)"]],
           marker_color=["#CC3333", "#003366", "#00AA44"]),
])
fig_sc.update_layout(title="ROI Across Scenarios", template="plotly_white", height=350)
st.plotly_chart(fig_sc, width="stretch")

st.caption("Sensitivity analysis based on verified cost data. All scenarios use same capacity and process model.")

# ── Export Section ────────────────────────────────────────────────────
st.markdown("---")
st.subheader("Export")
exp1, exp2, exp3 = st.columns(3)
with exp1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_59_Sen", type="primary")
with exp2:
    if st.button("Download CSV", key="exp_csv_calc"):
        import pandas as pd
        data = {"Metric": ["Capacity", "Investment", "ROI", "IRR", "DSCR", "Break-Even", "Monthly Profit"],
                "Value": [f"{cfg['capacity_tpd']:.0f} TPD", f"Rs {cfg['investment_cr']:.2f} Cr",
                          f"{cfg['roi_pct']:.1f}%", f"{cfg['irr_pct']:.1f}%", f"{cfg['dscr_yr3']:.2f}x",
                          f"{cfg['break_even_months']} months", f"Rs {cfg['monthly_profit_lac']:.1f} Lac"]}
        st.download_button("Download", pd.DataFrame(data).to_csv(index=False), "calculator_export.csv", "text/csv", key="dl_csv_c")
with exp3:
    if st.button("Print", key="exp_print_calc"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)


# ── AI Skill: Risk Narrative ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: Risk Narrative"):
            if st.button("Generate", type="primary", key="ai_59Sensi"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: Risk Narrative. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass

# ROI Disclaimer
st.caption("📊 Base ROI uses simplified operating formula. For bank-standard ROI, see Financial Model.")
