"""
Product Grades — VG10/VG20/VG30/VG40 Specification Matrix + Lab Standards
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
from state_manager import init_state, get_config
import pandas as pd
import plotly.express as px

st.set_page_config(page_title="Product Grades", page_icon="🧪", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass

st.title("Product Grades & Specifications")
st.markdown("**Bio-Bitumen VG Grade Matrix | IS:73 Compliance | MoRTH 2026 Standards**")
st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# VG GRADE SPECIFICATION MATRIX (from uploaded image)
# ═══════════════════════════════════════════════════════════════════
st.subheader("Bio-Bitumen VG Grade Specification Matrix")

specs = pd.DataFrame([
    {"Property": "Penetration at 25°C (0.1mm)", "Test Method": "IS 1203", "VG10": "80-100", "VG20": "60-80", "VG30": "50-70", "VG40": "40-60"},
    {"Property": "Absolute Viscosity at 60°C (Poise)", "Test Method": "IS 1206 Part 2", "VG10": "800-1200", "VG20": "1600-2400", "VG30": "2400-3600", "VG40": "3200-4800"},
    {"Property": "Kinematic Viscosity at 135°C (cSt)", "Test Method": "IS 1206 Part 3", "VG10": "min 250", "VG20": "min 300", "VG30": "min 350", "VG40": "min 400"},
    {"Property": "Flash Point (°C)", "Test Method": "IS 1209", "VG10": "min 220", "VG20": "min 220", "VG30": "min 220", "VG40": "min 220"},
    {"Property": "Solubility in TCE (%)", "Test Method": "IS 1216", "VG10": "min 99", "VG20": "min 99", "VG30": "min 99", "VG40": "min 99"},
    {"Property": "Softening Point (°C)", "Test Method": "IS 1205", "VG10": "40-50", "VG20": "45-55", "VG30": "47-58", "VG40": "50-60"},
    {"Property": "Ductility at 25°C (cm)", "Test Method": "IS 1208", "VG10": "min 75", "VG20": "min 50", "VG30": "min 40", "VG40": "min 25"},
    {"Property": "Specific Gravity at 27°C", "Test Method": "IS 1202", "VG10": "min 0.99", "VG20": "min 0.99", "VG30": "min 0.99", "VG40": "min 0.99"},
    {"Property": "Loss on Heating (%)", "Test Method": "IS 9382", "VG10": "max 1.0", "VG20": "max 1.0", "VG30": "max 1.0", "VG40": "max 0.5"},
    {"Property": "Retained Penetration after RTFO (%)", "Test Method": "IS 1203", "VG10": "min 47", "VG20": "min 47", "VG30": "min 47", "VG40": "min 55"},
])

st.dataframe(specs, width="stretch", hide_index=True)

st.markdown("""
**Standards Reference:**
- IS:73 — Paving Bitumen Specification
- MoRTH Specification for Road & Bridge Works, 6th Revision (2026)
- IRC SP:53 — Guidelines on Use of Modified Bitumen
""")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# GRADE SELECTION GUIDE
# ═══════════════════════════════════════════════════════════════════
st.subheader("Grade Selection Guide")

grade_guide = pd.DataFrame([
    {"Grade": "VG10", "Climate": "Cold regions (<25°C avg)", "Application": "Surface dressing, spray applications", "States": "J&K, HP, UK, NE states"},
    {"Grade": "VG20", "Climate": "Moderate (25-35°C avg)", "Application": "Light to medium traffic roads", "States": "North India plains"},
    {"Grade": "VG30", "Climate": "Hot (>35°C avg)", "Application": "Heavy traffic highways, NHAI roads", "States": "Central, West, South India"},
    {"Grade": "VG40", "Climate": "Very hot / high stress", "Application": "Intersections, toll plazas, heavy load areas", "States": "Rajasthan, Gujarat, Maharashtra"},
])
st.dataframe(grade_guide, width="stretch", hide_index=True)

# Visual comparison
fig = px.bar(grade_guide, x="Grade", y=[1, 2, 3, 4],
             title="Grade vs Performance Level",
             labels={"value": "Performance Index"},
             color_discrete_sequence=["#003366"])
fig.update_layout(template="plotly_white", height=300, showlegend=False)
st.plotly_chart(fig, width="stretch")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# PRODUCT PORTFOLIO
# ═══════════════════════════════════════════════════════════════════
st.subheader("Complete Product Portfolio & Pricing")

products = pd.DataFrame([
    {"Product": "Anantams Bio-Binder (KrishiBind)", "Grade": "Proprietary", "Selling Price (Rs/MT)": "65,000+", "Target": "Premium NHAI projects", "Margin": "High"},
    {"Product": "Green-Pave Hybrid VG-30", "Grade": "VG30", "Selling Price (Rs/MT)": "48,000-54,000", "Target": "Standard highway construction", "Margin": "Medium"},
    {"Product": "Bio-Emulsion (SS/RS)", "Grade": "SS-1 / RS-1", "Selling Price (Rs/MT)": "38,000-46,000", "Target": "Surface dressing, patch repair", "Margin": "Medium"},
    {"Product": "Bio-Modified VG-40", "Grade": "VG40", "Selling Price (Rs/MT)": "48,000-55,000", "Target": "Heavy-duty intersections", "Margin": "Medium-High"},
    {"Product": "Bio-CRMB 55/60", "Grade": "CRMB", "Selling Price (Rs/MT)": "52,000-65,000", "Target": "Modified bitumen requirements", "Margin": "High"},
    {"Product": "Bio-PMB (Polymer Modified)", "Grade": "PMB", "Selling Price (Rs/MT)": "60,000-68,000", "Target": "Expressways, bridges", "Margin": "High"},
    {"Product": "Biochar (Agriculture)", "Grade": "Standard", "Selling Price (Rs/MT)": "12,000-15,000", "Target": "Farmers, nurseries", "Margin": "Low"},
    {"Product": "Activated Carbon", "Grade": "Industrial", "Selling Price (Rs/MT)": "35,000+", "Target": "Water treatment, industrial", "Margin": "High"},
])
st.dataframe(products, width="stretch", hide_index=True)

# Revenue mix chart
import plotly.graph_objects as go
fig2 = go.Figure(data=[go.Bar(
    x=["Bio-Binder", "VG-30", "VG-40", "CRMB", "PMB", "Emulsion", "Biochar", "Act. Carbon"],
    y=[65000, 51000, 51500, 58500, 64000, 42000, 13500, 35000],
    marker_color=["#003366", "#006699", "#004c8c", "#0088cc", "#00aadd", "#FF8800", "#228B22", "#333333"],
)])
fig2.update_layout(title="Product Selling Price (Rs/MT)", template="plotly_white",
                    yaxis_title="Rs/MT", height=400)
st.plotly_chart(fig2, width="stretch")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# MARKET DEMAND-SUPPLY GAP
# ═══════════════════════════════════════════════════════════════════
st.subheader("India Bitumen Market — Demand-Supply Gap")

col1, col2 = st.columns(2)
with col1:
    st.markdown("""
    | Parameter | Value |
    |-----------|-------|
    | India Bitumen Demand | ~8 Million MT/year |
    | Domestic Production | ~5.5 Million MT/year |
    | **Supply Gap** | **~2.5 Million MT/year** |
    | Annual Import Cost | **Rs 15,000-20,000 Crore** |
    | Import Dependency | **30-35%** |
    | Growth Rate | 8-10% CAGR |
    """)

with col2:
    gap_years = [2020, 2021, 2022, 2023, 2024, 2025, 2026]
    demand = [6.5, 6.8, 7.2, 7.5, 7.8, 8.0, 8.5]
    production = [4.8, 5.0, 5.1, 5.2, 5.3, 5.5, 5.6]
    imports = [d - p for d, p in zip(demand, production)]

    fig3 = go.Figure()
    fig3.add_trace(go.Scatter(x=gap_years, y=demand, name="Demand", fill="tozeroy",
                               line=dict(color="#CC3333", width=2)))
    fig3.add_trace(go.Scatter(x=gap_years, y=production, name="Domestic Production", fill="tozeroy",
                               line=dict(color="#006699", width=2)))
    fig3.update_layout(title="Bitumen Demand vs Production (Million MT)",
                        template="plotly_white", height=350, yaxis_title="Million MT")
    st.plotly_chart(fig3, width="stretch")

st.success("""
**Bio-Bitumen Opportunity:**
- If India replaces even **15% of imports** with bio-bitumen, annual saving: **Rs 4,000-4,500 Crore**
- Conventional bitumen: ~Rs 48-50/kg | Bio-bitumen: ~Rs 35-45/kg
- NHAI green mandate creates **guaranteed demand**
- Farmer crop residue monetization = social impact + ESG compliance
""")

st.caption("Sources: PPAC, MoPNG, IBEF Infrastructure Report, PetroBazaar, CSIR-CRRI research")


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    if st.button("Download Excel", type="primary", key="exp_xl_52Pro"):
        try:
            import io
            from openpyxl import Workbook
            _wb = Workbook()
            _ws = _wb.active
            _ws.title = "Export"
            _ws.cell(row=1, column=1, value="Bio Bitumen Export")
            _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
            _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
            _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
            _buf = io.BytesIO()
            _wb.save(_buf)
            _buf.seek(0)
            st.download_button("Download", _buf.getvalue(), "export.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_xl_52Pro")
        except Exception as _e:
            st.error(f"Export failed: {_e}")
with _ex2:
    if st.button("Print", key="exp_prt_52Pro"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)


# ── AI Assist ────────────────────────────────────────────────────
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("AI Assist"):
            if st.button("Generate AI Summary", type="primary", key="ai_52Pro"):
                with st.spinner("AI working..."):
                    _p = f"Summarize this section for a {cfg.get('capacity_tpd',20):.0f} TPD bio-bitumen plant in {cfg.get('state','')}. Investment Rs {cfg.get('investment_cr',8):.2f} Cr. Professional consultant format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 800)
                if _r:
                    st.markdown(_r)
except Exception:
    pass
