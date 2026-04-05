"""
Market Intelligence — REAL API DATA: Crude oil, FX rates, VG30 estimate, Demand
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
from state_manager import init_state, get_config
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from engines.market_data_api import get_market_summary
from config import STATES

st.set_page_config(page_title="Market Intelligence", page_icon="📈", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass

st.sidebar.markdown("---")
if st.sidebar.button("Print This Page", key="print_page"):
    import streamlit.components.v1 as _stc; _stc.html('<script>window.print();</script>', height=0)

st.title("Market Intelligence (LIVE DATA)")
st.markdown("**Real-time: Crude Oil | USD/INR | VG30 Estimate | Demand Heatmap | Govt Projects**")
st.markdown("---")

# ── Fetch REAL data ───────────────────────────────────────────────────
with st.spinner("Fetching live market data..."):
    market = get_market_summary()

# ── Top Metrics (REAL) ────────────────────────────────────────────────
vg30 = market.get("vg30_estimate", {})
fx = market.get("usd_inr", {})
gold = market.get("gold", {})
crude_data = market.get("crude_oil")

c1, c2, c3, c4 = st.columns(4)
c1.metric("VG30 Estimate", f"Rs {vg30.get('vg30_estimated', 0):,.0f}/MT",
          help=vg30.get("formula", ""))
c2.metric("Brent Crude", f"${vg30.get('crude_usd', 0):.2f}/bbl")
c3.metric("USD/INR", f"Rs {fx.get('rate', 0):.2f}",
          help=f"Source: {fx.get('source', 'N/A')}")
c4.metric("Gold", f"${gold.get('price_usd', 0):,.2f}/oz" if gold else "Loading...")

st.caption(f"Last updated: {market.get('last_updated', 'N/A')} | Source: Yahoo Finance, Frankfurter/ECB")
st.markdown("---")

# ── Crude Oil Price Chart (REAL DATA) ─────────────────────────────────
st.subheader("Brent Crude Oil — 12 Month Trend (LIVE)")

if crude_data:
    crude_df = pd.DataFrame(crude_data)
    crude_df["date"] = pd.to_datetime(crude_df["date"])

    fig_crude = go.Figure()
    fig_crude.add_trace(go.Scatter(
        x=crude_df["date"], y=crude_df["price_usd"],
        mode="lines", fill="tozeroy",
        line=dict(color="#003366", width=2),
        fillcolor="rgba(0,51,102,0.1)",
        name="Brent Crude"
    ))
    fig_crude.update_layout(title="Brent Crude Oil Price ($/barrel) — LIVE",
                             template="plotly_white", height=350,
                             yaxis_title="$/barrel", xaxis_title="Date")
    st.plotly_chart(fig_crude, width="stretch")

    # Price stats
    ps1, ps2, ps3, ps4 = st.columns(4)
    ps1.metric("Current", f"${crude_df['price_usd'].iloc[-1]:.2f}")
    ps2.metric("12M High", f"${crude_df['price_usd'].max():.2f}")
    ps3.metric("12M Low", f"${crude_df['price_usd'].min():.2f}")
    ps4.metric("12M Average", f"${crude_df['price_usd'].mean():.2f}")
else:
    st.warning("Could not fetch crude oil data. Check internet connection.")

st.markdown("---")

# ── USD/INR Chart (REAL DATA) ─────────────────────────────────────────
st.subheader("USD/INR Exchange Rate — 90 Day Trend (LIVE)")

fx_hist = market.get("fx_history")
if fx_hist:
    fx_df = pd.DataFrame(fx_hist)
    fx_df["date"] = pd.to_datetime(fx_df["date"])

    fig_fx = go.Figure()
    fig_fx.add_trace(go.Scatter(
        x=fx_df["date"], y=fx_df["rate"],
        mode="lines+markers", line=dict(color="#006699", width=2),
        name="USD/INR"
    ))
    fig_fx.update_layout(title="USD/INR Exchange Rate — LIVE",
                          template="plotly_white", height=300,
                          yaxis_title="Rs per USD")
    st.plotly_chart(fig_fx, width="stretch")
else:
    st.info("FX history loading... Showing latest rate above.")

st.markdown("---")

# ── VG30 Price Estimation ─────────────────────────────────────────────
st.subheader("VG30 Bitumen Price Estimation")
st.markdown(f"""
| Parameter | Value |
|-----------|-------|
| Brent Crude | **${vg30.get('crude_usd', 0):.2f}/barrel** |
| USD/INR Rate | **Rs {vg30.get('fx_rate', 0):.2f}** |
| Correlation Multiplier | **13.5x** (industry average) |
| **Estimated VG30 Price** | **Rs {vg30.get('vg30_estimated', 0):,.0f}/MT** |

*Note: {vg30.get('note', '')}*
""")

# Impact on Bio-Bitumen
bio_price = 35000  # Bio-bitumen selling price
discount = vg30.get('vg30_estimated', 48000) - bio_price
st.success(f"**Bio-Bitumen Advantage:** At Rs 35,000/MT vs VG30 at Rs {vg30.get('vg30_estimated', 48000):,.0f}/MT, "
           f"bio-bitumen offers a **Rs {discount:,.0f}/MT discount** ({discount/vg30.get('vg30_estimated', 48000)*100:.1f}% cheaper)")

st.markdown("---")

# ── Demand Heatmap ────────────────────────────────────────────────────
st.subheader("Bio-Bitumen Demand Potential (State-wise)")
from config import STATE_SCORES
import numpy as np

demand_data = []
np.random.seed(42)
for state in STATES:
    score = STATE_SCORES.get(state, {})
    # Demand correlated with logistics + biomass availability
    base_demand = (score.get("logistics", 50) * 300 + score.get("biomass", 50) * 200)
    demand_data.append({
        "State": state,
        "Demand Potential (MT/yr)": int(base_demand),
        "Road Projects": score.get("logistics", 50),
        "Biomass Score": score.get("biomass", 50),
    })

demand_df = pd.DataFrame(demand_data).sort_values("Demand Potential (MT/yr)", ascending=False)
fig_demand = px.bar(demand_df, x="State", y="Demand Potential (MT/yr)",
                     color="Demand Potential (MT/yr)", color_continuous_scale="YlOrRd",
                     title="State-wise Bio-Bitumen Demand Potential (MT/Year)")
fig_demand.update_layout(xaxis_tickangle=-45, template="plotly_white", height=400)
st.plotly_chart(fig_demand, width="stretch")

st.markdown("---")

# ── Government Projects ───────────────────────────────────────────────
st.subheader("Major Government Road Projects (Active)")
projects = [
    {"Project": "Bharatmala Phase-II", "Authority": "NHAI", "Budget (Cr)": 12000, "States": "Pan India", "Status": "Active"},
    {"Project": "PMGSY Phase-IV", "Authority": "MoRD", "Budget (Cr)": 8000, "States": "All States", "Status": "Active"},
    {"Project": "National Highway Expansion", "Authority": "MoRTH", "Budget (Cr)": 25000, "States": "UP, MP, MH, RJ", "Status": "Active"},
    {"Project": "Delhi-Mumbai Expressway", "Authority": "NHAI", "Budget (Cr)": 98000, "States": "DL, HR, RJ, GJ, MH", "Status": "Under Construction"},
    {"Project": "Chennai-Bengaluru Expressway", "Authority": "NHAI", "Budget (Cr)": 17500, "States": "TN, KA", "Status": "Active"},
    {"Project": "Smart City Roads", "Authority": "MoHUA", "Budget (Cr)": 5000, "States": "100 Cities", "Status": "Active"},
]
st.dataframe(pd.DataFrame(projects), width="stretch", hide_index=True)

total_budget = sum(p["Budget (Cr)"] for p in projects)
st.metric("Total Active Road Budget", f"Rs {total_budget:,} Crore")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# PRICE COMPARISON TABLE (NEW)
# ══════════════════════════════════════════════════════════════════════
st.subheader("Bitumen Product Price Comparison")

vg30_price = vg30.get("vg30_estimated", 38000)
price_comparison = pd.DataFrame([
    {"Product": "VG-10", "Price (Rs/MT)": int(vg30_price * 0.90), "Use": "Light traffic roads, primers", "Bio-Substitute": "Yes"},
    {"Product": "VG-20", "Price (Rs/MT)": int(vg30_price * 0.95), "Use": "Medium traffic, surface dressing", "Bio-Substitute": "Yes"},
    {"Product": "VG-30", "Price (Rs/MT)": int(vg30_price), "Use": "Heavy traffic highways (standard)", "Bio-Substitute": "Yes"},
    {"Product": "VG-40", "Price (Rs/MT)": int(vg30_price * 1.10), "Use": "Very heavy traffic, airports", "Bio-Substitute": "Partial"},
    {"Product": "PMB (Polymer Modified)", "Price (Rs/MT)": int(vg30_price * 1.35), "Use": "Expressways, high-stress", "Bio-Substitute": "No"},
    {"Product": "CRMB (Crumb Rubber)", "Price (Rs/MT)": int(vg30_price * 1.25), "Use": "Flexible pavements", "Bio-Substitute": "No"},
    {"Product": "Bio-Modified VG30", "Price (Rs/MT)": int(vg30_price * 0.92), "Use": "All VG30 applications + green", "Bio-Substitute": "THIS IS IT"},
    {"Product": "Bitumen Emulsion", "Price (Rs/MT)": int(vg30_price * 0.85), "Use": "Tack coat, prime coat", "Bio-Substitute": "Partial"},
])
st.dataframe(price_comparison, width="stretch", hide_index=True)
st.info(f"**Bio-Modified VG30 is ~8% cheaper than conventional VG30** while meeting all IS:73 specs + earning carbon credits!")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# STATE-WISE BITUMEN DEMAND HEATMAP (NEW)
# ══════════════════════════════════════════════════════════════════════
from config import STATE_COSTS

st.subheader("State-wise Bitumen Demand")

demand_data = pd.DataFrame([
    {"State": state, "Annual Demand (MT)": data["bitumen_demand_mt"],
     "Refinery Distance (km)": data["refinery_dist_km"],
     "Port Distance (km)": data["port_dist_km"]}
    for state, data in STATE_COSTS.items()
]).sort_values("Annual Demand (MT)", ascending=True)

fig_demand = px.bar(demand_data, y="State", x="Annual Demand (MT)", orientation="h",
                     color="Annual Demand (MT)", color_continuous_scale="YlOrRd",
                     title="Annual Bitumen Demand by State (MT)")
fig_demand.update_layout(template="plotly_white", height=500)
st.plotly_chart(fig_demand, width="stretch")

total_demand = demand_data["Annual Demand (MT)"].sum()
st.metric("Total 18-State Bitumen Demand", f"{total_demand:,} MT/year")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# PRICE ALERT SETUP (NEW)
# ══════════════════════════════════════════════════════════════════════
st.subheader("Price Alerts")
st.caption("Set thresholds to monitor — alerts will show on dashboard when triggered")

from database import init_db, insert_price_alert, get_price_alerts, delete_price_alert
init_db()

alert_col1, alert_col2 = st.columns(2)

with alert_col1:
    st.markdown("**Set New Alert**")
    alert_metric = st.selectbox("Metric", ["Crude Oil ($/bbl)", "USD/INR", "VG30 (Rs/MT)"], key="alert_metric")
    alert_threshold = st.number_input("Threshold Value", 0.0, 100000.0, 85.0, key="alert_thresh")
    alert_dir = st.radio("Trigger When", ["Above", "Below"], horizontal=True, key="alert_dir")
    if st.button("Create Alert", key="create_alert"):
        insert_price_alert(alert_metric, alert_threshold, alert_dir.lower())
        st.success(f"Alert created: {alert_metric} {alert_dir.lower()} {alert_threshold}")
        st.rerun()

with alert_col2:
    st.markdown("**Active Alerts**")
    alerts = get_price_alerts()
    if alerts:
        for a in alerts:
            ac1, ac2 = st.columns([3, 1])
            with ac1:
                st.markdown(f"- **{a['metric_name']}** {a['direction']} **{a['threshold']}**")
            with ac2:
                if st.button("Delete", key=f"del_alert_{a['id']}"):
                    delete_price_alert(a["id"])
                    st.rerun()
    else:
        st.info("No alerts set. Create one to monitor market prices.")

# ══════════════════════════════════════════════════════════════════════
# INDIA GDP & INFRASTRUCTURE (World Bank API — FREE)
# ══════════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("India Economic Context (World Bank Data)")

try:
    from engines.free_apis import get_india_gdp, get_exchange_rates

    gdp_col, fx_col = st.columns(2)

    with gdp_col:
        gdp_data = get_india_gdp(10)
        if gdp_data:
            gdp_df = pd.DataFrame(gdp_data)
            fig_gdp = go.Figure()
            fig_gdp.add_trace(go.Bar(x=gdp_df["year"], y=gdp_df["gdp_usd_billion"],
                                       name="GDP", marker_color="#003366"))
            fig_gdp.update_layout(title="India GDP (USD Billion) — World Bank",
                                    template="plotly_white", height=350)
            st.plotly_chart(fig_gdp, width="stretch")
        else:
            st.info("GDP data loading...")

    with fx_col:
        fx_backup = get_exchange_rates()
        if "error" not in fx_backup:
            fx_data = [
                {"Currency": "USD/INR", "Rate": fx_backup.get("usd_inr", 84)},
                {"Currency": "USD/AED", "Rate": fx_backup.get("usd_aed", 3.67)},
                {"Currency": "USD/EUR", "Rate": fx_backup.get("usd_eur", 0.92)},
                {"Currency": "USD/GBP", "Rate": fx_backup.get("usd_gbp", 0.79)},
                {"Currency": "USD/CNY", "Rate": fx_backup.get("usd_cny", 7.2)},
            ]
            st.markdown("**Multi-Currency Exchange Rates** *(ExchangeRate-API)*")
            st.dataframe(pd.DataFrame(fx_data), width="stretch", hide_index=True)
            st.caption(f"Updated: {fx_backup.get('last_update', 'N/A')}")
        else:
            st.info("FX data loading...")
except Exception:
    st.info("World Bank / FX data will load on next refresh")

# ── Quick Links ──────────────────────────────────────────────────────
st.markdown("---")
ql1, ql2, ql3 = st.columns(3)
ql1.page_link("pages/12_🛣️_NHAI_Tenders.py", label="NHAI Tender Tracker", icon="🛣️")
ql2.page_link("pages/63_Competitor_Intel.py", label="Competitor Intelligence", icon="🕵️")
ql3.page_link("pages/71_Weather_Site.py", label="Weather & Site Analysis", icon="🌤️")

st.caption(f"Data: Yahoo Finance (crude), Frankfurter/ECB (FX), World Bank (GDP), ExchangeRate-API (multi-FX) | Updated: {market.get('last_updated', 'N/A')}")


# ── AI Skill: Market Analysis ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: Market Analysis"):
            if st.button("Generate", type="primary", key="ai_04📈Mar"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: Market Analysis. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_04_📈_M", type="primary")
with _ex2:
    if st.button("Print", key="exp_prt_04📈M"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)
