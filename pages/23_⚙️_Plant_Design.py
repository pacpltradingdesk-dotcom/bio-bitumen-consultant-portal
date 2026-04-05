"""
Plant Design — Complete 15-Zone Layout with 82-Item BOQ
========================================================
Auto-scales everything from capacity: equipment, areas, costs, utilities.
All data flows from single source: state_manager.calculate_boq(tpd)
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.express as px
from state_manager import get_config, update_field, init_state, calculate_boq, format_inr

st.set_page_config(page_title="Plant Design", page_icon="⚙️", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass

st.title("Plant Design & Layout — Complete BOQ")
st.markdown("**15 zones | 82 items | Gate to Gate | Auto-scales with capacity**")
st.markdown("---")

# ── Capacity Selector ────────────────────────────────────────────────
st.subheader("Select Plant Capacity")
preset_options = [5, 10, 15, 20, 25, 30, 40, 50, 100]
col_cap, col_custom = st.columns(2)

with col_cap:
    current_idx = preset_options.index(int(cfg["capacity_tpd"])) if int(cfg["capacity_tpd"]) in preset_options else 3
    preset = st.selectbox("Standard Capacities (TPD)", preset_options, index=current_idx)
    if preset != cfg["capacity_tpd"]:
        update_field("capacity_tpd", float(preset))
        cfg = get_config()

with col_custom:
    custom = st.number_input("Or enter custom TPD", min_value=3.0, max_value=100.0,
                              value=float(cfg["capacity_tpd"]), step=5.0)
    if custom != cfg["capacity_tpd"]:
        update_field("capacity_tpd", custom)
        cfg = get_config()

tpd = cfg["capacity_tpd"]
st.markdown("---")

# ── Get Complete BOQ ─────────────────────────────────────────────────
boq = calculate_boq(tpd)
total_cost_lac = sum(i["amount_lac"] for i in boq)

# Group by category
categories = {}
for item in boq:
    cat = item["category"]
    if cat not in categories:
        categories[cat] = []
    categories[cat].append(item)

cat_totals = {cat: sum(i["amount_lac"] for i in items) for cat, items in categories.items()}

# ── KPIs ─────────────────────────────────────────────────────────────
k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("Capacity", f"{tpd:.0f} TPD")
k2.metric("Total Items", f"{len(boq)}")
k3.metric("Zones", f"{len(categories)}")
k4.metric("BOQ Total", format_inr(total_cost_lac * 1e5))
k5.metric("In Crores", f"₹ {total_cost_lac/100:.2f} Cr")

st.markdown("---")

# ── Tabs ─────────────────────────────────────────────────────────────
tab_zones, tab_boq, tab_layout, tab_process, tab_utils = st.tabs([
    "📋 Zone-wise Summary", "📊 Full 82-Item BOQ", "🗺️ Layout Visualization",
    "⚙️ Process Flow", "🔧 Utilities"
])

# ══════════════════════════════════════════════════════════════════════
# TAB: ZONE-WISE SUMMARY
# ══════════════════════════════════════════════════════════════════════
with tab_zones:
    st.subheader(f"15 Zones — {tpd:.0f} TPD Plant")

    zone_rows = []
    for cat in sorted(categories):
        items = categories[cat]
        zone_rows.append({
            "Zone": cat,
            "Items": len(items),
            "Cost (Lac)": f"₹ {cat_totals[cat]:.1f}",
            "% of Total": f"{cat_totals[cat]/total_cost_lac*100:.1f}%",
        })
    st.dataframe(pd.DataFrame(zone_rows), use_container_width=True, hide_index=True)

    # Zone cost pie chart
    fig = px.pie(pd.DataFrame(zone_rows), names="Zone", values=[cat_totals[c] for c in sorted(categories)],
                 title=f"Cost Distribution — {tpd:.0f} TPD (₹ {total_cost_lac/100:.2f} Cr)",
                 color_discrete_sequence=px.colors.qualitative.Set3)
    fig.update_traces(textposition="inside", textinfo="percent+label")
    fig.update_layout(height=500, template="plotly_white")
    st.plotly_chart(fig, use_container_width=True)

    # Show each zone expandable
    for cat in sorted(categories):
        items = categories[cat]
        with st.expander(f"{cat} — {len(items)} items — ₹ {cat_totals[cat]:.1f} Lac"):
            rows = []
            for i in items:
                rows.append({
                    "Item": i["item"],
                    "Specification": i["spec"],
                    "Qty": i["qty"],
                    "Unit": i["unit"],
                    "Rate (Lac)": f"₹ {i['rate_lac']:.1f}",
                    "Amount (Lac)": f"₹ {i['amount_lac']:.1f}",
                })
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

# ══════════════════════════════════════════════════════════════════════
# TAB: FULL 82-ITEM BOQ
# ══════════════════════════════════════════════════════════════════════
with tab_boq:
    st.subheader(f"Complete Bill of Quantities — {tpd:.0f} TPD ({len(boq)} items)")

    all_rows = []
    for idx, item in enumerate(boq, 1):
        all_rows.append({
            "#": idx,
            "Zone": item["category"],
            "Item": item["item"],
            "Specification": item["spec"],
            "Qty": item["qty"],
            "Unit": item["unit"],
            "Rate (₹ Lac)": round(item["rate_lac"], 1),
            "Amount (₹ Lac)": round(item["amount_lac"], 1),
        })
    st.dataframe(pd.DataFrame(all_rows), use_container_width=True, hide_index=True, height=600)

    st.markdown(f"""
    | | Amount |
    |---|---|
    | **Total BOQ Cost** | **₹ {total_cost_lac:.1f} Lac (₹ {total_cost_lac/100:.2f} Cr)** |
    | GST @ 18% (on machinery) | ₹ {total_cost_lac * 0.12:.1f} Lac |
    | Contingency @ 5% | ₹ {total_cost_lac * 0.05:.1f} Lac |
    | **Grand Total (incl. GST + Contingency)** | **₹ {total_cost_lac * 1.17:.1f} Lac (₹ {total_cost_lac * 1.17/100:.2f} Cr)** |
    """)

    # Export BOQ as Excel
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "BOQ"
        ws.cell(row=1, column=1, value=f"Bill of Quantities — {tpd:.0f} TPD Bio-Bitumen Plant")
        ws.cell(row=2, column=1, value=f"Total: Rs {total_cost_lac:.1f} Lac ({total_cost_lac/100:.2f} Cr)")
        headers = ["#", "Zone", "Item", "Specification", "Qty", "Unit", "Rate (Lac)", "Amount (Lac)"]
        hf = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        hfont = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.fill = hf
            c.font = hfont
        for row_idx, item in enumerate(all_rows, 5):
            for col, key in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col, value=item.get(key, ""))
        buf = io.BytesIO()
        wb.save(buf)
        _xl = buf.getvalue()
    except Exception:
        _xl = None
    if _xl:
        st.download_button(f"Download BOQ Excel ({len(boq)} items)",
                           _xl, f"BOQ_{tpd:.0f}TPD_{len(boq)}_items.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           type="primary", key="dl_boq_xl")

# ══════════════════════════════════════════════════════════════════════
# TAB: LAYOUT VISUALIZATION
# ══════════════════════════════════════════════════════════════════════
with tab_layout:
    st.subheader(f"Plant Layout — {tpd:.0f} TPD")

    # Area calculations based on capacity (15 zones)
    zone_areas = {
        "A. Gate & Security": round(max(500, tpd * 15), 0),
        "B. RM Receiving & Storage": round(tpd * 100, 0),
        "C. Pre-Processing": round(tpd * 40, 0),
        "D. Reactor Zone": round(tpd * 60, 0),
        "E. Oil Recovery": round(tpd * 35, 0),
        "F. Blending Section": round(tpd * 30, 0),
        "G. Storage Tanks": round(tpd * 50, 0),
        "H. Packing & Dispatch": round(tpd * 45, 0),
        "I. Electrical Substation": round(max(300, tpd * 15), 0),
        "J. Utility Area": round(tpd * 30, 0),
        "K. Laboratory": round(max(300, tpd * 12), 0),
        "L. Safety & Environmental": round(tpd * 20, 0),
        "M. Civil (Office/Canteen/Toilet)": round(max(600, tpd * 25), 0),
        "N. Parking & Roads": round(tpd * 40, 0),
        "O. Green Belt & Buffer": round(tpd * 30, 0),
    }
    total_sqft = sum(zone_areas.values())
    total_acres = total_sqft / 43560

    a1, a2, a3 = st.columns(3)
    a1.metric("Total Area", f"{total_sqft:,.0f} sq ft")
    a2.metric("In Acres", f"{total_acres:.2f} acres")
    a3.metric("In Sq Meters", f"{total_sqft * 0.093:.0f} sqm")

    area_df = pd.DataFrame([{"Zone": k, "Area (sq ft)": v, "% of Total": round(v/total_sqft*100, 1)}
                             for k, v in zone_areas.items()])
    c1, c2 = st.columns(2)
    with c1:
        st.dataframe(area_df, use_container_width=True, hide_index=True)
    with c2:
        fig2 = px.treemap(area_df, path=["Zone"], values="Area (sq ft)",
                          color="Area (sq ft)", color_continuous_scale="Blues",
                          title=f"Layout — {tpd:.0f} TPD ({total_sqft:,.0f} sq ft)")
        fig2.update_layout(height=500)
        st.plotly_chart(fig2, use_container_width=True)

# ══════════════════════════════════════════════════════════════════════
# TAB: PROCESS FLOW
# ══════════════════════════════════════════════════════════════════════
with tab_process:
    oil_y = cfg.get("bio_oil_yield_pct", 32)
    char_y = cfg.get("bio_char_yield_pct", 28)
    syn_y = cfg.get("syngas_yield_pct", 22)
    loss_y = cfg.get("process_loss_pct", 18)

    st.subheader(f"Process Flow — {tpd:.0f} TPD Feed")
    st.markdown(f"""
```
  GATE IN → WEIGHBRIDGE → UNLOADING RAMP
       ↓
  RM STORAGE (Open + Closed Shed)
       ↓
  SHREDDER → MAGNETIC SEPARATOR → DRYER → PELLETIZER
       ↓
  PYROLYSIS REACTOR ({int(tpd/max(1,int(tpd/10)))} MT each × {max(1,int(tpd/10))} units, 450-550°C)
       ├──→ BIO-OIL ({oil_y}%) → CONDENSER → OIL TANK → BLENDING with VG30
       │                                                    ↓
       │                                              HIGH SHEAR MIXER → COLLOID MILL
       │                                                    ↓
       │                                              FINISHED BITUMEN TANK (heated)
       │                                                    ↓
       │                                              DRUM FILLING / TANKER LOADING
       │
       ├──→ BIO-CHAR ({char_y}%) → COOLER → SILO → BAGGING (50kg HDPE)
       │
       └──→ SYNGAS ({syn_y}%) → BURNER (captive fuel for reactor heating)

  Loss: {loss_y}% (moisture, flue gas, ash)

  DISPATCH BAY → WEIGHBRIDGE → GATE OUT
```
    """)

    p1, p2, p3, p4 = st.columns(4)
    p1.metric("Bio-Oil Output", f"{tpd * oil_y/100:.1f} T/day")
    p2.metric("Bio-Char Output", f"{tpd * char_y/100:.1f} T/day")
    p3.metric("Syngas", f"{tpd * syn_y/100:.1f} T/day")
    p4.metric("Process Loss", f"{tpd * loss_y/100:.1f} T/day")

# ══════════════════════════════════════════════════════════════════════
# TAB: UTILITIES
# ══════════════════════════════════════════════════════════════════════
with tab_utils:
    st.subheader("Utility Requirements")
    u1, u2, u3, u4 = st.columns(4)
    u1.metric("Power", f"{cfg.get('power_kw', int(tpd*5)):.0f} kW")
    u2.metric("Water", f"{max(5, int(tpd * 1.5)):,} KLD")
    u3.metric("Compressed Air", f"{max(50, int(tpd * 5)):,} CFM")
    u4.metric("DG Backup", f"{max(50, int(tpd*5))} kVA")

    st.markdown("---")
    st.subheader("Manpower Requirement")
    staff = cfg.get("staff", max(15, int(tpd * 1.2)))
    staff_data = [
        {"Role": "Plant Manager", "Count": 1, "Monthly (₹)": "60,000-80,000"},
        {"Role": "Shift Supervisor", "Count": max(2, int(tpd/10)), "Monthly (₹)": "30,000-40,000"},
        {"Role": "Reactor Operators", "Count": max(4, int(tpd * 0.3)), "Monthly (₹)": "18,000-25,000"},
        {"Role": "Lab Technician", "Count": max(1, int(tpd/20)), "Monthly (₹)": "20,000-30,000"},
        {"Role": "Electrician", "Count": max(1, int(tpd/20)), "Monthly (₹)": "18,000-25,000"},
        {"Role": "Fitter/Mechanic", "Count": max(1, int(tpd/15)), "Monthly (₹)": "18,000-25,000"},
        {"Role": "Helpers/Labour", "Count": max(6, int(tpd * 0.5)), "Monthly (₹)": "12,000-15,000"},
        {"Role": "Security Guard", "Count": max(2, int(tpd/10)), "Monthly (₹)": "12,000-15,000"},
        {"Role": "Driver (Loader)", "Count": max(1, int(tpd/20)), "Monthly (₹)": "15,000-18,000"},
        {"Role": "Office Staff (Accts/Admin)", "Count": max(2, int(tpd/10)), "Monthly (₹)": "18,000-25,000"},
    ]
    total_staff = sum(s["Count"] for s in staff_data)
    st.dataframe(pd.DataFrame(staff_data), use_container_width=True, hide_index=True)
    st.metric("Total Staff", f"{total_staff} persons")

# ── AI Recommendations ──────────────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("AI: Equipment & Layout Recommendations"):
            if st.button("Generate AI Recommendations", type="primary", key="ai_plant_rec"):
                with st.spinner("AI analyzing plant design..."):
                    prompt = (f"For a {tpd:.0f} TPD bio-bitumen plant in {cfg.get('state', 'Maharashtra')}:\n"
                              f"Investment: Rs {cfg.get('investment_cr', 8):.2f} Cr | BOQ: {len(boq)} items across 15 zones\n"
                              f"Total BOQ cost: Rs {total_cost_lac:.1f} Lac\n"
                              f"Top 3 cost zones: {', '.join(sorted(cat_totals, key=cat_totals.get, reverse=True)[:3])}\n"
                              f"Recommend: 1) Equipment optimization, 2) Layout improvements, "
                              f"3) Cost reduction opportunities, 4) Safety compliance gaps")
                    result, provider = ask_ai(prompt, "Senior plant design engineer for bio-bitumen.", 1200)
                if result:
                    st.markdown(result)
except Exception:
    pass

st.markdown("---")
st.caption("All values auto-update when capacity changes. BOQ feeds into Financial Model and DPR Generator.")
