"""
Training & Knowledge Base — SOPs, Training Modules, FAQs
==========================================================
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
from state_manager import init_state, get_config
from config import COMPANY, TRAINING_MODULES, FOUR_STAGES

st.set_page_config(page_title="Training & Knowledge Base", page_icon="📚", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass


st.title("Training & Knowledge Base")
st.markdown("**8 Training Modules | SOPs | FAQs | Plant Operations Guide**")
st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# TRAINING MODULE OVERVIEW
# ══════════════════════════════════════════════════════════════════════
st.subheader("Training Modules")

total_hrs = sum(m["duration_hrs"] for m in TRAINING_MODULES)
tm1, tm2, tm3 = st.columns(3)
tm1.metric("Total Modules", len(TRAINING_MODULES))
tm2.metric("Total Hours", f"{total_hrs} hrs")
tm3.metric("Categories", len(set(m["category"] for m in TRAINING_MODULES)))

st.markdown("---")

# ── Module Cards ─────────────────────────────────────────────────────
cat_filter = st.selectbox("Filter by Category",
    ["All"] + sorted(set(m["category"] for m in TRAINING_MODULES)), key="train_cat")

filtered = TRAINING_MODULES if cat_filter == "All" else [m for m in TRAINING_MODULES if m["category"] == cat_filter]

cat_colors = {
    "Safety": "#CC3333", "Operations": "#003366", "Quality": "#006699",
    "Supply Chain": "#FF8800", "Commercial": "#00AA44", "Regulatory": "#AA3366",
    "Finance": "#009966",
}

for module in filtered:
    color = cat_colors.get(module["category"], "#666666")
    with st.expander(f"{'🔴' if module['category']=='Safety' else '🔵' if module['category']=='Operations' else '🟢' if module['category']=='Quality' else '🟡'} {module['module']} — {module['duration_hrs']} hrs | {module['audience']}"):
        mc1, mc2 = st.columns([2, 1])
        with mc1:
            st.markdown(f"**Category:** {module['category']}")
            st.markdown(f"**Duration:** {module['duration_hrs']} hours")
            st.markdown(f"**Target Audience:** {module['audience']}")
            st.markdown("**Topics Covered:**")
            for topic in module["topics"]:
                st.markdown(f"- {topic}")
        with mc2:
            st.markdown(f"""
            <div style="background: {color}15; border: 1px solid {color}; border-radius: 10px;
                        padding: 15px; text-align: center;">
                <h3 style="color: {color}; margin: 0;">{module['duration_hrs']}h</h3>
                <p style="margin: 0;">{module['category']}</p>
            </div>
            """, unsafe_allow_html=True)

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# STANDARD OPERATING PROCEDURES (SOPs)
# ══════════════════════════════════════════════════════════════════════
st.subheader("Standard Operating Procedures (SOPs)")

sop_tab1, sop_tab2, sop_tab3, sop_tab4 = st.tabs([
    "Reactor Startup", "Quality Testing", "Emergency Shutdown", "Daily Checklist"
])

with sop_tab1:
    st.markdown("""
    ### SOP-001: Pyrolysis Reactor Startup Procedure

    **Pre-Startup Checks:**
    1. Verify all safety interlocks are functional
    2. Check reactor temperature sensors (min 3 points)
    3. Confirm nitrogen purge line pressure (2-3 bar)
    4. Verify condenser cooling water flow
    5. Check gas flare system is operational

    **Startup Sequence:**
    1. Start cooling water circulation → confirm flow at 50 KL/hr
    2. Begin nitrogen purge of reactor → hold 15 minutes
    3. Start thermic fluid heater → target 200°C initial
    4. Begin biomass feed at 30% rate
    5. Ramp temperature: 200°C → 350°C over 45 minutes
    6. Increase feed to 70% at 350°C
    7. Ramp to operating temp (450-550°C) over 30 minutes
    8. Increase feed to 100% at operating temperature
    9. Monitor bio-oil collection in condenser
    10. Confirm syngas flow to flare/fuel system

    **Critical Parameters:**
    - Reactor pressure: < 0.5 bar (atmospheric)
    - Temperature: 450-550°C (never exceed 600°C)
    - Feed rate: As per capacity chart
    - Condenser outlet temp: < 40°C
    """)

with sop_tab2:
    st.markdown("""
    ### SOP-002: Bio-Bitumen Quality Testing

    **Daily Tests (Every Batch):**
    1. **Penetration Test** (IS 1203) — Target: 50-70 (0.1mm)
       - Sample at 25°C, 100g, 5 sec
    2. **Softening Point** (IS 1205) — Target: 47-58°C
       - Ring & Ball method
    3. **Specific Gravity** (IS 1202) — Target: min 0.99

    **Weekly Tests:**
    4. **Ductility** (IS 1208) — Target: min 40 cm at 25°C
    5. **Flash Point** (IS 1209) — Target: min 220°C
    6. **Viscosity at 60°C** (IS 1206) — Target: 2400-3600 Poise

    **Monthly Tests:**
    7. **Solubility in TCE** (IS 1216) — Target: min 99%
    8. **Loss on Heating** (IS 9382) — Target: max 1.0%

    **Record Keeping:**
    - Log all results in Quality Register
    - Retain samples for 90 days
    - Flag any out-of-spec results immediately
    """)

with sop_tab3:
    st.markdown("""
    ### SOP-003: Emergency Shutdown Procedure

    **Level 1 — Controlled Shutdown (Equipment Malfunction):**
    1. Stop biomass feed immediately
    2. Reduce thermic fluid heater to minimum
    3. Maintain nitrogen purge
    4. Allow reactor to cool below 200°C (2-3 hours)
    5. Close all product outlets
    6. Notify Plant Manager

    **Level 2 — Emergency Shutdown (Fire/Leak/Injury):**
    1. Press EMERGENCY STOP button (red mushroom switch)
    2. Activate nitrogen flooding system
    3. Close all manual valves
    4. Sound fire alarm
    5. Evacuate to assembly point
    6. Call: Fire Dept, Ambulance, Plant Manager
    7. DO NOT re-enter until all-clear given

    **Emergency Contacts:**
    - Fire Department: 101
    - Ambulance: 108
    - Plant Manager: [Posted at control room]
    - PESO Inspector: [Posted at gate]
    """)

with sop_tab4:
    st.markdown("""
    ### SOP-004: Daily Operations Checklist

    **Morning Shift Start (6:00 AM):**
    - [ ] Check reactor temperature log (overnight)
    - [ ] Verify cooling water levels
    - [ ] Inspect safety equipment (extinguishers, PPE)
    - [ ] Check biomass inventory (min 3-day stock)
    - [ ] Review previous shift handover notes
    - [ ] Test gas detector calibration
    - [ ] Confirm DG set fuel level

    **Production Monitoring (Every 2 Hours):**
    - [ ] Record reactor temperature (3 points)
    - [ ] Record bio-oil collection rate
    - [ ] Check condenser performance
    - [ ] Monitor stack emissions (visual)
    - [ ] Log power consumption

    **End of Day (6:00 PM):**
    - [ ] Record daily production figures
    - [ ] Update quality test results
    - [ ] Report any maintenance issues
    - [ ] Complete shift handover log
    - [ ] Check security and lock hazardous areas
    """)

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# FAQ SECTION
# ══════════════════════════════════════════════════════════════════════
st.subheader("Frequently Asked Questions (FAQs)")

faqs = [
    ("What is bio-modified bitumen?", "Bio-modified bitumen is conventional VG-30 bitumen blended with 15-30% bio-oil derived from pyrolysis of agricultural waste. It meets all IS:73 specifications and is approved by CSIR-CRRI."),
    ("How is it different from polymer-modified bitumen (PMB)?", "PMB uses synthetic polymers (SBS, EVA), while bio-bitumen uses natural bio-oil from agro-waste. Bio-bitumen is cheaper, greener, and earns carbon credits. PMB is still needed for extreme-stress applications."),
    ("What temperature does pyrolysis operate at?", "450-550°C in an oxygen-free reactor. Below 400°C gives low yield. Above 600°C risks equipment damage and produces more gas than oil."),
    ("How much bio-oil does 1 MT of biomass produce?", "Approximately 40% bio-oil, 30% biochar, 25% syngas, and 5% loss. So 1 MT biomass → 400 kg bio-oil → enough to modify ~1.5-2 MT of bitumen."),
    ("Is CSIR-CRRI license mandatory?", "Not mandatory for production, but strongly recommended for market credibility. NHAI/PWD prefer CSIR-validated products. License fee: negotiable with CSIR."),
    ("What is the minimum investment?", "Rs 1.5 Crore for a 5 MT/day plant. This includes basic reactor, blending unit, and testing lab. For commercial viability, 20 TPD (Rs 8 Cr) is recommended."),
    ("Can I use any agricultural waste?", "Rice straw, wheat straw, sugarcane bagasse, cotton stalk, and groundnut shell work best. Moisture content should be <15%. Pelletization improves yield and consistency."),
    ("How long does plant setup take?", "12-18 months from scratch (new investor). 30-90 days for existing bitumen/pyrolysis operators adding bio-bitumen capability."),
    ("What are the main risks?", "Key risks: crude oil price drops (reduces cost advantage), regulatory delays (CTE/PESO), feedstock quality variation, and initial market adoption. All are manageable with proper planning."),
    ("Who are the buyers?", "NHAI contractors (2,758 in our network), state PWD departments, municipal corporations, and private road builders. We provide full buyer linkage support."),
]

for question, answer in faqs:
    with st.expander(f"❓ {question}"):
        st.markdown(answer)

st.markdown("---")

# ── 4 Stages Quick Reference ────────────────────────────────────────
st.subheader("4 Stages of Bio-Bitumen — Quick Reference")
for stage in FOUR_STAGES:
    with st.expander(f"Stage {stage['stage']}: {stage['name']}"):
        st.markdown(f"**Description:** {stage['description']}")
        st.markdown(f"**CAPEX:** {stage['capex']} | **Manpower:** {stage['manpower']}")
        if "space" in stage:
            st.markdown(f"**Space Required:** {stage['space']}")

st.markdown("---")
st.caption(f"{COMPANY['name']} | Training & Knowledge Management")


# ── AI Skill: Training Schedule ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: Training Schedule"):
            if st.button("Generate", type="primary", key="ai_69Train"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: Training Schedule. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    if st.button("Download Excel", type="primary", key="exp_xl_69Tra"):
        try:
            import io
            from openpyxl import Workbook
            _wb = Workbook()
            _ws = _wb.active
            _ws.title = "Export"
            _ws.cell(row=1, column=1, value="Bio Bitumen Export")
            _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
            _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
            _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
            _buf = io.BytesIO()
            _wb.save(_buf)
            _buf.seek(0)
            st.download_button("Download", _buf.getvalue(), "export.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_xl_69Tra")
        except Exception as _e:
            st.error(f"Export failed: {_e}")
with _ex2:
    if st.button("Print", key="exp_prt_69Tra"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)
