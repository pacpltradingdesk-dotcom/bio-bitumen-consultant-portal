"""
Buyers Network — Contractors, Govt tenders, Bulk buyers for Bio-Bitumen
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
from state_manager import init_state, get_config
import pandas as pd
import plotly.express as px
from config import STATES

st.set_page_config(page_title="Buyers Network", page_icon="🤝", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass

st.title("Buyers & Customer Network")
st.markdown("**Contractors | Government tenders | Bulk buyers | NHAI/PWD demand**")
st.markdown("---")

tab_contractors, tab_tenders, tab_bulk = st.tabs(["Road Contractors", "Government Tenders", "Bulk Buyers"])

with tab_contractors:
    st.subheader("Major Road Construction Contractors (Pan India)")
    contractors = [
        {"Company": "Larsen & Toubro (L&T)", "Type": "EPC Major", "Annual Demand (MT)": 50000, "States": "Pan India", "Contact": "procurement@lnt.com"},
        {"Company": "IRB Infrastructure", "Type": "BOT Developer", "Annual Demand (MT)": 30000, "States": "MH, GJ, RJ, KA", "Contact": "procurement@irb.co.in"},
        {"Company": "Dilip Buildcon", "Type": "Highway Contractor", "Annual Demand (MT)": 25000, "States": "MP, GJ, RJ, UP", "Contact": "purchase@dilipbuildcon.com"},
        {"Company": "Ashoka Buildcon", "Type": "BOT + EPC", "Annual Demand (MT)": 20000, "States": "MH, MP, KA", "Contact": "supply@ashokabuildcon.com"},
        {"Company": "PNC Infratech", "Type": "Highway EPC", "Annual Demand (MT)": 18000, "States": "UP, MP, RJ, HR", "Contact": "materials@pncinfra.com"},
        {"Company": "Sadbhav Engineering", "Type": "Road + Mining", "Annual Demand (MT)": 15000, "States": "GJ, MH, RJ", "Contact": "procurement@sadbhav.com"},
        {"Company": "KNR Constructions", "Type": "Highway EPC", "Annual Demand (MT)": 12000, "States": "AP, TS, KA", "Contact": "purchase@knrcl.com"},
        {"Company": "G R Infraprojects", "Type": "Highway Builder", "Annual Demand (MT)": 15000, "States": "RJ, UP, MP", "Contact": "supply@grinfra.com"},
        {"Company": "HG Infra Engineering", "Type": "Road EPC", "Annual Demand (MT)": 10000, "States": "RJ, UP, GJ", "Contact": "procurement@hginfra.com"},
        {"Company": "Gayatri Projects", "Type": "EPC", "Annual Demand (MT)": 8000, "States": "AP, TS, TN", "Contact": "purchase@gayatri.co.in"},
        {"Company": "State PWD Contractors", "Type": "Various", "Annual Demand (MT)": 100000, "States": "All States", "Contact": "Via state PWD tenders"},
        {"Company": "Border Roads Organization", "Type": "Defence", "Annual Demand (MT)": 20000, "States": "J&K, NE, HP, UK", "Contact": "bro.procurement@gov.in"},
    ]
    st.dataframe(pd.DataFrame(contractors), width="stretch", hide_index=True)

    # Demand by company chart
    c_df = pd.DataFrame(contractors)
    fig = px.bar(c_df.sort_values("Annual Demand (MT)", ascending=True),
                  x="Annual Demand (MT)", y="Company", orientation="h",
                  title="Annual Bitumen Demand by Contractor",
                  color="Annual Demand (MT)", color_continuous_scale="Blues")
    fig.update_layout(template="plotly_white", height=450)
    st.plotly_chart(fig, width="stretch")

with tab_tenders:
    st.subheader("Active Government Road Tenders")
    tenders = [
        {"Authority": "NHAI", "Project": "NH-44 Widening (Delhi-Agra)", "Budget (Cr)": 3500, "Bitumen Req (MT)": 8000, "State": "UP/HR", "Status": "Active"},
        {"Authority": "NHAI", "Project": "Bharatmala Phase-II Pkg 12", "Budget (Cr)": 2800, "Bitumen Req (MT)": 6000, "State": "MP", "Status": "Active"},
        {"Authority": "State PWD", "Project": "Maharashtra Rural Roads", "Budget (Cr)": 1500, "Bitumen Req (MT)": 4000, "State": "Maharashtra", "Status": "Bidding"},
        {"Authority": "State PWD", "Project": "UP State Highways", "Budget (Cr)": 2000, "Bitumen Req (MT)": 5000, "State": "UP", "Status": "Active"},
        {"Authority": "NHAI", "Project": "Chennai-Bengaluru Expressway", "Budget (Cr)": 5000, "Bitumen Req (MT)": 12000, "State": "TN/KA", "Status": "Active"},
        {"Authority": "MoRD", "Project": "PMGSY Phase-IV", "Budget (Cr)": 8000, "Bitumen Req (MT)": 20000, "State": "Pan India", "Status": "Ongoing"},
        {"Authority": "NHAI", "Project": "Delhi-Mumbai Expressway (Remaining)", "Budget (Cr)": 4000, "Bitumen Req (MT)": 10000, "State": "RJ/GJ/MH", "Status": "Active"},
    ]
    st.dataframe(pd.DataFrame(tenders), width="stretch", hide_index=True)

    total_demand = sum(t["Bitumen Req (MT)"] for t in tenders)
    st.metric("Total Tender Demand", f"{total_demand:,} MT", f"Rs {sum(t['Budget (Cr)'] for t in tenders):,} Cr budget")

with tab_bulk:
    st.subheader("Bulk Buyer Categories")
    buyers = [
        {"Category": "NHAI Contractors", "Annual Volume": "2-3 Lakh MT", "Price Sensitivity": "Medium", "Payment": "30-45 days", "Bio-Bitumen Interest": "High (green mandate)"},
        {"Category": "State PWD", "Annual Volume": "1-2 Lakh MT per state", "Price Sensitivity": "High", "Payment": "45-60 days", "Bio-Bitumen Interest": "Medium"},
        {"Category": "Municipal Corporations", "Annual Volume": "5,000-20,000 MT", "Price Sensitivity": "Medium", "Payment": "30 days", "Bio-Bitumen Interest": "High (smart city)"},
        {"Category": "Private Developers", "Annual Volume": "2,000-10,000 MT", "Price Sensitivity": "Low", "Payment": "15-30 days", "Bio-Bitumen Interest": "Medium"},
        {"Category": "Defence (BRO)", "Annual Volume": "20,000+ MT", "Price Sensitivity": "Low", "Payment": "30 days", "Bio-Bitumen Interest": "High"},
        {"Category": "Airport Authority", "Annual Volume": "5,000-15,000 MT", "Price Sensitivity": "Low", "Payment": "30 days", "Bio-Bitumen Interest": "High (LEED)"},
    ]
    st.dataframe(pd.DataFrame(buyers), width="stretch", hide_index=True)

    st.markdown("""
    ### Why Bio-Bitumen Has Strong Demand:
    - **Government Green Mandate:** NHAI/MoRTH increasingly requiring eco-friendly road materials
    - **Carbon Credit Potential:** Bio-bitumen qualifies for carbon credits under UNFCCC
    - **Smart City Program:** 100 smart cities mandate sustainable construction materials
    - **LEED Certification:** Bio-bitumen helps contractors achieve green building ratings
    - **Cost Parity:** At Rs 35,000/MT, bio-bitumen is competitive with VG30 (Rs 48,000-52,000/MT)
    """)

st.caption("Sources: NHAI tender portal, GEM (Government e-Marketplace), State PWD websites, Industry reports")


# ── AI Skill: Buyer Pitch Script ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: Buyer Pitch Script"):
            if st.button("Generate", type="primary", key="ai_40Buyer"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: Buyer Pitch Script. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    if st.button("Download Excel", type="primary", key="exp_xl_40Buy"):
        try:
            import io
            from openpyxl import Workbook
            _wb = Workbook()
            _ws = _wb.active
            _ws.title = "Export"
            _ws.cell(row=1, column=1, value="Bio Bitumen Export")
            _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
            _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
            _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
            _buf = io.BytesIO()
            _wb.save(_buf)
            _buf.seek(0)
            st.download_button("Download", _buf.getvalue(), "export.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_xl_40Buy")
        except Exception as _e:
            st.error(f"Export failed: {_e}")
with _ex2:
    if st.button("Print", key="exp_prt_40Buy"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)
