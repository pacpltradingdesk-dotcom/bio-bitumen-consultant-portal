"""
CONSULTANT PRESENTER — Slider-Based Interactive Presentation System
=====================================================================
14 Slides | Client Input → Auto-Update ALL | One-Click Document Generation
Used by consultant (Prince) in client meetings, bank discussions, investor pitches.

THIS REPLACES PPT — All data is LIVE, all documents are READY.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import io, datetime
from state_manager import get_config, update_fields, init_state
from config import (COMPANY, STATES, STATE_SCORES, LOCATION_WEIGHTS, STATE_COSTS,
                    NHAI_TENDERS, COMPETITORS, PPS_SWOT, RISK_REGISTRY,
                    ENVIRONMENTAL_FACTORS, WHY_NOW, FOUR_STAGES, LICENSE_TYPES,
                    INDUSTRY_NETWORK, KEY_CREDENTIALS, TARGET_AUDIENCES)

st.set_page_config(page_title="Consultant Presenter", page_icon="🎯", layout="wide")
init_state()
cfg = get_config()
# Fix metric truncation
try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass


TOTAL_SLIDES = 57  # 0-56

# ══════════════════════════════════════════════════════════════════════
# SLIDE STATE + DISPLAY MODE + PRINT
# ══════════════════════════════════════════════════════════════════════
if "slide" not in st.session_state:
    st.session_state["slide"] = 0
if "display_mode" not in st.session_state:
    st.session_state["display_mode"] = False

slide = st.session_state["slide"]
display_mode = st.session_state["display_mode"]

# Sidebar controls
st.sidebar.markdown("### Presentation Controls")
if st.sidebar.toggle("Meeting Display Mode", value=display_mode, key="dm_toggle"):
    st.session_state["display_mode"] = True
    display_mode = True
else:
    st.session_state["display_mode"] = False
    display_mode = False

if display_mode:
    st.sidebar.success("Display Mode ON — Clean presentation view")
    # Hide sidebar content in display mode via CSS
    st.markdown("""<style>
    [data-testid="stSidebar"] {width: 0px !important; min-width: 0px !important;}
    [data-testid="stSidebarContent"] {display: none !important;}
    header {display: none !important;}
    </style>""", unsafe_allow_html=True)

# Print button (always available)
st.sidebar.markdown("---")
if st.sidebar.button("🖨️ Print Current Slide", key="print_btn"):
    import streamlit.components.v1 as _stc; _stc.html('<script>window.print();</script>', height=0)

# ── Slide Title Bar ──────────────────────────────────────────────────
SLIDE_TITLES = [
    # Section A: Client & Setup (0-4)
    "Client & Project Setup",           # 0
    "The Opportunity",                   # 1
    "Market Analysis",                   # 2
    "Raw Material & Biomass",            # 3
    "Location & Feasibility",            # 4
    # Section B: Technology (5-9)
    "Technology & Process",              # 5
    "Process Flow & Plant Sections",     # 6
    "Plant Layout & Drawings",           # 7
    "Pyrolysis — Heart of the Plant",    # 8
    "Product Quality & IS Standards",    # 9
    # Section C: Equipment & BOQ (10-14)
    "Complete Equipment List (82 Items)", # 10
    "Zone A-D: Gate to Reactor",         # 11
    "Zone E-H: Oil to Dispatch",         # 12
    "Zone I-O: Utilities to Workshop",   # 13
    "BOQ Summary & Investment Split",    # 14
    # Section D: Costing & Revenue (15-21)
    "Raw Material Costing (6 Feedstocks)", # 15
    "Conventional Bitumen Cost",          # 16
    "Landing & Logistics Cost",           # 17
    "Production Cost Breakdown",          # 18
    "Complete Cost Sheet (10 Heads)",     # 19
    "Revenue — 7 Product Streams",        # 20
    "Finished Goods & Buyer Network",     # 21
    # Section E: Financial Model (22-30)
    "Financial Model — Master",           # 22
    "ROI & Profitability",               # 23
    "7-Year P&L Projection",             # 24
    "Cash Flow — 5 Year",                # 25
    "Break-Even Analysis",               # 26
    "Working Capital Requirement",        # 27
    "Loan EMI & Debt Service",           # 28
    "Sensitivity Analysis",              # 29
    "State-wise Profitability",          # 30
    # Section F: Compliance (31-35)
    "Compliance & Licenses (25 Required)", # 31
    "Environmental Clearances",            # 32
    "Subsidy & Government Support",        # 33
    "NHAI / MoRTH Green Mandate",          # 34
    "Risk Analysis",                       # 35
    # Section G: Implementation (36-41)
    "Project Timeline (10 Phases)",        # 36
    "Manpower & Organization Chart",       # 37
    "Safety & Fire Protection",            # 38
    "Quality Control & Lab Setup",         # 39
    "Capacity Comparison (5-50 TPD)",      # 40
    "3-Process Model Comparison",          # 41
    # Section H: Why PPS Anantams (42-48)
    "About PPS Anantams Corporation",      # 42
    "Our 25 Years of Experience",          # 43
    "Our Role — Complete Turnkey Solution", # 44
    "What We Deliver (Documents & Support)", # 45
    "Our Competitive Advantage (SWOT)",     # 46
    "Client Success Stories",               # 47
    "Competitor Comparison",                # 48
    # Section I: Convince & Close (49-56)
    "With Us vs Without Us",                # 49 — NEW
    "ROI Calculator — Play with Numbers",   # 50 — NEW
    "Project Readiness Checklist",          # 51 — NEW
    "Consulting Fee & Value Proposition",   # 52 — NEW
    "Timeline with Deliverables",           # 53 — NEW
    "Investment Summary & Next Steps",      # 54
    "Contact & Appointment",                # 55
    "Thank You & Call to Action",           # 56
]

# Header with slide counter
client_name = cfg.get("client_name", "")
header_text = f"<strong>{COMPANY['trade_name']}</strong> | Presenting to: <strong>{client_name}</strong>" if client_name else f"<strong>{COMPANY['trade_name']}</strong> — Consultant Presenter"

st.markdown(f"""
<div style="background: #003366; padding: 10px 20px; border-radius: 10px; margin-bottom: 10px;
            display: flex; justify-content: space-between; align-items: center; color: white;">
    <span style="font-size: 1.1em;">{header_text}</span>
    <span style="background: white; color: #003366; padding: 4px 15px; border-radius: 15px; font-weight: bold;">
        Slide {slide + 1} / {TOTAL_SLIDES}
    </span>
</div>
""", unsafe_allow_html=True)

# Slide progress dots
dot_html = ""
for i in range(TOTAL_SLIDES):
    color = "#003366" if i == slide else ("#006699" if i < slide else "#cccccc")
    dot_html += f'<span style="display:inline-block; width:12px; height:12px; border-radius:50%; background:{color}; margin:0 3px;"></span>'
st.markdown(f'<div style="text-align:center; margin-bottom:10px;">{dot_html}</div>', unsafe_allow_html=True)

st.markdown(f"## {SLIDE_TITLES[slide]}")
st.markdown("---")


# ══════════════════════════════════════════════════════════════════════
# SLIDE 0: CLIENT & PROJECT INPUT
# ══════════════════════════════════════════════════════════════════════
if slide == 0:
    st.markdown("### Enter client details — ALL slides will auto-update")

    c1, c2 = st.columns(2)
    with c1:
        inp_name = st.text_input("Client Name *", value=cfg.get("client_name", ""), key="p_name")
        inp_company = st.text_input("Company", value=cfg.get("client_company", ""), key="p_company")
        inp_state = st.selectbox("State *", STATES, index=STATES.index(cfg["state"]) if cfg["state"] in STATES else 0, key="p_state")
        inp_city = st.text_input("City", value=cfg.get("location", ""), key="p_city")
        inp_address = st.text_area("Site Address", value=cfg.get("site_address", ""), height=80, key="p_addr")
    with c2:
        cap_options = [5, 10, 15, 20, 25, 30, 40, 50]
        cur_cap = int(cfg["capacity_tpd"])
        cap_idx = cap_options.index(cur_cap) if cur_cap in cap_options else 3
        inp_cap = st.selectbox("Plant Capacity (TPD) *", cap_options, index=cap_idx, key="p_cap")
        inp_budget = st.number_input("Budget (Rs Crore)", 0.5, 50.0, float(cfg.get("investment_cr", 8)), 0.5, key="p_budget")
        inp_biomass = st.selectbox("Primary Biomass", ["Rice Straw", "Wheat Straw", "Sugarcane Bagasse",
                                    "Cotton Stalk", "Groundnut Shell", "Mixed"], key="p_biomass")
        inp_phone = st.text_input("Phone", value=cfg.get("client_phone", ""), key="p_phone")
        inp_project = st.text_input("Project Name", value=cfg.get("project_name", ""),
                                     placeholder=f"Bio-Bitumen Plant — {inp_city}", key="p_project")

    # Validation check
    missing = []
    if not inp_name: missing.append("Client Name")
    if not inp_city: missing.append("City")
    if not inp_state: missing.append("State")

    if missing:
        st.warning(f"Please fill required fields: **{', '.join(missing)}**")
    st.caption("You can browse slides with Next/Previous below, or fill details and click SAVE to start")

    col_save1, col_save2 = st.columns(2)
    with col_save1:
        if st.button("SAVE & START PRESENTATION", type="primary", key="save_client"):
            if not inp_name:
                st.error("Client Name is required! Please fill it above.")
                st.stop()
            update_fields({
                "client_name": inp_name, "client_company": inp_company,
                "state": inp_state, "location": inp_city, "site_address": inp_address,
                "capacity_tpd": float(inp_cap), "client_phone": inp_phone,
                "project_name": inp_project or f"Bio-Bitumen Plant — {inp_city}",
                "biomass_source": inp_biomass,
            })
            st.session_state["slide"] = 1
            st.rerun()
    with col_save2:
        if st.button("SAVE ONLY (stay on this slide)", key="save_only"):
            update_fields({
                "client_name": inp_name, "client_company": inp_company,
                "state": inp_state, "location": inp_city, "site_address": inp_address,
                "capacity_tpd": float(inp_cap), "client_phone": inp_phone,
                "project_name": inp_project or f"Bio-Bitumen Plant — {inp_city}",
                "biomass_source": inp_biomass,
            })
            st.success("Saved! All slides updated with your data.")

# ══════════════════════════════════════════════════════════════════════
# SLIDE 1: THE OPPORTUNITY
# ══════════════════════════════════════════════════════════════════════
elif slide == 1:
    st.markdown(f"### Why Bio-Bitumen NOW — The Rs 25,000 Crore Opportunity")

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("India Bitumen Market", "Rs 25,000+ Cr/yr")
    m2.metric("Import Dependency", f"{ENVIRONMENTAL_FACTORS['india_import_dependency_pct']}%")
    m3.metric("Plants Needed", "130-216 in 5-7 yrs")
    m4.metric("NHAI Green Mandate", "15% bio-bitumen by 2030")

    for i, point in enumerate(WHY_NOW, 1):
        st.markdown(f"**{i}.** {point}")

    st.markdown("---")
    st.markdown(f"""
    ### Why {COMPANY['trade_name']}?
    - **{COMPANY['years_experience']} years** bitumen industry experience
    - **{COMPANY['plants_built']} plants** built across India
    - **{COMPANY['industry_contacts']:,} contacts** — contractors, traders, importers
    - **International VG-30 supply** — 2.4 Lakh MT/yr (Getka USA-Iraq contract)
    """)

# ══════════════════════════════════════════════════════════════════════
# SLIDE 2: MARKET ANALYSIS
# ══════════════════════════════════════════════════════════════════════
elif slide == 2:
    try:
        from engines.market_data_api import get_market_summary
        market = get_market_summary()
        vg30 = market.get("vg30_estimate", {})
        fx = market.get("usd_inr", {})
        crude = market.get("crude_oil")
        crude_price = crude.get("latest_price", 0) if isinstance(crude, dict) else 0

        mk1, mk2, mk3, mk4 = st.columns(4)
        mk1.metric("Crude Oil", f"${crude_price:.1f}/bbl")
        mk2.metric("USD/INR", f"Rs {fx.get('rate', 84):.2f}")
        mk3.metric("VG30 Price", f"Rs {vg30.get('vg30_estimated', 38000):,.0f}/MT")
        mk4.metric("Gold", f"${market.get('gold', {}).get('price_usd', 0):,.0f}")
    except Exception:
        st.info("Market data loading...")

    # NHAI Tenders
    open_tenders = [t for t in NHAI_TENDERS if t["status"] == "Open"]
    st.metric("Open NHAI/PWD Tenders", f"{len(open_tenders)} projects worth Rs {sum(t['budget_cr'] for t in open_tenders):,} Cr")

    tender_df = pd.DataFrame(open_tenders[:8])[["project", "state", "budget_cr", "bitumen_mt", "deadline"]]
    tender_df.columns = ["Project", "State", "Budget (Cr)", "Bitumen (MT)", "Deadline"]
    st.dataframe(tender_df, width="stretch", hide_index=True)

    st.page_link("pages/13_📈_Market.py", label="See Full Market Intelligence", icon="📈")

# ══════════════════════════════════════════════════════════════════════
# SLIDE 3: RAW MATERIAL
# ══════════════════════════════════════════════════════════════════════
elif slide == 3:
    st.markdown(f"### Biomass Sourcing for {cfg.get('location', 'Your')} Plant")

    try:
        from engines.agro_engine import calculate_procurement_cost, calculate_plant_requirement, get_quality_specs
        state_short = cfg.get("state", "Gujarat")[:2] if len(cfg.get("state", "")) > 2 else cfg.get("state", "default")

        req = calculate_plant_requirement(cfg["capacity_tpd"], cfg.get("biomass_source", "Rice Straw"))
        cost = calculate_procurement_cost(cfg.get("biomass_source", "Rice Straw"), state_short)

        r1, r2, r3, r4 = st.columns(4)
        r1.metric("Daily Biomass Need", f"{req.get('daily_biomass_mt', 50):.0f} MT")
        r2.metric("Annual Need", f"{req.get('annual_biomass_mt', 15000):,.0f} MT")
        r3.metric("Delivered Cost", f"Rs {cost.get('total_delivered_rs_mt', 3000):,.0f}/MT")
        r4.metric("90-Day Buffer", f"{req.get('buffer_90_days_mt', 4500):,.0f} MT")

        st.markdown(f"**Primary Source:** {cfg.get('biomass_source', 'Rice Straw')} — available from farmers within 50-100 km radius")
        st.markdown(f"**Output:** Bio-Oil {req.get('bio_oil_output_mt_day', 8):.1f} MT/day + Biochar {req.get('biochar_output_mt_day', 6):.1f} MT/day")
    except Exception:
        st.markdown("- Rice Straw, Sugarcane Bagasse, Cotton Stalk, Groundnut Shell")
        st.markdown(f"- Daily requirement: ~{cfg['capacity_tpd']*2.5:.0f} MT for {cfg['capacity_tpd']:.0f} TPD plant")

    st.page_link("pages/24_🌾_Raw_Material.py", label="See Full Raw Material Analysis", icon="🌾")

# ══════════════════════════════════════════════════════════════════════
# SLIDE 4: LOCATION & FEASIBILITY
# ══════════════════════════════════════════════════════════════════════
elif slide == 4:
    st.markdown(f"### Site: {cfg.get('location', 'TBD')}, {cfg.get('state', 'TBD')}")

    if cfg.get("site_address"):
        st.info(f"**Site Address:** {cfg['site_address']}")

    # State scores
    scores = STATE_SCORES.get(cfg.get("state", "Gujarat"), {})
    total = sum(scores.get(k, 50) * LOCATION_WEIGHTS[k] for k in LOCATION_WEIGHTS)

    s1, s2, s3, s4, s5 = st.columns(5)
    s1.metric("Overall Score", f"{total:.0f}/100")
    s2.metric("Biomass", f"{scores.get('biomass', 50)}/100")
    s3.metric("Subsidy", f"{scores.get('subsidy', 50)}/100")
    s4.metric("Logistics", f"{scores.get('logistics', 50)}/100")
    s5.metric("Power", f"{scores.get('power', 50)}/100")

    # State cost details
    sc = STATE_COSTS.get(cfg.get("state", "Gujarat"), {})
    st.markdown(f"""
    | Factor | Value |
    |--------|-------|
    | Power Rate | Rs {sc.get('power_rate', 7)}/kWh |
    | Labor Cost | Rs {sc.get('labor_daily', 450)}/day |
    | Land Cost | Rs {sc.get('land_lac_acre', 15)} Lac/acre |
    | Subsidy | {sc.get('subsidy_pct', 15)}% |
    | Bitumen Demand | {sc.get('bitumen_demand_mt', 200000):,} MT/yr |
    """)

    st.page_link("pages/12_📍_Location.py", label="See Full Location Analysis", icon="📍")

# ══════════════════════════════════════════════════════════════════════
# SLIDE 5: TECHNOLOGY
# ══════════════════════════════════════════════════════════════════════
elif slide == 5:
    st.markdown("### CSIR-CRRI Licensed Technology — Proven & Certified")

    for stage in FOUR_STAGES:
        with st.expander(f"Stage {stage['stage']}: {stage['name']}", expanded=(stage['stage'] <= 2)):
            st.markdown(f"{stage['description']}")
            st.markdown(f"**CAPEX:** {stage['capex']} | **Manpower:** {stage['manpower']}")

    st.markdown("""
    **Key Facts:**
    - Technology by CSIR-CRRI / CSIR-IIP (Government of India)
    - KrishiBind bio-binder — 22% less rutting, 18% better fatigue life
    - Meets all IS:73 and MoRTH 2026 specifications
    - India became FIRST country to commercially produce bio-bitumen (Jan 2026)
    """)

    st.page_link("pages/20_Technology.py", label="See Technology Details", icon="🔬")

# ══════════════════════════════════════════════════════════════════════
# SLIDE 6: PROCESS FLOW
# ══════════════════════════════════════════════════════════════════════
elif slide == 6:
    st.markdown(f"### 13-Step Manufacturing Process — {cfg['capacity_tpd']:.0f} TPD")

    steps = [
        ("1. Biomass Receiving", "Truck unloading, weighbridge, quality check"),
        ("2. Size Reduction", "Shredder + Hammer Mill → 10-30mm"),
        ("3. Drying", "Rotary dryer → Moisture <15%"),
        ("4. Pelletization", "Uniform pellets for consistent feed"),
        ("5. Pyrolysis", "450-550°C in oxygen-free reactor → Bio-Oil 40% + Biochar 30% + Syngas 25%"),
        ("6. Bio-Oil Condensation", "Condenser HE-101 → Liquid bio-oil collection"),
        ("7. Bio-Oil Storage", "Intermediate storage tank"),
        ("8. VG-30 Heating", "Heat conventional bitumen to 160-180°C"),
        ("9. Blending", "High shear mixer — 15-30% bio-oil + 70-85% VG-30"),
        ("10. Quality Testing", "Penetration, softening point, viscosity (IS:73)"),
        ("11. Storage", "Heated bio-bitumen storage tanks"),
        ("12. Packaging", "Steel drums (181 kg) or bulk tanker"),
        ("13. Dispatch", "To NHAI/PWD contractors"),
    ]

    for step, desc in steps:
        st.markdown(f"**{step}** — {desc}")

    st.page_link("pages/22_Process_Flow.py", label="See Detailed Process Flow", icon="🔄")

# ══════════════════════════════════════════════════════════════════════
# SLIDE 7: PLANT LAYOUT
# ══════════════════════════════════════════════════════════════════════
elif slide == 7:
    st.markdown(f"### Plant Layout — {cfg['capacity_tpd']:.0f} TPD")

    # Show drawing if available
    from pathlib import Path
    draw_dir = Path(__file__).parent.parent / "data" / "all_drawings"
    cap_str = f"{int(cfg['capacity_tpd'])}TPD"
    layout_files = list(draw_dir.glob(f"*Layout*{cap_str}*")) if draw_dir.exists() else []

    if layout_files:
        st.image(str(layout_files[0]), caption=f"Plant Layout — {cap_str}", use_container_width=True)
    else:
        st.info(f"Drawing available for {cap_str} — see Drawings page")

    # Area breakdown
    sections = ["Raw Material Area (15%)", "Processing Area (20%)", "Reactor Section (15%)",
                "Blending Section (15%)", "Storage Area (10%)", "Pollution Control (5%)",
                "Dispatch Area (5%)", "Utility Area (5%)", "Admin & Lab (10%)"]
    for s in sections:
        st.markdown(f"- {s}")

    # Show area pie chart
    area_data = pd.DataFrame([
        {"Section": "Raw Material", "Pct": 15}, {"Section": "Processing", "Pct": 20},
        {"Section": "Reactor", "Pct": 15}, {"Section": "Blending", "Pct": 15},
        {"Section": "Storage", "Pct": 10}, {"Section": "Pollution Control", "Pct": 5},
        {"Section": "Dispatch", "Pct": 5}, {"Section": "Utility", "Pct": 5},
        {"Section": "Admin & Lab", "Pct": 10},
    ])
    fig_area = px.pie(area_data, names="Section", values="Pct", title="Plant Area Allocation",
                       color_discrete_sequence=["#003366","#006699","#CC3333","#FF8800","#0088cc","#00AA44","#AA3366","#009966","#666666"])
    fig_area.update_layout(template="plotly_white", height=300)
    st.plotly_chart(fig_area, width="stretch")

    st.page_link("pages/50_📐_Drawings.py", label="See All Engineering Drawings", icon="📐")

# ══════════════════════════════════════════════════════════════════════
# SLIDES 8-51: EXPANDED SLIDES (replaced old slides 8-14)
# ══════════════════════════════════════════════════════════════════════

# ── Slide 8: Pyrolysis Detail ──
elif slide == 8:
    st.markdown(f"### Pyrolysis Reactor — The Heart of {cfg['capacity_tpd']:.0f} TPD Plant")
    try:
        from engines.plant_engineering import compute_all
        comp = compute_all(cfg)
        p1, p2, p3 = st.columns(3)
        p1.metric("Reactor Diameter", f"{comp['reactor_dia_m']}m")
        p2.metric("Reactor Height", f"{comp['reactor_ht_m']}m")
        p3.metric("No. of Reactors", f"{comp['reactor_qty']}")
    except Exception:
        pass
    st.markdown(f"""
    **Operating Conditions:**
    - Temperature: {cfg.get('pyrolysis_temp_C', 500)}°C (oxygen-free)
    - Pressure: Atmospheric to slight vacuum
    - Material: AISI 310S stainless steel inner shell
    - Insulation: 100mm mineral wool + SS304 cladding
    - Safety: PSV at 0.3 barg, N₂ purge, CO detector, auto-shutdown

    **Outputs per day ({cfg['capacity_tpd']:.0f} TPD feed):**
    - Bio-Oil: {cfg['capacity_tpd']*cfg.get('bio_oil_yield_pct',32)/100:.1f} T/day ({cfg.get('bio_oil_yield_pct',32)}%)
    - Bio-Char: {cfg['capacity_tpd']*cfg.get('bio_char_yield_pct',28)/100:.1f} T/day ({cfg.get('bio_char_yield_pct',28)}%)
    - Syngas: {cfg['capacity_tpd']*cfg.get('syngas_yield_pct',22)/100:.1f} T/day ({cfg.get('syngas_yield_pct',22)}% — used as fuel)
    """)

# ── Slide 9: Product Quality ──
elif slide == 9:
    st.markdown("### Product Quality — IS:73 VG30/VG40 Certified")
    st.markdown("""
    | Test | IS:73 Requirement | Bio-Bitumen Result | Status |
    |------|-------------------|-------------------|--------|
    | Penetration (25°C) | 50-70 dmm | 55-65 dmm | PASS |
    | Softening Point | Min 47°C | 52-58°C | PASS |
    | Kinematic Viscosity (135°C) | Min 350 cSt | 380-450 cSt | PASS |
    | Ductility (25°C) | Min 40 cm | 75-100 cm | PASS |
    | Flash Point | Min 220°C | 280-310°C | PASS |
    | Solubility in TCE | Min 99% | 99.5%+ | PASS |
    """)
    st.success("Bio-bitumen meets ALL IS:73 and MoRTH 2026 specifications. CSIR-CRRI certified.")

# ── Slide 10: Complete Equipment List ──
elif slide == 10:
    st.markdown(f"### 82 Equipment Items — 15 Zones — {cfg['capacity_tpd']:.0f} TPD")
    try:
        from state_manager import calculate_boq
        boq = calculate_boq(cfg['capacity_tpd'])
        cats = {}
        for item in boq:
            c = item['category']
            cats[c] = cats.get(c, 0) + item['amount_lac']
        total = sum(cats.values())
        st.metric("Total BOQ", f"₹ {total/100:.2f} Cr ({len(boq)} items)")
        rows = [{"Zone": c, "Items": len([i for i in boq if i['category']==c]),
                 "Cost (Lac)": f"₹ {v:.1f}"} for c, v in sorted(cats.items())]
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    except Exception:
        st.info("BOQ data loading...")

# ── Slides 11-13: Zone walkthrough ──
elif slide == 11:
    st.markdown("### Zone A-D: Gate → Receiving → Pre-Processing → Reactor")
    zones_ad = [
        ("A. Gate & Security", "Main gate, weighbridge, CCTV, guard booth, boom barrier"),
        ("B. RM Receiving", "Unloading ramp, covered storage, belt conveyor, front-end loader"),
        ("C. Pre-Processing", "Shredder, rotary dryer, pelletizer, magnetic separator"),
        ("D. Reactor Zone", "Pyrolysis reactor, feed hopper, char discharge, syngas burner, PLC/SCADA"),
    ]
    for name, items in zones_ad:
        st.markdown(f"**{name}:** {items}")

elif slide == 12:
    st.markdown("### Zone E-H: Oil Recovery → Blending → Storage → Dispatch")
    zones_eh = [
        ("E. Oil Recovery", "Shell & tube condenser, cooling tower, oil-water separator, collection tanks"),
        ("F. Blending", "Bitumen heating tank, high shear mixer, colloid mill, inline viscosity meter"),
        ("G. Storage", "Finished bitumen tanks (heated), biochar silo, bio-oil surplus tank"),
        ("H. Packing & Dispatch", "Drum filling, bagging machine, tanker loading arm, forklift, parking"),
    ]
    for name, items in zones_eh:
        st.markdown(f"**{name}:** {items}")

elif slide == 13:
    st.markdown("### Zone I-O: Utilities → Lab → Safety → Office → Workshop")
    zones_io = [
        ("I. Electrical", "HT/LT transformer, DG set, MCC panels, earthing, street lighting"),
        ("J. Utilities", "Compressor, water supply, ETP, RO plant, nitrogen generator"),
        ("K. Laboratory", "Penetration, softening point, viscosity, ductility, flash point, Marshall"),
        ("L. Safety", "Fire hydrant ring main, bag filter, scrubber, gas detection, PPE"),
        ("M. Civil", "Plant building, office, lab, control room, canteen, compound wall, roads"),
        ("N. Office", "Furniture, IT equipment, AC, signage, lockers, ERP software"),
        ("O. Maintenance", "Workshop tools, spare parts store, overhead crane, welding"),
    ]
    for name, items in zones_io:
        st.markdown(f"**{name}:** {items}")

# ── Slide 14: BOQ Summary ──
elif slide == 14:
    st.markdown(f"### BOQ Investment Split — {cfg['capacity_tpd']:.0f} TPD")
    try:
        from state_manager import calculate_boq
        boq = calculate_boq(cfg['capacity_tpd'])
        total_lac = sum(i['amount_lac'] for i in boq)
        inv = cfg.get('investment_cr', total_lac/100)
        st.markdown(f"""
        | Component | Amount |
        |-----------|--------|
        | Equipment (82 items) | ₹ {total_lac:.0f} Lac |
        | GST @ 18% on machinery | ₹ {total_lac*0.12:.0f} Lac |
        | Working Capital | ₹ {cfg.get('working_capital_lac', total_lac*0.08):.0f} Lac |
        | Pre-operative Expenses | ₹ {total_lac*0.04:.0f} Lac |
        | Contingency 5% | ₹ {total_lac*0.05:.0f} Lac |
        | **Total Project Cost** | **₹ {inv:.2f} Cr** |
        """)
    except Exception:
        st.metric("Investment", f"₹ {cfg.get('investment_cr', 8):.2f} Cr")

# ── Slides 15-21: Costing & Revenue ──
elif slide == 15:
    st.markdown(f"### Raw Material — 6 Feedstock Blended Cost")
    feedstocks = [
        ("Rice Straw (Loose)", cfg.get('price_rice_straw_loose', 1200), "35%"),
        ("Rice Straw (Baled)", cfg.get('price_rice_straw_baled', 2700), "20%"),
        ("Wheat Straw", cfg.get('price_wheat_straw', 1700), "15%"),
        ("Bagasse", cfg.get('price_bagasse', 1000), "10%"),
        ("Lignin (Kraft)", cfg.get('price_lignin', 4000), "5%"),
        ("Other Agro Waste", cfg.get('price_other_agro_waste', 900), "15%"),
    ]
    rows = [{"Feedstock": n, "₹/Tonne": f"₹{p:,}", "Mix %": m} for n, p, m in feedstocks]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

elif slide == 16:
    bit_price = cfg.get('price_conv_bitumen', 45750)
    st.markdown("### Conventional Bitumen — Landed Cost")
    st.markdown(f"""
    | Component | ₹/Tonne |
    |-----------|---------|
    | Base Price (ex-IOCL VG30) | ₹{bit_price:,} |
    | GST @ 18% | ₹{bit_price*0.18:,.0f} |
    | Road Freight | ₹{cfg.get('bitumen_transport', 650):,} |
    | Tanker Hire | ₹180 |
    | Storage/Transit Loss 1.5% | ₹{bit_price*0.015:,.0f} |
    | **Landed Cost** | **₹{bit_price + bit_price*0.18 + 650 + 180 + bit_price*0.015:,.0f}** |
    """)

elif slide == 17:
    st.markdown("### Landing & Logistics — Farm Gate to Plant Gate")
    items = [("Baling & Collection", cfg.get('landing_baling', 350)),
             ("Primary Transport", cfg.get('landing_primary_transport', 250)),
             ("Depot Storage", cfg.get('landing_depot_storage', 300)),
             ("Long Haul", cfg.get('landing_long_haul', 480)),
             ("Load/Unload", cfg.get('landing_load_unload', 140)),
             ("Testing & Misc", cfg.get('landing_testing_misc', 65))]
    total = sum(v for _, v in items) * 1.02
    rows = [{"Component": n, "₹/Tonne": f"₹{v:,}"} for n, v in items]
    rows.append({"Component": "Contingency 2%", "₹/Tonne": f"₹{total - sum(v for _,v in items):,.0f}"})
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    st.metric("Total Landing Cost", f"₹{total:,.0f}/T")

elif slide == 18:
    st.markdown("### Production Cost — Daily Breakdown")
    elec = cfg.get('electricity_rate', 7.5) * cfg.get('electricity_kwh_day', 1200)
    diesel = cfg.get('diesel_rate', 92) * cfg.get('diesel_litres_day', 120)
    syngas_cr = cfg['capacity_tpd'] * cfg.get('syngas_yield_pct', 22) / 100 * 1000
    labour = cfg.get('labour_daily_cost', 18000) * 1.2
    items = [("Electricity", elec), ("Diesel", diesel), ("Less: Syngas Credit", -syngas_cr),
             ("Labour (incl PF/ESI)", labour), ("Overheads", cfg.get('overheads_daily_cost', 12000)),
             ("Chemicals", cfg.get('chemicals_daily_cost', 2500))]
    for n, v in items:
        st.markdown(f"- **{n}:** ₹{v:,.0f}/day")
    st.metric("Net Production Cost/Day", f"₹{sum(v for _,v in items):,.0f}")

elif slide == 19:
    st.markdown(f"### Complete Cost Sheet — 10 Heads — {cfg['capacity_tpd']:.0f} TPD")
    try:
        from engines.detailed_costing import calculate_complete_cost_sheet
        cs = calculate_complete_cost_sheet(cfg)
        d = cs['blend_total_tpd'] if cs['blend_total_tpd'] > 0 else 1
        for head, cost in cs['cost_heads']:
            st.markdown(f"- {head}: **₹{cost/d:,.0f}/T** (₹{cost:,.0f}/day)")
        st.markdown(f"**Net Cost/T: ₹{cs['net_cpt']:,}** | Sale: ₹{cs['sale_price_pt']:,} | Margin: ₹{cs['margin_pt']:,} ({cs['margin_pct']:.1f}%)")
    except Exception:
        st.info("Cost sheet loading...")

elif slide == 20:
    st.markdown("### Revenue — 7 Product Streams")
    try:
        from engines.detailed_costing import calculate_complete_cost_sheet
        cs = calculate_complete_cost_sheet(cfg)
        for item in cs['revenue']['items']:
            st.markdown(f"- **{item['product']}**: {item['qty_tpd']} T/day × ₹{item['price']:,} = **₹{item['daily_rev']:,}/day**")
        st.metric("Total Daily Revenue", f"₹{cs['revenue']['total_daily']:,}")
        st.metric("Annual Revenue", f"₹{cs['revenue']['total_daily']*cfg.get('working_days',300)/1e7:.2f} Cr")
    except Exception:
        st.info("Revenue data loading...")

elif slide == 21:
    st.markdown("### Finished Goods — Pan India Buyer Network")
    buyers = [
        ("Bio-Bitumen VG30", "NHAI, State PWDs, Highway Contractors", "70% of revenue"),
        ("Bio-Bitumen VG40", "Airport runways, Heavy-duty overlay", "15%"),
        ("Bio-Char (Agriculture)", "Farmers, Fertilizer companies", "8%"),
        ("Bio-Char (Industrial)", "Water treatment, Carbon sequestration", "3%"),
        ("Bio-Oil (Fuel)", "Boilers, Brick kilns", "2%"),
        ("Carbon Credits", "Voluntary carbon market", "2%"),
    ]
    for product, buyer, pct in buyers:
        st.markdown(f"**{product}** — {buyer} ({pct})")

# ── Slides 22-30: Financial Details ──
elif slide == 22:
    # This was slide 8 — Financial Model (keep original content)
    st.markdown(f"### Financial Model — {cfg['capacity_tpd']:.0f} TPD | Rs {cfg['investment_cr']:.2f} Crore")
    f1, f2, f3, f4, f5 = st.columns(5)
    f1.metric("Investment", f"Rs {cfg['investment_cr']:.2f} Cr")
    f2.metric("Loan", f"Rs {cfg.get('loan_cr', cfg['investment_cr']*0.6):.2f} Cr")
    f3.metric("Equity", f"Rs {cfg.get('equity_cr', cfg['investment_cr']*0.4):.2f} Cr")
    f4.metric("EMI/Month", f"Rs {cfg.get('emi_lac_mth', 0):.2f} Lac")
    f5.metric("Rev Yr5", f"Rs {cfg.get('revenue_yr5_lac', 0):.0f} Lac")
    if cfg.get("roi_timeline"):
        roi_df = pd.DataFrame(cfg["roi_timeline"])
        fig = go.Figure()
        fig.add_trace(go.Bar(x=roi_df["Year"], y=roi_df["Revenue (Lac)"], name="Revenue", marker_color="#003366"))
        fig.add_trace(go.Bar(x=roi_df["Year"], y=roi_df["Variable Cost (Lac)"], name="Cost", marker_color="#CC3333"))
        fig.add_trace(go.Scatter(x=roi_df["Year"], y=roi_df["PAT (Lac)"], name="Net Profit", mode="lines+markers", line=dict(color="#00AA44", width=3)))
        fig.update_layout(title="7-Year Revenue vs Cost vs Profit", barmode="group", template="plotly_white", height=350)
        st.plotly_chart(fig, use_container_width=True)

elif slide == 23:
    # ROI slide (was slide 9)
    st.markdown(f"### Return on Investment — {cfg['capacity_tpd']:.0f} TPD")
    g1, g2, g3 = st.columns(3)
    for col, val, title in [(g1, cfg.get('roi_pct',0), "ROI %"), (g2, cfg.get('irr_pct',0), "IRR %"), (g3, cfg.get('break_even_months',0), "Break-Even (Months)")]:
        with col:
            fig = go.Figure(go.Indicator(mode="gauge+number", value=val, title={"text": title},
                gauge={"axis": {"range": [0, 60 if 'ROI' in title else (80 if 'IRR' in title else 60)]}, "bar": {"color": "#003366"}}))
            fig.update_layout(height=220, margin=dict(t=50, b=10))
            st.plotly_chart(fig, use_container_width=True)

elif slide == 24:
    st.markdown("### 7-Year P&L Projection")
    if cfg.get("roi_timeline"):
        st.dataframe(pd.DataFrame(cfg["roi_timeline"]), use_container_width=True, hide_index=True)
    else:
        st.info("Set capacity in Financial Model first")

elif slide == 25:
    st.markdown("### Cash Flow — 5 Year with Capacity Ramp-Up")
    st.markdown("60% → 75% → 85% → 90% → 95% utilization schedule")
    try:
        from engines.dpr_financial_engine import calculate_5year_cashflow
        from engines.detailed_costing import calculate_complete_cost_sheet
        cs = calculate_complete_cost_sheet(cfg)
        cf = calculate_5year_cashflow(cfg, cs)
        for y in cf['years']:
            st.markdown(f"**Year {y['year']}** ({y['utilization_pct']}): Revenue ₹{y['revenue_cr']:.2f} Cr | PAT ₹{y['pat_cr']:.2f} Cr")
        st.metric("Payback", f"Year {cf['payback_year']}")
    except Exception:
        st.info("Cash flow data loading...")

elif slide == 26:
    st.markdown("### Break-Even Analysis")
    try:
        from engines.dpr_financial_engine import calculate_break_even
        from engines.detailed_costing import calculate_complete_cost_sheet
        cs = calculate_complete_cost_sheet(cfg)
        be = calculate_break_even(cfg, cs)
        b1, b2, b3, b4 = st.columns(4)
        b1.metric("BE Tonnes/Year", f"{be['be_tonnes_annual']:,}")
        b2.metric("BE Days", f"{be['be_days']}")
        b3.metric("BE Capacity %", f"{be['be_pct']:.1f}%")
        b4.metric("Margin of Safety", f"{be['margin_of_safety']:.1f}%")
    except Exception:
        st.info("Break-even loading...")

elif slide == 27:
    st.markdown("### Working Capital Requirement")
    try:
        from engines.dpr_financial_engine import calculate_working_capital
        from engines.detailed_costing import calculate_complete_cost_sheet
        cs = calculate_complete_cost_sheet(cfg)
        wc = calculate_working_capital(cfg, cs)
        st.metric("Net Working Capital", f"₹ {wc['net_wc_lac']:.1f} Lac")
        st.metric("Current Ratio", f"{wc['current_ratio']:.2f}")
        for item in wc['items']:
            st.markdown(f"- {item['component']}: ₹{abs(item['amount']):,.0f} ({item['days']} days)")
    except Exception:
        st.info("Working capital loading...")

elif slide == 28:
    st.markdown("### Loan EMI & Debt Service")
    loan = cfg.get('loan_cr', cfg['investment_cr'] * 0.6)
    emi = cfg.get('emi_lac_mth', 0)
    rate = cfg.get('interest_rate', 0.115)
    st.markdown(f"""
    | Parameter | Value |
    |-----------|-------|
    | Loan Amount | ₹ {loan:.2f} Cr |
    | Interest Rate | {rate*100:.1f}% p.a. |
    | Tenure | {cfg.get('emi_tenure_months', 84)} months (7 years) |
    | Monthly EMI | ₹ {emi:.2f} Lac |
    | Annual Debt Service | ₹ {emi*12:.2f} Lac |
    | DSCR Year 3 | {cfg.get('dscr_yr3', 0):.2f}x |
    | CGTMSE Eligible | {'Yes (< ₹5 Cr)' if loan*100 < 500 else 'No (> ₹5 Cr)'} |
    """)

elif slide == 29:
    st.markdown("### Sensitivity Analysis — What If Prices Change?")
    st.markdown("Impact of ±20% change on key variables:")
    variables = [("Conv. Bitumen Price", "Largest cost driver — 60% of gross cost"),
                 ("Sale Price VG30", "Most critical revenue driver"),
                 ("Plant Capacity", "Volume sensitivity"),
                 ("Bio-Oil Yield %", "Process efficiency impact"),
                 ("Rice Straw Price", "Raw material risk")]
    for var, impact in variables:
        st.markdown(f"- **{var}** — {impact}")

elif slide == 30:
    st.markdown("### State-wise Profitability Comparison")
    states_compare = ["Punjab", "Gujarat", "Maharashtra", "Uttar Pradesh", "Tamil Nadu"]
    try:
        from engines.detailed_costing import calculate_complete_cost_sheet
        for s in states_compare:
            t = dict(cfg); t['state'] = s
            tcs = calculate_complete_cost_sheet(t)
            color = "🟢" if tcs['margin_pt'] > 3000 else ("🟡" if tcs['margin_pt'] > 0 else "🔴")
            st.markdown(f"{color} **{s}**: Cost ₹{tcs['net_cpt']:,}/T | Margin ₹{tcs['margin_pt']:,}/T ({tcs['margin_pct']:.1f}%)")
    except Exception:
        st.info("State comparison loading...")

# ── Slides 31-35: Compliance ──
elif slide == 31:
    mandatory = [lt for lt in LICENSE_TYPES if lt.get("mandatory")]
    c1, c2 = st.columns(2)
    c1.metric("Total Licenses", len(LICENSE_TYPES))
    c2.metric("Mandatory", len(mandatory))
    for lt in mandatory[:15]:
        st.markdown(f"- **{lt['name']}** — {lt['authority']} (~{lt['typical_days']} days)")

elif slide == 32:
    st.markdown("### Environmental Clearances & Green Certifications")
    st.markdown(f"""
    - **CTE (Consent to Establish)** — State Pollution Board — 30-60 days
    - **CTO (Consent to Operate)** — After plant ready — 15-30 days
    - **EIA (if > 50 TPD)** — MoEFCC — 90-180 days
    - **Carbon Credits** — Voluntary market — ₹{cfg.get('sale_carbon_credit', 12500):,}/credit
    - **CO₂ Saved** — {cfg['capacity_tpd']*300*0.35:.0f} tonnes/year
    - **Green Building Certification** — Optional, adds value to brand
    """)

elif slide == 33:
    # Same as old slide 11 (Subsidy)
    schemes = [
        ("MNRE Waste-to-Wealth", "25% capital subsidy", "Bio-bitumen eligible"),
        (f"State MSME ({cfg.get('state', 'Gujarat')})", f"{STATE_COSTS.get(cfg.get('state','Gujarat'),{}).get('subsidy_pct',15)}%", "State industrial policy"),
        ("CGTMSE Guarantee", "Collateral-free up to ₹5 Cr", "MSME manufacturing"),
        ("Carbon Credits", f"₹{cfg['capacity_tpd']*300*0.35*12500/100000:.1f} Lac/yr", "Voluntary market"),
    ]
    for name, benefit, detail in schemes:
        st.markdown(f"**{name}** — {benefit} ({detail})")

elif slide == 34:
    st.markdown("### NHAI / MoRTH Green Mandate")
    open_tenders = [t for t in NHAI_TENDERS if t["status"] == "Open"]
    st.metric("Open Tenders", f"{len(open_tenders)} projects")
    st.markdown("""
    - MoRTH Specification 2026 Section 519 allows bio-modified bitumen
    - NHAI target: 15% green bitumen by 2030
    - First commercial use: Nagpur-Mansar NH44 (Dec 2024)
    - GeM portal registration mandatory for government supply
    """)

elif slide == 35:
    # Risk Analysis (was slide 13)
    risk_df = pd.DataFrame(RISK_REGISTRY)
    risk_df["score"] = risk_df["probability"] * risk_df["impact"]
    for _, r in risk_df.nlargest(6, "score").iterrows():
        sev = "🔴" if r["score"] >= 15 else ("🟠" if r["score"] >= 10 else "🟡")
        st.markdown(f"{sev} **{r['category']}**: {r['risk'][:60]} — *{r['mitigation'][:60]}*")

# ── Slides 36-41: Implementation ──
elif slide == 36:
    # Timeline (was slide 12)
    st.markdown("### 10-Phase Implementation — 12 to 18 Months")
    phases = [("DPR & Feasibility", "Month 1"), ("Company Setup", "Month 1-2"),
              ("Land & Approvals", "Month 2-4"), ("Environmental", "Month 3-7"),
              ("Bank Loan", "Month 3-5"), ("Engineering", "Month 4-6"),
              ("Equipment Order", "Month 5-8"), ("Civil Construction", "Month 6-10"),
              ("Installation", "Month 10-12"), ("Commercial Production", "Month 13+")]
    for name, period in phases:
        st.markdown(f"**{name}** — {period}")

elif slide == 37:
    st.markdown(f"### Manpower — {cfg['capacity_tpd']:.0f} TPD Plant")
    staff_data = [
        ("Plant Manager", 1, "₹60-80K"), ("Shift Supervisors", max(2, int(cfg['capacity_tpd']/10)), "₹30-40K"),
        ("Reactor Operators", max(4, int(cfg['capacity_tpd']*0.3)), "₹18-25K"),
        ("Lab Technician", max(1, int(cfg['capacity_tpd']/20)), "₹20-30K"),
        ("Electrician", max(1, int(cfg['capacity_tpd']/20)), "₹18-25K"),
        ("Helpers/Labour", max(6, int(cfg['capacity_tpd']*0.5)), "₹12-15K"),
        ("Security", max(2, int(cfg['capacity_tpd']/10)), "₹12-15K"),
        ("Office Staff", max(2, int(cfg['capacity_tpd']/10)), "₹18-25K"),
    ]
    total = sum(c for _, c, _ in staff_data)
    for role, count, salary in staff_data:
        st.markdown(f"- **{role}**: {count} nos — {salary}/month")
    st.metric("Total Staff", f"{total} persons")

elif slide == 38:
    st.markdown("### Safety & Fire Protection (IS 14489 / NBC 2016)")
    st.markdown("""
    - Fire hydrant ring main: DN150, 45m spacing (IS 3844)
    - Reactor to boundary: MIN 15m clearance
    - Control room: 30m from reactor (blast-resistant)
    - Gas detectors: CO/H₂S/CH₄ at reactor + tank area
    - Emergency assembly points: 2 locations, 50m from hazard
    - PPE zones: Hard hat, fire-resistant coverall, safety shoes
    - Fire water tank: 50-100 m³ with diesel pump backup
    """)

elif slide == 39:
    st.markdown("### Quality Control & Lab Setup")
    lab_equip = ["Penetration Tester (IS 1203)", "Softening Point (IS 1205)", "Viscosity Bath (IS 1206)",
                 "Ductility Machine (IS 1208)", "Flash Point (IS 1209)", "Marshall Stability",
                 "Muffle Furnace", "Moisture Oven + Balance"]
    for e in lab_equip:
        st.markdown(f"- {e}")
    st.success("All testing per IS:73 / MoRTH 2026 standards — NABL accreditation recommended")

elif slide == 40:
    st.markdown("### Capacity Comparison — 5 to 50 TPD")
    caps = [5, 10, 20, 25, 50]
    from master_data_loader import get_plant
    rows = []
    for c in caps:
        key = f"{c:02d}MT"
        p = get_plant(key)
        rows.append({"Capacity": f"{c} TPD", "Investment": f"₹{p.get('inv_cr',0)} Cr",
                     "IRR": f"{p.get('irr_pct',0)}%", "Staff": p.get('staff',0)})
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

elif slide == 41:
    st.markdown("### 3-Process Model Comparison")
    st.markdown("""
    | Model | Process | Investment | Best For |
    |-------|---------|-----------|----------|
    | **Process 1** | Full Chain (Biomass→Bitumen) | ₹3-10 Cr | New investors, maximum ROI |
    | **Process 2** | Blending Only (Buy oil→Blend) | ₹0.8-2 Cr | Existing bitumen plants |
    | **Process 3** | Raw Output (Oil + Char) | ₹1-3 Cr | Biomass companies |
    """)

# ── Slides 42-48: Why PPS Anantams ──
elif slide == 42:
    st.markdown("### About PPS Anantams Corporation Pvt Ltd")
    st.markdown(f"""
    - **Founder:** {COMPANY['owner']}
    - **Experience:** {COMPANY.get('years_experience', 25)} years in bitumen industry
    - **Plants Built:** {COMPANY.get('plants_built', 10)} across India
    - **Network:** {COMPANY.get('industry_contacts', 4452):,} industry contacts
    - **HQ:** {COMPANY.get('hq', 'Vadodara, Gujarat')}
    - **International:** VG-30 supply — 2.4 Lakh MT/yr (Getka USA-Iraq)
    - **GST:** {COMPANY.get('gst', '')}
    """)

elif slide == 43:
    st.markdown("### Our 25 Years of Experience")
    try:
        for cred in KEY_CREDENTIALS[:8]:
            st.markdown(f"- **{cred['category']}**: {cred['detail']}")
    except Exception:
        st.markdown("- International bitumen trading (USA, Iraq, UAE)")
        st.markdown("- 10 plants built across Gujarat, Maharashtra, MP, UP")
        st.markdown("- CSIR-CRRI technology partnership")

elif slide == 44:
    st.markdown("### Our Role — Complete Turnkey Solution")
    services = [
        ("DPR & Feasibility Study", "Bank-ready Detailed Project Report with all financials"),
        ("Site Selection & Land", "Location analysis, land identification, soil testing"),
        ("Government Approvals", "All 25 licenses — CTE, CTO, PESO, Fire NOC, Factory License"),
        ("Equipment Procurement", "Vendor selection, negotiation, quality inspection"),
        ("Civil Construction", "Foundation, building, compound wall, roads — turnkey"),
        ("Installation & Commissioning", "Equipment erection, piping, electrical, testing"),
        ("Training & SOPs", "Operator training, safety drills, quality procedures"),
        ("Marketing & Sales", "Buyer network, NHAI registration, GeM portal setup"),
    ]
    for name, desc in services:
        st.markdown(f"**{name}** — {desc}")

elif slide == 45:
    st.markdown("### What We Deliver — Documents & Support")
    docs = [
        "Detailed Project Report (DPR) — 100+ pages, bank-ready",
        "117+ Engineering Drawings — All capacities",
        "Financial Model — 7-year P&L, cash flow, sensitivity",
        "Bank Loan Proposal — SBI MSME format",
        "Investor Pitch Deck — 10 slides + financials",
        "Environmental Impact Assessment",
        "Equipment Specifications & BOQ",
        "Government Application Forms (state-wise)",
        "Quality Test Reports & IS:73 Compliance",
        "Operations Manual & Safety SOPs",
    ]
    for d in docs:
        st.markdown(f"- {d}")
    st.success(f"**Total: 2,800+ documents** ready for {cfg['capacity_tpd']:.0f} TPD plant")

elif slide == 46:
    st.markdown("### Our Competitive Advantage (SWOT)")
    try:
        for category, items in PPS_SWOT.items():
            icon = {"Strengths": "💪", "Weaknesses": "⚠️", "Opportunities": "🎯", "Threats": "🛡️"}.get(category, "")
            with st.expander(f"{icon} {category}"):
                for item in items:
                    st.markdown(f"- {item}")
    except Exception:
        st.markdown("- 25 years experience | 10 plants | 4,452 contacts")

elif slide == 47:
    st.markdown("### Client Success Stories")
    st.markdown("""
    - **Nagpur-Mansar NH44** — First commercial bio-bitumen road in India (Dec 2024)
    - **Gujarat Industrial Estate** — 20 TPD plant, ROI 35%, operational in 14 months
    - **UP Expressway** — VG40 bio-bitumen for heavy traffic overlay
    - **NHAI Green Pilot** — 5 km test track, MoRTH approved specification
    """)

elif slide == 48:
    st.markdown("### How We Compare — Competitor Analysis")
    try:
        for comp in COMPETITORS[:5]:
            st.markdown(f"**{comp['name']}** — {comp.get('focus', '')} | Revenue: {comp.get('revenue', 'N/A')}")
        st.success(f"**PPS Anantams Advantage:** Only consultant offering COMPLETE turnkey solution — DPR to dispatch. Others do only partial.")
    except Exception:
        st.markdown("We are the only company offering end-to-end consulting + execution.")

# ══════════════════════════════════════════════════════════════════════
# SLIDES 49-53: CONVINCE & CLOSE (5 NEW SLIDES)
# ══════════════════════════════════════════════════════════════════════

# ── Slide 49: WITH US vs WITHOUT US ──
elif slide == 49:
    st.markdown("### With PPS Anantams vs Without Us")
    inv = cfg.get("investment_cr", 10)
    st.markdown(f"""
    | Parameter | Without Consultant | With PPS Anantams |
    |---|---|---|
    | **DPR Preparation** | Rs 3-5 Lac (external CA) | Included in fee |
    | **Technology Selection** | Trial & error, wrong choice risk | CSIR-CRRI proven, pre-validated |
    | **Machinery Procurement** | Overpay 15-25% (no benchmarks) | Best price from 4,452 contacts |
    | **Bank Loan Process** | 6-12 months, multiple rejections | 3-4 months, bank-ready DPR |
    | **Government Approvals** | 8-15 months (wrong sequence) | 4-6 months (parallel processing) |
    | **Construction Mistakes** | Rs 20-50 Lac wasted on rework | Zero rework — pre-engineered layout |
    | **Wrong Equipment** | Rs 30-80 Lac loss (wrong spec) | Exact specs from 10 plant experience |
    | **Production Start** | 24-30 months | 12-15 months |
    | **First Year Issues** | Low yield, quality rejection | 85%+ utilization, IS:73 certified |
    | **TOTAL EXTRA COST** | **Rs {inv*0.3:.1f}-{inv*0.5:.1f} Cr wasted** | **Rs {inv*0.10:.1f} Cr consulting fee** |
    """)
    st.error(f"**Without us:** You risk Rs {inv*0.3:.1f}-{inv*0.5:.1f} Cr in mistakes, delays, and wrong decisions")
    st.success(f"**With us:** You pay Rs {inv*0.10:.1f} Cr (10%) and save Rs {inv*0.2:.1f}-{inv*0.4:.1f} Cr + 12 months time")

# ── Slide 50: ROI CALCULATOR (Client plays with sliders) ──
elif slide == 50:
    st.markdown("### Play with Numbers — See YOUR Return")
    st.caption("Move the sliders to see how your investment performs")

    rc1, rc2 = st.columns(2)
    with rc1:
        r_cap = st.slider("Your Plant Capacity (TPD)", 5, 100, int(cfg.get("capacity_tpd", 25)), 5, key="roi_play_cap")
        r_price = st.slider("Selling Price (Rs/MT)", 30000, 55000, int(cfg.get("selling_price_per_mt", 44000)), 1000, key="roi_play_price")
        r_days = st.slider("Working Days/Year", 250, 330, int(cfg.get("working_days", 300)), 10, key="roi_play_days")
    with rc2:
        r_cost = st.slider("Variable Cost (Rs/MT)", 15000, 35000, 22000, 1000, key="roi_play_cost")
        r_inv = st.slider("Investment (Rs Crore)", 2.0, 30.0, float(cfg.get("investment_cr", 10)), 0.5, key="roi_play_inv")

    # Calculate live
    output_per_day = r_cap * 0.4  # 40% yield
    annual_output = output_per_day * r_days
    annual_revenue = annual_output * r_price
    annual_cost = annual_output * r_cost
    annual_profit = annual_revenue - annual_cost
    monthly_profit = annual_profit / 12
    roi = annual_profit / (r_inv * 1e7) * 100 if r_inv > 0 else 0
    payback = (r_inv * 1e7) / annual_profit if annual_profit > 0 else 999

    rk1, rk2, rk3, rk4 = st.columns(4)
    rk1.metric("Annual Revenue", f"Rs {annual_revenue/1e7:.2f} Cr")
    rk2.metric("Annual Profit", f"Rs {annual_profit/1e7:.2f} Cr",
               delta="Profitable" if annual_profit > 0 else "Loss",
               delta_color="normal" if annual_profit > 0 else "inverse")
    rk3.metric("ROI", f"{roi:.1f}%")
    rk4.metric("Payback", f"{payback:.1f} years" if payback < 100 else "N/A")

    st.metric("Monthly Profit to Your Pocket", f"Rs {monthly_profit/1e5:.1f} Lac/month")

# ── Slide 51: PROJECT READINESS CHECKLIST ──
elif slide == 51:
    st.markdown("### Project Readiness — What You Need vs What We Handle")

    checklist = [
        ("Land (owned or leased)", "YOU", "Identify 1-3 acre industrial plot"),
        ("Company Registration (Pvt Ltd / LLP)", "WE HELP", "ROC filing, MOA/AOA drafting"),
        ("GST + PAN + TAN Registration", "WE HANDLE", "Online filing within 7 days"),
        ("Udyam (MSME) Registration", "WE HANDLE", "Instant online registration"),
        ("DPR (Detailed Project Report)", "WE DELIVER", "21-section bank-ready DPR"),
        ("Bank Loan Application + CMA Data", "WE PREPARE", "SBI/SIDBI format, CA certified"),
        ("CTE (Pollution Board)", "WE APPLY", "Complete application with process flow"),
        ("Factory License", "WE APPLY", "Form + plan + machinery list"),
        ("Fire NOC", "WE APPLY", "Fire safety plan + equipment list"),
        ("Machinery Procurement", "WE COORDINATE", "OEM selection, negotiation, inspection"),
        ("Civil Construction", "WE SUPERVISE", "Contractor selection, quality check"),
        ("Installation & Commissioning", "WE MANAGE", "Erection, piping, testing, trial run"),
        ("Quality Testing (IS:73)", "WE ARRANGE", "CSIR-CRRI lab testing"),
        ("NHAI / GeM Registration", "WE DO", "Supplier registration for govt supply"),
        ("Buyer Network Setup", "WE CONNECT", "4,452 contacts — contractors, traders"),
    ]

    for item, who, detail in checklist:
        color = "#00AA44" if "WE" in who else "#FF8800"
        icon = "✅" if "WE" in who else "📋"
        st.markdown(f"{icon} **{item}** — <span style='color:{color}; font-weight:bold'>{who}</span> — {detail}", unsafe_allow_html=True)

    we_count = sum(1 for _, w, _ in checklist if "WE" in w)
    st.success(f"**We handle {we_count} out of {len(checklist)} steps.** You focus on land and capital — we do everything else.")

# ── Slide 52: CONSULTING FEE STRUCTURE ──
elif slide == 52:
    inv = cfg.get("investment_cr", 10)
    fee = inv * 0.10  # 10% of project cost
    st.markdown("### Our Consulting Fee — Transparent & Value-Based")

    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #003366, #006699); padding: 25px; border-radius: 15px; color: white; text-align: center;">
        <h2 style="color: white;">Consulting Fee: 10% of Project Cost</h2>
        <p style="font-size: 2em; color: #ffcc00; font-weight: bold;">Rs {fee:.2f} Crore</p>
        <p style="color: #99ccff;">(For Rs {inv:.2f} Crore project)</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### Fee Bifurcation")
    st.markdown(f"""
    | Phase | Deliverable | Fee % | Amount |
    |---|---|---|---|
    | **Phase 1:** Feasibility + DPR | Complete DPR, financial model, drawings | 20% | Rs {fee*0.20:.2f} Cr |
    | **Phase 2:** Bank Loan + Approvals | Loan sanction, CTE, Factory License, Fire NOC | 20% | Rs {fee*0.20:.2f} Cr |
    | **Phase 3:** Procurement + Construction | Equipment order, civil work, quality check | 30% | Rs {fee*0.30:.2f} Cr |
    | **Phase 4:** Commissioning + Handover | Installation, trial run, IS:73 testing, buyer setup | 30% | Rs {fee*0.30:.2f} Cr |
    | **TOTAL** | **Complete Turnkey Consulting** | **100%** | **Rs {fee:.2f} Cr** |
    """)

    st.markdown("---")
    st.markdown("### What Happens If You Go Without Consultant?")
    st.markdown(f"""
    | Risk | Typical Loss |
    |---|---|
    | Wrong technology selection | Rs {inv*0.05:.1f}-{inv*0.10:.1f} Cr |
    | Overpaying for machinery (no benchmarks) | Rs {inv*0.04:.1f}-{inv*0.08:.1f} Cr |
    | Bank loan rejection + reapplication | 6-12 months delay |
    | Wrong sequence of approvals | 4-8 months delay |
    | Construction rework (wrong layout) | Rs {inv*0.03:.1f}-{inv*0.05:.1f} Cr |
    | Low production yield (no expert commissioning) | Rs {inv*0.02:.1f}-{inv*0.04:.1f} Cr/year |
    | **TOTAL RISK WITHOUT CONSULTANT** | **Rs {inv*0.15:.1f}-{inv*0.30:.1f} Cr + 12-18 months delay** |
    """)

    st.success(f"**Our fee: Rs {fee:.2f} Cr | Risk without us: Rs {inv*0.15:.1f}-{inv*0.30:.1f} Cr** — You SAVE by hiring us")

# ── Slide 53: TIMELINE WITH DELIVERABLES ──
elif slide == 53:
    st.markdown("### Month-by-Month Timeline with Deliverables")

    timeline = [
        ("Month 1", "DPR + Feasibility", [
            "Complete DPR (21 sections, 35+ pages)",
            "Financial Model (7-year P&L, DSCR, IRR)",
            "BOQ with machinery quotations",
            "Site evaluation report",
        ]),
        ("Month 2", "Company + Bank", [
            "Company registration (if needed)",
            "GST + Udyam + PAN + TAN",
            "Bank loan application with CMA data",
            "CA certifications",
        ]),
        ("Month 3-4", "Approvals (Parallel)", [
            "CTE application to State PCB",
            "Factory License application",
            "Fire NOC application",
            "Electricity HT connection sanction",
            "Bank loan sanction expected",
        ]),
        ("Month 4-6", "Engineering + Procurement", [
            "Detailed engineering drawings (9 types)",
            "Equipment ordering (3-4 OEM quotes each)",
            "Civil contractor finalization",
            "Foundation + civil work begins",
        ]),
        ("Month 6-10", "Construction", [
            "Civil construction (PEB/RCC)",
            "Equipment delivery + installation",
            "Piping + electrical work",
            "Pollution control equipment",
            "Progress reports every 15 days",
        ]),
        ("Month 10-12", "Commissioning", [
            "Mechanical completion",
            "Electrical testing + energizing",
            "Cold commissioning (water run)",
            "Hot commissioning (trial run)",
            "IS:73 quality testing at CSIR-CRRI lab",
            "CTO (Consent to Operate) from PCB",
        ]),
        ("Month 13+", "Production + Sales", [
            "Commercial production begins",
            "GeM portal + NHAI registration",
            "First customer orders",
            "Buyer network activation (4,452 contacts)",
            "Monthly review + support for 6 months",
        ]),
    ]

    for month, phase, deliverables in timeline:
        with st.expander(f"**{month}: {phase}**", expanded=(month == "Month 1")):
            for d in deliverables:
                st.markdown(f"- {d}")

# ── Slides 54-56: Original final slides (renumbered) ──
elif slide == 54:
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #003366, #006699); padding: 25px; border-radius: 15px; color: white;">
        <h2 style="color: white; margin: 0;">{cfg.get('project_name', 'Bio-Bitumen Plant')}</h2>
        <p style="color: #99ccff;">For: {cfg.get('client_name', 'Client')} | {cfg.get('client_company', '')}</p>
        <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin-top: 15px;">
            <div><span style="font-size: 1.8em; font-weight: bold;">{cfg['capacity_tpd']:.0f}</span><br>MT/Day</div>
            <div><span style="font-size: 1.8em; font-weight: bold;">Rs {cfg['investment_cr']:.1f}Cr</span><br>Investment</div>
            <div><span style="font-size: 1.8em; font-weight: bold;">{cfg.get('roi_pct', 0):.0f}%</span><br>ROI</div>
            <div><span style="font-size: 1.8em; font-weight: bold;">{cfg.get('break_even_months', 0)}</span><br>Month BE</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

elif slide == 55:
    st.markdown(f"""
    <div style="background: #003366; padding: 30px; border-radius: 15px; color: white; text-align: center;">
        <h2 style="color: white;">Ready to Start?</h2>
        <p style="font-size: 1.3em; color: #99ccff;">{COMPANY['owner']}</p>
        <p style="font-size: 1.5em; margin-top: 15px;">📞 {COMPANY.get('phone', '')}</p>
        <p>📧 {COMPANY.get('email', '')}</p>
        <p>📍 {COMPANY.get('hq', 'Vadodara, Gujarat')}</p>
    </div>
    """, unsafe_allow_html=True)

elif slide == 56:
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #003366, #00AA44); padding: 40px; border-radius: 15px; color: white; text-align: center;">
        <h1 style="color: white; font-size: 2.5em;">Thank You</h1>
        <p style="font-size: 1.3em; color: #ccffcc;">{cfg.get('client_name', 'Dear Client')}</p>
        <p style="font-size: 1.1em; margin-top: 20px; color: white;">
            We look forward to building India's green infrastructure together.
        </p>
        <p style="margin-top: 25px; font-size: 0.9em; color: #ccccff;">
            📞 {COMPANY.get('phone', '')} | 📧 {COMPANY.get('email', '')}
        </p>
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════
# NAVIGATION BAR (Bottom of every slide)
# ══════════════════════════════════════════════════════════════════════
st.markdown("---")

def _go_next():
    st.session_state["slide"] = min(st.session_state.get("slide", 0) + 1, TOTAL_SLIDES - 1)

def _go_prev():
    st.session_state["slide"] = max(st.session_state.get("slide", 0) - 1, 0)

def _go_start():
    st.session_state["slide"] = 0

def _go_end():
    st.session_state["slide"] = TOTAL_SLIDES - 1

def _jump_to():
    st.session_state["slide"] = st.session_state.get("jump_slide", 0)

nav1, nav2, nav3, nav4, nav5 = st.columns([1, 1, 4, 1, 1])

with nav1:
    if slide > 0:
        st.button("◀ Previous", key="prev", use_container_width=True, on_click=_go_prev)

with nav2:
    if slide > 0:
        st.button("⏮ Start", key="first", use_container_width=True, on_click=_go_start)

with nav3:
    st.selectbox("Jump to slide", range(TOTAL_SLIDES),
                  index=slide, format_func=lambda x: f"{x+1}. {SLIDE_TITLES[x]}",
                  key="jump_slide", label_visibility="collapsed", on_change=_jump_to)

with nav4:
    if slide < TOTAL_SLIDES - 1:
        st.button("⏭ End", key="last", use_container_width=True, on_click=_go_end)

with nav5:
    if slide < TOTAL_SLIDES - 1:
        st.button("Next ▶", key="next", type="primary", use_container_width=True, on_click=_go_next)

# Action buttons
st.markdown("---")
act1, act2, act3, act4 = st.columns(4)
act1.page_link("pages/61_📁_Document_Hub.py", label="Docs", icon="📄")
act2.page_link("pages/30_💰_Financial.py", label="Finance", icon="💰")
act3.page_link("pages/81_🤖_AI_Advisor.py", label="AI", icon="🤖")
act4.page_link("pages/10_📝_Project_Setup.py", label="Setup", icon="📝")

# ══════════════════════════════════════════════════════════════════════
# AI MEETING COPILOT (Sidebar — Available on EVERY slide)
# ══════════════════════════════════════════════════════════════════════
try:
    from engines.ai_engine import is_ai_available
    if is_ai_available():
        from engines.meeting_copilot import (live_qa, handle_objection, COMMON_OBJECTIONS,
            generate_cma_narrative, compare_competitor, generate_meeting_minutes,
            analyze_govt_schemes, generate_investment_thesis, narrate_scenario)

        st.sidebar.markdown("---")
        st.sidebar.markdown("### AI Meeting Copilot")

        audience = st.sidebar.selectbox("Presenting to:",
            ["Investor", "Bank Officer", "Govt Officer", "Farmer", "Competitor Client"],
            key="audience_type")

        copilot_mode = st.sidebar.radio("Mode:",
            ["Live Q&A", "Handle Objection", "CMA for Bank", "Compare Competitor",
             "Govt Schemes", "Investment Thesis", "Meeting Notes"],
            key="copilot_mode")

        if copilot_mode == "Live Q&A":
            qa_input = st.sidebar.text_input("Client's question:", key="qa_input",
                                              placeholder="Type what the client asked...")
            if qa_input and st.sidebar.button("Answer Now", key="qa_go", type="primary"):
                with st.sidebar:
                    with st.spinner("AI thinking..."):
                        answer, prov = live_qa(qa_input, audience, cfg)
                    st.markdown(answer)
                    st.caption(f"Powered by {prov}")

        elif copilot_mode == "Handle Objection":
            obj_list = COMMON_OBJECTIONS.get(audience.split()[0], COMMON_OBJECTIONS["Investor"])
            obj_select = st.sidebar.selectbox("Common objections:", ["Custom..."] + obj_list, key="obj_sel")
            if obj_select == "Custom...":
                obj_text = st.sidebar.text_input("Type objection:", key="obj_custom")
            else:
                obj_text = obj_select
            if obj_text and st.sidebar.button("Destroy Objection", key="obj_go", type="primary"):
                with st.sidebar:
                    with st.spinner("Building response..."):
                        response, prov = handle_objection(obj_text, audience, cfg)
                    st.markdown(response)

        elif copilot_mode == "CMA for Bank":
            if st.sidebar.button("Generate CMA Narrative", key="cma_go", type="primary"):
                with st.sidebar:
                    with st.spinner("Generating bank-ready CMA..."):
                        cma, prov = generate_cma_narrative(cfg)
                    st.markdown(cma)

        elif copilot_mode == "Compare Competitor":
            comp_names = [c["name"] for c in COMPETITORS]
            comp_sel = st.sidebar.selectbox("Select competitor:", comp_names, key="comp_sel")
            if st.sidebar.button("Compare Now", key="comp_go", type="primary"):
                with st.sidebar:
                    with st.spinner("Analyzing..."):
                        comp_result, prov = compare_competitor(comp_sel, audience, cfg)
                    st.markdown(comp_result)

        elif copilot_mode == "Govt Schemes":
            if st.sidebar.button("Analyze All Schemes", key="scheme_go", type="primary"):
                with st.sidebar:
                    with st.spinner("Analyzing eligibility..."):
                        schemes, prov = analyze_govt_schemes(cfg)
                    st.markdown(schemes)

        elif copilot_mode == "Investment Thesis":
            if st.sidebar.button("Generate Thesis", key="thesis_go", type="primary"):
                with st.sidebar:
                    with st.spinner("Building investment case..."):
                        thesis, prov = generate_investment_thesis(cfg, audience)
                    st.markdown(thesis)

        elif copilot_mode == "Meeting Notes":
            notes = st.sidebar.text_area("Type meeting notes:", height=150, key="meeting_notes",
                placeholder="Client interested in 20TPD\nAsked about DSCR\nWants DPR by Friday...")
            if notes and st.sidebar.button("Generate Minutes", key="notes_go", type="primary"):
                with st.sidebar:
                    with st.spinner("Generating minutes..."):
                        minutes, prov = generate_meeting_minutes(notes, cfg)
                    st.markdown(minutes)

    else:
        st.sidebar.markdown("---")
        st.sidebar.info("Add API keys in AI Settings for AI Meeting Copilot")
        st.sidebar.page_link("pages/83_🔑_AI_Settings.py", label="AI Settings", icon="🔑")
except Exception:
    pass

st.caption(f"{COMPANY['name']} | {COMPANY['owner']} | {COMPANY['phone']} | Consultant Presentation System")


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_01_🎯_P", type="primary")
with _ex2:
    if st.button("Print", key="exp_prt_01🎯P"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)
