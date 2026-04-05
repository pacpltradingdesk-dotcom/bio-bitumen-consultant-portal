"""
Environmental Impact Dashboard — CO2 Savings, Carbon Credits, ESG Score
========================================================================
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from state_manager import get_config, init_state
from config import COMPANY, ENVIRONMENTAL_FACTORS

st.set_page_config(page_title="Environmental Impact", page_icon="🌱", layout="wide")
init_state()
cfg = get_config()
# Fix metric truncation
try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass

ef = ENVIRONMENTAL_FACTORS

st.title("Environmental Impact Dashboard")
st.markdown(f"**CO2 Savings | Carbon Credits | Stubble Burning | ESG Score — {cfg['capacity_tpd']:.0f} TPD Plant**")
st.markdown("---")

# ── Calculations ─────────────────────────────────────────────────────
tpd = cfg["capacity_tpd"]
working_days = 300
annual_output_mt = tpd * working_days
co2_saved_annual = annual_output_mt * ef["co2_saved_per_mt_bio_bitumen"]
stubble_diverted = annual_output_mt * ef["stubble_diverted_per_mt_output"]
carbon_credit_usd = co2_saved_annual * ef["carbon_credit_rate_usd"]
carbon_credit_inr = carbon_credit_usd * ef["usd_inr_for_calc"]

# ══════════════════════════════════════════════════════════════════════
# TOP IMPACT METRICS
# ══════════════════════════════════════════════════════════════════════
st.markdown(f"""
<div style="background: linear-gradient(135deg, #006633 0%, #009966 50%, #00cc88 100%);
            padding: 25px 35px; border-radius: 15px; color: white; margin-bottom: 20px;">
    <h2 style="color: white; margin: 0;">Annual Environmental Impact — {tpd:.0f} TPD Plant</h2>
    <div style="display: flex; gap: 40px; margin-top: 15px; flex-wrap: wrap;">
        <div><span style="font-size: 2.2em; font-weight: bold;">{co2_saved_annual:,.0f}</span><br>Tonnes CO2 Saved/Year</div>
        <div><span style="font-size: 2.2em; font-weight: bold;">{stubble_diverted:,.0f}</span><br>MT Stubble Diverted</div>
        <div><span style="font-size: 2.2em; font-weight: bold;">Rs {carbon_credit_inr/100000:.1f}L</span><br>Carbon Credit Revenue</div>
        <div><span style="font-size: 2.2em; font-weight: bold;">{ef['water_saved_pct_vs_conventional']}%</span><br>Less Water Usage</div>
    </div>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════
# CO2 SAVINGS CALCULATOR
# ══════════════════════════════════════════════════════════════════════
st.subheader("1. CO2 Savings Calculator")

calc1, calc2 = st.columns([1, 2])
with calc1:
    calc_tpd = st.slider("Plant Capacity (TPD)", 5, 100, int(tpd), 5, key="env_tpd")
    calc_days = st.slider("Working Days/Year", 200, 365, 300, 10, key="env_days")
    bio_pct = st.slider("Bio-Bitumen Blend %", 10, 40, 20, 5, key="env_blend")

with calc2:
    calc_output = calc_tpd * calc_days
    calc_co2 = calc_output * ef["co2_saved_per_mt_bio_bitumen"] * (bio_pct / 20)
    calc_credit = calc_co2 * ef["carbon_credit_rate_usd"] * ef["usd_inr_for_calc"]

    # Gauge for CO2
    fig_co2 = go.Figure(go.Indicator(
        mode="gauge+number",
        value=calc_co2,
        title={"text": "Annual CO2 Saved (Tonnes)"},
        number={"suffix": " T"},
        gauge={"axis": {"range": [0, 15000]},
               "bar": {"color": "#006633"},
               "steps": [{"range": [0, 3000], "color": "#ccffe6"},
                          {"range": [3000, 8000], "color": "#99ffcc"},
                          {"range": [8000, 15000], "color": "#66ff99"}]},
    ))
    fig_co2.update_layout(height=280, margin=dict(t=60, b=10))
    st.plotly_chart(fig_co2, width="stretch")

    r1, r2, r3 = st.columns(3)
    r1.metric("Annual Output", f"{calc_output:,} MT")
    r2.metric("Carbon Credits", f"Rs {calc_credit/100000:.1f} Lac/yr")
    r3.metric("Equivalent Trees", f"{calc_co2 * 50:,.0f}")  # ~50 trees per tonne CO2

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# STUBBLE BURNING IMPACT
# ══════════════════════════════════════════════════════════════════════
st.subheader("2. Stubble Burning Reduction")

stub1, stub2 = st.columns(2)

with stub1:
    st.markdown(f"""
    **India's Crop Residue Problem:**
    - Annual stubble burned: **{ef['annual_stubble_burning_india_mt']/10000000:.0f} Crore MT**
    - Causes: Air pollution (Delhi AQI crisis), soil degradation, health issues
    - Government spends Rs 1,000+ Cr/year on anti-burning campaigns

    **Bio-Bitumen Solution:**
    - Your {calc_tpd} TPD plant diverts: **{calc_tpd * calc_days * ef['stubble_diverted_per_mt_output']:,.0f} MT/year**
    - Equivalent to **{calc_tpd * calc_days * ef['stubble_diverted_per_mt_output'] / 5:.0f} acres** of farmland cleared
    - Farmers earn Rs 2,000-3,000/MT instead of burning for free
    """)

with stub2:
    # Stubble impact visualization
    stubble_data = pd.DataFrame([
        {"Use": "Bio-Bitumen Feedstock", "MT": calc_tpd * calc_days * ef["stubble_diverted_per_mt_output"]},
        {"Use": "Biochar (Soil Amendment)", "MT": calc_tpd * calc_days * 0.3 * ef["stubble_diverted_per_mt_output"]},
        {"Use": "Syngas (Captive Fuel)", "MT": calc_tpd * calc_days * 0.25 * ef["stubble_diverted_per_mt_output"]},
    ])
    fig_stub = px.bar(stubble_data, x="Use", y="MT", color="Use",
                       title="Crop Residue Utilization (MT/Year)",
                       color_discrete_sequence=["#006633", "#009966", "#00cc88"])
    fig_stub.update_layout(template="plotly_white", height=350, showlegend=False)
    st.plotly_chart(fig_stub, width="stretch")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# CARBON CREDIT REVENUE
# ══════════════════════════════════════════════════════════════════════
st.subheader("3. Carbon Credit Revenue Projection")

years = list(range(1, 8))
util_rates = [0.40, 0.55, 0.70, 0.80, 0.85, 0.90, 0.90]
credit_data = []
cumulative = 0
for yr, util in zip(years, util_rates):
    annual_co2 = calc_tpd * calc_days * util * ef["co2_saved_per_mt_bio_bitumen"] * (bio_pct / 20)
    annual_rev = annual_co2 * ef["carbon_credit_rate_usd"] * ef["usd_inr_for_calc"]
    cumulative += annual_rev
    credit_data.append({
        "Year": yr, "CO2 Saved (T)": round(annual_co2, 0),
        "Revenue (Rs Lac)": round(annual_rev / 100000, 1),
        "Cumulative (Rs Lac)": round(cumulative / 100000, 1),
    })

credit_df = pd.DataFrame(credit_data)

fig_credit = go.Figure()
fig_credit.add_trace(go.Bar(x=credit_df["Year"], y=credit_df["Revenue (Rs Lac)"],
                              name="Annual Revenue", marker_color="#006633"))
fig_credit.add_trace(go.Scatter(x=credit_df["Year"], y=credit_df["Cumulative (Rs Lac)"],
                                  name="Cumulative", mode="lines+markers",
                                  line=dict(color="#00cc88", width=3)))
fig_credit.update_layout(title="7-Year Carbon Credit Revenue (Rs Lac)",
                           template="plotly_white", height=400,
                           xaxis_title="Year", yaxis_title="Rs Lac")
st.plotly_chart(fig_credit, width="stretch")

st.dataframe(credit_df, width="stretch", hide_index=True)

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# ESG SCORE
# ══════════════════════════════════════════════════════════════════════
st.subheader("4. ESG Compliance Score")

esg_categories = ["Carbon Reduction", "Waste Utilization", "Water Efficiency",
                    "Energy Efficiency", "Farmer Livelihood", "Green Mandate Compliance"]
esg_scores = [85, 90, 70, 75, 88, 80]
esg_avg = sum(esg_scores) / len(esg_scores)

e1, e2 = st.columns([1, 1])

with e1:
    fig_esg = go.Figure(go.Scatterpolar(
        r=esg_scores + [esg_scores[0]],
        theta=esg_categories + [esg_categories[0]],
        fill="toself", line=dict(color="#006633", width=2),
        fillcolor="rgba(0,102,51,0.2)",
    ))
    fig_esg.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
        title=f"ESG Score: {esg_avg:.0f}/100",
        template="plotly_white", height=400,
    )
    st.plotly_chart(fig_esg, width="stretch")

with e2:
    st.markdown(f"""
    **ESG Score: {esg_avg:.0f}/100** {'🟢 Excellent' if esg_avg >= 80 else '🟡 Good' if esg_avg >= 60 else '🔴 Needs Work'}

    | Category | Score | Status |
    |----------|-------|--------|
    | Carbon Reduction | {esg_scores[0]}/100 | {'Pass' if esg_scores[0] >= 70 else 'Improve'} |
    | Waste Utilization | {esg_scores[1]}/100 | {'Pass' if esg_scores[1] >= 70 else 'Improve'} |
    | Water Efficiency | {esg_scores[2]}/100 | {'Pass' if esg_scores[2] >= 70 else 'Improve'} |
    | Energy Efficiency | {esg_scores[3]}/100 | {'Pass' if esg_scores[3] >= 70 else 'Improve'} |
    | Farmer Livelihood | {esg_scores[4]}/100 | {'Pass' if esg_scores[4] >= 70 else 'Improve'} |
    | Green Mandate | {esg_scores[5]}/100 | {'Pass' if esg_scores[5] >= 70 else 'Improve'} |

    **NHAI Green Mandate:**
    - Target: {ef['nhai_green_mandate_replacement_pct']}% bio-bitumen by 2030
    - Your compliance: **Ready**
    """)

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# BIO-BITUMEN vs CONVENTIONAL COMPARISON
# ══════════════════════════════════════════════════════════════════════
st.subheader("5. Bio-Bitumen vs Conventional — Environmental Comparison")

compare_data = pd.DataFrame([
    {"Parameter": "CO2 Emissions (kg/MT)", "Conventional": 450, "Bio-Bitumen": 290},
    {"Parameter": "Water Usage (KL/MT)", "Conventional": 3.5, "Bio-Bitumen": 3.0},
    {"Parameter": "Energy (kWh/MT)", "Conventional": 120, "Bio-Bitumen": 108},
    {"Parameter": "Waste Generated (kg/MT)", "Conventional": 50, "Bio-Bitumen": 15},
    {"Parameter": "Renewable Content (%)", "Conventional": 0, "Bio-Bitumen": 20},
])

fig_comp = go.Figure()
fig_comp.add_trace(go.Bar(x=compare_data["Parameter"], y=compare_data["Conventional"],
                            name="Conventional Bitumen", marker_color="#CC3333"))
fig_comp.add_trace(go.Bar(x=compare_data["Parameter"], y=compare_data["Bio-Bitumen"],
                            name="Bio-Bitumen", marker_color="#006633"))
fig_comp.update_layout(title="Environmental Footprint Comparison",
                         barmode="group", template="plotly_white", height=400)
st.plotly_chart(fig_comp, width="stretch")

st.markdown("---")
st.caption(f"{COMPANY['name']} | {COMPANY['owner']} | {COMPANY['phone']} | Green Technology Division")

# ── Export Section ────────────────────────────────────────────────────
st.markdown("---")
st.subheader("Export")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_65_Env", type="primary")
with _ex2:
    if st.button("Print Page", key="exp_print_analysis"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)


# ── AI Skill: ESG Report Section ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: ESG Report Section"):
            if st.button("Generate", type="primary", key="ai_65Envir"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: ESG Report Section. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass
