"""
Risk Assessment Matrix — Heatmap, Mitigation Strategies, Risk Score
====================================================================
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import numpy as np
from state_manager import init_state, get_config
from config import COMPANY, RISK_REGISTRY

st.set_page_config(page_title="Risk Assessment", page_icon="⚠️", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass


st.title("Risk Assessment Matrix")
st.markdown("**20 Project Risks | Probability x Impact Heatmap | Mitigation Strategies**")
st.markdown("---")

# ── Prepare Data ─────────────────────────────────────────────────────
risk_df = pd.DataFrame(RISK_REGISTRY)
risk_df["score"] = risk_df["probability"] * risk_df["impact"]
risk_df["severity"] = risk_df["score"].apply(
    lambda x: "Critical" if x >= 15 else ("High" if x >= 10 else ("Medium" if x >= 6 else "Low")))

# ══════════════════════════════════════════════════════════════════════
# TOP METRICS
# ══════════════════════════════════════════════════════════════════════
critical = len(risk_df[risk_df["severity"] == "Critical"])
high = len(risk_df[risk_df["severity"] == "High"])
medium = len(risk_df[risk_df["severity"] == "Medium"])
low = len(risk_df[risk_df["severity"] == "Low"])
avg_score = risk_df["score"].mean()
overall_rating = "Low Risk" if avg_score < 6 else ("Moderate Risk" if avg_score < 10 else "High Risk")

m1, m2, m3, m4, m5 = st.columns(5)
m1.metric("Total Risks", len(risk_df))
m2.metric("Critical/High", f"{critical + high}")
m3.metric("Medium", medium)
m4.metric("Low", low)
m5.metric("Overall Rating", overall_rating)

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# RISK HEATMAP (Probability x Impact Matrix)
# ══════════════════════════════════════════════════════════════════════
st.subheader("Risk Heatmap — Probability vs Impact")

# Build 5x5 matrix
matrix = np.zeros((5, 5))
risk_labels = [["" for _ in range(5)] for _ in range(5)]
for _, r in risk_df.iterrows():
    p = r["probability"] - 1
    i = r["impact"] - 1
    matrix[p][i] += 1
    short_name = r["risk"][:25] + "..."
    if risk_labels[p][i]:
        risk_labels[p][i] += f"\n{short_name}"
    else:
        risk_labels[p][i] = short_name

# Color matrix (score = p*i)
score_matrix = np.array([[((p+1)*(i+1)) for i in range(5)] for p in range(5)])

fig_heat = go.Figure(data=go.Heatmap(
    z=score_matrix,
    x=["1-Very Low", "2-Low", "3-Medium", "4-High", "5-Very High"],
    y=["1-Rare", "2-Unlikely", "3-Possible", "4-Likely", "5-Almost Certain"],
    text=[[f"{int(matrix[p][i])} risks" if matrix[p][i] > 0 else "" for i in range(5)] for p in range(5)],
    texttemplate="%{text}",
    colorscale=[[0, "#ccffcc"], [0.3, "#ffffcc"], [0.6, "#ffcc99"], [1.0, "#ff6666"]],
    showscale=True,
    colorbar=dict(title="Risk Score"),
))
fig_heat.update_layout(
    title="Risk Matrix (Probability x Impact)",
    xaxis_title="IMPACT",
    yaxis_title="PROBABILITY",
    template="plotly_white",
    height=450,
)
st.plotly_chart(fig_heat, width="stretch")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# RISK BY CATEGORY
# ══════════════════════════════════════════════════════════════════════
cat_col, sev_col = st.columns(2)

with cat_col:
    st.subheader("Risks by Category")
    cat_data = risk_df.groupby("category").agg(
        Count=("risk", "count"),
        Avg_Score=("score", "mean"),
    ).reset_index().sort_values("Avg_Score", ascending=False)

    fig_cat = px.bar(cat_data, x="category", y="Count", color="Avg_Score",
                      title="Risk Count by Category", text="Count",
                      color_continuous_scale=["#00AA44", "#FF8800", "#CC3333"])
    fig_cat.update_layout(template="plotly_white", height=400)
    st.plotly_chart(fig_cat, width="stretch")

with sev_col:
    st.subheader("Severity Distribution")
    sev_data = risk_df["severity"].value_counts().reset_index()
    sev_data.columns = ["Severity", "Count"]
    color_map = {"Critical": "#CC3333", "High": "#FF8800", "Medium": "#FFcc00", "Low": "#00AA44"}

    fig_sev = px.pie(sev_data, names="Severity", values="Count",
                      title="Risk Severity Distribution",
                      color="Severity", color_discrete_map=color_map)
    fig_sev.update_layout(template="plotly_white", height=400)
    st.plotly_chart(fig_sev, width="stretch")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# DETAILED RISK TABLE WITH FILTERS
# ══════════════════════════════════════════════════════════════════════
st.subheader("Detailed Risk Register")

f1, f2 = st.columns(2)
with f1:
    filter_cat = st.multiselect("Filter by Category", risk_df["category"].unique().tolist(), key="risk_cat")
with f2:
    filter_sev = st.multiselect("Filter by Severity", ["Critical", "High", "Medium", "Low"], key="risk_sev")

filtered = risk_df.copy()
if filter_cat:
    filtered = filtered[filtered["category"].isin(filter_cat)]
if filter_sev:
    filtered = filtered[filtered["severity"].isin(filter_sev)]

filtered_sorted = filtered.sort_values("score", ascending=False)

for _, r in filtered_sorted.iterrows():
    sev = r["severity"]
    icon = {"Critical": "🔴", "High": "🟠", "Medium": "🟡", "Low": "🟢"}[sev]
    score = r["score"]

    with st.expander(f"{icon} [{sev}] {r['category']} — {r['risk'][:70]} (Score: {score}/25)"):
        rc1, rc2 = st.columns(2)
        with rc1:
            st.markdown(f"**Category:** {r['category']}")
            st.markdown(f"**Probability:** {r['probability']}/5")
            st.markdown(f"**Impact:** {r['impact']}/5")
            st.markdown(f"**Risk Score:** {score}/25")
        with rc2:
            st.markdown(f"**Mitigation Strategy:**")
            st.info(r["mitigation"])

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# RISK SCORE SUMMARY
# ══════════════════════════════════════════════════════════════════════
st.subheader("Risk Score Summary")

fig_bar = px.bar(filtered_sorted, y="risk", x="score", orientation="h",
                  color="category", title="All Risks Ranked by Score",
                  labels={"risk": "Risk", "score": "Score (P x I)"},
                  color_discrete_sequence=["#003366", "#CC3333", "#FF8800", "#006699", "#00AA44", "#AA3366"])
fig_bar.update_layout(template="plotly_white", height=600, yaxis={"categoryorder": "total ascending"})
st.plotly_chart(fig_bar, width="stretch")

st.markdown("---")
st.caption(f"{COMPANY['name']} | Risk Assessment Module")

# ── Export Section ────────────────────────────────────────────────────
st.markdown("---")
st.subheader("Export")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_66_Ris", type="primary")
with _ex2:
    if st.button("Print Page", key="exp_print_analysis"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)


# ── AI Skill: Risk Mitigation Plan ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: Risk Mitigation Plan"):
            if st.button("Generate", type="primary", key="ai_66Risk"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: Risk Mitigation Plan. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass


# ── Next Steps Navigation ──
try:
    from engines.page_navigation import add_next_steps
    add_next_steps(st, "42")
except Exception:
    pass
