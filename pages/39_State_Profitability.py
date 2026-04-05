"""
State-wise Profitability — ALL 18 states ranked by profit + heatmap
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from state_manager import get_config, init_state
from engines.three_process_model import state_profitability
from config import STATE_COSTS

st.set_page_config(page_title="State Profitability", page_icon="📍", layout="wide")
init_state()
cfg = get_config()
# Fix metric truncation
try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass


st.page_link("pages/30_💰_Financial.py", label="← Back to Financial Model", icon="💰")

st.title("State-wise Profitability Analysis")
st.markdown(f"**All 18 states ranked — {cfg['capacity_tpd']:.0f} TPD | Which state gives BEST ROI?**")
st.markdown("---")

# ── Inputs ────────────────────────────────────────────────────────────
col1, col2 = st.columns(2)
with col1:
    tpd = st.number_input("Capacity (TPD)", 5.0, 100.0, float(cfg["capacity_tpd"]), 5.0, key="sp_tpd")
with col2:
    process = st.selectbox("Process Model", [1, 2, 3],
                            format_func=lambda x: {1: "Process 1: Full Chain", 2: "Process 2: Blending Only", 3: "Process 3: Raw Output"}[x])

# ── Calculate for ALL states ──────────────────────────────────────────
with st.spinner("Calculating profitability for 18 states..."):
    results = state_profitability(tpd, process)

df = pd.DataFrame(results)

# ── Top Metrics ───────────────────────────────────────────────────────
best = results[0]
worst = results[-1]

m1, m2, m3, m4 = st.columns(4)
m1.metric("Best State", best["State"], f"ROI: {best['ROI (%)']:.1f}%")
m2.metric("Best Profit/MT", f"Rs {best['Profit/MT']:,.0f}", best["State"])
m3.metric("Worst State", worst["State"], f"ROI: {worst['ROI (%)']:.1f}%")
m4.metric("Avg ROI", f"{df['ROI (%)'].mean():.1f}%")

st.markdown("---")

# ── Heatmap (ROI by state) ────────────────────────────────────────────
st.subheader("Profitability Heatmap — All 18 States")

fig_map = px.bar(df, x="State", y="ROI (%)", color="ROI (%)",
                  color_continuous_scale="RdYlGn", title="State-wise ROI (%) — Highest to Lowest",
                  text="ROI (%)")
fig_map.update_traces(texttemplate="%{text:.1f}%", textposition="outside")
fig_map.update_layout(xaxis_tickangle=-45, template="plotly_white", height=450)
st.plotly_chart(fig_map, width="stretch")

st.markdown("---")

# ── Full State Table ──────────────────────────────────────────────────
st.subheader("Complete State-wise Report")
st.dataframe(df, width="stretch", hide_index=True)

st.markdown("---")

# ── Top 5 vs Bottom 5 ────────────────────────────────────────────────
col_top, col_bot = st.columns(2)

with col_top:
    st.subheader("TOP 5 Most Profitable States")
    for i, r in enumerate(results[:5]):
        st.markdown(f"**{i+1}. {r['State']}** — ROI: {r['ROI (%)']:.1f}% | Profit: Rs {r['Profit/MT']:,.0f}/MT | Risk: {r['Risk']}")

with col_bot:
    st.subheader("BOTTOM 5 States")
    for r in results[-5:]:
        st.markdown(f"- **{r['State']}** — ROI: {r['ROI (%)']:.1f}% | Cost: Rs {r['Cost/MT']:,.0f}/MT | Risk: {r['Risk']}")

st.markdown("---")

# ── Cost Drivers Comparison ───────────────────────────────────────────
st.subheader("Cost Drivers — What Makes a State Profitable?")

col_d1, col_d2 = st.columns(2)
with col_d1:
    fig_bio = px.bar(df, x="State", y="Biomass Cost (Rs/MT)", color="Biomass Cost (Rs/MT)",
                      color_continuous_scale="Greens_r", title="Biomass Cost (Lower = Better)")
    fig_bio.update_layout(xaxis_tickangle=-45, template="plotly_white", height=350)
    st.plotly_chart(fig_bio, width="stretch")

with col_d2:
    fig_pow = px.bar(df, x="State", y="Power Rate (Rs/kWh)", color="Power Rate (Rs/kWh)",
                      color_continuous_scale="Reds_r", title="Power Rate (Lower = Better)")
    fig_pow.update_layout(xaxis_tickangle=-45, template="plotly_white", height=350)
    st.plotly_chart(fig_pow, width="stretch")

# Demand chart
fig_dem = px.bar(df, x="State", y="Demand (MT/yr)", color="Demand (MT/yr)",
                  color_continuous_scale="Blues", title="Bitumen Demand (Higher = More Market)")
fig_dem.update_layout(xaxis_tickangle=-45, template="plotly_white", height=350)
st.plotly_chart(fig_dem, width="stretch")

st.markdown("---")

# ── Recommendation ────────────────────────────────────────────────────
st.subheader("Recommendations")

best3 = results[:3]
st.success(f"""
**Best States for {tpd:.0f} TPD Bio-Bitumen Plant (Process {process}):**

1. **{best3[0]['State']}** — ROI: {best3[0]['ROI (%)']:.1f}% | Profit: Rs {best3[0]['Profit/MT']:,.0f}/MT | Biomass: Rs {best3[0]['Biomass Cost (Rs/MT)']:,.0f}/MT
2. **{best3[1]['State']}** — ROI: {best3[1]['ROI (%)']:.1f}% | Profit: Rs {best3[1]['Profit/MT']:,.0f}/MT
3. **{best3[2]['State']}** — ROI: {best3[2]['ROI (%)']:.1f}% | Profit: Rs {best3[2]['Profit/MT']:,.0f}/MT

**Key factors:** Low biomass cost + high govt subsidy + strong road demand
""")

st.markdown("""
| Category | Best State | Why |
|----------|-----------|-----|
| **Lowest Investment** | States with 20-25% subsidy | Chhattisgarh, MP, Rajasthan, UP |
| **Highest Profit** | Lowest biomass cost | UP, Bihar, Punjab, MP |
| **Fastest Setup** | Best logistics + infrastructure | Gujarat, Maharashtra, Karnataka |
| **Govt Support** | Highest subsidy + policy push | Assam (25%), MP (25%), Rajasthan (25%) |
""")

st.caption("Sources: State DISCOM tariffs 2025-26, Labour Bureau Min Wages, State Industrial Policy documents, IBEF, MoRTH road construction data")

# ── Export Section ────────────────────────────────────────────────────
st.markdown("---")
st.subheader("Export")
exp1, exp2, exp3 = st.columns(3)
with exp1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_58_Sta", type="primary")
with exp2:
    if st.button("Download CSV", key="exp_csv_calc"):
        import pandas as pd
        data = {"Metric": ["Capacity", "Investment", "ROI", "IRR", "DSCR", "Break-Even", "Monthly Profit"],
                "Value": [f"{cfg['capacity_tpd']:.0f} TPD", f"Rs {cfg['investment_cr']:.2f} Cr",
                          f"{cfg['roi_pct']:.1f}%", f"{cfg['irr_pct']:.1f}%", f"{cfg['dscr_yr3']:.2f}x",
                          f"{cfg['break_even_months']} months", f"Rs {cfg['monthly_profit_lac']:.1f} Lac"]}
        st.download_button("Download", pd.DataFrame(data).to_csv(index=False), "calculator_export.csv", "text/csv", key="dl_csv_c")
with exp3:
    if st.button("Print", key="exp_print_calc"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)


# ── AI Skill: State Investment Brief ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: State Investment Brief"):
            if st.button("Generate", type="primary", key="ai_58State"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: State Investment Brief. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass
