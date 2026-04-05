"""
Bio Bitumen Consultant Portal — Send to Customer
Email and WhatsApp communication hub for sending packages to customers.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
from state_manager import init_state, get_config
import pandas as pd
from database import (init_db, get_all_customers, get_customer,
                       get_packages_for_customer, get_communications_for_customer,
                       insert_communication)
from master_data_loader import get_plant
from engines.email_engine import send_email, generate_email_body, get_email_config
from engines.whatsapp_engine import generate_whatsapp_message, get_whatsapp_link
from config import COMPANY, CAPACITY_KEYS, CAPACITY_LABELS

st.set_page_config(page_title="Send to Customer", page_icon="📧", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass

init_db()
st.title("Send to Customer")

# ── Select Customer ───────────────────────────────────────────────────
customers = get_all_customers()
if not customers:
    st.warning("No customers found. Add customers in **Customer Manager** first.")
    st.stop()

cust_map = {c["id"]: f"{c['name']} ({c.get('company', 'N/A')})" for c in customers}
selected_id = st.selectbox("Select Customer", options=list(cust_map.keys()),
                            format_func=lambda x: cust_map[x])
customer = get_customer(selected_id)

st.info(f"**{customer['name']}** | {customer.get('email', 'No email')} | {customer.get('phone', 'No phone')} | {customer.get('state', 'N/A')}")

# ── Select Package (optional) ────────────────────────────────────────
packages = get_packages_for_customer(selected_id)
selected_pkg = None
if packages:
    pkg_map = {p["id"]: f"#{p['id']} — {p['capacity']} ({p['recipient_type']}) — {p['created_at']}" for p in packages}
    pkg_id = st.selectbox("Select Package (optional)", options=[None] + list(pkg_map.keys()),
                           format_func=lambda x: pkg_map.get(x, "None — send without package"))
    if pkg_id:
        selected_pkg = next(p for p in packages if p["id"] == pkg_id)

# ── Capacity for message content ─────────────────────────────────────
cap_key = (selected_pkg["capacity"] if selected_pkg else
           customer.get("interested_capacity", "20MT"))
if cap_key not in CAPACITY_KEYS:
    cap_key = "20MT"
plant = get_plant(cap_key)

st.markdown("---")

# ── Tabs: Email / WhatsApp / Log ──────────────────────────────────────
tab_email, tab_whatsapp, tab_settings, tab_log = st.tabs(
    ["Email", "WhatsApp", "Email Settings", "Communication Log"])

# ══════════════════════════════════════════════════════════════════════
# TAB: EMAIL
# ══════════════════════════════════════════════════════════════════════
with tab_email:
    st.subheader("Send Email")

    config = get_email_config()
    if not config["sender_email"]:
        st.warning("Email not configured. Go to **Email Settings** tab to set up.")

    to_email = st.text_input("To", value=customer.get("email", ""), key="email_to")
    subject = st.text_input("Subject",
                             value=f"Bio-Modified Bitumen {plant['label']} — Project Information | {COMPANY['trade_name']}",
                             key="email_subject")

    # Generate default body
    default_body = generate_email_body(customer, plant, COMPANY)
    body = st.text_area("Email Body (HTML)", value=default_body, height=300, key="email_body")

    # Attachments
    attachments = []
    if selected_pkg and selected_pkg.get("output_folder"):
        pkg_dir = selected_pkg["output_folder"]
        if os.path.exists(pkg_dir):
            pkg_files = [os.path.join(pkg_dir, f) for f in os.listdir(pkg_dir) if not f.startswith(".")]
            st.markdown(f"**Attachments from package:** {len(pkg_files)} files")
            for f in pkg_files[:10]:
                st.caption(f"  - {os.path.basename(f)}")
            if len(pkg_files) > 10:
                st.caption(f"  ... and {len(pkg_files) - 10} more")
            attachments = pkg_files

    if st.button("Send Email", type="primary", key="send_email_btn"):
        if not to_email:
            st.error("Please enter a recipient email address.")
        else:
            with st.spinner("Sending email..."):
                success, message = send_email(to_email, subject, body, attachments)
                if success:
                    st.success(message)
                    insert_communication({
                        "customer_id": selected_id,
                        "package_id": selected_pkg["id"] if selected_pkg else None,
                        "channel": "email",
                        "subject": subject,
                        "content_summary": f"Email to {to_email} with {len(attachments)} attachments",
                        "attachments": [os.path.basename(a) for a in attachments],
                        "status": "sent",
                    })
                else:
                    st.error(message)
                    insert_communication({
                        "customer_id": selected_id,
                        "package_id": selected_pkg["id"] if selected_pkg else None,
                        "channel": "email",
                        "subject": subject,
                        "content_summary": f"Failed: {message}",
                        "status": "failed",
                    })


# ══════════════════════════════════════════════════════════════════════
# TAB: WHATSAPP
# ══════════════════════════════════════════════════════════════════════
with tab_whatsapp:
    st.subheader("Send via WhatsApp")

    phone = st.text_input("WhatsApp Number", value=customer.get("whatsapp", "") or customer.get("phone", ""),
                           placeholder="+91 XXXXX XXXXX", key="wa_phone")

    # Generate message
    wa_message = generate_whatsapp_message(customer, plant, COMPANY)
    edited_message = st.text_area("Message", value=wa_message, height=400, key="wa_msg")

    if phone:
        wa_link = get_whatsapp_link(phone, edited_message)
        st.markdown(f"[Open WhatsApp Web]({wa_link})")

        if st.button("Log WhatsApp Message", type="primary", key="wa_log_btn"):
            insert_communication({
                "customer_id": selected_id,
                "package_id": selected_pkg["id"] if selected_pkg else None,
                "channel": "whatsapp",
                "subject": f"WhatsApp to {phone}",
                "content_summary": edited_message[:200],
                "status": "sent",
            })
            st.success("WhatsApp communication logged. Click the link above to send via WhatsApp Web.")
    else:
        st.warning("Enter a phone number to generate the WhatsApp link.")


# ══════════════════════════════════════════════════════════════════════
# TAB: EMAIL SETTINGS
# ══════════════════════════════════════════════════════════════════════
with tab_settings:
    st.subheader("Email Configuration")
    st.markdown("Configure your SMTP settings for sending emails. For Gmail, use an App Password.")

    with st.form("email_settings"):
        smtp_server = st.text_input("SMTP Server", value=st.session_state.get("smtp_server", "smtp.gmail.com"))
        smtp_port = st.number_input("SMTP Port", value=st.session_state.get("smtp_port", 587))
        sender_email = st.text_input("Your Email", value=st.session_state.get("sender_email", ""))
        sender_password = st.text_input("App Password", type="password",
                                         value=st.session_state.get("sender_password", ""))
        sender_name = st.text_input("Sender Name", value=st.session_state.get("sender_name", "PPS Anantams"))

        if st.form_submit_button("Save Settings"):
            st.session_state["smtp_server"] = smtp_server
            st.session_state["smtp_port"] = smtp_port
            st.session_state["sender_email"] = sender_email
            st.session_state["sender_password"] = sender_password
            st.session_state["sender_name"] = sender_name
            st.success("Email settings saved for this session.")


# ══════════════════════════════════════════════════════════════════════
# TAB: COMMUNICATION LOG
# ══════════════════════════════════════════════════════════════════════
with tab_log:
    st.subheader("Communication History")

    comms = get_communications_for_customer(selected_id)
    if comms:
        log_data = []
        for c in comms:
            log_data.append({
                "Date": c.get("sent_at", ""),
                "Channel": c.get("channel", ""),
                "Subject": c.get("subject", ""),
                "Status": c.get("status", ""),
                "Summary": c.get("content_summary", "")[:80],
            })
        st.dataframe(pd.DataFrame(log_data), width="stretch", hide_index=True)
    else:
        st.info("No communications logged for this customer yet.")


# ══════════════════════════════════════════════════════════════════════
# TAB: AUTO-SEQUENCE FOLLOW-UPS
# ══════════════════════════════════════════════════════════════════════
tab_auto = st.tabs(["Auto Follow-Up Sequence"])
if tab_auto:
    pass

st.markdown("---")
st.subheader("Auto Follow-Up Sequence Planner")
st.markdown("""
**Automated communication sequence for each customer:**
""")

sequence_data = pd.DataFrame([
    {"Day": "Day 0", "Action": "Send initial package (DPR + Financial + Pitch)", "Channel": "Email + WhatsApp", "Status": "Manual trigger above"},
    {"Day": "Day 1", "Action": "WhatsApp: 'Did you receive the documents?'", "Channel": "WhatsApp", "Status": "Reminder set"},
    {"Day": "Day 3", "Action": "Follow-up email: Key highlights + ROI summary", "Channel": "Email", "Status": "Template ready"},
    {"Day": "Day 7", "Action": "Call + WhatsApp: 'Any questions? Ready for discussion?'", "Channel": "Phone + WhatsApp", "Status": "Reminder set"},
    {"Day": "Day 14", "Action": "Send updated pricing / market data", "Channel": "Email", "Status": "Auto from dashboard"},
    {"Day": "Day 21", "Action": "Invitation for site visit / video call", "Channel": "Email + WhatsApp", "Status": "Template ready"},
    {"Day": "Day 30", "Action": "Final follow-up: Special offer / urgency", "Channel": "Email + WhatsApp", "Status": "Template ready"},
])
st.dataframe(sequence_data, width="stretch", hide_index=True)

# Quick follow-up templates
st.markdown("---")
st.subheader("Quick Follow-Up Templates")

templates = {
    "Day 1 — Delivery Confirmation": f"Dear {{customer_name}},\n\nThis is {COMPANY['owner']} from {COMPANY['trade_name']}.\n\nI sent you the Bio-Bitumen plant project details yesterday. Did you receive the documents?\n\nKey highlights:\n- Investment: Rs {{investment}} Crore\n- ROI: {{roi}}%\n- Break-even: {{months}} months\n\nPlease let me know if you have any questions.\n\nRegards,\n{COMPANY['owner']}\n{COMPANY['phone']}",
    "Day 3 — ROI Highlight": f"Dear {{customer_name}},\n\nQuick reminder about the Bio-Bitumen opportunity:\n\n- India imports 49% of bitumen (Rs 25,000 Cr/year)\n- 130-216 plants needed in next 5-7 years\n- Your ROI: {{roi}}% with break-even in {{months}} months\n\nWould you like to schedule a 30-minute discussion?\n\nBest regards,\n{COMPANY['owner']}",
    "Day 7 — Discussion Invite": f"Dear {{customer_name}},\n\nFollowing up on the Bio-Bitumen project proposal.\n\nI have complete project documentation ready:\n- DPR with financial model\n- 117 engineering drawings\n- Bank-ready loan proposal\n- NHAI compliance roadmap\n\nShall we schedule a call this week?\n\n{COMPANY['owner']}\n{COMPANY['phone']}",
    "Day 14 — Market Update": f"Dear {{customer_name}},\n\nMarket update: VG30 bitumen price is Rs {{vg30_price}}/MT.\nBio-bitumen at Rs 35,000/MT offers {{discount}}% cost advantage.\n\nNHAI green mandate is creating strong demand.\n\nLet's discuss how to capitalize on this opportunity.\n\n{COMPANY['owner']}",
}

for title, template in templates.items():
    with st.expander(title):
        filled = template.replace("{customer_name}", customer.get("name", "Sir/Madam") if customer else "Sir/Madam")
        filled = filled.replace("{investment}", f"{cfg['investment_cr']:.2f}")
        filled = filled.replace("{roi}", f"{cfg['roi_pct']:.1f}")
        filled = filled.replace("{months}", f"{cfg['break_even_months']}")
        filled = filled.replace("{vg30_price}", "50,000")
        filled = filled.replace("{discount}", "30")
        st.text_area("Template (editable)", value=filled, height=200, key=f"tmpl_{title[:10]}")
        if customer and customer.get("phone"):
            from engines.whatsapp_engine import get_whatsapp_link
            wa_link = get_whatsapp_link(customer["phone"], filled)
            st.markdown(f"[Send via WhatsApp]({wa_link})")



# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    if st.button("Download Excel", type="primary", key="exp_xl_43Sen"):
        try:
            import io
            from openpyxl import Workbook
            _wb = Workbook()
            _ws = _wb.active
            _ws.title = "Export"
            _ws.cell(row=1, column=1, value="Bio Bitumen Export")
            _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
            _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
            _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
            _buf = io.BytesIO()
            _wb.save(_buf)
            _buf.seek(0)
            st.download_button("Download", _buf.getvalue(), "export.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_xl_43Sen")
        except Exception as _e:
            st.error(f"Export failed: {_e}")
with _ex2:
    if st.button("Print", key="exp_prt_43Sen"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)


# ── AI Assist ────────────────────────────────────────────────────
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("AI Assist"):
            if st.button("Generate AI Summary", type="primary", key="ai_43Sen"):
                with st.spinner("AI working..."):
                    _p = f"Summarize this section for a {cfg.get('capacity_tpd',20):.0f} TPD bio-bitumen plant in {cfg.get('state','')}. Investment Rs {cfg.get('investment_cr',8):.2f} Cr. Professional consultant format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 800)
                if _r:
                    st.markdown(_r)
except Exception:
    pass
