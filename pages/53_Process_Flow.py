"""
Process Flow — 13-step detailed flow with plant layout sections
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from state_manager import get_config, init_state

st.set_page_config(page_title="Process Flow", page_icon="🔄", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass

st.title("Process Flow & Plant Sections")
st.markdown(f"**Complete 13-Step Bio-Bitumen Manufacturing Process — {cfg['capacity_tpd']:.0f} TPD**")
st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# PLANT SECTIONS (from uploaded 30TPD layout images)
# ═══════════════════════════════════════════════════════════════════
st.subheader("Plant Layout Sections (from 30 TPD Reference Design)")

sections = pd.DataFrame([
    {"#": 1, "Section": "RAW MATERIAL RECEIVING AREA", "Color": "Green", "Key Equipment": "Truck Access Road, Weigh Bridge, Open Storage Yard, Covered Shed", "Area": "~15% of plot"},
    {"#": 2, "Section": "RAW MATERIAL PROCESSING", "Color": "Yellow", "Key Equipment": "Biomass Receiving Hopper, Belt Conveyor, Shredder, Hammer Mill, Rotary Dryer, Cyclone Dust Collector, Storage Silo", "Area": "~20% of plot"},
    {"#": 3, "Section": "PYROLYSIS REACTOR SECTION", "Color": "Red (Core)", "Key Equipment": "Pyrolysis Reactor R-101, Thermic Fluid Heater, Bio-Oil Condenser HE-101, Gas Cooling Tower, 5m Safety Zone", "Area": "~15% of plot"},
    {"#": 4, "Section": "BIO-OIL & BITUMEN PROCESSING", "Color": "Orange", "Key Equipment": "Bio-Oil Intermediate Storage, Bitumen Heating Tank, High Shear Mixer, Colloid Mill, Additive Dosing Tanks, Bio-Bitumen Blending Tank", "Area": "~15% of plot"},
    {"#": 5, "Section": "STORAGE AREA", "Color": "Blue", "Key Equipment": "Heated Bitumen Storage T-201, Bio-Oil Storage T-201, Containment Bund Wall, 4m Fire Spacing, Pipe Racks", "Area": "~10% of plot"},
    {"#": 6, "Section": "POLLUTION CONTROL AREA", "Color": "Brown", "Key Equipment": "Bag Filter, Gas Scrubber, 20m Chimney Stack (minimum height)", "Area": "~5% of plot"},
    {"#": 7, "Section": "PRODUCT DISPATCH AREA", "Color": "Purple", "Key Equipment": "Tanker Loading Platform, Automated Drum Filling Machine, Bitumen Drum Storage, Forklift Area, Truck Parking", "Area": "~10% of plot"},
    {"#": 8, "Section": "UTILITY AREA", "Color": "Gray", "Key Equipment": "DG Backup Generator, PLC Electrical Control Room, Air Compressor, Cooling Tower, Water Tank", "Area": "~5% of plot"},
    {"#": 9, "Section": "ADMIN & SAFETY", "Color": "White", "Key Equipment": "Office Building, Fire Water Tank, Emergency Assembly Point, Fire Extinguishers", "Area": "~5% of plot"},
])
st.dataframe(sections, width="stretch", hide_index=True)

# Section distribution chart
fig_sections = go.Figure(data=[go.Bar(
    x=sections["Section"], y=[15, 20, 15, 15, 10, 5, 10, 5, 5],
    marker_color=["#228B22", "#FFD700", "#CC0000", "#FF8C00", "#4169E1",
                   "#8B4513", "#800080", "#808080", "#FFFFFF"],
    text=[f"{v}%" for v in [15, 20, 15, 15, 10, 5, 10, 5, 5]],
    textposition="outside",
)])
fig_sections.update_layout(title="Plant Area Distribution by Section (%)",
                            template="plotly_white", height=400,
                            xaxis_tickangle=-45, yaxis_title="% of Total Plot")
st.plotly_chart(fig_sections, width="stretch")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# 13-STEP PROCESS FLOW
# ═══════════════════════════════════════════════════════════════════
st.subheader("Complete 13-Step Manufacturing Process")

steps = [
    {"Step": 1, "Process": "BIOMASS COLLECTION", "Section": "Raw Material Receiving", "Description": "Agro-waste (rice straw, bagasse, cotton stalk) collected from farmers within 50 km radius via truck", "Key Equipment": "Truck, Weigh Bridge"},
    {"Step": 2, "Process": "SORTING & INSPECTION", "Section": "Raw Material Receiving", "Description": "Visual inspection, moisture check, foreign material removal", "Key Equipment": "Manual sorting station"},
    {"Step": 3, "Process": "DRYING", "Section": "Raw Material Processing", "Description": "Reduce moisture to <10% using rotary dryer with hot air from syngas combustion", "Key Equipment": "Rotary Dryer (insulated)"},
    {"Step": 4, "Process": "SHREDDING / PELLETIZING", "Section": "Raw Material Processing", "Description": "Size reduction to 10mm diameter, 20-30mm length pellets via hammer mill", "Key Equipment": "Biomass Shredder, Hammer Mill, Storage Silo"},
    {"Step": 5, "Process": "PYROLYSIS", "Section": "Pyrolysis Reactor", "Description": "Thermal decomposition at 400-550°C in oxygen-free reactor. Biomass converts to vapors + char", "Key Equipment": "Pyrolysis Reactor R-101, Thermic Fluid Heater"},
    {"Step": 6, "Process": "CONDENSATION", "Section": "Pyrolysis Reactor", "Description": "Hot vapors cooled through 4 condensers in series. Splits into organic (heavy) and aqueous (light) fractions", "Key Equipment": "Bio-Oil Condenser HE-101, Gas Cooling Tower"},
    {"Step": 7, "Process": "HEAVY FRACTION STORAGE", "Section": "Bio-Oil Processing", "Description": "Organic heavy fraction (bio-oil) stored in intermediate tank at controlled temperature", "Key Equipment": "Bio-Oil Intermediate Storage Tank"},
    {"Step": 8, "Process": "ADDITIVE BLENDING", "Section": "Bio-Oil Processing", "Description": "Lime-based additives added for moisture resistance and ageing resistance improvement", "Key Equipment": "Additive Dosing Tanks"},
    {"Step": 9, "Process": "BITUMEN BLENDING", "Section": "Bio-Oil Processing", "Description": "Bio-oil blended with heated conventional bitumen/hard pitch using high-shear mixer at 150-170°C", "Key Equipment": "Bitumen Heating Tank, High Shear Mixer, Colloid Mill"},
    {"Step": 10, "Process": "QUALITY TESTING", "Section": "Lab", "Description": "Penetration, viscosity, softening point, flash point, ductility tested per IS:73 / VG grade specs", "Key Equipment": "Bitumen Testing Lab"},
    {"Step": 11, "Process": "PRODUCT STORAGE", "Section": "Storage Area", "Description": "Finished bio-bitumen stored in heated storage tanks within bund wall enclosure", "Key Equipment": "Heated Bitumen Storage T-201, Bund Wall"},
    {"Step": 12, "Process": "PACKING / LOADING", "Section": "Product Dispatch", "Description": "Bulk loading into tankers or automated drum filling (200L drums)", "Key Equipment": "Tanker Loading Platform, Drum Filling Machine"},
    {"Step": 13, "Process": "DISPATCH", "Section": "Product Dispatch", "Description": "Weighed at weigh bridge, documentation, dispatch to NHAI/PWD construction sites", "Key Equipment": "Weigh Bridge, Truck Parking"},
]
steps_df = pd.DataFrame(steps)
st.dataframe(steps_df, width="stretch", hide_index=True)

# Sankey-style flow
fig_flow = go.Figure(data=[go.Sankey(
    node=dict(
        pad=15, thickness=20,
        line=dict(color="black", width=0.5),
        label=["Biomass\nInput", "Drying\n& Shredding", "Pyrolysis\nReactor",
               "Bio-Oil\n(40%)", "Biochar\n(30%)", "Syngas\n(25%)", "Loss\n(5%)",
               "Additive\nBlending", "Bitumen\nBlending", "Bio-Modified\nBitumen",
               "Quality\nTesting", "Dispatch"],
        color=["#228B22", "#FFD700", "#CC0000", "#FF8C00", "#8B4513",
               "#4169E1", "#CCCCCC", "#FF8C00", "#800080", "#003366",
               "#006699", "#00AA44"],
    ),
    link=dict(
        source=[0, 1, 2, 2, 2, 2, 3, 7, 8, 9, 10],
        target=[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11],
        value=[100, 100, 40, 30, 25, 5, 40, 40, 40, 40, 40],
    )
)])
fig_flow.update_layout(title="Material Flow: Biomass → Bio-Modified Bitumen",
                         height=500, template="plotly_white")
st.plotly_chart(fig_flow, width="stretch")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# SAFETY ZONES (from layout images)
# ═══════════════════════════════════════════════════════════════════
st.subheader("Safety & Compliance Zones")
st.markdown("""
| Zone | Requirement | Standard |
|------|-------------|----------|
| **Reactor Safety Zone** | 5m clear zone around pyrolysis reactor | PESO / Factory Act |
| **Fire Spacing** | 4m between storage tanks | IS:2190 / TAC norms |
| **Bund Wall** | Reinforced concrete, 110% of tank capacity | OISD-117 |
| **Chimney Height** | 20m minimum for gas scrubber stack | State PCB norms |
| **Equipment Clearance** | 2m between processing equipment | Factory Act |
| **Emergency Exit** | Clearly marked, unobstructed | Factory Act |
| **Assembly Point** | Designated area near entry gate | Safety manual |
| **Fire Extinguishers** | At every 15m along internal roads | IS:2190 |
| **Internal Roads** | 6m wide concrete | Heavy vehicle turning radius |
""")

# Plot dimensions
st.subheader(f"Plot Size Required — {cfg['capacity_tpd']:.0f} TPD Plant")
plot_length = max(40, cfg['capacity_tpd'] * 2)
plot_width = max(25, cfg['capacity_tpd'] * 1.2)
st.metric("Plot Dimensions", f"{plot_length:.0f}m x {plot_width:.0f}m ({plot_length * plot_width:.0f} sqm / {plot_length * plot_width * 10.764 / 43560:.2f} acres)")
st.caption("Based on 30 TPD reference: 60m x 35m (2,100 sqm). Scaled linearly for other capacities.")


# ── AI Skill: Process Optimization ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: Process Optimization"):
            if st.button("Generate", type="primary", key="ai_53Proce"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: Process Optimization. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    if st.button("Download Excel", type="primary", key="exp_xl_53Pro"):
        try:
            import io
            from openpyxl import Workbook
            _wb = Workbook()
            _ws = _wb.active
            _ws.title = "Export"
            _ws.cell(row=1, column=1, value="Bio Bitumen Export")
            _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
            _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
            _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
            _buf = io.BytesIO()
            _wb.save(_buf)
            _buf.seek(0)
            st.download_button("Download", _buf.getvalue(), "export.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_xl_53Pro")
        except Exception as _e:
            st.error(f"Export failed: {_e}")
with _ex2:
    if st.button("Print", key="exp_prt_53Pro"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)
