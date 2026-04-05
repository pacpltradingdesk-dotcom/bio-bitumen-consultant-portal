"""
Financial Model — FULLY EDITABLE inputs → ALL outputs auto-update → Graphs everywhere
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from state_manager import get_config, update_field, update_fields, init_state

st.set_page_config(page_title="Financial Model", page_icon="💰", layout="wide")
init_state()
st.sidebar.markdown("---")
if st.sidebar.button("🖨️ Print This Page", key="print_page"):
    import streamlit.components.v1 as stc
    stc.html("<script>window.print();</script>", height=0)

st.title("Financial Model (Auto-Updating)")
st.markdown("**Edit ANY input below → ALL calculations, charts, and reports auto-update instantly**")

# Live Data Badge
try:
    from engines.live_calculation_engine import get_live_market_inputs, calculate_live_vg30_price
    live = get_live_market_inputs()
    vg30_live = calculate_live_vg30_price(live["crude_oil_usd"], live["usd_inr"])
    is_live = live["data_freshness"] == "live"
    badge_color = "#00AA44" if is_live else "#FF8800"
    st.markdown(f"""
    <div style="background: {badge_color}15; border: 1px solid {badge_color}; border-radius: 8px;
                padding: 8px 15px; display: inline-flex; gap: 20px; align-items: center; margin-bottom: 10px;">
        <span style="color: {badge_color}; font-weight: bold;">{'LIVE DATA' if is_live else 'CACHED DATA'}</span>
        <span>Crude: ${live['crude_oil_usd']:.1f}</span>
        <span>FX: Rs {live['usd_inr']:.2f}</span>
        <span>VG30: Rs {vg30_live['vg30_estimated']:,.0f}</span>
        <span>GDP: {live['gdp_growth_pct']:.1f}%</span>
    </div>
    """, unsafe_allow_html=True)
except Exception:
    pass

st.markdown("---")

cfg = get_config()

# ═══════════════════════════════════════════════════════════════════
# INPUT SHEET (LEFT) → AUTO OUTPUTS (RIGHT)
# ═══════════════════════════════════════════════════════════════════
col_input, col_output = st.columns([1, 1])

with col_input:
    st.subheader("EDITABLE INPUTS")

    st.markdown("**Production**")
    tpd = st.number_input("Capacity (TPD)", 3.0, 100.0, float(cfg["capacity_tpd"]), 5.0, key="fin_tpd")
    days = st.number_input("Working Days/Year", 250, 365, int(cfg["working_days"]), 10, key="fin_days")

    st.markdown("**Revenue**")
    sell_price = st.number_input("Selling Price (Rs/MT)", 20000, 60000, int(cfg["selling_price_per_mt"]), 1000, key="fin_sell",
                                  help="Price at which bio-modified bitumen is sold to NHAI/PWD contractors")
    biochar_p = st.number_input("Biochar Price (Rs/MT)", 1000, 10000, int(cfg["biochar_price_per_mt"]), 500, key="fin_char",
                                 help="Revenue from biochar byproduct — sold as soil amendment or carbon black substitute")
    syngas_val = st.number_input("Syngas Revenue (Rs/MT)", 0, 5000, int(cfg.get("syngas_value_per_mt", 1250)), 250, key="fin_syngas",
                                  help="Value of syngas used as captive fuel — reduces power cost")

    st.markdown("**Costs (Rs per MT output)**")
    raw_cost = st.number_input("Raw Material", 3000, 15000, int(cfg["raw_material_cost_per_mt"]), 500, key="fin_raw")
    power_cost = st.number_input("Power & Fuel", 2000, 8000, int(cfg["power_cost_per_mt"]), 500, key="fin_power")
    labour_cost = st.number_input("Labour", 1000, 6000, int(cfg["labour_cost_per_mt"]), 500, key="fin_labour")
    transport = st.number_input("Transport", 500, 5000, int(cfg["transport_cost_per_mt"]), 500, key="fin_trans")

    st.markdown("**Finance**")
    interest = st.slider("Interest Rate (%)", 8.0, 14.0, float(cfg["interest_rate"]*100), 0.5, key="fin_int")
    equity_pct = st.slider("Equity (%)", 20, 50, int(cfg["equity_ratio"]*100), 5, key="fin_eq")

    st.markdown("**Advanced Parameters**")
    with st.expander("Land, Inflation, WC, Moratorium, Carbon Credits"):
        adv1, adv2 = st.columns(2)
        with adv1:
            land_rate = st.number_input("Land Cost (Rs Lac/acre)", 1, 100, int(cfg["land_cost_per_acre"]), 1, key="fin_land")
            inflation = st.slider("Inflation Rate (%)", 0.0, 10.0, float(cfg["inflation_rate"]*100), 0.5, key="fin_infl")
            rev_growth = st.slider("Revenue Growth (%/yr)", 0.0, 10.0, float(cfg["revenue_growth_rate"]*100), 0.5, key="fin_revgr")
            wc_months = st.number_input("Working Capital (months)", 1, 12, int(cfg["working_capital_months"]), 1, key="fin_wc")
        with adv2:
            moratorium = st.number_input("Moratorium (months)", 0, 24, int(cfg["moratorium_months"]), 3, key="fin_mora")
            salvage = st.slider("Salvage Value (%)", 0, 30, int(cfg["salvage_value_pct"]*100), 5, key="fin_salv")
            bio_blend = st.slider("Bio-Blend (%)", 10, 40, int(cfg["bio_blend_pct"]), 5, key="fin_blend")
            shifts = st.selectbox("Shifts/Day", [1, 2, 3], index=[1,2,3].index(int(cfg["num_shifts"])), key="fin_shifts")
            carbon_rate = st.number_input("Carbon Credit (USD/tCO2)", 5, 50, int(cfg["carbon_credit_rate_usd"]), 1, key="fin_carbon")

    # Apply ALL changes
    changes = {}
    if tpd != cfg["capacity_tpd"]: changes["capacity_tpd"] = tpd
    if days != cfg["working_days"]: changes["working_days"] = days
    if sell_price != cfg["selling_price_per_mt"]: changes["selling_price_per_mt"] = sell_price
    if biochar_p != cfg["biochar_price_per_mt"]: changes["biochar_price_per_mt"] = biochar_p
    if raw_cost != cfg["raw_material_cost_per_mt"]: changes["raw_material_cost_per_mt"] = raw_cost
    if power_cost != cfg["power_cost_per_mt"]: changes["power_cost_per_mt"] = power_cost
    if labour_cost != cfg["labour_cost_per_mt"]: changes["labour_cost_per_mt"] = labour_cost
    if transport != cfg["transport_cost_per_mt"]: changes["transport_cost_per_mt"] = transport
    if abs(interest/100 - cfg["interest_rate"]) > 0.001: changes["interest_rate"] = interest/100
    if abs(equity_pct/100 - cfg["equity_ratio"]) > 0.01: changes["equity_ratio"] = equity_pct/100
    if land_rate != cfg["land_cost_per_acre"]: changes["land_cost_per_acre"] = land_rate
    if abs(inflation/100 - cfg["inflation_rate"]) > 0.001: changes["inflation_rate"] = inflation/100
    if abs(rev_growth/100 - cfg["revenue_growth_rate"]) > 0.001: changes["revenue_growth_rate"] = rev_growth/100
    if wc_months != cfg["working_capital_months"]: changes["working_capital_months"] = wc_months
    if moratorium != cfg["moratorium_months"]: changes["moratorium_months"] = moratorium
    if abs(salvage/100 - cfg["salvage_value_pct"]) > 0.01: changes["salvage_value_pct"] = salvage/100
    if bio_blend != cfg["bio_blend_pct"]: changes["bio_blend_pct"] = bio_blend
    if shifts != cfg["num_shifts"]: changes["num_shifts"] = shifts
    if carbon_rate != cfg["carbon_credit_rate_usd"]: changes["carbon_credit_rate_usd"] = carbon_rate
    if syngas_val != cfg.get("syngas_value_per_mt", 1250): changes["syngas_value_per_mt"] = syngas_val

    # Input validation
    if sell_price <= 0:
        st.error("Selling price must be greater than zero!")
        st.stop()
    if tpd <= 0:
        st.error("Capacity must be greater than zero!")
        st.stop()
    if sell_price < cfg.get("total_variable_cost_per_mt", 0):
        st.warning(f"Selling price (Rs {sell_price:,}) is BELOW variable cost (Rs {cfg.get('total_variable_cost_per_mt',0):,}) — project will make a LOSS!")

    if changes:
        update_fields(changes)
        cfg = get_config()

with col_output:
    st.subheader("AUTO CALCULATIONS")

    st.markdown("**Production**")
    oc1, oc2 = st.columns(2)
    output_per_day = cfg["capacity_tpd"] * 0.40
    annual_prod = output_per_day * cfg["working_days"]
    oc1.metric("Daily Output", f"{output_per_day:.1f} MT")
    oc2.metric("Annual Production", f"{annual_prod:,.0f} MT")

    st.markdown("**Revenue (Year 5 @ 85% utilization)**")
    or1, or2 = st.columns(2)
    or1.metric("Annual Revenue", f"Rs {cfg['revenue_yr5_lac']:.0f} Lac")
    or2.metric("Monthly Revenue", f"Rs {cfg['revenue_yr5_lac']/12:.1f} Lac")

    st.markdown("**Costs**")
    ot1, ot2 = st.columns(2)
    ot1.metric("Variable Cost/MT", f"Rs {cfg['total_variable_cost_per_mt']:,.0f}",
               help="Sum of all variable costs: Raw Material + Power + Labour + Chemical + Packaging + Transport + QC + Misc")
    ot2.metric("Profit/MT", f"Rs {cfg['profit_per_mt']:,.0f}",
               help="Selling Price minus Total Variable Cost per MT")

    # Fix #11: Show full cost breakdown
    with st.expander("Cost Breakdown (all components visible)"):
        cost_items = [
            ("Raw Material", cfg.get("raw_material_cost_per_mt", 0)),
            ("Power & Fuel", cfg.get("power_cost_per_mt", 0)),
            ("Labour", cfg.get("labour_cost_per_mt", 0)),
            ("Chemicals", cfg.get("chemical_cost_per_mt", 0)),
            ("Packaging", cfg.get("packaging_cost_per_mt", 0)),
            ("Transport", cfg.get("transport_cost_per_mt", 0)),
            ("QC & Testing", cfg.get("qc_cost_per_mt", 0)),
            ("Miscellaneous", cfg.get("misc_cost_per_mt", 0)),
        ]
        for name, val in cost_items:
            st.markdown(f"- {name}: **Rs {val:,}/MT**")
        st.markdown(f"**TOTAL: Rs {sum(v for _,v in cost_items):,}/MT**")
        model_total = cfg.get("total_variable_cost_per_mt", 0)
        visible_total = sum(v for _, v in cost_items)
        if model_total != visible_total:
            st.warning(f"Model uses Rs {model_total:,}/MT but visible inputs sum to Rs {visible_total:,}/MT. Difference: Rs {model_total - visible_total:,}")

    # Fix #12: CAPEX Investment Breakdown
    with st.expander("Investment Breakdown (CAPEX details)"):
        p = cfg.get("plant_data", {})
        capex_items = [
            ("Civil & Building", p.get("civil_lac", cfg["investment_lac"] * 0.15)),
            ("Machinery & Equipment", p.get("mach_lac", cfg["investment_lac"] * 0.55)),
            ("GST on Machinery (18%)", p.get("gst_mach_lac", cfg["investment_lac"] * 0.10)),
            ("Working Capital", cfg.get("working_capital_lac", cfg["investment_lac"] * 0.08)),
            ("Interest During Construction", p.get("idc_lac", cfg["investment_lac"] * 0.04)),
            ("Pre-operative Expenses", p.get("preop_lac", cfg["investment_lac"] * 0.03)),
            ("Contingency (5%)", p.get("cont_lac", cfg["investment_lac"] * 0.03)),
            ("Security Deposit", p.get("sec_lac", cfg["investment_lac"] * 0.02)),
        ]
        for name, val in capex_items:
            st.markdown(f"- {name}: **Rs {val:.1f} Lac**")
        capex_total = sum(v for _, v in capex_items)
        st.markdown(f"**TOTAL CAPEX: Rs {capex_total:.1f} Lac (Rs {capex_total/100:.2f} Cr)**")
        if abs(capex_total - cfg["investment_lac"]) > 10:
            st.info(f"Model investment: Rs {cfg['investment_lac']:.1f} Lac. Breakdown total: Rs {capex_total:.1f} Lac. Difference due to interpolation adjustments.")

    st.markdown("**Key Metrics**")
    ok1, ok2, ok3 = st.columns(3)
    ok1.metric("ROI", f"{cfg['roi_pct']:.1f}%", help="Return on Investment — Annual Profit / Total Investment × 100")
    ok2.metric("IRR", f"{cfg['irr_pct']:.1f}%", help="Internal Rate of Return — Discount rate where NPV = 0")
    ok3.metric("Break-Even", f"{cfg['break_even_months']} mo", help="Months to recover total investment from net profits")

    om1, om2, om3 = st.columns(3)
    om1.metric("CAPEX", f"Rs {cfg['investment_cr']:.2f} Cr", help="Total Capital Expenditure — Plant + Machinery + WC + Pre-operative")
    om2.metric("EMI/Month", f"Rs {cfg['emi_lac_mth']:.2f} Lac", help="Equated Monthly Installment — Fixed monthly loan repayment")
    om3.metric("Monthly Profit", f"Rs {cfg['monthly_profit_lac']:.1f} Lac", help="Net Profit After Tax — Year 5 at 85% utilization")

    # NEW — Advanced Financial Metrics
    st.markdown("**Bank-Grade Metrics**")
    bk1, bk2, bk3 = st.columns(3)
    npv_v = cfg.get('npv_lac', 0)
    bk1.metric("NPV", f"Rs {npv_v:.0f}L" if abs(npv_v) < 100 else f"Rs {npv_v/100:.1f}Cr",
               help="Net Present Value — Today's worth of all future profits. Positive = good investment")
    bk2.metric("D/E Ratio", f"{cfg.get('debt_equity_ratio', 0):.2f}x",
               help="Debt-Equity Ratio — Loan ÷ Equity. Lower = safer for lender")
    bk3.metric("Current Ratio", f"{cfg.get('current_ratio', 0):.2f}",
               help="Current Assets ÷ Current Liabilities — measures short-term liquidity")

    bk4, bk5, bk6 = st.columns(3)
    nw = cfg.get('net_worth_yr5_lac', 0)
    bk4.metric("Net Worth Yr5", f"Rs {nw:.0f}L" if nw < 100 else f"Rs {nw/100:.1f}Cr",
               help="Total equity + retained profits built up by Year 5")
    cc = cfg.get('carbon_credit_annual_lac', 0)
    bk5.metric("Carbon Credits", f"Rs {cc:.1f}L/yr",
               help="Annual revenue from selling CO2 offset certificates on carbon market")
    wc = cfg.get('working_capital_lac', 0)
    bk6.metric("Working Capital", f"Rs {wc:.0f}L",
               help="Cash needed for daily operations — typically 2-3 months of costs")

    # DSCR Schedule — Fix G3: Color each year individually, not blanket red
    dscr_sched = cfg.get("dscr_schedule", [])
    if dscr_sched:
        st.markdown("**DSCR Schedule (all 7 years)**")
        dscr_parts = []
        for i, d in enumerate(dscr_sched):
            if d >= 1.5:
                dscr_parts.append(f'<span style="color:#00AA44;font-weight:bold">Yr{i+1}: {d:.2f}x</span>')
            elif d >= 1.0:
                dscr_parts.append(f'<span style="color:#FF8800;font-weight:bold">Yr{i+1}: {d:.2f}x</span>')
            else:
                dscr_parts.append(f'<span style="color:#CC3333;font-weight:bold">Yr{i+1}: {d:.2f}x</span>')
        st.markdown(" | ".join(dscr_parts), unsafe_allow_html=True)
        st.caption("Green ≥ 1.5x (healthy) | Orange 1.0-1.5x (watch) | Red < 1.0x (risk)")

    # H7: CGTMSE notice
    loan_lac_val = cfg.get("loan_cr", 0) * 100
    if loan_lac_val > 500:
        remain = loan_lac_val - 500
        st.info(f"Loan exceeds Rs 5 Cr CGTMSE limit. Consider: Rs 5 Cr under CGTMSE (collateral-free) + Rs {remain:.0f}L with collateral.")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# CHARTS
# ═══════════════════════════════════════════════════════════════════
if cfg["roi_timeline"]:
    roi_df = pd.DataFrame(cfg["roi_timeline"])

    # Chart 1: Revenue vs Cost vs Profit
    st.subheader("7-Year Projection")
    fig1 = go.Figure()
    fig1.add_trace(go.Bar(x=roi_df["Year"], y=roi_df["Revenue (Lac)"], name="Revenue", marker_color="#003366"))
    fig1.add_trace(go.Bar(x=roi_df["Year"], y=roi_df["Variable Cost (Lac)"], name="Variable Cost", marker_color="#CC3333"))
    fig1.add_trace(go.Bar(x=roi_df["Year"], y=roi_df["Fixed Cost (Lac)"], name="Fixed Cost", marker_color="#FF8800"))
    fig1.add_trace(go.Scatter(x=roi_df["Year"], y=roi_df["PAT (Lac)"], name="Net Profit",
                               mode="lines+markers", line=dict(color="#00AA44", width=3)))
    fig1.update_layout(barmode="stack", template="plotly_white", height=400,
                        xaxis_title="Year", yaxis_title="Rs Lakhs")
    st.plotly_chart(fig1, width="stretch")

    # Chart 2: DSCR Trend
    col_dscr, col_cashflow = st.columns(2)
    with col_dscr:
        fig2 = go.Figure()
        fig2.add_trace(go.Bar(x=roi_df["Year"], y=roi_df["DSCR"], name="DSCR",
                               marker_color=["#CC3333" if d < 1.5 else "#006699" for d in roi_df["DSCR"]]))
        fig2.add_hline(y=1.5, line_dash="dash", line_color="red", annotation_text="Bank Minimum (1.5x)")
        fig2.update_layout(title="DSCR Trend", template="plotly_white", height=350, yaxis_title="DSCR Ratio")
        st.plotly_chart(fig2, width="stretch")

    with col_cashflow:
        fig3 = go.Figure()
        fig3.add_trace(go.Scatter(x=roi_df["Year"], y=roi_df["Cash Accrual (Lac)"],
                                   name="Cash Accrual", fill="tozeroy", line=dict(color="#003366")))
        fig3.add_trace(go.Scatter(x=roi_df["Year"], y=[cfg["emi_lac_mth"]*12]*7,
                                   name="Annual EMI", line=dict(color="#CC3333", dash="dash")))
        fig3.update_layout(title="Cash Accrual vs EMI", template="plotly_white", height=350, yaxis_title="Rs Lakhs")
        st.plotly_chart(fig3, width="stretch")

    # Full P&L Table
    st.markdown("---")
    st.subheader("Detailed P&L Statement")
    st.dataframe(roi_df, width="stretch", hide_index=True)

    # NEW — Cash Flow Statement
    if cfg.get("cash_flow_statement"):
        st.markdown("---")
        st.subheader("Cash Flow Statement (7 Years)")
        cf_df = pd.DataFrame(cfg["cash_flow_statement"])
        st.dataframe(cf_df, width="stretch", hide_index=True)

        fig_cf = go.Figure()
        fig_cf.add_trace(go.Bar(x=cf_df["Year"], y=cf_df["Operating (Lac)"], name="Operating", marker_color="#003366"))
        fig_cf.add_trace(go.Bar(x=cf_df["Year"], y=cf_df["Financing (Lac)"], name="Financing", marker_color="#CC3333"))
        fig_cf.add_trace(go.Scatter(x=cf_df["Year"], y=cf_df["Net Cash (Lac)"], name="Net Cash",
                                     mode="lines+markers", line=dict(color="#00AA44", width=3)))
        fig_cf.update_layout(title="Cash Flow Statement (Rs Lac)", barmode="group",
                              template="plotly_white", height=350)
        st.plotly_chart(fig_cf, width="stretch")

    # NEW — Balance Sheet
    if cfg.get("balance_sheet"):
        st.markdown("---")
        st.subheader("Balance Sheet Summary (Year 5)")
        bs_df = pd.DataFrame(cfg["balance_sheet"])
        st.dataframe(bs_df, width="stretch", hide_index=True)

    # NEW — Loan Repayment Schedule
    if cfg.get("cash_flow_statement"):
        st.markdown("---")
        st.subheader("Loan Outstanding Schedule")
        loan_df = pd.DataFrame(cfg["cash_flow_statement"])[["Year", "Loan Outstanding (Lac)"]]
        fig_loan = go.Figure()
        fig_loan.add_trace(go.Bar(x=loan_df["Year"], y=loan_df["Loan Outstanding (Lac)"],
                                    name="Loan Outstanding", marker_color="#FF8800"))
        fig_loan.update_layout(title="Loan Repayment Progress (Rs Lac)", template="plotly_white", height=300)
        st.plotly_chart(fig_loan, width="stretch")

    # Sensitivity Matrix
    st.markdown("---")
    st.subheader("Sensitivity Analysis (EBITDA Yr5 in Rs Lakhs)")
    if cfg["sensitivity_matrix"]:
        fig_heat = go.Figure(data=go.Heatmap(
            z=cfg["sensitivity_matrix"],
            x=["Low Sell Price", "Base Price", "High Sell Price"],
            y=["Low Cost", "Base Cost", "High Cost"],
            text=[[f"Rs {v:.0f}L" for v in row] for row in cfg["sensitivity_matrix"]],
            texttemplate="%{text}", colorscale="RdYlGn",
        ))
        fig_heat.update_layout(title="EBITDA Sensitivity (Green = Profitable, Red = Risk)",
                                height=350, template="plotly_white")
        st.plotly_chart(fig_heat, width="stretch")

    # Monthly P&L
    if cfg["monthly_pnl"]:
        st.markdown("---")
        st.subheader("Monthly P&L (Year 5 Steady State)")
        pnl = cfg["monthly_pnl"]
        pnl_df = pd.DataFrame([{"Item": k, "Rs Lakhs/Month": v} for k, v in pnl.items()])
        fig_pnl = px.bar(pnl_df, x="Item", y="Rs Lakhs/Month", color="Item",
                          color_discrete_sequence=["#003366", "#CC3333", "#FF8800", "#006699", "#AA3366", "#00AA44"])
        fig_pnl.update_layout(template="plotly_white", height=350, showlegend=False)
        st.plotly_chart(fig_pnl, width="stretch")

# ═══════════════════════════════════════════════════════════════════
# EXPORT SECTION
# ═══════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("Export Financial Model")

exp1, exp2, exp3 = st.columns(3)

# Excel Export
with exp1:
    if st.button("Export to Excel", type="primary"):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "Summary"
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

        summary_rows = [
            ["Bio Bitumen Financial Model", ""],
            ["Capacity", f"{cfg['capacity_tpd']:.0f} TPD"],
            ["Working Days", f"{cfg['working_days']}"],
            ["Investment", f"Rs {cfg['investment_cr']:.2f} Crore"],
            ["Loan", f"Rs {cfg['loan_cr']:.2f} Crore"],
            ["Equity", f"Rs {cfg['equity_cr']:.2f} Crore"],
            ["Monthly EMI", f"Rs {cfg['emi_lac_mth']:.2f} Lakhs"],
            ["Revenue Yr5", f"Rs {cfg['revenue_yr5_lac']:.0f} Lakhs"],
            ["Profit per MT", f"Rs {cfg['profit_per_mt']:,.0f}"],
            ["ROI", f"{cfg['roi_pct']:.1f}%"],
            ["IRR", f"{cfg['irr_pct']:.1f}%"],
            ["DSCR Yr3", f"{cfg['dscr_yr3']:.2f}x"],
            ["Break-Even", f"{cfg['break_even_months']} months"],
        ]
        for i, (a, b) in enumerate(summary_rows):
            ws1.cell(row=i+1, column=1, value=a).font = Font(bold=True)
            ws1.cell(row=i+1, column=2, value=b)

        # Sheet 2: 7-Year P&L
        if cfg["roi_timeline"]:
            ws2 = wb.create_sheet("P&L 7-Year")
            headers = list(cfg["roi_timeline"][0].keys())
            for j, h in enumerate(headers):
                cell = ws2.cell(row=1, column=j+1, value=h)
                cell.font = header_font
                cell.fill = header_fill
            for i, row_data in enumerate(cfg["roi_timeline"]):
                for j, h in enumerate(headers):
                    ws2.cell(row=i+2, column=j+1, value=row_data[h])

        # Sheet 3: Sensitivity
        if cfg["sensitivity_matrix"]:
            ws3 = wb.create_sheet("Sensitivity")
            ws3.cell(row=1, column=1, value="Cost \\ Price")
            labels_col = ["Low Price", "Base Price", "High Price"]
            labels_row = ["Low Cost", "Base Cost", "High Cost"]
            for j, lbl in enumerate(labels_col):
                ws3.cell(row=1, column=j+2, value=lbl).font = header_font
            for i, lbl in enumerate(labels_row):
                ws3.cell(row=i+2, column=1, value=lbl).font = Font(bold=True)
                for j, val in enumerate(cfg["sensitivity_matrix"][i]):
                    ws3.cell(row=i+2, column=j+2, value=val)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        st.download_button(
            label="Download Excel",
            data=buffer.getvalue(),
            file_name=f"Financial_Model_{cfg['capacity_tpd']:.0f}TPD.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# PDF Export
with exp2:
    if st.button("Export to PDF", type="primary", key="exp_pdf"):
        with st.spinner("Generating PDF..."):
            try:
                from engines.report_generator_engine import generate_dpr_pdf
                from config import COMPANY as _CO
                pdf_path = os.path.join(os.path.dirname(__file__), "..", "data", "test_outputs", "_fin_export.pdf")
                os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
                generate_dpr_pdf(pdf_path, cfg, _CO)
                with open(pdf_path, "rb") as _pf:
                    st.download_button("Download PDF", _pf.read(),
                        f"Financial_{cfg['capacity_tpd']:.0f}TPD.pdf", "application/pdf", key="dl_fin_pdf")
            except Exception as e:
                st.error(f"PDF generation failed: {e}")

# Print
with exp3:
    if st.button("Print Page", key="exp_print"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# SCENARIO MANAGER (NEW)
# ═══════════════════════════════════════════════════════════════════
st.subheader("Scenario Manager")
st.caption("Save current configuration as a named scenario for quick recall")

from database import save_configuration, get_configurations, init_db
import json
init_db()

scen_col1, scen_col2 = st.columns(2)

with scen_col1:
    st.markdown("**Save Current Scenario**")
    scenario_name = st.text_input("Scenario Name", placeholder="e.g., Conservative / Optimistic / Bank Presentation", key="scen_name")
    if st.button("Save Scenario", key="save_scen", type="primary") and scenario_name:
        config_snapshot = {
            "capacity_tpd": cfg["capacity_tpd"],
            "working_days": cfg["working_days"],
            "selling_price_per_mt": cfg["selling_price_per_mt"],
            "biochar_price_per_mt": cfg["biochar_price_per_mt"],
            "raw_material_cost_per_mt": cfg["raw_material_cost_per_mt"],
            "power_cost_per_mt": cfg["power_cost_per_mt"],
            "labour_cost_per_mt": cfg["labour_cost_per_mt"],
            "transport_cost_per_mt": cfg["transport_cost_per_mt"],
            "interest_rate": cfg["interest_rate"],
            "equity_ratio": cfg["equity_ratio"],
            "investment_cr": cfg["investment_cr"],
            "roi_pct": cfg["roi_pct"],
            "irr_pct": cfg["irr_pct"],
            "monthly_profit_lac": cfg["monthly_profit_lac"],
            "break_even_months": cfg["break_even_months"],
            "dscr_yr3": cfg["dscr_yr3"],
            "profit_per_mt": cfg["profit_per_mt"],
        }
        save_configuration(None, scenario_name, json.dumps(config_snapshot))
        st.success(f"Scenario '{scenario_name}' saved!")

with scen_col2:
    st.markdown("**Saved Scenarios**")
    saved = get_configurations()
    if saved:
        for s in saved[:5]:
            raw = s.get("config_json", {})
            snap = json.loads(raw) if isinstance(raw, str) else (raw if isinstance(raw, dict) else {})
            with st.expander(f"{s['name']} — {snap.get('capacity_tpd', '?')} TPD | ROI: {snap.get('roi_pct', '?')}%"):
                for k, v in snap.items():
                    st.text(f"{k}: {v}")
                if st.button(f"Load '{s['name']}'", key=f"load_scen_{s['id']}"):
                    load_changes = {}
                    for k in ["capacity_tpd", "working_days", "selling_price_per_mt", "biochar_price_per_mt",
                              "raw_material_cost_per_mt", "power_cost_per_mt", "labour_cost_per_mt",
                              "transport_cost_per_mt", "interest_rate", "equity_ratio"]:
                        if k in snap:
                            load_changes[k] = snap[k]
                    if load_changes:
                        update_fields(load_changes)
                        st.success(f"Loaded scenario '{s['name']}'! Refresh to see updated values.")
                        st.rerun()
    else:
        st.info("No saved scenarios yet. Save your first scenario above.")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# CUMULATIVE CASH FLOW CHART (NEW)
# ═══════════════════════════════════════════════════════════════════
if cfg.get("roi_timeline"):
    st.subheader("Cumulative Cash Flow & Payback")
    roi_df2 = pd.DataFrame(cfg["roi_timeline"])
    investment_lac = cfg["investment_cr"] * 100
    cumulative = []
    running_total = -investment_lac  # Start negative (initial investment)
    for _, row in roi_df2.iterrows():
        running_total += row.get("Cash Accrual (Lac)", row.get("PAT (Lac)", 0))
        cumulative.append(running_total)

    fig_cum = go.Figure()
    colors_cum = ["#CC3333" if v < 0 else "#00AA44" for v in cumulative]
    fig_cum.add_trace(go.Bar(x=roi_df2["Year"], y=cumulative, name="Cumulative Cash Flow",
                               marker_color=colors_cum))
    fig_cum.add_hline(y=0, line_dash="dash", line_color="black", annotation_text="Payback Point")
    fig_cum.update_layout(title="Cumulative Cash Flow (Rs Lac) — Green = Payback Achieved",
                           template="plotly_white", height=400, yaxis_title="Rs Lac")
    st.plotly_chart(fig_cum, width="stretch")

# ═══════════════════════════════════════════════════════════════════
# AI FINANCIAL ANALYSIS (NEW — powered by OpenAI/Claude)
# ═══════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("AI Financial Analysis")

try:
    from engines.ai_engine import is_ai_available, ai_financial_analysis
    if is_ai_available():
        if st.button("Generate AI Financial Commentary", type="primary", key="ai_fin"):
            with st.spinner("AI analyzing your financial model..."):
                analysis, provider = ai_financial_analysis(cfg)
            if analysis:
                st.markdown("---")
                st.markdown(analysis)
                st.caption(f"Powered by {provider.upper()}")
            else:
                st.warning("AI analysis could not be generated. Check API settings.")
    else:
        st.info("Add API keys in AI Settings to enable AI-powered financial analysis.")
        st.page_link("pages/17_🔑_AI_Settings.py", label="Go to AI Settings", icon="🔑")
except Exception:
    pass

# Quick links
st.markdown("---")
st.subheader("Related Tools")
ql1, ql2, ql3, ql4 = st.columns(4)
ql1.page_link("pages/60_ROI_Quick_Calc.py", label="ROI Calc", icon="🎯")
ql2.page_link("pages/61_Loan_EMI.py", label="EMI Calc", icon="🏦")
ql3.page_link("pages/62_Capacity_Compare.py", label="Compare", icon="⚖️")
ql4.page_link("pages/17_🔑_AI_Settings.py", label="AI Setup", icon="🔑")

st.caption("All values auto-calculate when ANY input changes. This model is bank-ready and investor-grade.")
