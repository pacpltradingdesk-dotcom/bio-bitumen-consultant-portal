"""
Lab Testing & Validation — DSR, Marshall, Rheology, Field Trials
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
from state_manager import init_state, get_config
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

st.set_page_config(page_title="Lab Testing", page_icon="🔬", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass

st.title("Lab Testing & Performance Validation")
st.markdown("**CSIR-CRRI Validated | IS:73 Compliant | Field Trial Proven**")
st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# TEST RESULTS SUMMARY
# ═══════════════════════════════════════════════════════════════════
st.subheader("Test Results — Bio-Bitumen VG30 (Representative)")

results = pd.DataFrame([
    {"Test": "Penetration at 25°C", "Method": "IS 1203", "Required": "50-70 (0.1mm)", "Bio-Bitumen Result": "62", "Status": "PASS"},
    {"Test": "Absolute Viscosity at 60°C", "Method": "IS 1206-P2", "Required": "2400-3600 Poise", "Bio-Bitumen Result": "2850", "Status": "PASS"},
    {"Test": "Kinematic Viscosity at 135°C", "Method": "IS 1206-P3", "Required": "min 350 cSt", "Bio-Bitumen Result": "412", "Status": "PASS"},
    {"Test": "Flash Point", "Method": "IS 1209", "Required": "min 220°C", "Bio-Bitumen Result": "285°C", "Status": "PASS"},
    {"Test": "Solubility in TCE", "Method": "IS 1216", "Required": "min 99%", "Bio-Bitumen Result": "99.4%", "Status": "PASS"},
    {"Test": "Softening Point", "Method": "IS 1205", "Required": "47-58°C", "Bio-Bitumen Result": "52°C", "Status": "PASS"},
    {"Test": "Ductility at 25°C", "Method": "IS 1208", "Required": "min 40 cm", "Bio-Bitumen Result": "68 cm", "Status": "PASS"},
    {"Test": "Specific Gravity", "Method": "IS 1202", "Required": "min 0.99", "Bio-Bitumen Result": "1.02", "Status": "PASS"},
    {"Test": "Loss on Heating (RTFO)", "Method": "IS 9382", "Required": "max 1.0%", "Bio-Bitumen Result": "0.6%", "Status": "PASS"},
    {"Test": "Retained Penetration after RTFO", "Method": "IS 1203", "Required": "min 47%", "Bio-Bitumen Result": "72%", "Status": "PASS"},
])

# Color-code PASS/FAIL
st.dataframe(results, width="stretch", hide_index=True)

pass_count = (results["Status"] == "PASS").sum()
total = len(results)
st.success(f"**{pass_count}/{total} tests PASSED** — Bio-Bitumen VG30 meets IS:73 specification")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# PERFORMANCE COMPARISON CHARTS
# ═══════════════════════════════════════════════════════════════════
st.subheader("Performance Comparison: Bio-Bitumen vs Conventional")

col1, col2 = st.columns(2)

with col1:
    # Penetration comparison
    fig1 = go.Figure(data=[
        go.Bar(name="Conventional VG30", x=["Penetration\n(0.1mm)", "Viscosity\n(Poise)", "Flash Point\n(°C)"],
               y=[65, 2900, 260], marker_color="#CC3333"),
        go.Bar(name="Bio-Bitumen VG30", x=["Penetration\n(0.1mm)", "Viscosity\n(Poise)", "Flash Point\n(°C)"],
               y=[62, 2850, 285], marker_color="#003366"),
    ])
    fig1.update_layout(barmode="group", title="Key Properties Comparison", template="plotly_white", height=350)
    st.plotly_chart(fig1, width="stretch")

with col2:
    # Marshall Stability
    fig2 = go.Figure(data=[
        go.Bar(name="Conventional", x=["Marshall\nStability (kN)", "Flow\n(mm)", "Air Voids\n(%)"],
               y=[12.5, 3.2, 4.5], marker_color="#CC3333"),
        go.Bar(name="Bio-Modified", x=["Marshall\nStability (kN)", "Flow\n(mm)", "Air Voids\n(%)"],
               y=[14.2, 3.0, 4.2], marker_color="#003366"),
    ])
    fig2.update_layout(barmode="group", title="Marshall Test Comparison", template="plotly_white", height=350)
    st.plotly_chart(fig2, width="stretch")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# ADVANCED TESTING
# ═══════════════════════════════════════════════════════════════════
st.subheader("Advanced Performance Testing")

advanced = pd.DataFrame([
    {"Test Type": "DSR (Dynamic Shear Rheometer)", "Parameter": "G*/sin(delta) at 64°C", "Requirement": "min 1.0 kPa", "Result": "2.8 kPa", "Status": "PASS"},
    {"Test Type": "DSR (after RTFO)", "Parameter": "G*/sin(delta) at 64°C", "Requirement": "min 2.2 kPa", "Result": "4.1 kPa", "Status": "PASS"},
    {"Test Type": "BBR (Bending Beam)", "Parameter": "Stiffness at -12°C", "Requirement": "max 300 MPa", "Result": "220 MPa", "Status": "PASS"},
    {"Test Type": "Storage Stability", "Parameter": "Softening point diff (top-bottom)", "Requirement": "max 3°C", "Result": "1.8°C", "Status": "PASS"},
    {"Test Type": "Moisture Susceptibility", "Parameter": "TSR (Tensile Strength Ratio)", "Requirement": "min 80%", "Result": "88%", "Status": "PASS"},
    {"Test Type": "Rutting Resistance", "Parameter": "Rut depth at 10,000 passes", "Requirement": "max 4mm", "Result": "2.8mm", "Status": "PASS"},
    {"Test Type": "Fatigue Life", "Parameter": "Cycles to failure at 400μ strain", "Requirement": ">100,000", "Result": "185,000", "Status": "PASS"},
])
st.dataframe(advanced, width="stretch", hide_index=True)

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# FIELD TRIALS
# ═══════════════════════════════════════════════════════════════════
st.subheader("Field Trial Records")

trials = pd.DataFrame([
    {"Trial": "CSIR-CRRI Campus Road", "Location": "New Delhi", "Length": "200m", "Bio-Bitumen %": "15%", "Grade": "VG30", "Year": 2023, "Status": "Performing well after 2 years"},
    {"Trial": "NHAI Test Section", "Location": "UP", "Length": "500m", "Bio-Bitumen %": "20%", "Grade": "VG30", "Year": 2024, "Status": "Under monitoring"},
    {"Trial": "State Highway Pilot", "Location": "MP", "Length": "1 km", "Bio-Bitumen %": "15%", "Grade": "VG30", "Year": 2024, "Status": "Completed successfully"},
    {"Trial": "IIT Roorkee Test Track", "Location": "Uttarakhand", "Length": "100m", "Bio-Bitumen %": "30%", "Grade": "VG30", "Year": 2023, "Status": "Lab + field data published"},
])
st.dataframe(trials, width="stretch", hide_index=True)

st.info("""
**Key Field Trial Findings:**
- Bio-bitumen with 15-30% replacement shows **equal or better** rutting resistance
- Moisture susceptibility improved due to lime-based additives
- No cracking or stripping observed after 2+ years in service
- Validated by CSIR-CRRI, IIT Roorkee, and NHAI inspection teams
""")

st.caption("Source: CSIR-CRRI research publications, Dr. Ambika Behl technical sessions, IIT Roorkee reports")


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_55Lab", type="primary")
with _ex2:
    if st.button("Print", key="exp_prt_55Lab"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)


# ── AI Assist ────────────────────────────────────────────────────
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("AI Assist"):
            if st.button("Generate AI Summary", type="primary", key="ai_55Lab"):
                with st.spinner("AI working..."):
                    _p = f"Summarize this section for a {cfg.get('capacity_tpd',20):.0f} TPD bio-bitumen plant in {cfg.get('state','')}. Investment Rs {cfg.get('investment_cr',8):.2f} Cr. Professional consultant format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 800)
                if _r:
                    st.markdown(_r)
except Exception:
    pass
