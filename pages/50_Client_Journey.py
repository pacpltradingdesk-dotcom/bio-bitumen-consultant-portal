"""
CLIENT JOURNEY — Dynamic 8-Step Flow Based on Client Type
===========================================================
4 Client Types → Same 8 Steps → Different Content → All Documents Ready
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from state_manager import get_config, update_field, init_state
from config import (COMPANY, FOUR_STAGES, TARGET_AUDIENCES, CONSULTING_SERVICES,
                    KEY_CREDENTIALS, INDUSTRY_NETWORK, WHY_NOW, CAREER_TRACK)

st.set_page_config(page_title="Client Journey", page_icon="🧭", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass


# ═══════════════════════════════════════════════════════════════════════
# CLIENT TYPE SELECTOR
# ═══════════════════════════════════════════════════════════════════════
st.title("Client Journey — Step by Step")
st.markdown(f"### {COMPANY['trade_name']} | {COMPANY['owner']} | {COMPANY['experience']}")
st.markdown("---")

CLIENT_TYPES = {
    "new_investor": {
        "icon": "🧑‍🌾", "name": "New Investor / Farmer Group / Society",
        "tagline": "New plant setup → Full system → Ready business",
        "stages": "ALL 4 STAGES",
        "investment": "Rs 2-6 Cr (full chain)",
        "color": "#003366",
    },
    "bitumen_owner": {
        "icon": "🛢️", "name": "Existing Bitumen Plant Owner",
        "tagline": "Convert or upgrade → Reduce cost → Increase margin",
        "stages": "STAGE 1-2 (You handle 3-4)",
        "investment": "Rs 80L-2 Cr (add-on)",
        "color": "#CC6600",
    },
    "biomass_company": {
        "icon": "🌾", "name": "Biomass / Agro Waste Company",
        "tagline": "Convert waste → Value product → High margin",
        "stages": "STAGE 2-3-4 (You handle 1)",
        "investment": "Rs 1-2 Cr (pyrolysis + blending)",
        "color": "#228B22",
    },
    "pyrolysis_owner": {
        "icon": "🔥", "name": "Existing Pyrolysis Plant Owner",
        "tagline": "Upgrade from low-value oil → High-value bitumen",
        "stages": "STAGE 3-4 (You handle 1-2)",
        "investment": "Rs 40-80L (blending + testing + marketing)",
        "color": "#CC0000",
    },
}

# Client selector
st.subheader("Select Your Client Type")
cols = st.columns(4)
selected_type = st.session_state.get("client_type", "new_investor")

for i, (key, ct) in enumerate(CLIENT_TYPES.items()):
    with cols[i]:
        if st.button(f"{ct['icon']} {ct['name'].split('/')[0].strip()}", key=f"ct_{key}",
                      type="primary" if selected_type == key else "secondary",
                      width="stretch"):
            st.session_state["client_type"] = key
            selected_type = key
            st.rerun()

ct = CLIENT_TYPES[selected_type]
st.markdown(f"## {ct['icon']} {ct['name']}")
st.markdown(f"**{ct['tagline']}** | Stages: {ct['stages']} | Investment: {ct['investment']}")
st.markdown("---")

# ═══════════════════════════════════════════════════════════════════════
# 8-STEP JOURNEY (Content changes based on client type)
# ═══════════════════════════════════════════════════════════════════════

# Step progress bar
steps = ["Opportunity", "Feasibility", "Solution", "Financial", "Implementation", "Compliance", "Execution", "Profit"]
step_idx = st.session_state.get("journey_step", 0)

step_cols = st.columns(8)
for i, step in enumerate(steps):
    with step_cols[i]:
        color = ct["color"] if i <= step_idx else "#CCCCCC"
        st.markdown(f"<div style='text-align:center;padding:8px;background:{color};color:white;border-radius:8px;font-size:11px;font-weight:bold;'>{i+1}. {step}</div>",
                    unsafe_allow_html=True)

nav1, nav2, nav3 = st.columns([1, 6, 1])
with nav1:
    if step_idx > 0 and st.button("← Back"):
        st.session_state["journey_step"] = step_idx - 1
        st.rerun()
with nav3:
    if step_idx < 7 and st.button("Next →"):
        st.session_state["journey_step"] = step_idx + 1
        st.rerun()

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════════
# STEP CONTENT — Changes by client type
# ═══════════════════════════════════════════════════════════════════════

# ── STEP 1: OPPORTUNITY ──────────────────────────────────────────────
if step_idx == 0:
    st.header(f"Step 1: Market Opportunity")

    if selected_type == "new_investor":
        st.markdown("""
        ### Why Bio-Bitumen NOW?
        - India became **FIRST country** to commercially produce bio-bitumen (Jan 2026)
        - **130-216 plants needed** across India in next 5-7 years
        - India imports **49% of bitumen** (Rs 25,000 Cr/year)
        - Government target: **100% import replacement** in 10 years
        - **NHAI green mandate** creates guaranteed demand
        """)
    elif selected_type == "bitumen_owner":
        st.markdown("""
        ### Your Plant Can Earn MORE
        - Current VG30 margin: **Rs 2,000-3,000/MT** (thin, volatile)
        - With bio-blending: **Rs 5,000-8,000/MT** (stable, premium)
        - **15-30% bio-oil** replaces crude-based bitumen
        - Lower raw material cost = **higher profit per ton**
        - NHAI preference for **green bitumen** = more contracts
        """)
    elif selected_type == "biomass_company":
        st.markdown("""
        ### Your Waste Has HIGH VALUE
        - Rice straw currently: **Rs 2,000/MT** (or burned!)
        - Converted to bio-bitumen: **Rs 35,000-45,000/MT product**
        - Value multiplication: **15-20x** from waste to product
        - Solves stubble burning problem = **government support**
        - Carbon credits potential = **additional revenue**
        """)
    elif selected_type == "pyrolysis_owner":
        st.markdown("""
        ### Upgrade Your Output Value
        - Current pyrolysis oil selling: **Rs 15-20/litre** (commodity)
        - Bio-bitumen selling: **Rs 35,000-45,000/MT** (premium)
        - Same oil + blending = **3x-5x value increase**
        - Direct access to **road construction market**
        - Prince provides **VG30 supply** for blending
        """)

    # Market data chart (common)
    fig = go.Figure()
    years = [2022, 2023, 2024, 2025, 2026, 2027]
    demand = [7.2, 7.5, 7.8, 8.0, 8.5, 9.0]
    production = [5.1, 5.2, 5.3, 5.5, 5.6, 5.8]
    fig.add_trace(go.Bar(x=years, y=demand, name="Demand (Mn MT)", marker_color='#CC3333'))
    fig.add_trace(go.Bar(x=years, y=production, name="Production (Mn MT)", marker_color='#003366'))
    fig.update_layout(title="India Bitumen Demand vs Production Gap", barmode="group",
                       template="plotly_white", height=350)
    st.plotly_chart(fig, width="stretch")

    st.markdown("**Why NOW facts:**")
    for fact in WHY_NOW:
        st.markdown(f"- {fact}")

# ── STEP 2: FEASIBILITY ─────────────────────────────────────────────
elif step_idx == 1:
    st.header("Step 2: Feasibility Analysis")

    if selected_type == "new_investor":
        st.markdown("### Full Feasibility — Location, Raw Material, Market")
        st.page_link("pages/05_📍_Location.py", label="📍 Location Feasibility (18 States Scored)", icon="📍")
        st.page_link("pages/06_🌾_Raw_Material.py", label="🌾 Raw Material Analysis", icon="🌾")
        st.markdown(f"""
        **For your {cfg['capacity_tpd']:.0f} TPD plant:**
        - Land required: **{int(cfg['capacity_tpd'] * 520):,} sq ft**
        - Raw material: **{cfg['capacity_tpd']:.0f} MT/day** biomass
        - Staff: **{cfg['staff']} persons**
        - Power: **{cfg['power_kw']:.0f} kW**
        """)
    elif selected_type == "bitumen_owner":
        st.markdown("""
        ### Upgrade Feasibility — Minimal Investment
        - **No new land needed** — use existing plant
        - **Existing tanks** can be modified for blending
        - Only need: Blending unit + Heating system + Bio-oil storage
        - Timeline: **30-60 days** from order to production
        - Investment: **Rs 40-80 Lakh** (add-on to existing)
        """)
    elif selected_type == "biomass_company":
        st.markdown("""
        ### Your Raw Material = Your Strength
        - You already have **biomass supply** (near-zero cost)
        - You already have **land and sheds**
        - You need: Pyrolysis reactor + Blending unit
        - Your margin will be **HIGHEST** because input cost is lowest
        """)
        st.page_link("pages/06_🌾_Raw_Material.py", label="🌾 See Biomass Availability Data", icon="🌾")
    elif selected_type == "pyrolysis_owner":
        st.markdown("""
        ### You're Already 60% There
        - You have: Pyrolysis reactor ✅ Bio-oil production ✅
        - You need: Blending unit + VG30 supply + Quality lab
        - Investment: **Rs 40-80 Lakh** only
        - Prince arranges: **VG30 bitumen supply** (int'l network, 2.4L MT/yr)
        """)

# ── STEP 3: TECHNICAL SOLUTION ───────────────────────────────────────
elif step_idx == 2:
    st.header("Step 3: Technical Solution")

    if selected_type == "new_investor":
        st.markdown("### Complete 4-Stage Plant")
        for stage in FOUR_STAGES:
            with st.expander(f"STAGE {stage['stage']}: {stage['name']}", expanded=stage['stage'] == 1):
                st.markdown(stage['description'])
                st.markdown(f"**CAPEX:** {stage['capex']} | **Manpower:** {stage['manpower']} | **Space:** {stage.get('space', stage.get('network', 'N/A'))}")
        st.page_link("pages/07_⚙️_Plant_Design.py", label="⚙️ Full Plant Design (5-100 TPD)", icon="⚙️")
        st.page_link("pages/53_Process_Flow.py", label="🔄 13-Step Process Flow", icon="🔄")

    elif selected_type == "bitumen_owner":
        st.markdown("""
        ### Retrofit Solution — Add Bio-Blending
        **What changes in your plant:**
        1. **Add bio-oil receiving tank** (10-50 KL based on scale)
        2. **Add high-shear mixer** for uniform blending
        3. **Modify heating system** for bio-oil pre-heating (230-250°C)
        4. **Add quality testing** equipment (penetration, softening point)
        5. **Existing storage tanks** — no change needed
        6. **Existing dispatch** — no change needed

        **Blending ratio:** 15-30% bio-oil with your existing VG-30
        **Result:** Bio-Modified Bitumen meeting NHAI specs
        """)

    elif selected_type == "biomass_company":
        st.markdown("""
        ### Biomass → Bio-Oil → Bio-Bitumen
        **Your journey:**
        1. Your existing biomass → **Pelletization** (if not already pellets)
        2. Pellets → **Pyrolysis reactor** (450-550°C) → Bio-oil + Biochar + Syngas
        3. Bio-oil → **Blending with VG-30** (Prince supplies VG-30)
        4. Blended product → **Testing & certification**
        5. Certified bio-bitumen → **Sell to NHAI contractors**

        **Your advantage:** Raw material cost near ZERO = HIGHEST margin in the chain
        """)
        st.page_link("pages/51_Technology.py", label="🔬 CSIR-CRRI Technology Details", icon="🔬")

    elif selected_type == "pyrolysis_owner":
        st.markdown("""
        ### Upgrade: Oil → Bitumen
        **You already have:** Bio-oil production ✅
        **You need to add:**
        1. **Bio-oil oxidation unit** (230-250°C upgrading)
        2. **VG-30 receiving & heating tank** (Prince supplies VG-30 from int'l network)
        3. **Blending tank with high-shear mixer**
        4. **Quality testing lab** (penetration, softening point, ductility)
        5. **Storage & dispatch** for finished bio-bitumen

        **Prince provides VG-30 supply:** 2.4 Lakh MT/yr contract with Getka Energy (Iraq/USA)
        """)

    st.page_link("pages/08_📐_Drawings.py", label="📐 View All Engineering Drawings (117)", icon="📐")

# ── STEP 4: FINANCIAL BENEFIT ────────────────────────────────────────
elif step_idx == 3:
    st.header("Step 4: Financial Benefit")

    if selected_type == "new_investor":
        st.markdown(f"""
        ### Full Investment — {cfg['capacity_tpd']:.0f} TPD Plant
        | Parameter | Value |
        |-----------|-------|
        | Total Investment | **Rs {cfg['investment_cr']:.2f} Crore** |
        | Bank Loan (60%) | **Rs {cfg['loan_cr']:.2f} Crore** |
        | Equity (40%) | **Rs {cfg['equity_cr']:.2f} Crore** |
        | Monthly EMI | **Rs {cfg['emi_lac_mth']:.2f} Lakhs** |
        | Revenue Yr5 | **Rs {cfg['revenue_yr5_lac']:.0f} Lakhs** |
        | Profit/MT | **Rs {cfg['profit_per_mt']:,.0f}** |
        | ROI | **{cfg['roi_pct']:.1f}%** |
        | Break-Even | **{cfg['break_even_months']} months** |
        """)

    elif selected_type == "bitumen_owner":
        st.markdown("""
        ### Upgrade ROI — Bio-Blending Addition
        | Parameter | Before (VG30 only) | After (Bio-Blend) |
        |-----------|--------------------|--------------------|
        | Selling Price | Rs 50,000/MT | Rs 48,000/MT (competitive) |
        | Raw Material Cost | Rs 42,000/MT | Rs 35,000/MT (bio-oil cheaper) |
        | **Margin/MT** | **Rs 8,000** | **Rs 13,000** |
        | **Margin Increase** | — | **+62%** |

        - Upgrade investment: **Rs 40-80 Lakh**
        - Payback: **6-12 months**
        - Additional profit: **Rs 5,000/MT x volume**
        """)

    elif selected_type == "biomass_company":
        st.markdown("""
        ### Value Addition — Waste to Product
        | Parameter | Current (Biomass) | After (Bio-Bitumen) |
        |-----------|-------------------|--------------------|
        | Product Value | Rs 2,000/MT | Rs 35,000+/MT |
        | **Value Multiplication** | — | **15-20x** |
        | Raw Material Cost | Near ZERO | Near ZERO (own supply) |
        | Processing Cost | — | Rs 15,000/MT |
        | **Net Profit/MT** | **Rs 1,000** | **Rs 18,000+** |
        """)

    elif selected_type == "pyrolysis_owner":
        st.markdown("""
        ### Value Upgrade — Oil to Bitumen
        | Parameter | Current (Oil) | After (Bio-Bitumen) |
        |-----------|---------------|--------------------|
        | Product | Pyrolysis Oil | Bio-Modified Bitumen |
        | Selling Price | Rs 15-20/L (~Rs 18,000/MT) | Rs 35,000-45,000/MT |
        | **Value Increase** | — | **2x-3x** |
        | Upgrade Cost | — | Rs 40-80 Lakh |
        | Payback | — | **8-12 months** |
        """)

    st.page_link("pages/09_💰_Financial.py", label="💰 Open Full Financial Model (Editable)", icon="💰")

    # Common: 7-year chart
    if cfg.get("roi_timeline"):
        roi_df = pd.DataFrame(cfg["roi_timeline"])
        fig = go.Figure()
        fig.add_trace(go.Bar(x=roi_df["Year"], y=roi_df["Revenue (Lac)"], name="Revenue", marker_color='#003366'))
        fig.add_trace(go.Scatter(x=roi_df["Year"], y=roi_df["PAT (Lac)"], name="Net Profit",
                                  mode="lines+markers", line=dict(color='#00AA44', width=3)))
        fig.update_layout(title="7-Year Projection", template="plotly_white", height=350, barmode="group")
        st.plotly_chart(fig, width="stretch")

# ── STEP 5: IMPLEMENTATION PLAN ──────────────────────────────────────
elif step_idx == 4:
    st.header("Step 5: Implementation Plan")

    if selected_type in ("new_investor", "biomass_company"):
        st.markdown("""
        ### 12-18 Month Implementation Timeline
        | Phase | Duration | Activities |
        |-------|----------|------------|
        | 1. Pre-Feasibility | Month 0-1 | DPR, site selection, feasibility |
        | 2. Company Setup | Month 1-2 | ROC, GST, PAN, Udyam |
        | 3. Land & Approvals | Month 1-4 | Land, building plan, CTE |
        | 4. Bank Loan | Month 2-5 | DPR submission, sanction |
        | 5. Design | Month 3-6 | Engineering, P&ID, civil |
        | 6. Procurement | Month 4-8 | Equipment orders, delivery |
        | 7. Construction | Month 5-10 | Civil, structural, erection |
        | 8. Commissioning | Month 9-12 | Installation, trial run |
        | 9. CTO | Month 12-13 | Consent to Operate |
        | 10. Production | Month 13+ | Commercial operations, ramp-up |
        """)
    elif selected_type == "bitumen_owner":
        st.markdown("""
        ### 30-60 Day Upgrade Plan
        | Week | Activity |
        |------|----------|
        | Week 1-2 | Design review, equipment ordering |
        | Week 3-4 | Civil modification for blending unit |
        | Week 5-6 | Equipment installation |
        | Week 7 | Piping, electrical connections |
        | Week 8 | Trial run, quality testing |
        | Week 9+ | Commercial production |
        """)
    elif selected_type == "pyrolysis_owner":
        st.markdown("""
        ### 45-90 Day Upgrade Plan
        | Week | Activity |
        |------|----------|
        | Week 1-2 | Design, VG-30 supply arrangement |
        | Week 3-4 | Blending unit procurement |
        | Week 5-8 | Installation, piping |
        | Week 9-10 | Lab setup, quality testing |
        | Week 11-12 | Trial production, NHAI spec testing |
        | Week 13+ | Commercial production + sales |
        """)

    st.page_link("pages/54_Timeline.py", label="📅 Detailed Project Timeline", icon="📅")

# ── STEP 6: COMPLIANCE ───────────────────────────────────────────────
elif step_idx == 5:
    st.header("Step 6: Compliance & Licenses")

    if selected_type in ("new_investor", "biomass_company"):
        st.markdown("### 25 Licenses Required — All Tracked")
        st.markdown("""
        **Critical (do first):** GST, Udyam MSME, CTE from PCB, Factory License, Fire NOC
        **Parallel:** PESO, Electricity HT, EPF/ESI, Professional Tax
        **Post-commissioning:** CTO, BIS Certification, NABL
        """)
    elif selected_type in ("bitumen_owner", "pyrolysis_owner"):
        st.markdown("""
        ### Minimal Additional Compliance
        You already have most licenses. Additional needed:
        - **CTE Amendment** (for bio-blending process addition)
        - **PESO review** (if new petroleum storage)
        - **Product BIS certification** (for bio-bitumen grade)
        - **NHAI material approval** (for road project supply)
        """)

    st.page_link("pages/11_📋_Compliance.py", label="📋 Full Compliance Tracker (25 Licenses)", icon="📋")

# ── STEP 7: EXECUTION ────────────────────────────────────────────────
elif step_idx == 6:
    st.header("Step 7: Execution — What Prince Provides")

    st.markdown("### At EVERY Stage, Prince Provides:")

    svc_cols = st.columns(3)
    services = list(CONSULTING_SERVICES.items())
    for i, (name, items) in enumerate(services):
        with svc_cols[i % 3]:
            st.markdown(f"**{name}**")
            for item in items:
                st.markdown(f"- {item}")

    st.markdown("---")
    st.markdown("### Prince's Track Record")
    career_df = pd.DataFrame(CAREER_TRACK)
    st.dataframe(career_df, width="stretch", hide_index=True)

    st.markdown("### Industry Network")
    net = INDUSTRY_NETWORK
    nc = st.columns(6)
    nc[0].metric("Contractors", f"{net['contractors']:,}")
    nc[1].metric("Traders", f"{net['traders']:,}")
    nc[2].metric("Importers", f"{net['importers']:,}")
    nc[3].metric("Transporters", f"{net['transporters']:,}")
    nc[4].metric("Manufacturers", f"{net['manufacturers']:,}")
    nc[5].metric("Total Network", f"{net['total']:,}")

    # Fee structure based on client type
    audience = None
    for ta in TARGET_AUDIENCES:
        if selected_type == "new_investor" and "New Investor" in ta["type"]:
            audience = ta
        elif selected_type == "bitumen_owner" and "Bitumen" in ta["type"]:
            audience = ta
        elif selected_type == "pyrolysis_owner" and "Pyrolysis" in ta["type"]:
            audience = ta
        elif selected_type == "biomass_company" and "Pellet" in ta["type"]:
            audience = ta

    if audience:
        st.markdown("---")
        st.markdown(f"### Consulting Fee Structure — {audience['type']}")
        st.markdown(f"- **DPR:** {audience['fee_dpr']}")
        st.markdown(f"- **Setup:** {audience['fee_setup']}")
        st.markdown(f"- **Retainer:** {audience['fee_retainer']}")

# ── STEP 8: FINAL OUTPUT / PROFIT ────────────────────────────────────
elif step_idx == 7:
    st.header("Step 8: Final Output — Ready to Start")

    st.success(f"""
    ### Everything is READY for {ct['name']}

    **Documents Ready:**
    - DPR (Detailed Project Report) — auto-generated
    - Financial Model (Excel) — with 7-year projection
    - Investor Pitch (PowerPoint) — 8 slides
    - Bank Proposal (Word) — with security details
    - Engineering Drawings — 117 drawings, 13 types
    - Compliance Checklist — 25 licenses tracked
    - Vendor Database — 20+ verified suppliers
    - Buyer Network — 2,758 contractors + 994 traders

    **Investment:** {ct['investment']}
    **Timeline:** {'12-18 months' if selected_type in ('new_investor', 'biomass_company') else '30-90 days'}
    """)

    st.markdown("### Generate Your Documents NOW")
    doc_cols = st.columns(4)
    doc_cols[0].page_link("pages/44_DPR_Generator.py", label="📄 Generate DPR", icon="📄")
    doc_cols[1].page_link("pages/09_💰_Financial.py", label="💰 Financial Model", icon="💰")
    doc_cols[2].page_link("pages/08_📐_Drawings.py", label="📐 All Drawings", icon="📐")
    doc_cols[3].page_link("pages/15_🤖_AI_Advisor.py", label="🤖 Ask AI", icon="🤖")

    st.markdown("---")
    st.markdown(f"""
    ### Ready to Start? Contact:
    **{COMPANY['owner']}** | **{COMPANY['phone']}** | **{COMPANY['email']}**

    *{COMPANY['usp']}*
    """)

    # Key credentials
    st.markdown("### Why Choose Prince?")
    for cred in KEY_CREDENTIALS:
        st.markdown(f"- {cred}")

st.markdown("---")
st.caption(f"{COMPANY['name']} | {COMPANY['owner']} | {COMPANY['phone']} | {COMPANY['hq']}")


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    if st.button("Download Excel", type="primary", key="exp_xl_50Cli"):
        try:
            import io
            from openpyxl import Workbook
            _wb = Workbook()
            _ws = _wb.active
            _ws.title = "Export"
            _ws.cell(row=1, column=1, value="Bio Bitumen Export")
            _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
            _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
            _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
            _buf = io.BytesIO()
            _wb.save(_buf)
            _buf.seek(0)
            st.download_button("Download", _buf.getvalue(), "export.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_xl_50Cli")
        except Exception as _e:
            st.error(f"Export failed: {_e}")
with _ex2:
    if st.button("Print", key="exp_prt_50Cli"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)


# ── AI Assist ────────────────────────────────────────────────────
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("AI Assist"):
            if st.button("Generate AI Summary", type="primary", key="ai_50Cli"):
                with st.spinner("AI working..."):
                    _p = f"Summarize this section for a {cfg.get('capacity_tpd',20):.0f} TPD bio-bitumen plant in {cfg.get('state','')}. Investment Rs {cfg.get('investment_cr',8):.2f} Cr. Professional consultant format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 800)
                if _r:
                    st.markdown(_r)
except Exception:
    pass
