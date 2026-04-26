"""
Bio Bitumen Master Consulting System — Location & Feasibility Module
====================================================================
State-wise feasibility scoring, subsidy tracking, logistics analysis.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.express as px
from config import STATES, STATE_SCORES, LOCATION_WEIGHTS
from state_manager import get_config, update_field, init_state

st.set_page_config(page_title="Location & Feasibility", page_icon="📍", layout="wide")
init_state()
try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass
st.sidebar.markdown("---")
if st.sidebar.button("Print This Page", key="print_page"):
    import streamlit.components.v1 as _stc; _stc.html('<script>window.print();</script>', height=0)

st.title("Location & Feasibility Analysis")
st.markdown("**Pan-India state-wise scoring for Bio-Bitumen plant setup**")
st.markdown("---")

# ── Calculate Scores ──────────────────────────────────────────────────
state_data = []
for state in STATES:
    scores = STATE_SCORES.get(state, {})
    total = sum(scores.get(k, 50) * LOCATION_WEIGHTS[k] for k in LOCATION_WEIGHTS)
    state_data.append({
        "State": state,
        "Biomass (25%)": scores.get("biomass", 50),
        "Subsidy (20%)": scores.get("subsidy", 50),
        "Logistics (20%)": scores.get("logistics", 50),
        "Power (15%)": scores.get("power", 50),
        "Land Cost (10%)": scores.get("land_cost", 50),
        "Season (10%)": scores.get("season", 50),
        "Total Score": round(total, 1),
        "Rating": "Excellent" if total >= 75 else "Good" if total >= 65 else "Fair" if total >= 55 else "Below Avg",
    })

df = pd.DataFrame(state_data).sort_values("Total Score", ascending=False)

# ── Top Metrics ───────────────────────────────────────────────────────
top3 = df.head(3)
c1, c2, c3 = st.columns(3)
for i, (_, row) in enumerate(top3.iterrows()):
    medals = ["1st", "2nd", "3rd"]
    with [c1, c2, c3][i]:
        st.metric(f"{medals[i]} Best State", row["State"], f"Score: {row['Total Score']}/100")

st.markdown("---")

# ── Tabs ──────────────────────────────────────────────────────────────
tab_rank, tab_detail, tab_compare = st.tabs(["State Rankings", "State Detail", "Factor Analysis"])

with tab_rank:
    st.subheader("All 18 States — Ranked by Feasibility Score")

    # Color-coded bar chart
    fig = px.bar(df, x="State", y="Total Score", color="Rating",
                  color_discrete_map={"Excellent": "#006400", "Good": "#228B22", "Fair": "#FFA500", "Below Avg": "#CC3333"},
                  title="State-wise Feasibility Score (0-100)")
    fig.update_layout(xaxis_tickangle=-45)
    st.plotly_chart(fig, width="stretch")

    # Full table
    st.dataframe(df, width="stretch", hide_index=True)

with tab_detail:
    st.subheader("State Detail View")
    selected_state = st.selectbox("Select State", options=STATES,
                                   index=STATES.index(get_config()["state"]) if get_config()["state"] in STATES else 0)

    state_row = df[df["State"] == selected_state].iloc[0] if not df[df["State"] == selected_state].empty else None
    if state_row is not None:
        st.markdown(f"### {selected_state} — Score: **{state_row['Total Score']}/100** ({state_row['Rating']})")

        d1, d2, d3 = st.columns(3)
        d1.metric("Biomass Availability", f"{state_row['Biomass (25%)']}/100")
        d2.metric("Govt Subsidy", f"{state_row['Subsidy (20%)']}/100")
        d3.metric("Logistics", f"{state_row['Logistics (20%)']}/100")

        d4, d5, d6 = st.columns(3)
        d4.metric("Power Availability", f"{state_row['Power (15%)']}/100")
        d5.metric("Land Cost (lower=better)", f"{state_row['Land Cost (10%)']}/100")
        d6.metric("Season Favorability", f"{state_row['Season (10%)']}/100")

        # Radar chart
        categories = ["Biomass", "Subsidy", "Logistics", "Power", "Land Cost", "Season"]
        values = [state_row[f"Biomass (25%)"], state_row[f"Subsidy (20%)"],
                  state_row[f"Logistics (20%)"], state_row[f"Power (15%)"],
                  state_row[f"Land Cost (10%)"], state_row[f"Season (10%)"]]
        fig_radar = px.line_polar(
            r=values + [values[0]],
            theta=categories + [categories[0]],
            title=f"{selected_state} — Factor Profile",
        )
        fig_radar.update_traces(fill="toself", fillcolor="rgba(0,51,102,0.2)", line_color="#003366")
        st.plotly_chart(fig_radar, width="stretch")

        if st.button(f"Set {selected_state} as project location"):
            update_field("state", selected_state)
            st.success(f"Project location set to {selected_state}")

with tab_compare:
    st.subheader("Factor-wise Comparison")

    factor = st.selectbox("Select Factor", ["Biomass (25%)", "Subsidy (20%)", "Logistics (20%)",
                                              "Power (15%)", "Land Cost (10%)", "Season (10%)"])
    fig_factor = px.bar(df.sort_values(factor, ascending=False), x="State", y=factor,
                         color=factor, color_continuous_scale="Greens",
                         title=f"{factor} — State Comparison")
    fig_factor.update_layout(xaxis_tickangle=-45)
    st.plotly_chart(fig_factor, width="stretch")

# ── Export Section ────────────────────────────────────────────────────
st.markdown("---")
st.subheader("Export")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_05_📍_L", type="primary")
with _ex2:
    if st.button("Print Page", key="exp_print_analysis"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)


# ── AI Skill: Site Selection Advice ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: Site Selection Advice"):
            if st.button("Generate", type="primary", key="ai_05📍Loc"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: Site Selection Advice. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass


# ── Live Weather & Site Conditions ────────────────────────────────────
st.markdown("---")
st.subheader("🌤️ Live Site Weather & Conditions")
st.caption("Real-time data via Open-Meteo (free, no key) — updates every 10 min")

try:
    from engines.free_apis import get_weather_current, get_weather_forecast, get_india_holidays
    _city  = cfg.get("location", "") or cfg.get("state", "Delhi")
    _state = cfg.get("state", "")

    _wc1, _wc2, _wc3 = st.columns([2, 2, 1])

    with _wc1:
        _w = get_weather_current(_city)
        if "error" not in _w:
            st.markdown(f"""
<div style="background:#1E1B14;border:1px solid #3A3520;border-radius:10px;padding:14px;">
  <div style="font-size:1.6rem;font-weight:800;color:#E8B547;">{_w.get('temperature_c','?')}°C</div>
  <div style="color:#9A8A6A;font-size:13px;">{_w.get('condition','')}</div>
  <div style="color:#C8B88A;font-size:12px;margin-top:6px;">
    💧 Humidity: {_w.get('humidity_pct','?')}%&nbsp;&nbsp;
    🌬️ Wind: {_w.get('wind_kmh','?')} km/h
  </div>
  <div style="color:#7A6A4A;font-size:11px;margin-top:4px;">{_city}, {_state}</div>
</div>
""", unsafe_allow_html=True)
        else:
            st.info(f"Weather unavailable: {_w.get('error','')}")

    with _wc2:
        _forecast = get_weather_forecast(_city, days=5)
        if _forecast and "error" not in _forecast:
            import pandas as _pd
            _fdf = _pd.DataFrame(_forecast[:5])
            if not _fdf.empty and "date" in _fdf.columns:
                import plotly.graph_objects as _go
                _fig_w = _go.Figure()
                _fig_w.add_trace(_go.Bar(
                    x=_fdf["date"], y=_fdf.get("temp_max_c", _fdf.get("temperature_c", [])),
                    name="Max °C", marker_color="#E8B547", opacity=0.8
                ))
                if "temp_min_c" in _fdf.columns:
                    _fig_w.add_trace(_go.Bar(
                        x=_fdf["date"], y=_fdf["temp_min_c"],
                        name="Min °C", marker_color="#74C0FC", opacity=0.7
                    ))
                _fig_w.update_layout(
                    title="5-Day Forecast", height=200,
                    margin=dict(t=30, b=10, l=30, r=10),
                    paper_bgcolor="#1E1B14", plot_bgcolor="#1E1B14",
                    font_color="#C8B88A", showlegend=True,
                    legend=dict(font=dict(color="#9A8A6A", size=10)),
                    barmode="group",
                )
                st.plotly_chart(_fig_w, use_container_width=True,
                                config={"displayModeBar": False})

    with _wc3:
        st.markdown("**Operational Impact**")
        _temp = _w.get("temperature_c", 25) if "error" not in _w else 25
        _hum  = _w.get("humidity_pct", 50) if "error" not in _w else 50
        if _temp > 40:
            st.warning("🔥 High temp — extra cooling needed for biomass storage")
        elif _temp < 10:
            st.warning("❄️ Cold — viscosity checks critical for bitumen output")
        else:
            st.success("✅ Temp optimal for plant operations")
        if _hum > 75:
            st.warning("💧 High humidity — biomass drying time +20%")
        elif _hum < 30:
            st.success("✅ Low humidity — fast biomass drying")

        # Upcoming holidays
        _yr = __import__('datetime').datetime.now().year
        _hols = get_india_holidays(_yr)
        if _hols:
            _upcoming = [h for h in _hols
                         if h["date"] >= __import__('datetime').datetime.now().strftime("%Y-%m-%d")][:3]
            if _upcoming:
                st.markdown("**Upcoming Holidays:**")
                for _h in _upcoming:
                    st.markdown(f"- {_h['date'][5:]} {_h['name_en']}")

except Exception as _ex:
    st.info(f"Live weather: {_ex}")


# ── Next Steps Navigation ──
try:
    from engines.page_navigation import add_next_steps
    add_next_steps(st, "12")
except Exception:
    pass
