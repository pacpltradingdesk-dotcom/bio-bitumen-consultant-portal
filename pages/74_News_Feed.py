"""
Industry News Feed — Curated Bio-Bitumen & Infrastructure Updates
==================================================================
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
from state_manager import init_state, get_config
from config import COMPANY, INDUSTRY_NEWS, WHY_NOW

st.set_page_config(page_title="Industry News", page_icon="📰", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass


st.title("Industry News & Market Updates")
st.markdown("**Bio-Bitumen | NHAI | Government Policy | Technology | Market Prices**")
st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# CATEGORY FILTER
# ══════════════════════════════════════════════════════════════════════
categories = sorted(set(n["category"] for n in INDUSTRY_NEWS))
all_cats = ["All"] + categories

filter_col, search_col = st.columns([1, 2])
with filter_col:
    selected_cat = st.selectbox("Filter by Category", all_cats, key="news_cat")
with search_col:
    search_term = st.text_input("Search News", placeholder="Search by keyword...", key="news_search")

filtered = INDUSTRY_NEWS.copy()
if selected_cat != "All":
    filtered = [n for n in filtered if n["category"] == selected_cat]
if search_term:
    term = search_term.lower()
    filtered = [n for n in filtered if term in n["title"].lower() or term in n["summary"].lower()]

# ── Category counts ──────────────────────────────────────────────────
cat_counts = {}
for n in INDUSTRY_NEWS:
    cat_counts[n["category"]] = cat_counts.get(n["category"], 0) + 1

cols = st.columns(len(categories))
for i, cat in enumerate(categories):
    cols[i].metric(cat, cat_counts.get(cat, 0))

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# NEWS FEED
# ══════════════════════════════════════════════════════════════════════
st.subheader(f"Latest Updates ({len(filtered)} articles)")

cat_colors = {
    "Government": "#003366", "Technology": "#006699", "Market": "#FF8800",
    "Policy": "#00AA44", "Industry": "#AA3366", "Price": "#CC3333",
}

for news in filtered:
    color = cat_colors.get(news["category"], "#666666")
    st.markdown(f"""
    <div style="border-left: 4px solid {color}; padding: 12px 20px; margin-bottom: 15px;
                background: {color}08; border-radius: 0 8px 8px 0;">
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <span style="background: {color}; color: white; padding: 2px 10px; border-radius: 12px;
                         font-size: 0.8em;">{news['category']}</span>
            <span style="color: #888; font-size: 0.85em;">{news['date']}</span>
        </div>
        <h4 style="margin: 8px 0 5px 0; color: #222;">{news['title']}</h4>
        <p style="color: #555; margin: 0; font-size: 0.95em;">{news['summary']}</p>
        <p style="color: #999; font-size: 0.8em; margin: 5px 0 0 0;">Source: {news['source']}</p>
    </div>
    """, unsafe_allow_html=True)

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# WHY BIO-BITUMEN NOW
# ══════════════════════════════════════════════════════════════════════
st.subheader("Why Bio-Bitumen NOW — Key Market Drivers")

for i, point in enumerate(WHY_NOW, 1):
    st.markdown(f"""
    <div style="background: linear-gradient(90deg, #003366{hex(20 + i*8)[2:]}, transparent);
                padding: 10px 15px; border-radius: 8px; margin-bottom: 8px; color: white;">
        <strong>{i}.</strong> {point}
    </div>
    """, unsafe_allow_html=True)

st.markdown("---")

# ── Market Commentary ────────────────────────────────────────────────
st.subheader("Market Commentary")
st.markdown("""
**Q1 2026 Outlook:**

The bio-bitumen market is at an inflection point. With CSIR-CRRI's technology transfer to 14+ companies
and Minister Gadkari's announcement of India becoming the first country to commercially produce bio-bitumen,
the market is set for exponential growth.

**Key Trends:**
- VG30 prices at 3-year highs (Rs 40,000+/MT) — increasing bio-bitumen cost advantage
- NHAI green mandates accelerating — 15% bio-bitumen target by 2030
- Carbon credit framework now operational — additional revenue stream
- Equipment costs dropping — pelletizer prices down 30% in 2 years
- State subsidies expanding — MNRE Waste-to-Wealth now covers bio-bitumen

**PPS Anantams Position:** With 25 years experience, 4,452 contacts, and VG-30 import contract,
we are uniquely positioned to serve the 130-216 new plants needed in the next 5-7 years.
""")

st.markdown("---")
st.caption(f"{COMPANY['name']} | Industry Intelligence Division")


# ── AI Skill: Industry Brief ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: Industry Brief"):
            if st.button("Generate", type="primary", key="ai_68News"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: Industry Brief. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_68New", type="primary")
with _ex2:
    if st.button("Print", key="exp_prt_68New"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)


# ── Next Steps Navigation ──
try:
    from engines.page_navigation import add_next_steps
    add_next_steps(st, "74")
except Exception:
    pass
