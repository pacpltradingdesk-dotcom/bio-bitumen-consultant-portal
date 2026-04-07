"""
Capacity Comparison — Side-by-Side Analysis of Multiple Plant Sizes
====================================================================
Select 2-3 capacities, see all metrics compared with charts and radar overlay.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from state_manager import get_config, init_state
from engines.three_process_model import calculate_process
from config import COMPANY, CAPACITY_LABELS

st.set_page_config(page_title="Capacity Comparison", page_icon="⚖️", layout="wide")
init_state()
cfg = get_config()
# Fix metric truncation
try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass


st.page_link("pages/30_💰_Financial.py", label="← Back to Financial Model", icon="💰")

st.title("Capacity Comparison Tool")
st.markdown("**Compare 2-3 plant capacities side-by-side — Investment, ROI, Profit, Break-Even**")
st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# SELECT CAPACITIES
# ══════════════════════════════════════════════════════════════════════
capacity_options = [5, 10, 15, 20, 30, 40, 50]

sel_col, proc_col, budget_col = st.columns(3)
with sel_col:
    selected_caps = st.multiselect("Select Capacities to Compare (2-3)",
                                    capacity_options, default=[10, 20, 40], max_selections=3,
                                    key="compare_caps")
with proc_col:
    process = st.selectbox("Process Model", [1, 2, 3],
                            format_func=lambda x: {1: "Full Chain", 2: "Blending Only", 3: "Raw Output"}[x],
                            key="compare_process")
with budget_col:
    budget_cr = st.number_input("Your Budget (Rs Cr)", 0.5, 50.0, 8.0, 0.5, key="compare_budget")

if len(selected_caps) < 2:
    st.warning("Please select at least 2 capacities to compare.")
    st.stop()

# ── Calculate for each capacity ──────────────────────────────────────
results = {}
for cap in selected_caps:
    results[cap] = calculate_process(process, float(cap))

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# SIDE-BY-SIDE METRICS
# ══════════════════════════════════════════════════════════════════════
st.subheader("Side-by-Side Comparison")

cols = st.columns(len(selected_caps))
for i, cap in enumerate(selected_caps):
    r = results[cap]
    with cols[i]:
        investment = r.get("capex_lac", 0) / 100
        within_budget = investment <= budget_cr
        color = "#00AA44" if within_budget else "#CC3333"

        st.markdown(f"""
        <div style="background: {color}11; border: 2px solid {color}; border-radius: 12px;
                    padding: 20px; text-align: center;">
            <h2 style="color: {color}; margin: 0;">{cap} TPD</h2>
            <p style="color: #666;">{'Within Budget' if within_budget else 'Over Budget'}</p>
        </div>
        """, unsafe_allow_html=True)

        st.metric("Investment", f"Rs {investment:.1f} Cr")
        st.metric("ROI", f"{r.get('roi_pct', 0):.1f}%")
        st.metric("IRR", f"{r.get('irr_pct', 0):.1f}%")
        st.metric("Break-Even", f"{r.get('break_even_months', 0)} months")
        st.metric("Monthly Profit", f"Rs {r.get('pat_yr5_lac', 0)/12:.1f} Lac")
        st.metric("Profit/MT", f"Rs {r.get('profit_per_mt', 0):,.0f}")
        st.metric("DSCR Yr3", f"{r.get('dscr_yr3', 0):.2f}x")
        st.metric("EMI/Month", f"Rs {r.get('emi_lac_mth', 0):.2f} Lac")
        st.metric("Staff Required", f"{r.get('staff', 0)}")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# GROUPED BAR CHARTS
# ══════════════════════════════════════════════════════════════════════
st.subheader("Visual Comparison")

chart_col1, chart_col2 = st.columns(2)

with chart_col1:
    # Investment & Revenue comparison
    comp_df = pd.DataFrame([
        {"Capacity": f"{cap} TPD",
         "Investment (Lac)": results[cap].get("capex_lac", 0),
         "Annual Revenue (Lac)": results[cap].get("revenue_yr5_lac", results[cap].get("annual_revenue_lac", 0)),
         "Annual Profit (Lac)": results[cap].get("pat_yr5_lac", 0)}
        for cap in selected_caps
    ])

    fig_bar = go.Figure()
    fig_bar.add_trace(go.Bar(x=comp_df["Capacity"], y=comp_df["Investment (Lac)"],
                              name="Investment", marker_color="#003366"))
    fig_bar.add_trace(go.Bar(x=comp_df["Capacity"], y=comp_df["Annual Revenue (Lac)"],
                              name="Annual Revenue", marker_color="#006699"))
    fig_bar.add_trace(go.Bar(x=comp_df["Capacity"], y=comp_df["Annual Profit (Lac)"],
                              name="Annual Profit", marker_color="#00AA44"))
    fig_bar.update_layout(title="Investment vs Revenue vs Profit (Rs Lac)",
                           barmode="group", template="plotly_white", height=400)
    st.plotly_chart(fig_bar, width="stretch")

with chart_col2:
    # ROI / IRR / DSCR comparison
    perf_df = pd.DataFrame([
        {"Capacity": f"{cap} TPD",
         "ROI %": results[cap].get("roi_pct", 0),
         "IRR %": results[cap].get("irr_pct", 0),
         "DSCR": results[cap].get("dscr_yr3", 0) * 10}  # Scale for visibility
        for cap in selected_caps
    ])

    fig_perf = go.Figure()
    fig_perf.add_trace(go.Bar(x=perf_df["Capacity"], y=perf_df["ROI %"],
                               name="ROI %", marker_color="#003366"))
    fig_perf.add_trace(go.Bar(x=perf_df["Capacity"], y=perf_df["IRR %"],
                               name="IRR %", marker_color="#FF8800"))
    fig_perf.update_layout(title="Performance Metrics Comparison",
                            barmode="group", template="plotly_white", height=400)
    st.plotly_chart(fig_perf, width="stretch")

# ══════════════════════════════════════════════════════════════════════
# RADAR CHART OVERLAY
# ══════════════════════════════════════════════════════════════════════
st.subheader("Radar Chart — Overall Comparison")

radar_categories = ["ROI", "IRR", "DSCR", "Profit/MT", "Break-Even\n(inverse)"]
colors_radar = ["#003366", "#006699", "#0088cc"]

fig_radar = go.Figure()
for i, cap in enumerate(selected_caps):
    r = results[cap]
    # Normalize to 0-100 scale
    roi_norm = min(r.get("roi_pct", 0) / 50 * 100, 100)
    irr_norm = min(r.get("irr_pct", 0) / 40 * 100, 100)
    dscr_norm = min(r.get("dscr_yr3", 0) / 3 * 100, 100)
    profit_norm = min(r.get("profit_per_mt", 0) / 15000 * 100, 100)
    be_norm = max(100 - r.get("break_even_months", 48) / 60 * 100, 0)

    values = [roi_norm, irr_norm, dscr_norm, profit_norm, be_norm]
    values.append(values[0])  # Close the polygon

    fig_radar.add_trace(go.Scatterpolar(
        r=values,
        theta=radar_categories + [radar_categories[0]],
        fill="toself",
        name=f"{cap} TPD",
        line=dict(color=colors_radar[i]),
        opacity=0.6,
    ))

fig_radar.update_layout(
    polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
    title="Multi-Dimensional Comparison",
    template="plotly_white",
    height=500,
    showlegend=True,
)
st.plotly_chart(fig_radar, width="stretch")

# ══════════════════════════════════════════════════════════════════════
# RECOMMENDATION ENGINE
# ══════════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("Recommendation")

# Find best within budget
within_budget = {cap: r for cap, r in results.items() if r.get("capex_lac", 0) / 100 <= budget_cr}

if within_budget:
    best_cap = max(within_budget, key=lambda c: within_budget[c].get("roi_pct", 0))
    best_r = within_budget[best_cap]

    st.success(f"""
    **Recommended: {best_cap} TPD** (within your Rs {budget_cr:.1f} Cr budget)

    - Investment: Rs {best_r.get('capex_lac', 0)/100:.1f} Cr
    - ROI: {best_r.get('roi_pct', 0):.1f}%
    - Monthly Profit: Rs {best_r.get('pat_yr5_lac', 0)/12:.1f} Lac
    - Break-Even: {best_r.get('break_even_months', 0)} months
    - DSCR: {best_r.get('dscr_yr3', 0):.2f}x — {'Bankable' if best_r.get('dscr_yr3', 0) >= 1.5 else 'Needs improvement'}
    """)
else:
    st.warning(f"None of the selected capacities fit within Rs {budget_cr:.1f} Cr budget. Consider smaller capacities or higher equity.")

# ── Summary Table ────────────────────────────────────────────────────
st.markdown("---")
summary_data = []
for cap in selected_caps:
    r = results[cap]
    summary_data.append({
        "Capacity (TPD)": cap,
        "Investment (Cr)": round(r.get("capex_lac", 0) / 100, 1),
        "ROI %": round(r.get("roi_pct", 0), 1),
        "IRR %": round(r.get("irr_pct", 0), 1),
        "Break-Even (Months)": r.get("break_even_months", 0),
        "Monthly Profit (Lac)": round(r.get("pat_yr5_lac", 0) / 12, 1),
        "Profit/MT (Rs)": round(r.get("profit_per_mt", 0), 0),
        "DSCR Yr3": round(r.get("dscr_yr3", 0), 2),
        "EMI (Lac/Month)": round(r.get("emi_lac_mth", 0), 2),
        "Staff": r.get("staff", 0),
        "Within Budget": "Yes" if r.get("capex_lac", 0) / 100 <= budget_cr else "No",
    })

st.dataframe(pd.DataFrame(summary_data), width="stretch")

st.markdown("---")
st.caption(f"{COMPANY['name']} | {COMPANY['owner']} | {COMPANY['phone']}")

# ── Export Section ────────────────────────────────────────────────────
st.markdown("---")
st.subheader("Export")
exp1, exp2, exp3 = st.columns(3)
with exp1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_62_Cap", type="primary")
with exp2:
    if st.button("Download CSV", key="exp_csv_calc"):
        import pandas as pd
        data = {"Metric": ["Capacity", "Investment", "ROI", "IRR", "DSCR", "Break-Even", "Monthly Profit"],
                "Value": [f"{cfg['capacity_tpd']:.0f} TPD", f"Rs {cfg['investment_cr']:.2f} Cr",
                          f"{cfg['roi_pct']:.1f}%", f"{cfg['irr_pct']:.1f}%", f"{cfg['dscr_yr3']:.2f}x",
                          f"{cfg['break_even_months']} months", f"Rs {cfg['monthly_profit_lac']:.1f} Lac"]}
        st.download_button("Download", pd.DataFrame(data).to_csv(index=False), "calculator_export.csv", "text/csv", key="dl_csv_c")
with exp3:
    if st.button("Print", key="exp_print_calc"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)


# ── AI Skill: Capacity Recommendation ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: Capacity Recommendation"):
            if st.button("Generate", type="primary", key="ai_62Capac"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: Capacity Recommendation. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass

st.caption("📊 ROI uses operating-profit formula. For bank-standard, see Financial Model.")


# ── Next Steps Navigation ──
try:
    from engines.page_navigation import add_next_steps
    add_next_steps(st, "37")
except Exception:
    pass
