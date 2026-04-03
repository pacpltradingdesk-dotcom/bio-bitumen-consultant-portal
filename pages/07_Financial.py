"""
Financial Model — FULLY EDITABLE inputs → ALL outputs auto-update → Graphs everywhere
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from state_manager import get_config, update_field, update_fields, init_state

st.set_page_config(page_title="Financial Model", page_icon="💰", layout="wide")
init_state()
st.title("Financial Model (Auto-Updating)")
st.markdown("**Edit ANY input below → ALL calculations, charts, and reports auto-update instantly**")

# Live Data Badge
try:
    from engines.live_calculation_engine import get_live_market_inputs, calculate_live_vg30_price
    live = get_live_market_inputs()
    vg30_live = calculate_live_vg30_price(live["crude_oil_usd"], live["usd_inr"])
    is_live = live["data_freshness"] == "live"
    badge_color = "#00AA44" if is_live else "#FF8800"
    st.markdown(f"""
    <div style="background: {badge_color}15; border: 1px solid {badge_color}; border-radius: 8px;
                padding: 8px 15px; display: inline-flex; gap: 20px; align-items: center; margin-bottom: 10px;">
        <span style="color: {badge_color}; font-weight: bold;">{'LIVE DATA' if is_live else 'CACHED DATA'}</span>
        <span>Crude: ${live['crude_oil_usd']:.1f}</span>
        <span>FX: Rs {live['usd_inr']:.2f}</span>
        <span>VG30: Rs {vg30_live['vg30_estimated']:,.0f}</span>
        <span>GDP: {live['gdp_growth_pct']:.1f}%</span>
    </div>
    """, unsafe_allow_html=True)
except Exception:
    pass

st.markdown("---")

cfg = get_config()

# ═══════════════════════════════════════════════════════════════════
# INPUT SHEET (LEFT) → AUTO OUTPUTS (RIGHT)
# ═══════════════════════════════════════════════════════════════════
col_input, col_output = st.columns([1, 1])

with col_input:
    st.subheader("EDITABLE INPUTS")

    st.markdown("**Production**")
    tpd = st.number_input("Capacity (TPD)", 3.0, 100.0, float(cfg["capacity_tpd"]), 5.0, key="fin_tpd")
    days = st.number_input("Working Days/Year", 250, 365, int(cfg["working_days"]), 10, key="fin_days")

    st.markdown("**Revenue**")
    sell_price = st.number_input("Selling Price (Rs/MT)", 20000, 60000, int(cfg["selling_price_per_mt"]), 1000, key="fin_sell")
    biochar_p = st.number_input("Biochar Price (Rs/MT)", 1000, 10000, int(cfg["biochar_price_per_mt"]), 500, key="fin_char")

    st.markdown("**Costs (Rs per MT output)**")
    raw_cost = st.number_input("Raw Material", 3000, 15000, int(cfg["raw_material_cost_per_mt"]), 500, key="fin_raw")
    power_cost = st.number_input("Power & Fuel", 2000, 8000, int(cfg["power_cost_per_mt"]), 500, key="fin_power")
    labour_cost = st.number_input("Labour", 1000, 6000, int(cfg["labour_cost_per_mt"]), 500, key="fin_labour")
    transport = st.number_input("Transport", 500, 5000, int(cfg["transport_cost_per_mt"]), 500, key="fin_trans")

    st.markdown("**Finance**")
    interest = st.slider("Interest Rate (%)", 8.0, 14.0, float(cfg["interest_rate"]*100), 0.5, key="fin_int")
    equity_pct = st.slider("Equity (%)", 20, 50, int(cfg["equity_ratio"]*100), 5, key="fin_eq")

    # Apply changes
    changes = {}
    if tpd != cfg["capacity_tpd"]: changes["capacity_tpd"] = tpd
    if days != cfg["working_days"]: changes["working_days"] = days
    if sell_price != cfg["selling_price_per_mt"]: changes["selling_price_per_mt"] = sell_price
    if biochar_p != cfg["biochar_price_per_mt"]: changes["biochar_price_per_mt"] = biochar_p
    if raw_cost != cfg["raw_material_cost_per_mt"]: changes["raw_material_cost_per_mt"] = raw_cost
    if power_cost != cfg["power_cost_per_mt"]: changes["power_cost_per_mt"] = power_cost
    if labour_cost != cfg["labour_cost_per_mt"]: changes["labour_cost_per_mt"] = labour_cost
    if transport != cfg["transport_cost_per_mt"]: changes["transport_cost_per_mt"] = transport
    if abs(interest/100 - cfg["interest_rate"]) > 0.001: changes["interest_rate"] = interest/100
    if abs(equity_pct/100 - cfg["equity_ratio"]) > 0.01: changes["equity_ratio"] = equity_pct/100

    if changes:
        update_fields(changes)
        cfg = get_config()

with col_output:
    st.subheader("AUTO CALCULATIONS")

    st.markdown("**Production**")
    oc1, oc2 = st.columns(2)
    output_per_day = cfg["capacity_tpd"] * 0.40
    annual_prod = output_per_day * cfg["working_days"]
    oc1.metric("Daily Output", f"{output_per_day:.1f} MT")
    oc2.metric("Annual Production", f"{annual_prod:,.0f} MT")

    st.markdown("**Revenue (Year 5 @ 85% utilization)**")
    or1, or2 = st.columns(2)
    or1.metric("Annual Revenue", f"Rs {cfg['revenue_yr5_lac']:.0f} Lac")
    or2.metric("Monthly Revenue", f"Rs {cfg['revenue_yr5_lac']/12:.1f} Lac")

    st.markdown("**Costs**")
    ot1, ot2 = st.columns(2)
    ot1.metric("Variable Cost/MT", f"Rs {cfg['total_variable_cost_per_mt']:,.0f}")
    ot2.metric("Profit/MT", f"Rs {cfg['profit_per_mt']:,.0f}")

    st.markdown("**Key Metrics**")
    ok1, ok2, ok3 = st.columns(3)
    ok1.metric("ROI", f"{cfg['roi_pct']:.1f}%")
    ok2.metric("IRR", f"{cfg['irr_pct']:.1f}%")
    ok3.metric("Break-Even", f"{cfg['break_even_months']} mo")

    om1, om2, om3 = st.columns(3)
    om1.metric("Investment", f"Rs {cfg['investment_cr']:.2f} Cr")
    om2.metric("Monthly EMI", f"Rs {cfg['emi_lac_mth']:.2f} Lac")
    om3.metric("Monthly Profit", f"Rs {cfg['monthly_profit_lac']:.1f} Lac")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# CHARTS
# ═══════════════════════════════════════════════════════════════════
if cfg["roi_timeline"]:
    roi_df = pd.DataFrame(cfg["roi_timeline"])

    # Chart 1: Revenue vs Cost vs Profit
    st.subheader("7-Year Projection")
    fig1 = go.Figure()
    fig1.add_trace(go.Bar(x=roi_df["Year"], y=roi_df["Revenue (Lac)"], name="Revenue", marker_color="#003366"))
    fig1.add_trace(go.Bar(x=roi_df["Year"], y=roi_df["Variable Cost (Lac)"], name="Variable Cost", marker_color="#CC3333"))
    fig1.add_trace(go.Bar(x=roi_df["Year"], y=roi_df["Fixed Cost (Lac)"], name="Fixed Cost", marker_color="#FF8800"))
    fig1.add_trace(go.Scatter(x=roi_df["Year"], y=roi_df["PAT (Lac)"], name="Net Profit",
                               mode="lines+markers", line=dict(color="#00AA44", width=3)))
    fig1.update_layout(barmode="stack", template="plotly_white", height=400,
                        xaxis_title="Year", yaxis_title="Rs Lakhs")
    st.plotly_chart(fig1, width="stretch")

    # Chart 2: DSCR Trend
    col_dscr, col_cashflow = st.columns(2)
    with col_dscr:
        fig2 = go.Figure()
        fig2.add_trace(go.Bar(x=roi_df["Year"], y=roi_df["DSCR"], name="DSCR",
                               marker_color=["#CC3333" if d < 1.5 else "#006699" for d in roi_df["DSCR"]]))
        fig2.add_hline(y=1.5, line_dash="dash", line_color="red", annotation_text="Bank Minimum (1.5x)")
        fig2.update_layout(title="DSCR Trend", template="plotly_white", height=350, yaxis_title="DSCR Ratio")
        st.plotly_chart(fig2, width="stretch")

    with col_cashflow:
        fig3 = go.Figure()
        fig3.add_trace(go.Scatter(x=roi_df["Year"], y=roi_df["Cash Accrual (Lac)"],
                                   name="Cash Accrual", fill="tozeroy", line=dict(color="#003366")))
        fig3.add_trace(go.Scatter(x=roi_df["Year"], y=[cfg["emi_lac_mth"]*12]*7,
                                   name="Annual EMI", line=dict(color="#CC3333", dash="dash")))
        fig3.update_layout(title="Cash Accrual vs EMI", template="plotly_white", height=350, yaxis_title="Rs Lakhs")
        st.plotly_chart(fig3, width="stretch")

    # Full P&L Table
    st.markdown("---")
    st.subheader("Detailed P&L Statement")
    st.dataframe(roi_df, width="stretch", hide_index=True)

    # Sensitivity Matrix
    st.markdown("---")
    st.subheader("Sensitivity Analysis (EBITDA Yr5 in Rs Lakhs)")
    if cfg["sensitivity_matrix"]:
        fig_heat = go.Figure(data=go.Heatmap(
            z=cfg["sensitivity_matrix"],
            x=["Low Sell Price", "Base Price", "High Sell Price"],
            y=["Low Cost", "Base Cost", "High Cost"],
            text=[[f"Rs {v:.0f}L" for v in row] for row in cfg["sensitivity_matrix"]],
            texttemplate="%{text}", colorscale="RdYlGn",
        ))
        fig_heat.update_layout(title="EBITDA Sensitivity (Green = Profitable, Red = Risk)",
                                height=350, template="plotly_white")
        st.plotly_chart(fig_heat, width="stretch")

    # Monthly P&L
    if cfg["monthly_pnl"]:
        st.markdown("---")
        st.subheader("Monthly P&L (Year 5 Steady State)")
        pnl = cfg["monthly_pnl"]
        pnl_df = pd.DataFrame([{"Item": k, "Rs Lakhs/Month": v} for k, v in pnl.items()])
        fig_pnl = px.bar(pnl_df, x="Item", y="Rs Lakhs/Month", color="Item",
                          color_discrete_sequence=["#003366", "#CC3333", "#FF8800", "#006699", "#AA3366", "#00AA44"])
        fig_pnl.update_layout(template="plotly_white", height=350, showlegend=False)
        st.plotly_chart(fig_pnl, width="stretch")

# ═══════════════════════════════════════════════════════════════════
# EXPORT SECTION
# ═══════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("Export Financial Model")

exp1, exp2, exp3 = st.columns(3)

# Excel Export
with exp1:
    if st.button("Export to Excel", type="primary"):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "Summary"
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

        summary_rows = [
            ["Bio Bitumen Financial Model", ""],
            ["Capacity", f"{cfg['capacity_tpd']:.0f} TPD"],
            ["Working Days", f"{cfg['working_days']}"],
            ["Investment", f"Rs {cfg['investment_cr']:.2f} Crore"],
            ["Loan", f"Rs {cfg['loan_cr']:.2f} Crore"],
            ["Equity", f"Rs {cfg['equity_cr']:.2f} Crore"],
            ["Monthly EMI", f"Rs {cfg['emi_lac_mth']:.2f} Lakhs"],
            ["Revenue Yr5", f"Rs {cfg['revenue_yr5_lac']:.0f} Lakhs"],
            ["Profit per MT", f"Rs {cfg['profit_per_mt']:,.0f}"],
            ["ROI", f"{cfg['roi_pct']:.1f}%"],
            ["IRR", f"{cfg['irr_pct']:.1f}%"],
            ["DSCR Yr3", f"{cfg['dscr_yr3']:.2f}x"],
            ["Break-Even", f"{cfg['break_even_months']} months"],
        ]
        for i, (a, b) in enumerate(summary_rows):
            ws1.cell(row=i+1, column=1, value=a).font = Font(bold=True)
            ws1.cell(row=i+1, column=2, value=b)

        # Sheet 2: 7-Year P&L
        if cfg["roi_timeline"]:
            ws2 = wb.create_sheet("P&L 7-Year")
            headers = list(cfg["roi_timeline"][0].keys())
            for j, h in enumerate(headers):
                cell = ws2.cell(row=1, column=j+1, value=h)
                cell.font = header_font
                cell.fill = header_fill
            for i, row_data in enumerate(cfg["roi_timeline"]):
                for j, h in enumerate(headers):
                    ws2.cell(row=i+2, column=j+1, value=row_data[h])

        # Sheet 3: Sensitivity
        if cfg["sensitivity_matrix"]:
            ws3 = wb.create_sheet("Sensitivity")
            ws3.cell(row=1, column=1, value="Cost \\ Price")
            labels_col = ["Low Price", "Base Price", "High Price"]
            labels_row = ["Low Cost", "Base Cost", "High Cost"]
            for j, lbl in enumerate(labels_col):
                ws3.cell(row=1, column=j+2, value=lbl).font = header_font
            for i, lbl in enumerate(labels_row):
                ws3.cell(row=i+2, column=1, value=lbl).font = Font(bold=True)
                for j, val in enumerate(cfg["sensitivity_matrix"][i]):
                    ws3.cell(row=i+2, column=j+2, value=val)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        st.download_button(
            label="Download Excel",
            data=buffer.getvalue(),
            file_name=f"Financial_Model_{cfg['capacity_tpd']:.0f}TPD.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# PDF Export
with exp2:
    if st.button("Export to PDF"):
        st.page_link("pages/13_Client_Presentation.py", label="Go to DPR Generator", icon="📄")

# Print
with exp3:
    st.markdown("**Print:** Use browser Ctrl+P to print this page")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# SCENARIO MANAGER (NEW)
# ═══════════════════════════════════════════════════════════════════
st.subheader("Scenario Manager")
st.caption("Save current configuration as a named scenario for quick recall")

from database import save_configuration, get_configurations, init_db
import json
init_db()

scen_col1, scen_col2 = st.columns(2)

with scen_col1:
    st.markdown("**Save Current Scenario**")
    scenario_name = st.text_input("Scenario Name", placeholder="e.g., Conservative / Optimistic / Bank Presentation", key="scen_name")
    if st.button("Save Scenario", key="save_scen") and scenario_name:
        config_snapshot = {
            "capacity_tpd": cfg["capacity_tpd"],
            "working_days": cfg["working_days"],
            "selling_price_per_mt": cfg["selling_price_per_mt"],
            "biochar_price_per_mt": cfg["biochar_price_per_mt"],
            "raw_material_cost_per_mt": cfg["raw_material_cost_per_mt"],
            "power_cost_per_mt": cfg["power_cost_per_mt"],
            "labour_cost_per_mt": cfg["labour_cost_per_mt"],
            "transport_cost_per_mt": cfg["transport_cost_per_mt"],
            "interest_rate": cfg["interest_rate"],
            "equity_ratio": cfg["equity_ratio"],
            "investment_cr": cfg["investment_cr"],
            "roi_pct": cfg["roi_pct"],
            "irr_pct": cfg["irr_pct"],
            "monthly_profit_lac": cfg["monthly_profit_lac"],
        }
        save_configuration(None, scenario_name, json.dumps(config_snapshot))
        st.success(f"Scenario '{scenario_name}' saved!")

with scen_col2:
    st.markdown("**Saved Scenarios**")
    saved = get_configurations()
    if saved:
        for s in saved[:5]:
            snap = json.loads(s.get("config_json", "{}"))
            with st.expander(f"{s['name']} — {snap.get('capacity_tpd', '?')} TPD | ROI: {snap.get('roi_pct', '?')}%"):
                for k, v in snap.items():
                    st.text(f"{k}: {v}")
                if st.button(f"Load '{s['name']}'", key=f"load_scen_{s['id']}"):
                    load_changes = {}
                    for k in ["capacity_tpd", "working_days", "selling_price_per_mt", "biochar_price_per_mt",
                              "raw_material_cost_per_mt", "power_cost_per_mt", "labour_cost_per_mt",
                              "transport_cost_per_mt", "interest_rate", "equity_ratio"]:
                        if k in snap:
                            load_changes[k] = snap[k]
                    if load_changes:
                        update_fields(load_changes)
                        st.success(f"Loaded scenario '{s['name']}'! Refresh to see updated values.")
                        st.rerun()
    else:
        st.info("No saved scenarios yet. Save your first scenario above.")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# CUMULATIVE CASH FLOW CHART (NEW)
# ═══════════════════════════════════════════════════════════════════
if cfg.get("roi_timeline"):
    st.subheader("Cumulative Cash Flow & Payback")
    roi_df2 = pd.DataFrame(cfg["roi_timeline"])
    investment_lac = cfg["investment_cr"] * 100
    cumulative = []
    running_total = -investment_lac  # Start negative (initial investment)
    for _, row in roi_df2.iterrows():
        running_total += row.get("Cash Accrual (Lac)", row.get("PAT (Lac)", 0))
        cumulative.append(running_total)

    fig_cum = go.Figure()
    colors_cum = ["#CC3333" if v < 0 else "#00AA44" for v in cumulative]
    fig_cum.add_trace(go.Bar(x=roi_df2["Year"], y=cumulative, name="Cumulative Cash Flow",
                               marker_color=colors_cum))
    fig_cum.add_hline(y=0, line_dash="dash", line_color="black", annotation_text="Payback Point")
    fig_cum.update_layout(title="Cumulative Cash Flow (Rs Lac) — Green = Payback Achieved",
                           template="plotly_white", height=400, yaxis_title="Rs Lac")
    st.plotly_chart(fig_cum, width="stretch")

# Quick links
st.markdown("---")
st.subheader("Related Tools")
ql1, ql2, ql3 = st.columns(3)
ql1.page_link("pages/19_ROI_Quick_Calc.py", label="ROI Quick Calculator", icon="🎯")
ql2.page_link("pages/20_Loan_EMI_Calculator.py", label="Loan EMI Calculator", icon="🏦")
ql3.page_link("pages/21_Capacity_Compare.py", label="Capacity Comparison", icon="⚖️")

st.caption("All values auto-calculate when ANY input changes. This model is bank-ready and investor-grade.")
