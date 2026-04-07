"""
Bio Bitumen Consultant Portal — Customer Manager (CRM)
Add, edit, delete, and track customers through the sales pipeline.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
from state_manager import init_state, get_config
import pandas as pd
from database import (init_db, insert_customer, update_customer, delete_customer,
                       get_all_customers, search_customers, get_customer,
                       get_customer_count_by_status)
from config import CAPACITY_KEYS, CAPACITY_LABELS, STATES, CUSTOMER_STATUSES

st.set_page_config(page_title="Customer Manager", page_icon="👥", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass

init_db()
st.sidebar.markdown("---")
if st.sidebar.button("Print This Page", key="print_page"):
    import streamlit.components.v1 as _stc; _stc.html('<script>window.print();</script>', height=0)

st.title("Customer Manager")

# ── Pipeline Summary ──────────────────────────────────────────────────
status_counts = get_customer_count_by_status()
if status_counts:
    cols = st.columns(len(CUSTOMER_STATUSES))
    for i, status in enumerate(CUSTOMER_STATUSES):
        count = status_counts.get(status, 0)
        cols[i].metric(status, count)
    st.markdown("---")

# ── Tabs ──────────────────────────────────────────────────────────────
tab_list, tab_add = st.tabs(["Customer List", "Add New Customer"])

# ══════════════════════════════════════════════════════════════════════
# TAB: CUSTOMER LIST
# ══════════════════════════════════════════════════════════════════════
with tab_list:
    search_q = st.text_input("Search customers", placeholder="Search by name, company, email, phone, state...")

    if search_q:
        customers = search_customers(search_q)
    else:
        customers = get_all_customers()

    st.markdown(f"**{len(customers)} customers**")

    if customers:
        # Display as table
        display_data = []
        for c in customers:
            display_data.append({
                "ID": c["id"],
                "Name": c["name"],
                "Company": c.get("company", ""),
                "Phone": c.get("phone", ""),
                "State": c.get("state", ""),
                "Capacity": CAPACITY_LABELS.get(c.get("interested_capacity", ""), c.get("interested_capacity", "")),
                "Status": c["status"],
                "Budget (Cr)": c.get("budget_cr", 0),
            })
        st.dataframe(pd.DataFrame(display_data), use_container_width=True, hide_index=True)

        # Edit / Delete Section
        st.markdown("---")
        st.subheader("Edit / Delete Customer")

        customer_options = {c["id"]: f"{c['name']} ({c.get('company', 'N/A')})" for c in customers}
        selected_id = st.selectbox("Select customer to edit", options=list(customer_options.keys()),
                                    format_func=lambda x: customer_options[x])

        if selected_id:
            cust = get_customer(selected_id)
            if cust:
                with st.form(f"edit_customer_{selected_id}"):
                    st.markdown(f"**Editing: {cust['name']}**")
                    ec1, ec2 = st.columns(2)
                    with ec1:
                        e_name = st.text_input("Name", value=cust.get("name", ""))
                        e_company = st.text_input("Company", value=cust.get("company", ""))
                        e_email = st.text_input("Email", value=cust.get("email", ""))
                        e_phone = st.text_input("Phone", value=cust.get("phone", ""))
                        e_whatsapp = st.text_input("WhatsApp", value=cust.get("whatsapp", ""))
                    with ec2:
                        e_state = st.selectbox("State", options=[""] + STATES,
                                                index=(STATES.index(cust["state"]) + 1) if cust.get("state") in STATES else 0)
                        e_city = st.text_input("City", value=cust.get("city", ""))
                        e_capacity = st.selectbox("Interested Capacity",
                                                   options=[""] + CAPACITY_KEYS,
                                                   format_func=lambda x: CAPACITY_LABELS.get(x, "Select...") if x else "Select...",
                                                   index=(CAPACITY_KEYS.index(cust["interested_capacity"]) + 1)
                                                   if cust.get("interested_capacity") in CAPACITY_KEYS else 0)
                        e_budget = st.number_input("Budget (Crores)", value=float(cust.get("budget_cr", 0)),
                                                    min_value=0.0, step=0.5)
                        e_status = st.selectbox("Status", options=CUSTOMER_STATUSES,
                                                 index=CUSTOMER_STATUSES.index(cust["status"])
                                                 if cust.get("status") in CUSTOMER_STATUSES else 0)
                    e_notes = st.text_area("Notes", value=cust.get("notes", ""), height=80)

                    bcol1, bcol2 = st.columns(2)
                    submitted = bcol1.form_submit_button("Update Customer", type="primary")
                    if submitted:
                        update_customer(selected_id, {
                            "name": e_name, "company": e_company, "email": e_email,
                            "phone": e_phone, "whatsapp": e_whatsapp, "state": e_state,
                            "city": e_city, "interested_capacity": e_capacity,
                            "budget_cr": e_budget, "status": e_status, "notes": e_notes,
                        })
                        st.success(f"Updated {e_name}!")
                        st.rerun()

                if st.button(f"Delete {cust['name']}", type="secondary"):
                    delete_customer(selected_id)
                    st.warning(f"Deleted {cust['name']}")
                    st.rerun()
    else:
        st.info("No customers found. Add your first customer using the 'Add New Customer' tab.")


# ══════════════════════════════════════════════════════════════════════
# TAB: ADD NEW CUSTOMER
# ══════════════════════════════════════════════════════════════════════
with tab_add:
    st.subheader("Add New Customer")

    with st.form("add_customer_form", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            name = st.text_input("Name *", placeholder="Customer full name")
            company = st.text_input("Company", placeholder="Company name")
            email = st.text_input("Email", placeholder="email@example.com")
            phone = st.text_input("Phone", placeholder="+91 XXXXX XXXXX")
            whatsapp = st.text_input("WhatsApp", placeholder="+91 XXXXX XXXXX (leave blank = same as phone)")
        with c2:
            # Pincode auto-fill
            pincode = st.text_input("Pincode (auto-fills State/City)", placeholder="e.g. 390001", key="new_cust_pin")
            auto_state = ""
            auto_city = ""
            if pincode and len(pincode) == 6 and pincode.isdigit():
                try:
                    from engines.free_apis import lookup_pincode
                    pin_data = lookup_pincode(pincode)
                    if "error" not in pin_data:
                        auto_state = pin_data.get("state", "")
                        auto_city = pin_data.get("district", pin_data.get("city", ""))
                        st.caption(f"Auto-detected: {auto_city}, {auto_state}")
                except Exception:
                    pass

            state_default = STATES.index(auto_state) + 1 if auto_state in STATES else 0
            state = st.selectbox("State *", options=[""] + STATES, index=state_default,
                                 help="Select the customer's state")
            city = st.text_input("City", value=auto_city, placeholder="City name")

            # Default capacity from session cfg
            cfg_cap_key = f"{int(cfg.get('capacity_tpd', 20)):02d}MT"
            default_cap_idx = (CAPACITY_KEYS.index(cfg_cap_key) + 1) if cfg_cap_key in CAPACITY_KEYS else 0
            capacity = st.selectbox("Interested Capacity",
                                     options=[""] + CAPACITY_KEYS,
                                     format_func=lambda x: CAPACITY_LABELS.get(x, "Select...") if x else "Select...",
                                     index=default_cap_idx)
            budget = st.number_input("Budget (Crores)", min_value=0.0, step=0.5,
                                     value=float(cfg.get("investment_cr", 8)))
            status = st.selectbox("Status", options=CUSTOMER_STATUSES)
        notes = st.text_area("Notes", placeholder="Any additional details...", height=80)

        submitted = st.form_submit_button("Add Customer", type="primary")
        if submitted:
            if not name:
                st.error("Name is required!")
            else:
                # Use phone as WhatsApp if WhatsApp not provided
                wa = whatsapp if whatsapp else phone
                cid = insert_customer({
                    "name": name, "company": company, "email": email,
                    "phone": phone, "whatsapp": wa, "state": state,
                    "city": city, "interested_capacity": capacity,
                    "budget_cr": budget, "status": status, "notes": notes,
                })
                st.success(f"Added **{name}** (ID: {cid}) — State: {state or 'N/A'}, Capacity: {CAPACITY_LABELS.get(capacity, 'N/A')}")
                st.rerun()


# ── AI Skill: Client Outreach Email ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    ai_available = is_ai_available()
except Exception:
    ai_available = False

with st.expander("AI: Client Outreach Email"):
    if not ai_available:
        st.info("AI not available. Add OpenAI or Claude API keys in AI Settings page to enable this feature.")
    else:
        # Let user pick which customer to generate email for
        if customers:
            ai_cust_map = {c["id"]: f"{c['name']} ({c.get('company', 'N/A')})" for c in customers}
            ai_cust_id = st.selectbox("Generate email for", options=list(ai_cust_map.keys()),
                                       format_func=lambda x: ai_cust_map[x], key="ai_cust_select")
            ai_customer = get_customer(ai_cust_id)
        else:
            ai_customer = {"name": "Prospective Client", "company": ""}

        if st.button("Generate Outreach Email", type="primary", key="ai_14_gen"):
            with st.spinner("AI generating professional email..."):
                prompt = (
                    f"Write a professional client outreach email from {COMPANY['owner']} at {COMPANY['trade_name']} "
                    f"to {ai_customer.get('name', 'Sir/Madam')} at {ai_customer.get('company', 'their company')}.\n\n"
                    f"Project: Bio-Modified Bitumen Plant\n"
                    f"- Capacity: {cfg.get('capacity_tpd', 20):.0f} TPD\n"
                    f"- Investment: Rs {cfg.get('investment_cr', 8):.2f} Crore\n"
                    f"- ROI: {cfg.get('roi_pct', 20):.1f}%\n"
                    f"- IRR: {cfg.get('irr_pct', 26):.1f}%\n"
                    f"- Break-even: {cfg.get('break_even_months', 30)} months\n"
                    f"- Location: {cfg.get('location', '')}, {cfg.get('state', '')}\n\n"
                    f"Include specific financial numbers. Clear call to action. Under 200 words. Professional format."
                )
                result, provider = ask_ai(prompt, "You are a senior industrial consultant writing to a potential investor.", 1000)

            if result:
                st.markdown("**Generated Email:**")
                st.markdown(result)
                st.text_area("Copy/Edit", value=result, height=200, key="ai_email_output")
            else:
                st.error("AI could not generate email. Check API keys in AI Settings.")

        # Show last generated email if in session
        if "ai_email_output" in st.session_state and st.session_state.get("ai_email_output"):
            pass  # Already shown via text_area above


# ── Export: Download ACTUAL customer list as Excel ────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    all_customers = get_all_customers()
    if all_customers and st.button("Download Customer List (Excel)", type="primary", key="exp_xl_cust"):
        try:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Customers"

            # Header
            headers = ["ID", "Name", "Company", "Email", "Phone", "WhatsApp",
                       "State", "City", "Capacity", "Budget (Cr)", "Status", "Notes", "Created"]
            header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font

            # Data rows
            for row_idx, c in enumerate(all_customers, 2):
                ws.cell(row=row_idx, column=1, value=c["id"])
                ws.cell(row=row_idx, column=2, value=c.get("name", ""))
                ws.cell(row=row_idx, column=3, value=c.get("company", ""))
                ws.cell(row=row_idx, column=4, value=c.get("email", ""))
                ws.cell(row=row_idx, column=5, value=c.get("phone", ""))
                ws.cell(row=row_idx, column=6, value=c.get("whatsapp", ""))
                ws.cell(row=row_idx, column=7, value=c.get("state", ""))
                ws.cell(row=row_idx, column=8, value=c.get("city", ""))
                ws.cell(row=row_idx, column=9, value=CAPACITY_LABELS.get(c.get("interested_capacity", ""), c.get("interested_capacity", "")))
                ws.cell(row=row_idx, column=10, value=c.get("budget_cr", 0))
                ws.cell(row=row_idx, column=11, value=c.get("status", ""))
                ws.cell(row=row_idx, column=12, value=c.get("notes", ""))
                ws.cell(row=row_idx, column=13, value=c.get("created_at", ""))

            # Auto-width columns
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + min(col, 26))].width = 15

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.download_button("Download Excel", buf.getvalue(),
                f"Customer_List_{len(all_customers)}_records.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_xl_cust_file")
        except Exception as e:
            st.error(f"Export failed: {e}")
with _ex2:
    if st.button("Print", key="exp_prt_14C"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)


# ── Next Steps Navigation ──
try:
    from engines.page_navigation import add_next_steps
    add_next_steps(st, "11")
except Exception:
    pass
