"""
Loan EMI Calculator — Standalone Tool for Bank/Investor Discussions
====================================================================
EMI calculation, amortization schedule, multi-scenario comparison, CGTMSE check.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from state_manager import get_config, init_state
from config import COMPANY, EMI_PRESETS

st.set_page_config(page_title="Loan EMI Calculator", page_icon="🏦", layout="wide")
init_state()
cfg = get_config()

st.page_link("pages/09_💰_Financial.py", label="← Back to Financial Model", icon="💰")

st.title("Loan EMI Calculator")
st.markdown("**Calculate EMI, Total Interest, Amortization Schedule | Compare Loan Options | CGTMSE Eligibility**")
st.markdown("---")

# ── Helper Functions ─────────────────────────────────────────────────
def calc_emi(principal, annual_rate, tenure_months):
    if annual_rate <= 0 or tenure_months <= 0:
        return 0, 0, []
    r = annual_rate / 12 / 100
    emi = principal * r * (1 + r) ** tenure_months / ((1 + r) ** tenure_months - 1)
    total_payment = emi * tenure_months
    total_interest = total_payment - principal

    schedule = []
    balance = principal
    for month in range(1, tenure_months + 1):
        interest_part = balance * r
        principal_part = emi - interest_part
        balance -= principal_part
        if balance < 0:
            balance = 0
        schedule.append({
            "Month": month,
            "EMI": round(emi, 0),
            "Principal": round(principal_part, 0),
            "Interest": round(interest_part, 0),
            "Balance": round(balance, 0),
        })
    return emi, total_interest, schedule

# ══════════════════════════════════════════════════════════════════════
# LOAN PRESET SELECTOR
# ══════════════════════════════════════════════════════════════════════
st.subheader("1. Select Loan Type or Enter Custom")

preset_col, custom_col = st.columns([1, 1])

with preset_col:
    preset_names = ["Custom"] + [p["name"] for p in EMI_PRESETS]
    selected_preset = st.selectbox("Loan Preset", preset_names, key="loan_preset")

    if selected_preset != "Custom":
        preset = next(p for p in EMI_PRESETS if p["name"] == selected_preset)
        default_rate = preset["interest_pct"]
        default_tenure = preset["tenure_months"]
        default_max = preset["max_loan_cr"]
        st.info(f"**{preset['name']}** | Collateral: {preset['collateral']} | Processing Fee: {preset['processing_fee_pct']}%")
    else:
        default_rate = 11.5
        default_tenure = 84
        default_max = 25.0

with custom_col:
    loan_amount_cr = st.number_input("Loan Amount (Rs Crore)", 0.10, 100.0,
                                      min(float(cfg["investment_cr"]) * 0.7, default_max), 0.10, key="loan_amt")
    interest_rate = st.number_input("Interest Rate (%)", 5.0, 20.0, default_rate, 0.5, key="loan_rate")
    tenure_months = st.number_input("Tenure (Months)", 12, 240, default_tenure, 12, key="loan_tenure")

# ── Calculate ────────────────────────────────────────────────────────
principal_lac = loan_amount_cr * 100  # Convert Cr to Lac
emi, total_interest_lac, schedule = calc_emi(principal_lac, interest_rate, int(tenure_months))

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# RESULTS DASHBOARD
# ══════════════════════════════════════════════════════════════════════
st.subheader("2. Loan Summary")

r1, r2, r3, r4, r5 = st.columns(5)
r1.metric("Monthly EMI", f"Rs {emi:.2f} Lac")
r2.metric("Total Interest", f"Rs {total_interest_lac:.1f} Lac")
r3.metric("Total Payment", f"Rs {principal_lac + total_interest_lac:.1f} Lac")
r4.metric("Interest/Principal", f"{(total_interest_lac/principal_lac*100) if principal_lac > 0 else 0:.0f}%")
r5.metric("Effective Cost", f"Rs {(principal_lac + total_interest_lac)/100:.2f} Cr")

# ── Principal vs Interest Pie Chart ──────────────────────────────────
pie_col, bar_col = st.columns(2)

with pie_col:
    fig_pie = go.Figure(data=[go.Pie(
        labels=["Principal", "Interest"],
        values=[principal_lac, total_interest_lac],
        hole=0.5,
        marker=dict(colors=["#003366", "#CC3333"]),
        textinfo="label+percent",
    )])
    fig_pie.update_layout(title="Principal vs Interest Split", height=350, template="plotly_white")
    st.plotly_chart(fig_pie, width="stretch")

with bar_col:
    if schedule:
        # Yearly summary
        sched_df = pd.DataFrame(schedule)
        sched_df["Year"] = ((sched_df["Month"] - 1) // 12) + 1
        yearly = sched_df.groupby("Year").agg({"Principal": "sum", "Interest": "sum"}).reset_index()

        fig_bar = go.Figure()
        fig_bar.add_trace(go.Bar(x=yearly["Year"], y=yearly["Principal"], name="Principal", marker_color="#003366"))
        fig_bar.add_trace(go.Bar(x=yearly["Year"], y=yearly["Interest"], name="Interest", marker_color="#CC3333"))
        fig_bar.update_layout(title="Yearly Principal vs Interest (Rs Lac)", barmode="stack",
                               template="plotly_white", height=350, xaxis_title="Year", yaxis_title="Rs Lac")
        st.plotly_chart(fig_bar, width="stretch")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# AMORTIZATION SCHEDULE
# ══════════════════════════════════════════════════════════════════════
st.subheader("3. Amortization Schedule")

if schedule:
    sched_df = pd.DataFrame(schedule)

    # Balance decline chart
    fig_balance = go.Figure()
    fig_balance.add_trace(go.Scatter(x=sched_df["Month"], y=sched_df["Balance"],
                                      mode="lines", name="Outstanding Balance",
                                      line=dict(color="#003366", width=2), fill="tozeroy",
                                      fillcolor="rgba(0,51,102,0.1)"))
    fig_balance.update_layout(title="Loan Balance Over Time (Rs Lac)", template="plotly_white",
                               height=350, xaxis_title="Month", yaxis_title="Rs Lac")
    st.plotly_chart(fig_balance, width="stretch")

    with st.expander("View Full Monthly Schedule"):
        st.dataframe(sched_df, width="stretch", height=400)

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# MULTI-SCENARIO COMPARISON
# ══════════════════════════════════════════════════════════════════════
st.subheader("4. Compare Loan Scenarios")

comp_data = []
for preset in EMI_PRESETS:
    p_emi, p_interest, _ = calc_emi(principal_lac, preset["interest_pct"], preset["tenure_months"])
    comp_data.append({
        "Loan Type": preset["name"],
        "Interest %": preset["interest_pct"],
        "Tenure (Months)": preset["tenure_months"],
        "EMI (Lac)": round(p_emi, 2),
        "Total Interest (Lac)": round(p_interest, 1),
        "Total Cost (Lac)": round(principal_lac + p_interest, 1),
        "Collateral": preset["collateral"],
    })

comp_df = pd.DataFrame(comp_data)
st.dataframe(comp_df, width="stretch")

# ── Comparison Chart ─────────────────────────────────────────────────
fig_comp = go.Figure()
fig_comp.add_trace(go.Bar(x=comp_df["Loan Type"], y=comp_df["EMI (Lac)"], name="Monthly EMI", marker_color="#003366"))
fig_comp.add_trace(go.Bar(x=comp_df["Loan Type"], y=comp_df["Total Interest (Lac)"], name="Total Interest", marker_color="#CC3333"))
fig_comp.update_layout(title=f"Loan Comparison for Rs {loan_amount_cr:.1f} Cr",
                         barmode="group", template="plotly_white", height=400)
st.plotly_chart(fig_comp, width="stretch")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# CGTMSE ELIGIBILITY CHECK
# ══════════════════════════════════════════════════════════════════════
st.subheader("5. CGTMSE Eligibility Check")

cg1, cg2 = st.columns(2)
with cg1:
    st.markdown("""
    **CGTMSE (Credit Guarantee Fund Trust for MSEs)**
    - Collateral-free loans up to **Rs 5 Crore**
    - Available through all scheduled commercial banks
    - Annual guarantee fee: **1.5% of loan amount**
    - Processing time: **15-30 days**
    """)

    if loan_amount_cr <= 5.0:
        st.success(f"**ELIGIBLE** — Rs {loan_amount_cr:.1f} Cr is within Rs 5 Cr CGTMSE limit. No collateral needed!")
        annual_fee = loan_amount_cr * 1.5 / 100
        st.info(f"Annual CGTMSE guarantee fee: Rs {annual_fee:.2f} Lac/year")
    else:
        st.warning(f"**NOT ELIGIBLE** — Rs {loan_amount_cr:.1f} Cr exceeds Rs 5 Cr CGTMSE limit. Collateral required.")
        eligible_portion = 5.0
        remaining = loan_amount_cr - eligible_portion
        st.info(f"Split option: Rs {eligible_portion} Cr under CGTMSE (collateral-free) + Rs {remaining:.1f} Cr with collateral")

with cg2:
    st.markdown("""
    **Subsidy Impact Calculator**
    """)
    subsidy_pct = st.slider("Expected State Subsidy %", 0, 30, 20, 5, key="subsidy_pct")
    investment_cr = cfg["investment_cr"]
    subsidy_amount = investment_cr * subsidy_pct / 100
    effective_investment = investment_cr - subsidy_amount
    effective_loan = effective_investment * 0.7

    st.metric("Total Investment", f"Rs {investment_cr:.2f} Cr")
    st.metric("Subsidy Amount", f"Rs {subsidy_amount:.2f} Cr", delta=f"-{subsidy_pct}%")
    st.metric("Effective Investment", f"Rs {effective_investment:.2f} Cr")
    st.metric("Loan Needed (70%)", f"Rs {effective_loan:.2f} Cr")

    eff_emi, eff_int, _ = calc_emi(effective_loan * 100, interest_rate, int(tenure_months))
    st.metric("EMI with Subsidy", f"Rs {eff_emi:.2f} Lac", delta=f"-Rs {emi - eff_emi:.2f} Lac saved")

st.markdown("---")
st.caption(f"{COMPANY['name']} | {COMPANY['owner']} | {COMPANY['phone']}")

# ── Export Section ────────────────────────────────────────────────────
st.markdown("---")
st.subheader("Export")
exp1, exp2, exp3 = st.columns(3)
with exp1:
    if st.button("Download Excel", type="primary", key="exp_xl_calc"):
        try:
            import io, pandas as pd
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Calculator Output"
            ws.cell(row=1, column=1, value="Bio Bitumen Calculator Export")
            ws.cell(row=2, column=1, value=f"Capacity: {cfg['capacity_tpd']:.0f} TPD")
            ws.cell(row=3, column=1, value=f"Investment: Rs {cfg['investment_cr']:.2f} Cr")
            ws.cell(row=4, column=1, value=f"ROI: {cfg['roi_pct']:.1f}%")
            ws.cell(row=5, column=1, value=f"IRR: {cfg['irr_pct']:.1f}%")
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.download_button("Download", buf.getvalue(), "calculator_export.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_xl_c")
        except Exception as e:
            st.error(f"Export failed: {e}")
with exp2:
    if st.button("Download CSV", key="exp_csv_calc"):
        import pandas as pd
        data = {"Metric": ["Capacity", "Investment", "ROI", "IRR", "DSCR", "Break-Even", "Monthly Profit"],
                "Value": [f"{cfg['capacity_tpd']:.0f} TPD", f"Rs {cfg['investment_cr']:.2f} Cr",
                          f"{cfg['roi_pct']:.1f}%", f"{cfg['irr_pct']:.1f}%", f"{cfg['dscr_yr3']:.2f}x",
                          f"{cfg['break_even_months']} months", f"Rs {cfg['monthly_profit_lac']:.1f} Lac"]}
        st.download_button("Download", pd.DataFrame(data).to_csv(index=False), "calculator_export.csv", "text/csv", key="dl_csv_c")
with exp3:
    if st.button("Print", key="exp_print_calc"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)
