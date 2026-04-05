"""
System Health — 10-Component Monitoring + Auto-Updater + Self-Healing (UPGRADED)
==================================================================================
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
from state_manager import init_state, get_config
import pandas as pd
import plotly.graph_objects as go
from datetime import datetime, timezone, timedelta
from engines.self_healing_worker import (run_health_cycle, start_worker, stop_worker,
                                          get_repair_history, is_worker_running, get_health_status)
from engines.auto_updater import (start_updater, stop_updater, is_updater_running,
                                   get_update_log, run_full_update_cycle,
                                   cleanup_expired_cache, check_config_consistency,
                                   check_drawing_files, check_database_health)
from engines.deep_extractor import get_scan_stats
from engines.auto_doc_sync import get_synced_files, get_sync_log
from database import init_db
from config import COMPANY

IST = timezone(timedelta(hours=5, minutes=30))

st.set_page_config(page_title="System Health", page_icon="🏥", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass

init_db()
st.sidebar.markdown("---")
if st.sidebar.button("Print This Page", key="print_page"):
    import streamlit.components.v1 as _stc; _stc.html('<script>window.print();</script>', height=0)

st.title("System Health & Auto-Monitoring")
st.markdown("**10-Component Health Check | Auto-Updater | Self-Healing | Repair Log**")
st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# WORKER CONTROLS
# ══════════════════════════════════════════════════════════════════════
ctrl1, ctrl2, ctrl3, ctrl4 = st.columns(4)

with ctrl1:
    worker_on = is_worker_running()
    st.metric("Health Worker", "Running" if worker_on else "Stopped")
    if st.button("Start" if not worker_on else "Stop", key="worker_toggle"):
        if worker_on:
            stop_worker()
        else:
            start_worker(interval=300)
        st.rerun()

with ctrl2:
    updater_on = is_updater_running()
    st.metric("Auto-Updater", "Running" if updater_on else "Stopped")
    if st.button("Start" if not updater_on else "Stop", key="updater_toggle"):
        if updater_on:
            stop_updater()
        else:
            start_updater(interval=300)
        st.rerun()

with ctrl3:
    if st.button("Run Full Health Check", type="primary", key="run_health"):
        with st.spinner("Running 10-component health check..."):
            results, score = run_full_update_cycle()
        st.success(f"Health Score: {score}%")
        st.rerun()

with ctrl4:
    if st.button("Fix All Issues", key="fix_all"):
        with st.spinner("Auto-repairing..."):
            from engines.self_healing_worker import auto_repair
            results, _ = run_health_cycle()
            auto_repair(results)
            cleanup_expired_cache()
            check_database_health()
        st.success("Auto-repair complete!")
        st.rerun()

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# 10-COMPONENT HEALTH STATUS
# ══════════════════════════════════════════════════════════════════════
st.subheader("Health Status — 10 Components")

# Run health checks
health_result = get_health_status()
healer_results, healer_score = run_health_cycle()

# Additional checks
config_ok = check_config_consistency()
drawings_ok = check_drawing_files()
db_ok = check_database_health()

# Build 10-component status
components = {
    "Database": healer_results.get("database", "error"),
    "Master Data": healer_results.get("master_data", "error"),
    "Document Index": healer_results.get("doc_index", "error"),
    "Disk Space": healer_results.get("disk_space", "error"),
    "API Connectivity": healer_results.get("api", "error"),
    "Config Consistency": "ok" if config_ok else "warning",
    "Drawing Files": "ok" if drawings_ok else "warning",
    "Database Health": "ok" if db_ok else "error",
    "Health Worker": "ok" if worker_on else "stopped",
    "Auto-Updater": "ok" if updater_on else "stopped",
}

healthy = sum(1 for v in components.values() if v == "ok")
stopped = sum(1 for v in components.values() if v == "stopped")
warnings = sum(1 for v in components.values() if v == "warning")
errors = sum(1 for v in components.values() if v == "error")
overall_score = int(healthy / len(components) * 100)

# Score gauge
score_col, detail_col = st.columns([1, 2])

with score_col:
    color = "#00AA44" if overall_score >= 80 else ("#FF8800" if overall_score >= 60 else "#CC3333")
    fig_score = go.Figure(go.Indicator(
        mode="gauge+number",
        value=overall_score,
        title={"text": "Overall Health"},
        number={"suffix": "%"},
        gauge={"axis": {"range": [0, 100]},
               "bar": {"color": color},
               "steps": [{"range": [0, 60], "color": "#ffcccc"},
                          {"range": [60, 80], "color": "#ffffcc"},
                          {"range": [80, 100], "color": "#ccffcc"}]},
    ))
    fig_score.update_layout(height=250, margin=dict(t=50, b=10))
    st.plotly_chart(fig_score, width="stretch")

with detail_col:
    for comp, status in components.items():
        if status == "ok":
            icon, label = "✅", "OK"
        elif status == "stopped":
            icon, label = "⏸️", "Stopped — click Start"
        elif status == "warning":
            icon, label = "⚠️", "WARNING"
        else:
            icon, label = "❌", "ERROR"
        color = {"ok": "#00AA44", "stopped": "#888888", "warning": "#FF8800"}.get(status, "#CC3333")
        st.markdown(f"""
        <div style="display: flex; align-items: center; padding: 4px 0;">
            <span style="font-size: 1.2em; margin-right: 8px;">{icon}</span>
            <span style="flex: 1;">{comp}</span>
            <span style="color: {color}; font-weight: bold;">{label}</span>
        </div>
        """, unsafe_allow_html=True)

sm1, sm2, sm3 = st.columns(3)
sm1.metric("Healthy", healthy, delta=None)
sm2.metric("Warnings", warnings)
sm3.metric("Errors", errors)

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# DEEP SCAN + AUTO-SYNC
# ══════════════════════════════════════════════════════════════════════
scan_col, sync_col = st.columns(2)

with scan_col:
    st.subheader("Document Coverage")
    stats = get_scan_stats()
    if stats:
        ds1, ds2 = st.columns(2)
        ds1.metric("Files Scanned", f"{stats['total_files']:,}")
        ds2.metric("With Content", f"{stats['with_content']:,}")

        if stats.get("by_type"):
            import plotly.express as px
            type_df = pd.DataFrame([
                {"Type": k.replace(".", "").upper(), "Count": v}
                for k, v in stats["by_type"].items()
            ]).sort_values("Count", ascending=False).head(8)
            fig = px.bar(type_df, x="Type", y="Count", color="Count", color_continuous_scale="Blues")
            fig.update_layout(template="plotly_white", height=250, showlegend=False)
            st.plotly_chart(fig, width="stretch")
    else:
        st.info("Run deep scan to populate coverage data")

with sync_col:
    st.subheader("Auto-Synced Documents")
    synced = get_synced_files()
    if synced:
        st.success(f"**{len(synced)} documents** up-to-date")
        for f in synced[:6]:
            st.markdown(f"- **{f['name']}** — {f['size']/1024:.1f} KB")
    else:
        st.info("Change any input to trigger auto-sync")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# AUTO-UPDATER LOG
# ══════════════════════════════════════════════════════════════════════
st.subheader("Auto-Update Activity Log")

tab_updater, tab_repair = st.tabs(["Auto-Updater Log", "Self-Healing Repairs"])

with tab_updater:
    update_log = get_update_log(30)
    if update_log:
        log_df = pd.DataFrame(update_log)
        # Color code status
        st.dataframe(log_df, width="stretch", hide_index=True, height=300)
    else:
        st.info("Start the Auto-Updater to see activity log")

with tab_repair:
    history = get_repair_history()
    if history:
        repair_df = pd.DataFrame(history[:30])
        st.dataframe(repair_df, width="stretch", hide_index=True, height=300)
    else:
        st.info("No repair actions yet")

# ══════════════════════════════════════════════════════════════════════
# CACHE STATUS
# ══════════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("Cache & Data Files")

from pathlib import Path
cache_dir = Path(__file__).parent.parent / "data"
cache_files = []
if cache_dir.exists():
    for f in sorted(cache_dir.glob("*.json")):
        size_kb = f.stat().st_size / 1024
        age_hrs = (datetime.now().timestamp() - f.stat().st_mtime) / 3600
        cache_files.append({
            "File": f.name,
            "Size (KB)": round(size_kb, 1),
            "Age (Hours)": round(age_hrs, 1),
            "Status": "Fresh" if age_hrs < 2 else ("Stale" if age_hrs < 24 else "Expired"),
        })

if cache_files:
    cache_df = pd.DataFrame(cache_files)
    st.dataframe(cache_df, width="stretch", hide_index=True)
    total_cache = sum(f["Size (KB)"] for f in cache_files)
    st.metric("Total Cache Size", f"{total_cache:.0f} KB ({total_cache/1024:.1f} MB)")

    if st.button("Clean Expired Cache", key="clean_cache"):
        cleaned = cleanup_expired_cache(24)
        st.success(f"Cleaned {cleaned} expired files")
        st.rerun()

st.markdown("---")
st.caption(f"Last checked: {datetime.now(IST).strftime('%d %B %Y %H:%M IST')} | {COMPANY['name']}")


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_16_🏥_S", type="primary")
with _ex2:
    if st.button("Print", key="exp_prt_16🏥S"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)


# ── AI Assist ────────────────────────────────────────────────────
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("AI Assist"):
            if st.button("Generate AI Summary", type="primary", key="ai_16🏥S"):
                with st.spinner("AI working..."):
                    _p = f"Summarize this section for a {cfg.get('capacity_tpd',20):.0f} TPD bio-bitumen plant in {cfg.get('state','')}. Investment Rs {cfg.get('investment_cr',8):.2f} Cr. Professional consultant format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 800)
                if _r:
                    st.markdown(_r)
except Exception:
    pass
