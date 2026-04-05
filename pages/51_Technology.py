"""
Technology Validation — CSIR-CRRI/IIP License, KrishiBind, Process Summary
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.express as px
from state_manager import init_state, get_config

st.set_page_config(page_title="Technology", page_icon="🔬", layout="wide")
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass

st.title("Technology Validation & Process")
st.markdown("**CSIR-CRRI Licensed Technology | KrishiBind Bio-Binder | Proven at Lab & Field Level**")
st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# 1. TECHNOLOGY LICENSE
# ═══════════════════════════════════════════════════════════════════
st.subheader("1. Technology License & Certification")

lc1, lc2 = st.columns([1, 1])
with lc1:
    st.markdown("""
    | Parameter | Details |
    |-----------|---------|
    | **Technology** | A Process of Producing Bio-Bitumen from Pyrolysis of Lignocellulosic Biomass |
    | **Developed By** | CSIR-CRRI (Central Road Research Institute) / CSIR-IIP |
    | **Licensed To** | Rex Fuels Management Private Limited |
    | **Trade Name** | **KrishiBind** Bio-Binder |
    | **License Type** | Non-Exclusive Technology License |
    | **Validity** | Until 06 January 2031 |
    | **Patent Status** | CSIR Patent No. filed |
    | **Field Validation** | Lab + Field trials completed |
    | **Standards Compliance** | MoRTH 2026, IS:73, IRC SP:53 |
    """)

with lc2:
    st.markdown("""
    ### Key Credentials
    - CSIR-CRRI: India's premier road research institute
    - Dr. Ambika Behl: Lead researcher, bio-bitumen technology
    - Dr. Manoranjan Parida, IIT Roorkee: Technology validation
    - Field trials on NHAI road sections completed
    - Technology adopted by multiple industry participants across India

    ### Proof Documents Available
    - License Certificate (viewable in Files Library)
    - CSIR-CRRI Research Paper
    - IIT Roorkee Validation Report
    - Field Trial Documentation
    - Industry Participation Map
    """)

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# 2. PROCESS LOGIC (from uploaded product page)
# ═══════════════════════════════════════════════════════════════════
st.subheader("2. Technology Process Logic")

st.markdown("""
### Feedstock → Bio-Bitumen Process

```
STEP 1: FEEDSTOCK SELECTION
├── Rice straw / Wheat straw / Sugarcane bagasse / Cotton stalk
├── Lignocellulosic biomass (any agricultural waste)
└── Collection radius: 50 km from plant

STEP 2: FEEDSTOCK PREPARATION
├── Drying to <10% moisture
├── Shredding to small pieces
├── Pelletizing: Diameter ~10mm, Length ~20-30mm
└── Storage in covered shed

STEP 3: PYROLYSIS
├── Auger-type pyrolysis reactor
├── Temperature: 400-550°C
├── Oxygen-free (anaerobic) environment
├── Residence time: 30-60 minutes
└── Output: Vapors + Biochar + Syngas

STEP 4: CONDENSATION
├── 4 condensers in series
├── Vapors cooled progressively
├── Output splits into:
│   ├── ORGANIC FRACTION (heavy, viscous) → Used for bio-bitumen
│   └── AQUEOUS FRACTION (light) → Treated/recycled
└── Organic fraction = Bio-Oil

STEP 5: BIO-BITUMEN PREPARATION
├── Organic fraction (Bio-Oil) as base
├── Blending with hard pitch / conventional binder
├── Lime-based additives for moisture & ageing resistance
├── High-shear mixing at controlled temperature
├── Quality testing per VG grade specifications
└── Output: KrishiBind Bio-Modified Bitumen

STEP 6: PRODUCT GRADES
├── VG10 Bio-Bitumen
├── VG20 Bio-Bitumen
├── VG30 Bio-Bitumen (primary grade)
└── VG40 Bio-Bitumen
```
""")

# Process parameters table
st.markdown("### Key Process Parameters")
params = pd.DataFrame([
    {"Parameter": "Feedstock", "Value": "Rice straw, wheat straw, bagasse, cotton stalk", "Standard": "Lignocellulosic biomass"},
    {"Parameter": "Pellet Diameter", "Value": "~10 mm", "Standard": "CSIR-CRRI specification"},
    {"Parameter": "Pellet Length", "Value": "20-30 mm", "Standard": "CSIR-CRRI specification"},
    {"Parameter": "Moisture Content", "Value": "<10%", "Standard": "Pre-drying required"},
    {"Parameter": "Pyrolysis Temperature", "Value": "400-550°C", "Standard": "Controlled atmosphere"},
    {"Parameter": "Reactor Type", "Value": "Auger-type continuous", "Standard": "Industrial grade"},
    {"Parameter": "Condensers", "Value": "4 in series", "Standard": "Progressive cooling"},
    {"Parameter": "Bio-Oil Yield", "Value": "35-40% of biomass", "Standard": "Organic fraction"},
    {"Parameter": "Biochar Yield", "Value": "25-30%", "Standard": "Soil amendment grade"},
    {"Parameter": "Syngas Yield", "Value": "20-25%", "Standard": "Captive fuel use"},
    {"Parameter": "Bio-Bitumen Replacement", "Value": "15-30% of conventional bitumen", "Standard": "Lab + field validated"},
    {"Parameter": "Target Grade", "Value": "VG10 / VG20 / VG30 / VG40", "Standard": "IS:73 / MoRTH 2026"},
])
st.dataframe(params, width="stretch", hide_index=True)

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# 3. YIELD DIAGRAM
# ═══════════════════════════════════════════════════════════════════
st.subheader("3. Pyrolysis Yield Distribution")

yield_data = pd.DataFrame([
    {"Product": "Bio-Oil (Organic)", "Yield %": 40, "Use": "Bio-Bitumen blending"},
    {"Product": "Biochar", "Yield %": 30, "Use": "Soil amendment / Carbon black"},
    {"Product": "Syngas", "Yield %": 25, "Use": "Captive fuel (process heating)"},
    {"Product": "Loss", "Yield %": 5, "Use": "Moisture, volatiles"},
])

fig = px.pie(yield_data, names="Product", values="Yield %",
             title="Biomass Pyrolysis Yield Distribution",
             color_discrete_sequence=["#003366", "#006699", "#FF8800", "#CCCCCC"])
fig.update_traces(textposition="inside", textinfo="percent+label")
fig.update_layout(height=400, template="plotly_white")
st.plotly_chart(fig, width="stretch")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# 4. INDUSTRY PARTICIPATION
# ═══════════════════════════════════════════════════════════════════
st.subheader("4. Industry Participation & Adoption")
st.markdown("""
**Technology adopted by multiple companies across India:**
- Licensed through CSIR-CRRI technology transfer program
- Multiple non-exclusive licenses granted
- Field trials conducted on NHAI road sections
- Technology validated at IIT Roorkee
- Industry participation map covers 10+ states

**Key Adopters:**
- Rex Fuels Management Pvt Ltd (KrishiBind)
- PPS Anantams Corporation Pvt Ltd (Consulting)
- Multiple state-level road contractors exploring adoption
""")

# ═══════════════════════════════════════════════════════════════════
# 5. COMPETITIVE ADVANTAGE
# ═══════════════════════════════════════════════════════════════════
st.subheader("5. Bio-Bitumen vs Conventional Bitumen")

comp = pd.DataFrame([
    {"Parameter": "Raw Material", "Conventional Bitumen": "Crude oil (imported)", "Bio-Bitumen (KrishiBind)": "Agro waste (domestic)"},
    {"Parameter": "Cost (approx)", "Conventional Bitumen": "Rs 48,000-52,000/MT", "Bio-Bitumen (KrishiBind)": "Rs 35,000-45,000/MT"},
    {"Parameter": "Import Dependency", "Conventional Bitumen": "30-40% imported", "Bio-Bitumen (KrishiBind)": "0% imported"},
    {"Parameter": "Carbon Footprint", "Conventional Bitumen": "High (petroleum)", "Bio-Bitumen (KrishiBind)": "70% lower"},
    {"Parameter": "Farmer Benefit", "Conventional Bitumen": "None", "Bio-Bitumen (KrishiBind)": "Crop residue monetization"},
    {"Parameter": "Stubble Burning", "Conventional Bitumen": "No impact", "Bio-Bitumen (KrishiBind)": "Reduces stubble burning"},
    {"Parameter": "Performance", "Conventional Bitumen": "Standard", "Bio-Bitumen (KrishiBind)": "Equal or better (CSIR validated)"},
    {"Parameter": "Govt Policy", "Conventional Bitumen": "Standard", "Bio-Bitumen (KrishiBind)": "Green mandate, NHAI preference"},
])
st.dataframe(comp, width="stretch", hide_index=True)

st.caption("Source: CSIR-CRRI research, Dr. Ambika Behl technical session, IIT Roorkee validation report")


# ── AI Skill: Technology Explainer ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: Technology Explainer"):
            if st.button("Generate", type="primary", key="ai_51Techn"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: Technology Explainer. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_51Tec", type="primary")
with _ex2:
    if st.button("Print", key="exp_prt_51Tec"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)
