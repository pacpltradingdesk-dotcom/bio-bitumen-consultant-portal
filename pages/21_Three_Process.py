"""
Three Process Model Comparison — Side-by-side financial analysis
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from state_manager import get_config, init_state
from engines.three_process_model import compare_all_processes
from config import STATE_COSTS

st.set_page_config(page_title="3 Process Models", page_icon="🔄", layout="wide")
init_state()
cfg = get_config()
# Fix metric truncation
try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass


st.title("Three Process Model Comparison")
st.markdown("**Compare: Full Chain vs Blending Only vs Raw Output — Same capacity, different economics**")
st.markdown("---")

# ── Inputs ────────────────────────────────────────────────────────────
col1, col2, col3 = st.columns(3)
with col1:
    tpd = st.number_input("Capacity (TPD)", 5.0, 100.0, float(cfg["capacity_tpd"]), 5.0)
with col2:
    state = st.selectbox("State (for cost adjustment)", ["None (Default)"] + sorted(STATE_COSTS.keys()),
                          index=0)
with col3:
    st.markdown(f"**Working Days:** {cfg['working_days']} | **Interest:** {cfg['interest_rate']*100:.1f}%")

state_costs = STATE_COSTS.get(state, None) if state != "None (Default)" else None

# ── Calculate all 3 ───────────────────────────────────────────────────
results = compare_all_processes(tpd, state_costs)

# ── Key Metrics Comparison ────────────────────────────────────────────
st.subheader("Side-by-Side Comparison")

comp_data = []
for pid in [1, 2, 3]:
    r = results[pid]
    comp_data.append({
        "Process": r["process_short"],
        "Investment (Lac)": r["capex_lac"],
        "Revenue/MT (Rs)": f"{r['revenue_per_mt']:,.0f}",
        "Cost/MT (Rs)": f"{r['var_cost_per_mt']:,.0f}",
        "Profit/MT (Rs)": f"{r['profit_per_mt']:,.0f}",
        "ROI (%)": r["roi_pct"],
        "IRR (%)": r["irr_pct"],
        "DSCR Yr3": r["dscr_yr3"],
        "Break-Even (mo)": r["break_even_months"],
        "EMI (Lac/mo)": r["emi_lac_mth"],
        "Rev Yr5 (Lac)": r["rev_yr5_lac"],
        "PAT Yr5 (Lac)": r["pat_yr5_lac"],
        "Staff": r["staff"],
    })

st.dataframe(pd.DataFrame(comp_data), width="stretch", hide_index=True)

st.markdown("---")

# ── Visual Comparison Charts ──────────────────────────────────────────
col_c1, col_c2 = st.columns(2)

with col_c1:
    names = [r["process_short"] for r in results.values()]
    fig1 = go.Figure(data=[
        go.Bar(name="Investment (Lac)", x=names, y=[r["capex_lac"] for r in results.values()], marker_color="#003366"),
        go.Bar(name="Rev Yr5 (Lac)", x=names, y=[r["rev_yr5_lac"] for r in results.values()], marker_color="#006699"),
        go.Bar(name="PAT Yr5 (Lac)", x=names, y=[r["pat_yr5_lac"] for r in results.values()], marker_color="#00AA44"),
    ])
    fig1.update_layout(title="Investment vs Revenue vs Profit", barmode="group", template="plotly_white", height=400)
    st.plotly_chart(fig1, width="stretch")

with col_c2:
    fig2 = go.Figure(data=[
        go.Bar(name="ROI (%)", x=names, y=[r["roi_pct"] for r in results.values()], marker_color="#003366"),
        go.Bar(name="IRR (%)", x=names, y=[r["irr_pct"] for r in results.values()], marker_color="#006699"),
    ])
    fig2.update_layout(title="ROI vs IRR Comparison", barmode="group", template="plotly_white", height=400)
    st.plotly_chart(fig2, width="stretch")

# Per MT comparison
fig3 = go.Figure()
fig3.add_trace(go.Bar(name="Revenue/MT", x=names, y=[r["revenue_per_mt"] for r in results.values()], marker_color="#003366"))
fig3.add_trace(go.Bar(name="Cost/MT", x=names, y=[r["var_cost_per_mt"] for r in results.values()], marker_color="#CC3333"))
fig3.add_trace(go.Bar(name="Profit/MT", x=names, y=[r["profit_per_mt"] for r in results.values()], marker_color="#00AA44"))
fig3.update_layout(title="Per MT Economics (Rs)", barmode="group", template="plotly_white", height=400)
st.plotly_chart(fig3, width="stretch")

st.markdown("---")

# ── Detailed P&L for each ─────────────────────────────────────────────
st.subheader("Detailed 7-Year P&L for Each Process")
for pid in [1, 2, 3]:
    r = results[pid]
    with st.expander(f"Process {pid}: {r['process_name']}"):
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Investment", f"Rs {r['capex_lac']:.0f} Lac")
        m2.metric("ROI", f"{r['roi_pct']:.1f}%")
        m3.metric("IRR", f"{r['irr_pct']:.1f}%")
        m4.metric("Break-Even", f"{r['break_even_months']} months")
        st.dataframe(pd.DataFrame(r["timeline"]), width="stretch", hide_index=True)

st.markdown("---")
st.markdown("""
### Which Process to Choose?
| Situation | Best Process | Why |
|-----------|-------------|-----|
| New investor with land + funds | **Process 1 (Full Chain)** | Highest margin, full value capture |
| Existing bitumen company | **Process 2 (Blending)** | Lowest capex, fastest setup (30-60 days) |
| Existing pyrolysis operator | **Process 3 (Raw Output)** → then upgrade to Process 1 | Sell oil+char now, add blending later |
| Farmer cooperative | **Process 1 (Full Chain)** | Own biomass = near-zero raw material cost |
""")

st.caption("All calculations based on verified market data (March 2026). Source: MASTER_DATA_CORRECTED.py, Verified Supplier Research.")


# ── AI Skill: Process Comparison ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: Process Comparison"):
            if st.button("Generate", type="primary", key="ai_57Three"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: Process Comparison. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass

# ROI Disclaimer
st.info("📊 ROI shown here is **gross operating ROI** (EBITDA/Investment) before full loan repayment. "
        "For bank-standard net ROI after debt service, see Financial Model page.")


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_57Thr", type="primary")
with _ex2:
    if st.button("Print", key="exp_prt_57Thr"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)


# ── Next Steps Navigation ──
try:
    from engines.page_navigation import add_next_steps
    add_next_steps(st, "21")
except Exception:
    pass
