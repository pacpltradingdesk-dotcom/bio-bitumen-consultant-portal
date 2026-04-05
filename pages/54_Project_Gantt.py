"""
Project Gantt — Per-Customer Milestone Tracking with Interactive Timeline
==========================================================================
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
from database import (init_db, get_all_customers, get_milestones,
                       init_milestones_for_customer, update_milestone)
from state_manager import init_state, get_config
from config import COMPANY

st.set_page_config(page_title="Project Gantt", page_icon="📅", layout="wide")
init_db()
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass


st.title("Project Timeline & Gantt Chart")
st.markdown("**Per-Customer Milestone Tracking | 10-Phase Implementation | Status Updates**")
st.markdown("---")

customers = get_all_customers()

if not customers:
    st.warning("No customers in CRM yet.")
    st.markdown("**Demo Preview — Sample 20 TPD Project Timeline:**")
    import plotly.express as px
    import datetime, pandas as pd
    demo_data = [{"Task":"Pre-Feasibility","Start":"2026-06-01","Finish":"2026-07-01"},{"Task":"Company Setup","Start":"2026-07-01","Finish":"2026-08-01"},{"Task":"Land & Approvals","Start":"2026-07-15","Finish":"2026-10-15"},{"Task":"Bank Loan","Start":"2026-08-01","Finish":"2026-11-01"},{"Task":"Construction","Start":"2026-11-01","Finish":"2027-04-01"},{"Task":"Commissioning","Start":"2027-04-01","Finish":"2027-07-01"}]
    demo_df = pd.DataFrame(demo_data)
    demo_df["Start"]=pd.to_datetime(demo_df["Start"])
    demo_df["Finish"]=pd.to_datetime(demo_df["Finish"])
    fig=px.timeline(demo_df,x_start="Start",x_end="Finish",y="Task",title="Sample Project Timeline (Demo)")
    fig.update_yaxes(autorange="reversed")
    fig.update_layout(template="plotly_white",height=350)
    st.plotly_chart(fig,width="stretch")
    st.caption("Add a customer to create a real project timeline")
    st.page_link("pages/11_👥_Customers.py", label="Go to Customer Manager", icon="👥")
    st.stop()

# ══════════════════════════════════════════════════════════════════════
# SELECT CUSTOMER
# ══════════════════════════════════════════════════════════════════════
cust_names = {c["id"]: f"{c['name']} ({c.get('company', '')})" for c in customers}
selected_id = st.selectbox("Select Customer", list(cust_names.keys()),
                            format_func=lambda x: cust_names[x], key="gantt_cust")

selected_customer = next(c for c in customers if c["id"] == selected_id)
st.markdown(f"**Customer:** {selected_customer['name']} | **State:** {selected_customer.get('state', 'N/A')} | "
            f"**Capacity:** {selected_customer.get('interested_capacity', 'N/A')} | "
            f"**Status:** {selected_customer.get('status', 'New')}")

# ── Initialize milestones if none exist ──────────────────────────────
milestones = get_milestones(selected_id)

if not milestones:
    st.info("No milestones for this customer yet.")
    col_init1, col_init2 = st.columns(2)
    with col_init1:
        start_date = st.date_input("Project Start Date", datetime(2026, 5, 1), key="gantt_start")
    with col_init2:
        if st.button("Create Default Milestones (10-Phase Plan)", type="primary", key="init_ms"):
            init_milestones_for_customer(selected_id, start_date.strftime("%Y-%m-%d"))
            st.success("10 milestones created! Refresh to see Gantt chart.")
            st.rerun()
    st.stop()

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# PROGRESS METRICS
# ══════════════════════════════════════════════════════════════════════
total = len(milestones)
completed = sum(1 for m in milestones if m["status"] == "Completed")
in_progress = sum(1 for m in milestones if m["status"] == "In Progress")
delayed = sum(1 for m in milestones if m["status"] == "Delayed")
not_started = sum(1 for m in milestones if m["status"] == "Not Started")
progress_pct = (completed / total * 100) if total > 0 else 0

pm1, pm2, pm3, pm4, pm5 = st.columns(5)
pm1.metric("Total Milestones", total)
pm2.metric("Completed", completed)
pm3.metric("In Progress", in_progress)
pm4.metric("Delayed", delayed)
pm5.metric("Progress", f"{progress_pct:.0f}%")

st.progress(progress_pct / 100)
st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# GANTT CHART
# ══════════════════════════════════════════════════════════════════════
st.subheader("Project Gantt Chart")

gantt_data = []
for m in milestones:
    start = m.get("actual_start") or m.get("planned_start", "2026-05-01")
    end = m.get("actual_end") or m.get("planned_end", "2026-06-01")
    gantt_data.append({
        "Task": m["milestone_name"],
        "Start": start,
        "Finish": end,
        "Phase": m.get("phase", ""),
        "Status": m["status"],
    })

gantt_df = pd.DataFrame(gantt_data)
gantt_df["Start"] = pd.to_datetime(gantt_df["Start"])
gantt_df["Finish"] = pd.to_datetime(gantt_df["Finish"])

color_map = {
    "Not Started": "#cccccc",
    "In Progress": "#0088cc",
    "Completed": "#00AA44",
    "Delayed": "#CC3333",
    "On Hold": "#FF8800",
}

fig_gantt = px.timeline(gantt_df, x_start="Start", x_end="Finish", y="Task",
                          color="Status", color_discrete_map=color_map,
                          title=f"Project Timeline — {selected_customer['name']}")
fig_gantt.update_yaxes(autorange="reversed")
fig_gantt.update_layout(template="plotly_white", height=500)

# Add today line
today = datetime.now().strftime("%Y-%m-%d")
fig_gantt.add_vline(x=today, line_dash="dash", line_color="red",
                     annotation_text="Today", annotation_position="top left")

st.plotly_chart(fig_gantt, width="stretch")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# UPDATE MILESTONES
# ══════════════════════════════════════════════════════════════════════
st.subheader("Update Milestone Status")

status_options = ["Not Started", "In Progress", "Completed", "Delayed", "On Hold"]

for m in milestones:
    with st.expander(f"{'✅' if m['status']=='Completed' else '🔵' if m['status']=='In Progress' else '🔴' if m['status']=='Delayed' else '⚪'} {m['milestone_name']} — {m['status']}"):
        uc1, uc2, uc3 = st.columns(3)
        with uc1:
            new_status = st.selectbox("Status", status_options,
                                       index=status_options.index(m["status"]) if m["status"] in status_options else 0,
                                       key=f"ms_status_{m['id']}")
        with uc2:
            actual_start = st.date_input("Actual Start",
                                          value=datetime.strptime(m["actual_start"], "%Y-%m-%d") if m.get("actual_start") else None,
                                          key=f"ms_astart_{m['id']}")
        with uc3:
            actual_end = st.date_input("Actual End",
                                        value=datetime.strptime(m["actual_end"], "%Y-%m-%d") if m.get("actual_end") else None,
                                        key=f"ms_aend_{m['id']}")

        notes = st.text_input("Notes", value=m.get("notes", ""), key=f"ms_notes_{m['id']}")

        if st.button("Update", key=f"ms_update_{m['id']}"):
            update_data = {"status": new_status, "notes": notes}
            if actual_start:
                update_data["actual_start"] = actual_start.strftime("%Y-%m-%d")
            if actual_end:
                update_data["actual_end"] = actual_end.strftime("%Y-%m-%d")
            update_milestone(m["id"], update_data)
            st.success(f"Updated: {m['milestone_name']}")
            st.rerun()

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# PHASE SUMMARY
# ══════════════════════════════════════════════════════════════════════
st.subheader("Phase Summary")

phase_data = []
for m in milestones:
    phase_data.append({
        "Phase": m.get("phase", ""),
        "Milestone": m["milestone_name"],
        "Planned Start": m.get("planned_start", ""),
        "Planned End": m.get("planned_end", ""),
        "Status": m["status"],
    })

st.dataframe(pd.DataFrame(phase_data), width="stretch", hide_index=True)

# Health score
health = 100
for m in milestones:
    if m["status"] == "Delayed":
        health -= 15
    elif m["status"] == "On Hold":
        health -= 5
health = max(health, 0)

health_color = "#00AA44" if health >= 80 else ("#FF8800" if health >= 50 else "#CC3333")
health_label = "On Track" if health >= 80 else ("At Risk" if health >= 50 else "Critical")
st.markdown(f"""
<div style="background: {health_color}22; border-left: 4px solid {health_color}; padding: 15px; border-radius: 8px;">
    <h3 style="color: {health_color}; margin: 0;">Project Health: {health}% — {health_label}</h3>
</div>
""", unsafe_allow_html=True)

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# HOLIDAYS IN PROJECT PERIOD (from Free API)
# ══════════════════════════════════════════════════════════════════════
st.subheader("Public Holidays During Project")
try:
    from engines.free_apis import get_india_holidays
    holidays = get_india_holidays(2026)
    if holidays and milestones:
        project_start = milestones[0].get("planned_start", "2026-01-01")
        project_end = milestones[-1].get("planned_end", "2027-06-30")
        relevant = [h for h in holidays if project_start <= h["date"] <= project_end]
        if relevant:
            st.info(f"**{len(relevant)} holidays** fall during your project period ({project_start} to {project_end})")
            for h in relevant:
                st.markdown(f"- **{h['date']}** — {h['name']} ({h.get('name_en', '')})")
        else:
            st.success("No major holidays during project period")
except Exception:
    pass

st.markdown("---")
st.caption(f"{COMPANY['name']} | Project Management Module | Holiday data: Nager.Date (free)")


# ── AI Skill: Milestone Brief ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: Milestone Brief"):
            if st.button("Generate", type="primary", key="ai_64Proje"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: Milestone Brief. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_64Pro", type="primary")
with _ex2:
    if st.button("Print", key="exp_prt_64Pro"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)
