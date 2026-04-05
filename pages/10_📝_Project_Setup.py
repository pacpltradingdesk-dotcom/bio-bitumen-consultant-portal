"""
Project Setup — Enter Client & Project Info ONCE → Auto-Fills ALL Documents
=============================================================================
Fill this form → all DPR, Bank Proposal, Investor Pitch, Financial Model,
Quotation, PDF reports auto-populate with this information.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import datetime
from state_manager import get_config, update_field, update_fields, init_state
from config import COMPANY, STATES, CAPACITY_KEYS, CAPACITY_LABELS
from database import init_db, get_all_customers

st.set_page_config(page_title="Project Setup", page_icon="📝", layout="wide")
init_db()
init_state()
cfg = get_config()

try:
    from utils.page_helpers import fix_metric_truncation
    fix_metric_truncation()
except Exception:
    pass


st.sidebar.markdown("---")
if st.sidebar.button("Print This Page", key="print_page"):
    import streamlit.components.v1 as _stc; _stc.html('<script>window.print();</script>', height=0)

st.title("Project Setup — Client & Site Information")
st.markdown("**Fill this ONCE → Auto-populates ALL documents, calculations, and reports**")
st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# LOAD FROM EXISTING CUSTOMER (Optional)
# ══════════════════════════════════════════════════════════════════════
st.subheader("Quick Load from CRM")
customers = get_all_customers()

if customers:
    cust_map = {0: "-- Start Fresh --"}
    cust_map.update({c["id"]: f"{c['name']} ({c.get('company', '')})" for c in customers})
    load_cust = st.selectbox("Load from existing customer", list(cust_map.keys()),
                              format_func=lambda x: cust_map[x], key="load_from_crm")

    if load_cust and load_cust != 0:
        cust = next(c for c in customers if c["id"] == load_cust)
        if st.button("Load Customer Data", key="load_btn"):
            updates = {
                "client_name": cust.get("name", ""),
                "client_company": cust.get("company", ""),
                "client_email": cust.get("email", ""),
                "client_phone": cust.get("phone", ""),
                "state": cust.get("state", cfg["state"]),
                "location": cust.get("city", cfg["location"]),
            }
            update_fields(updates)
            st.success(f"Loaded: {cust['name']} ({cust.get('company', '')})")
            st.rerun()

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# CLIENT INFORMATION
# ══════════════════════════════════════════════════════════════════════
st.subheader("1. Client Information")

cl1, cl2 = st.columns(2)
with cl1:
    client_name = st.text_input("Client Name *", value=cfg.get("client_name", ""),
                                 placeholder="e.g., Raj Industries Pvt Ltd", key="ps_name")
    client_company = st.text_input("Company / Legal Entity", value=cfg.get("client_company", ""),
                                    placeholder="e.g., Raj Bio-Bitumen Manufacturing LLP", key="ps_company")
    client_email = st.text_input("Email", value=cfg.get("client_email", ""),
                                  placeholder="info@rajindustries.com", key="ps_email")
with cl2:
    client_phone = st.text_input("Phone", value=cfg.get("client_phone", ""),
                                  placeholder="+91 98765 43210", key="ps_phone")
    client_gst = st.text_input("GST Number", value=cfg.get("client_gst", ""),
                                placeholder="24XXXXX1234X1Z5", key="ps_gst")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# PROJECT INFORMATION
# ══════════════════════════════════════════════════════════════════════
st.subheader("2. Project Information")

pj1, pj2 = st.columns(2)
with pj1:
    project_name = st.text_input("Project Name *", value=cfg.get("project_name", ""),
                                  placeholder="e.g., Bio-Modified Bitumen Plant — Vadodara", key="ps_proj")
    project_id = st.text_input("Project ID / Reference", value=cfg.get("project_id", ""),
                                placeholder="e.g., PPS/2026/BIO-001", key="ps_projid")
with pj2:
    project_start = st.date_input("Proposed Start Date",
                                    value=datetime.date(2026, 6, 1), key="ps_start")
    project_completion = st.date_input("Target Completion",
                                        value=datetime.date(2027, 12, 1), key="ps_end")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# SITE LOCATION
# ══════════════════════════════════════════════════════════════════════
st.subheader("3. Site Location")

sl1, sl2 = st.columns(2)
with sl1:
    site_address = st.text_area("Site Address *", value=cfg.get("site_address", ""),
                                 placeholder="Plot No. 45, GIDC Industrial Estate,\nMakarpura Road, Vadodara — 390010",
                                 height=100, key="ps_addr")
    site_pincode = st.text_input("Pincode", value=cfg.get("site_pincode", ""),
                                  placeholder="390010", key="ps_pin")

    # Auto-fill from pincode
    if site_pincode and len(site_pincode) == 6 and site_pincode.isdigit():
        try:
            from engines.free_apis import lookup_pincode
            pin_data = lookup_pincode(site_pincode)
            if "error" not in pin_data:
                st.caption(f"Auto-detected: {pin_data.get('district', '')}, {pin_data.get('state', '')}")
        except Exception:
            pass

with sl2:
    state_idx = STATES.index(cfg.get("state", "Gujarat")) if cfg.get("state") in STATES else 0
    site_state = st.selectbox("State *", STATES, index=state_idx, key="ps_state")
    site_city = st.text_input("City / Town", value=cfg.get("location", ""), key="ps_city")
    site_district = st.text_input("District", value=cfg.get("site_district", ""), key="ps_dist")
    site_area = st.number_input("Site Area (Acres)", 0.0, 100.0,
                                 float(cfg.get("site_area_acres", 0)), 0.5, key="ps_area")
    site_ownership = st.selectbox("Land Ownership",
                                   ["Own", "Leased", "GIDC/MIDC Allotment", "To be acquired"], key="ps_own")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# PLANT CONFIGURATION
# ══════════════════════════════════════════════════════════════════════
st.subheader("4. Plant Configuration")

pc1, pc2, pc3 = st.columns(3)
with pc1:
    capacity = st.selectbox("Plant Capacity (TPD)",
                              [5, 10, 15, 20, 30, 40, 50],
                              index=[5,10,15,20,30,40,50].index(int(cfg["capacity_tpd"])) if int(cfg["capacity_tpd"]) in [5,10,15,20,30,40,50] else 3,
                              key="ps_cap")
with pc2:
    process_model = st.selectbox("Process Model",
                                  ["Full Chain (Biomass → Bitumen)",
                                   "Blending Only (Buy Bio-Oil)",
                                   "Raw Output (Oil + Char)"], key="ps_process")
with pc3:
    st.metric("Investment (auto)", f"Rs {cfg['investment_cr']:.2f} Cr")
    st.metric("ROI (auto)", f"{cfg['roi_pct']:.1f}%")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# TECHNICAL DETAILS
# ══════════════════════════════════════════════════════════════════════
st.subheader("5. Technical Details")

tc1, tc2, tc3 = st.columns(3)
with tc1:
    biomass_source = st.text_input("Primary Biomass Source",
                                    value=cfg.get("biomass_source", ""),
                                    placeholder="e.g., Rice straw from local FPOs", key="ps_biomass")
    site_contact = st.text_input("Site Contact Person",
                                  value=cfg.get("site_contact_person", ""),
                                  placeholder="Mr. Rajesh Patel (Site Engineer)", key="ps_contact")
with tc2:
    power_source = st.text_input("Power Source",
                                  value=cfg.get("power_source", ""),
                                  placeholder="e.g., GUVNL HT Connection + 100 kVA DG", key="ps_power")
    site_contact_phone = st.text_input("Site Contact Phone",
                                        value=cfg.get("site_contact_phone", ""),
                                        placeholder="+91 99999 99999", key="ps_cphone")
with tc3:
    water_source = st.text_input("Water Source",
                                  value=cfg.get("water_source", ""),
                                  placeholder="e.g., Bore well + Municipal tanker backup", key="ps_water")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# SAVE ALL
# ══════════════════════════════════════════════════════════════════════
if st.button("SAVE PROJECT SETUP", type="primary", key="save_project"):
    updates = {
        "client_name": client_name,
        "client_company": client_company,
        "client_email": client_email,
        "client_phone": client_phone,
        "client_gst": client_gst,
        "project_name": project_name,
        "project_id": project_id,
        "site_address": site_address,
        "site_pincode": site_pincode,
        "site_district": site_district,
        "site_area_acres": site_area,
        "site_ownership": site_ownership,
        "state": site_state,
        "location": site_city,
        "project_start_date": project_start.strftime("%Y-%m-%d"),
        "project_completion_target": project_completion.strftime("%Y-%m-%d"),
        "site_contact_person": site_contact,
        "site_contact_phone": site_contact_phone,
        "biomass_source": biomass_source,
        "power_source": power_source,
        "water_source": water_source,
    }
    if capacity != cfg["capacity_tpd"]:
        updates["capacity_tpd"] = float(capacity)

    update_fields(updates)
    st.success("Project Setup SAVED! All documents will now use this client & project information.")
    st.balloons()

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════
# PROJECT SUMMARY CARD
# ══════════════════════════════════════════════════════════════════════
st.subheader("Project Summary Card")

has_client = bool(cfg.get("client_name"))
has_project = bool(cfg.get("project_name"))
has_site = bool(cfg.get("site_address"))

if has_client or has_project:
    card_color = "#003366" if (has_client and has_project and has_site) else "#FF8800"
    completeness = sum([has_client, has_project, has_site, bool(cfg.get("site_pincode")),
                         bool(cfg.get("client_phone")), bool(cfg.get("biomass_source"))]) * 100 // 6

    st.markdown(f"""
    <div style="background: {card_color}10; border: 2px solid {card_color}; border-radius: 15px;
                padding: 25px; margin: 10px 0;">
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <h3 style="color: {card_color}; margin: 0;">{cfg.get('project_name', 'Untitled Project')}</h3>
            <span style="background: {card_color}; color: white; padding: 5px 15px; border-radius: 20px;">
                {completeness}% Complete
            </span>
        </div>
        <hr style="border-color: {card_color}33;">
        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 10px;">
            <div>
                <strong>Client:</strong> {cfg.get('client_name', 'Not set')}<br>
                <strong>Company:</strong> {cfg.get('client_company', 'Not set')}<br>
                <strong>Phone:</strong> {cfg.get('client_phone', 'Not set')}<br>
                <strong>GST:</strong> {cfg.get('client_gst', 'Not set')}
            </div>
            <div>
                <strong>Site:</strong> {cfg.get('site_address', 'Not set').replace(chr(10), ', ')}<br>
                <strong>Location:</strong> {cfg.get('location', '')}, {cfg.get('state', '')}<br>
                <strong>Area:</strong> {cfg.get('site_area_acres', 0)} Acres ({cfg.get('site_ownership', 'TBD')})<br>
                <strong>Pincode:</strong> {cfg.get('site_pincode', 'Not set')}
            </div>
        </div>
        <hr style="border-color: {card_color}33;">
        <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 10px;">
            <div><strong>Capacity:</strong> {cfg['capacity_tpd']:.0f} TPD</div>
            <div><strong>Investment:</strong> Rs {cfg['investment_cr']:.2f} Cr</div>
            <div><strong>ROI:</strong> {cfg['roi_pct']:.1f}%</div>
        </div>
    </div>
    """, unsafe_allow_html=True)
else:
    st.info("Fill the form above and click SAVE to see your Project Summary Card here.")

# Quick Links
st.markdown("---")
st.subheader("Next Steps")
n1, n2, n3, n4 = st.columns(4)
n1.page_link("pages/13_📁_Document_Hub.py", label="Generate Documents", icon="📁")
n2.page_link("pages/09_💰_Financial.py", label="Financial Model", icon="💰")
n3.page_link("pages/44_DPR_Generator.py", label="DPR Generator", icon="📄")
n4.page_link("pages/60_ROI_Quick_Calc.py", label="ROI Calculator", icon="🎯")

st.markdown("---")
st.caption(f"{COMPANY['name']} | Project Setup — All info auto-fills into documents")


# ── AI Skill: Smart Suggestions ──────────────────────────────────────
st.markdown("---")
try:
    from engines.ai_engine import is_ai_available, ask_ai
    if is_ai_available():
        with st.expander("🤖 AI: Smart Suggestions"):
            if st.button("Generate", type="primary", key="ai_03📝Pro"):
                with st.spinner("AI working..."):
                    _p = f"As a senior bio-bitumen consultant, generate: Smart Suggestions. "
                    _p += f"Plant: {cfg.get('capacity_tpd',20):.0f} TPD, Investment: Rs {cfg.get('investment_cr',8):.2f} Cr, "
                    _p += f"Location: {cfg.get('location','')}, {cfg.get('state','')}. "
                    _p += "Be specific with numbers. Professional format."
                    _r, _pv = ask_ai(_p, "Senior industrial consultant.", 1000)
                if _r:
                    st.markdown(_r)
except Exception:
    pass


# ── Export ────────────────────────────────────────────────────────
st.markdown("---")
_ex1, _ex2 = st.columns(2)
with _ex1:
    try:
        import io as _io
        from openpyxl import Workbook as _Wb
        _wb = _Wb()
        _ws = _wb.active
        _ws.title = "Export"
        _ws.cell(row=1, column=1, value="Bio Bitumen Export")
        _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
        _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
        _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _xl_data = _buf.getvalue()
    except Exception:
        _xl_data = None
    if _xl_data:
        st.download_button("Download Excel", _xl_data, "export.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xl_03_📝_P", type="primary")
with _ex2:
    if st.button("Print", key="exp_prt_03📝P"):
        import streamlit.components.v1 as _stc
        _stc.html("<script>window.print();</script>", height=0)
