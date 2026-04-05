"""
Fix ALL broken download buttons across all pages.
The anti-pattern: st.button("Download") wrapping st.download_button() —
the nested download_button disappears on Streamlit rerun.

Fix: Replace with direct st.download_button (no wrapper).
"""
import re
import glob
import os

# The broken pattern across ~40 pages looks like:
# if st.button("Download Excel", type="primary", key="exp_xl_XXX"):
#     try:
#         import io
#         from openpyxl import Workbook
#         _wb = Workbook()
#         _ws = _wb.active
#         _ws.title = "Export"
#         _ws.cell(row=1, column=1, value="Bio Bitumen Export")
#         _ws.cell(row=2, column=1, value=f"Capacity: {cfg.get('capacity_tpd',20):.0f} TPD")
#         _ws.cell(row=3, column=1, value=f"Investment: Rs {cfg.get('investment_cr',8):.2f} Cr")
#         _ws.cell(row=4, column=1, value=f"ROI: {cfg.get('roi_pct',0):.1f}%")
#         _buf = io.BytesIO()
#         _wb.save(_buf)
#         _buf.seek(0)
#         st.download_button("Download", _buf.getvalue(), "export.xlsx",
#             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_xl_XXX")
#     except Exception as _e:
#         st.error(f"Export failed: {_e}")

# We need a regex that matches the entire block and replaces it with a working version.

# The generic boilerplate export block
BROKEN_PATTERN = re.compile(
    r'if st\.button\("Download Excel", type="primary", key="exp_xl_(\w+)"\):\s*\n'
    r'\s+try:\s*\n'
    r'\s+import io\s*\n'
    r'\s+from openpyxl import Workbook\s*\n'
    r'\s+_wb = Workbook\(\)\s*\n'
    r'\s+_ws = _wb\.active\s*\n'
    r'\s+_ws\.title = "Export"\s*\n'
    r'\s+_ws\.cell\(row=1, column=1, value="Bio Bitumen Export"\)\s*\n'
    r'\s+_ws\.cell\(row=2, column=1, value=f"Capacity: \{cfg\.get\(\'capacity_tpd\',20\):.0f\} TPD"\)\s*\n'
    r'\s+_ws\.cell\(row=3, column=1, value=f"Investment: Rs \{cfg\.get\(\'investment_cr\',8\):.2f\} Cr"\)\s*\n'
    r'\s+_ws\.cell\(row=4, column=1, value=f"ROI: \{cfg\.get\(\'roi_pct\',0\):.1f\}%"\)\s*\n'
    r'\s+_buf = io\.BytesIO\(\)\s*\n'
    r'\s+_wb\.save\(_buf\)\s*\n'
    r'\s+_buf\.seek\(0\)\s*\n'
    r'\s+st\.download_button\("Download", _buf\.getvalue\(\), "export\.xlsx",\s*\n'
    r'\s+"application/vnd\.openxmlformats-officedocument\.spreadsheetml\.sheet", key="dl_xl_(\w+)"\)\s*\n'
    r'\s+except Exception as _e:\s*\n'
    r'\s+st\.error\(f"Export failed: \{_e\}"\)',
    re.MULTILINE
)

def make_replacement(match):
    key1 = match.group(1)
    key2 = match.group(2)
    # Use 4 spaces (standard indent for most pages, inside a with block)
    return f'''try:
    import io as _io
    from openpyxl import Workbook as _Wb
    _wb = _Wb()
    _ws = _wb.active
    _ws.title = "Export"
    _ws.cell(row=1, column=1, value="Bio Bitumen Export")
    _ws.cell(row=2, column=1, value=f"Capacity: {{cfg.get('capacity_tpd',20):.0f}} TPD")
    _ws.cell(row=3, column=1, value=f"Investment: Rs {{cfg.get('investment_cr',8):.2f}} Cr")
    _ws.cell(row=4, column=1, value=f"ROI: {{cfg.get('roi_pct',0):.1f}}%")
    _buf = _io.BytesIO()
    _wb.save(_buf)
    _xl_data = _buf.getvalue()
except Exception:
    _xl_data = None
if _xl_data:
    st.download_button("Download Excel", _xl_data, "export.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_xl_{key2}", type="primary")'''


def fix_file(filepath):
    """Fix broken download pattern in a single file."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    original = content

    # Fix the standard boilerplate pattern
    content = BROKEN_PATTERN.sub(make_replacement, content)

    if content != original:
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(content)
        return True
    return False


def main():
    pages_dir = os.path.join(os.path.dirname(__file__), 'pages')
    files = glob.glob(os.path.join(pages_dir, '*.py'))

    fixed = 0
    skipped = 0

    for filepath in sorted(files):
        fname = os.path.basename(filepath)
        if fix_file(filepath):
            print(f'  FIXED: {fname}')
            fixed += 1
        else:
            skipped += 1

    print(f'\n  {fixed} files fixed, {skipped} unchanged')


if __name__ == '__main__':
    main()
